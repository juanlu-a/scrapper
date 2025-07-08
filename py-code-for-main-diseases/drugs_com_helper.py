import pandas as pd
import requests
from bs4 import BeautifulSoup
import time
import random
from urllib.parse import urljoin
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import sys

class DrugsComHelper:
    def __init__(self):
        self.base_url = "https://www.drugs.com"
        
    def generate_urls_for_medications(self, medications):
        """Generate URLs for medications for manual checking"""
        urls = []
        for med in medications:
            clean_name = med.strip().lower().replace(' ', '-')
            # Remove common suffixes/prefixes
            clean_name = re.sub(r'\s+(tablets?|capsules?|mg|mL|oral|suspension|solution)', '', clean_name, flags=re.IGNORECASE)
            clean_name = clean_name.replace(' ', '-')
            
            # Generate possible URLs
            possible_urls = [
                f"{self.base_url}/{clean_name}.html",
                f"{self.base_url}/mtm/{clean_name}.html",
                f"{self.base_url}/otc/{clean_name}.html",
                f"{self.base_url}/pro/{clean_name}.html"
            ]
            
            urls.append({
                'medication': med,
                'primary_url': possible_urls[0],
                'side_effects_url': f"{self.base_url}/{clean_name}-side-effects.html",
                'search_url': f"{self.base_url}/search.php?searchterm={med.replace(' ', '+')}"
            })
        
        return urls
    
    def create_manual_guide(self, medications):
        """Create a manual guide for collecting side effects"""
        guide = []
        guide.append("DRUGS.COM SIDE EFFECTS MANUAL COLLECTION GUIDE")
        guide.append("=" * 60)
        guide.append("")
        guide.append("Instructions:")
        guide.append("1. Visit the URLs provided for each medication")
        guide.append("2. Look for the 'Side Effects' section or tab")
        guide.append("3. Copy all the side effects information")
        guide.append("4. Paste it into the Excel file in the 'Full Information' column")
        guide.append("")
        guide.append("URL Format:")
        guide.append("- Primary: https://www.drugs.com/{medication}.html")
        guide.append("- Side Effects: https://www.drugs.com/{medication}-side-effects.html")
        guide.append("- Search: https://www.drugs.com/search.php?searchterm={medication}")
        guide.append("")
        guide.append("MEDICATIONS TO PROCESS:")
        guide.append("=" * 40)
        
        for i, med in enumerate(medications[:20], 1):  # Show first 20 as example
            clean_name = med.strip().lower().replace(' ', '-')
            guide.append(f"\n{i}. {med}")
            guide.append(f"   Primary: {self.base_url}/{clean_name}.html")
            guide.append(f"   Side Effects: {self.base_url}/{clean_name}-side-effects.html")
            guide.append(f"   Search: {self.base_url}/search.php?searchterm={med.replace(' ', '+')}")
        
        if len(medications) > 20:
            guide.append(f"\n... and {len(medications) - 20} more medications")
        
        return '\n'.join(guide)

def create_excel_template_with_urls():
    """Create an Excel template with URLs for manual data collection"""
    
    # Load the existing Excel file
    excel_path = '/Users/juanlu/Documents/Wye/scrapper/Analysis/main_diseases_analysis_final.xlsx'
    
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found at {excel_path}")
        return
    
    # Load the workbook
    wb = load_workbook(excel_path)
    
    # Get the unique medications sheet
    if "All Unique Medications" not in wb.sheetnames:
        print("Error: 'All Unique Medications' sheet not found")
        return
    
    medications_ws = wb["All Unique Medications"]
    
    # Read medications from the sheet
    medications = []
    for row in medications_ws.iter_rows(min_row=9, max_col=1, values_only=True):
        if row[0] and row[0].strip():
            medications.append(row[0].strip())
    
    print(f"Found {len(medications)} medications to process")
    
    # Create helper instance
    helper = DrugsComHelper()
    
    # Add new column headers
    medications_ws['F8'] = 'DRUGS.COM URL'
    medications_ws['G8'] = 'SIDE EFFECTS URL'
    medications_ws['H8'] = 'FULL INFORMATION'
    
    # Style the headers
    for col in ['F', 'G', 'H']:
        cell = medications_ws[f'{col}8']
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Set column widths
    medications_ws.column_dimensions['F'].width = 40  # URLs
    medications_ws.column_dimensions['G'].width = 40  # Side effects URLs
    medications_ws.column_dimensions['H'].width = 60  # Full information
    
    # Generate URLs for each medication
    medication_urls = helper.generate_urls_for_medications(medications)
    
    # Add URLs to Excel
    for i, med_data in enumerate(medication_urls):
        row_num = 9 + i
        
        # Add URLs
        medications_ws[f'F{row_num}'] = med_data['primary_url']
        medications_ws[f'G{row_num}'] = med_data['side_effects_url']
        medications_ws[f'H{row_num}'] = f"Visit: {med_data['primary_url']} → Click 'Side Effects' tab → Copy all content here"
        
        # Add borders and formatting
        for col in ['F', 'G', 'H']:
            cell = medications_ws[f'{col}{row_num}']
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Alternate row colors
            if i % 2 == 0:
                cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    
    # Save the updated file
    wb.save(excel_path)
    print(f"✅ Updated Excel file with URLs: {excel_path}")
    
    # Create a manual guide
    guide_content = helper.create_manual_guide(medications)
    
    # Save the guide to a text file
    guide_path = '/Users/juanlu/Documents/Wye/scrapper/Analysis/drugs_com_manual_guide.txt'
    with open(guide_path, 'w', encoding='utf-8') as f:
        f.write(guide_content)
    
    print(f"✅ Created manual guide: {guide_path}")
    
    # Display first few medications as examples
    print("\n" + "="*60)
    print("SAMPLE MEDICATIONS TO PROCESS:")
    print("="*60)
    
    for i, med_data in enumerate(medication_urls[:5], 1):
        print(f"\n{i}. {med_data['medication']}")
        print(f"   Primary URL: {med_data['primary_url']}")
        print(f"   Side Effects: {med_data['side_effects_url']}")
        print(f"   Search: {med_data['search_url']}")
    
    print(f"\n... and {len(medications) - 5} more medications")
    print(f"\nAll URLs have been added to the Excel file.")
    print(f"Manual guide saved to: {guide_path}")

def try_automated_scraping():
    """Try automated scraping with better error handling"""
    
    excel_path = '/Users/juanlu/Documents/Wye/scrapper/Analysis/main_diseases_analysis_final.xlsx'
    wb = load_workbook(excel_path)
    medications_ws = wb["All Unique Medications"]
    
    # Read first 5 medications for testing
    medications = []
    for row in medications_ws.iter_rows(min_row=9, max_row=13, max_col=1, values_only=True):
        if row[0] and row[0].strip():
            medications.append(row[0].strip())
    
    print("Attempting automated scraping for first 5 medications...")
    
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
    })
    
    success_count = 0
    
    for med in medications:
        try:
            clean_name = med.strip().lower().replace(' ', '-')
            url = f"https://www.drugs.com/{clean_name}.html"
            
            print(f"Trying {med}: {url}")
            
            time.sleep(random.uniform(2, 4))  # Random delay
            
            response = session.get(url, timeout=10)
            
            if response.status_code == 200:
                print(f"  ✓ Success! Found page for {med}")
                success_count += 1
            else:
                print(f"  ✗ Failed: HTTP {response.status_code}")
                
        except Exception as e:
            print(f"  ✗ Error: {str(e)}")
    
    print(f"\nAutomated scraping results: {success_count}/{len(medications)} successful")
    
    if success_count == 0:
        print("❌ Automated scraping not working - website likely blocks requests")
        print("➡️  Using manual collection approach instead")
        return False
    else:
        print("✅ Automated scraping partially working")
        return True

if __name__ == "__main__":
    print("Drugs.com Side Effects Data Collection Tool")
    print("=" * 60)
    
    # First, try automated scraping
    print("Testing automated scraping...")
    automated_works = try_automated_scraping()
    
    print("\n" + "="*60)
    
    if not automated_works:
        print("Setting up manual collection approach...")
        create_excel_template_with_urls()
        
        print("\n" + "="*60)
        print("NEXT STEPS:")
        print("1. Open the updated Excel file")
        print("2. Use the provided URLs to manually collect side effects")
        print("3. Copy the side effects information into the 'Full Information' column")
        print("4. A manual guide has been created to help you")
    else:
        print("Automated scraping is working - would you like to proceed?")
        print("(This will take a while and may still get blocked)")
