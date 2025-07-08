from selenium import webdriver
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from webdriver_manager.firefox import GeckoDriverManager
import time
import random
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

class DrugsScraper:
    def __init__(self, headless=True):
        self.driver = self.setup_driver(headless)
        self.wait = WebDriverWait(self.driver, 15)
        
    def setup_driver(self, headless=True):
        """Set up Firefox driver with webdriver-manager"""
        firefox_options = FirefoxOptions()
        if headless:
            firefox_options.add_argument("--headless")
        
        # Add user agent
        firefox_options.set_preference("general.useragent.override", 
                                     "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
        
        # Use webdriver-manager to automatically handle GeckoDriver
        service = FirefoxService(GeckoDriverManager().install())
        driver = webdriver.Firefox(service=service, options=firefox_options)
        
        return driver
    
    def scrape_side_effects(self, medication, max_retries=2):
        """Scrape side effects for a specific medication with enhanced error handling"""
        print(f"\n🔍 Processing: {medication}")
        
        for attempt in range(max_retries):
            try:
                if attempt > 0:
                    print(f"  🔄 Retry attempt {attempt + 1}/{max_retries}")
                
                # Step 1: Navigate to drugs.com
                self.driver.get("https://www.drugs.com")
                time.sleep(4)
                
                # Step 2: Find search box and search
                search_selectors = [
                    "input[type='search']", 
                    "input[name='q']", 
                    "input[placeholder*='search']", 
                    "#livesearch",
                    "input[name='searchterm']"
                ]
                
                search_box = None
                for selector in search_selectors:
                    try:
                        search_box = self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, selector)))
                        if search_box:
                            break
                    except:
                        continue
                
                if not search_box:
                    print(f"  ❌ Could not find search box for {medication}")
                    continue
                
                search_box.clear()
                search_box.send_keys(medication)
                time.sleep(1)
                search_box.send_keys(Keys.RETURN)
                
                print(f"  ✅ Searched for: {medication}")
                time.sleep(5)
                
                # Step 3: Find main result
                main_result = self.find_main_result(medication)
                if not main_result:
                    print(f"  ❌ Could not find main result for {medication}")
                    continue
                
                # Step 4: Click on main result
                try:
                    self.driver.execute_script("arguments[0].scrollIntoView(true);", main_result)
                    time.sleep(2)
                    main_result.click()
                    print(f"  ✅ Clicked main result for {medication}")
                except Exception as e:
                    try:
                        self.driver.execute_script("arguments[0].click();", main_result)
                        print(f"  ✅ Clicked main result (JS) for {medication}")
                    except Exception as e2:
                        print(f"  ❌ Failed to click main result for {medication}: {str(e2)}")
                        continue
                
                time.sleep(5)
                
                # Step 5: Find and click side effects link
                side_effects_link = self.find_side_effects_link()
                if side_effects_link:
                    # Step 6: Click side effects link
                    try:
                        self.driver.execute_script("arguments[0].scrollIntoView(true);", side_effects_link)
                        time.sleep(2)
                        side_effects_link.click()
                        print(f"  ✅ Clicked side effects link for {medication}")
                        time.sleep(5)
                    except Exception as e:
                        try:
                            self.driver.execute_script("arguments[0].click();", side_effects_link)
                            print(f"  ✅ Clicked side effects link (JS) for {medication}")
                            time.sleep(5)
                        except Exception as e2:
                            print(f"  ❌ Failed to click side effects link for {medication}: {str(e2)}")
                            # Continue to extraction even if link click fails
                else:
                    print(f"  ℹ️  No side effects link found, checking current page for {medication}")
                
                # Step 7: Extract side effects content
                content = self.extract_side_effects_content(medication)
                
                if content and len(content) > 50:
                    print(f"  ✅ Successfully processed {medication} ({len(content)} characters)")
                    return content
                else:
                    print(f"  ❌ No substantial side effects content found for {medication}")
                    if attempt < max_retries - 1:
                        continue
                    else:
                        return f"No detailed side effects information found for {medication}"
                
            except Exception as e:
                error_msg = f"❌ Error processing {medication} (attempt {attempt + 1}): {str(e)}"
                print(error_msg)
                if attempt < max_retries - 1:
                    time.sleep(5)  # Wait before retry
                else:
                    return f"Error: {str(e)}"
        
        return f"Failed to process {medication} after {max_retries} attempts"
    
    def find_main_result(self, medication):
        """Find the main result for the medication with enhanced matching"""
        print(f"  🔍 Looking for main result for: {medication}")
        
        # Clean medication name for matching
        medication_clean = medication.lower().strip()
        medication_words = [word for word in medication_clean.split() if len(word) > 2]
        
        # Try different selectors to find results
        selectors = [
            "a[href*='.html']",
            ".search-results a",
            ".result a",
            ".drug-results a",
            "a[href*='/drug/']",
            "li a[href*='.html']",
            ".search-result a",
            ".results a",
            ".drug-result a"
        ]
        
        for selector in selectors:
            try:
                results = self.driver.find_elements(By.CSS_SELECTOR, selector)
                print(f"    Found {len(results)} results with selector: {selector}")
                
                for result in results:
                    try:
                        href = result.get_attribute('href')
                        if not href or '.html' not in href:
                            continue
                        
                        # Get text content
                        text = result.text.strip().lower()
                        
                        # Check if this looks like a main medication result
                        if not text:
                            continue
                        
                        # Look for star indicators or main drug page indicators
                        try:
                            parent = result.find_element(By.XPATH, "./..")
                            parent_text = parent.text.lower()
                            has_star = '★' in parent_text or 'star' in parent_text
                        except:
                            has_star = False
                        
                        # Check text matching
                        if len(medication_words) == 1:
                            # Single word medication
                            if medication_words[0] in text:
                                print(f"      ✅ Found single word match: {text[:50]}...")
                                return result
                        else:
                            # Multi-word medication - check for word matches
                            word_matches = sum(1 for word in medication_words if word in text)
                            if word_matches >= len(medication_words) * 0.6:  # Lowered threshold
                                print(f"      ✅ Found multi-word match: {text[:50]}...")
                                return result
                        
                        # If has star indicator, be more flexible
                        if has_star and any(word in text for word in medication_words):
                            print(f"      ✅ Found star result: {text[:50]}...")
                            return result
                    
                    except Exception as e:
                        continue
                
            except Exception as e:
                continue
        
        return None
    
    def find_side_effects_link(self):
        """Find the side effects navigation link with enhanced detection"""
        print(f"  🔍 Looking for side effects link...")
        
        time.sleep(3)
        
        # Try multiple approaches to find side effects link
        approaches = [
            # Approach 1: Direct text matching
            {
                'name': 'Direct text matching',
                'method': 'xpath',
                'selectors': [
                    "//a[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'side effects')]",
                    "//a[text()='Side effects']",
                    "//a[text()='side effects']",
                    "//a[text()='Side Effects']",
                    "//span[contains(text(), 'side effects')]/parent::a",
                    "//span[contains(text(), 'Side effects')]/parent::a"
                ]
            },
            # Approach 2: href containing side-effects
            {
                'name': 'Href containing side-effects',
                'method': 'css',
                'selectors': [
                    "a[href*='side-effects']",
                    "a[href*='#side-effects']",
                    "a[href*='sideeffects']",
                    "a[href*='adverse-reactions']"
                ]
            },
            # Approach 3: Navigation areas
            {
                'name': 'Navigation areas',
                'method': 'css',
                'selectors': [
                    "nav a",
                    ".nav a",
                    ".navigation a",
                    ".tabs a",
                    ".tab a",
                    ".menu a",
                    ".drug-nav a",
                    ".page-nav a",
                    "ul.nav a",
                    ".nav-tabs a"
                ]
            }
        ]
        
        for approach in approaches:
            print(f"    Trying approach: {approach['name']}")
            
            for selector in approach['selectors']:
                try:
                    if approach['method'] == 'xpath':
                        links = self.driver.find_elements(By.XPATH, selector)
                    else:
                        links = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    
                    print(f"      Found {len(links)} links with selector: {selector}")
                    
                    for link in links:
                        try:
                            text = link.text.strip().lower()
                            href = link.get_attribute('href')
                            
                            # Check if this looks like a side effects link
                            if approach['name'] == 'Navigation areas':
                                # For navigation areas, check text content
                                if 'side effect' in text or 'adverse' in text:
                                    print(f"      ✅ Found side effects link: {link.text}")
                                    return link
                            else:
                                # For other approaches, we already filtered
                                if text and href:
                                    print(f"      ✅ Found side effects link: {link.text}")
                                    return link
                        except Exception as e:
                            continue
                            
                except Exception as e:
                    continue
        
        # If no link found, try to find side effects content on current page
        print(f"    No side effects link found, checking current page...")
        return None
    
    def extract_side_effects_content(self, medication):
        """Extract side effects content from the current page"""
        print(f"  📄 Extracting side effects content for {medication}")
        
        try:
            content_parts = []
            content_found = False
            
            # Wait for page to load
            time.sleep(3)
            
            # Try multiple approaches to find side effects content
            
            # Approach 1: Look for side effects sections
            section_selectors = [
                "#side-effects",
                ".side-effects",
                "[id*='side-effects']",
                "[class*='side-effects']",
                "[id*='adverse']",
                "[class*='adverse']"
            ]
            
            for selector in section_selectors:
                try:
                    sections = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    for section in sections:
                        text = section.text.strip()
                        if text and len(text) > 50:
                            content_parts.append(f"--- Side Effects ---")
                            content_parts.append(text)
                            content_found = True
                            break
                except:
                    continue
            
            # Approach 2: Look for headings containing "side effects"
            if not content_found:
                try:
                    headings = self.driver.find_elements(By.XPATH, 
                        "//h1[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'side effects')] | "
                        "//h2[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'side effects')] | "
                        "//h3[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'side effects')]"
                    )
                    
                    for heading in headings:
                        content_parts.append(f"--- {heading.text} ---")
                        
                        # Find the following content
                        try:
                            next_sibling = heading.find_element(By.XPATH, "./following-sibling::*")
                            count = 0
                            while next_sibling and count < 5:  # Limit to avoid infinite loop
                                tag = next_sibling.tag_name.lower()
                                if tag in ['h1', 'h2', 'h3']:
                                    break
                                    
                                text = next_sibling.text.strip()
                                if text and len(text) > 20:
                                    content_parts.append(text)
                                    content_found = True
                                
                                try:
                                    next_sibling = next_sibling.find_element(By.XPATH, "./following-sibling::*")
                                    count += 1
                                except:
                                    break
                        except:
                            pass
                        
                        if content_found:
                            break
                except:
                    pass
            
            # Approach 3: Look for paragraphs with side effects keywords
            if not content_found:
                try:
                    keywords = ['side effect', 'adverse', 'reaction', 'emergency', 'call your doctor', 'serious', 'common', 'rare']
                    paragraphs = self.driver.find_elements(By.TAG_NAME, "p")
                    
                    for p in paragraphs:
                        text = p.text.strip()
                        if text and len(text) > 30:
                            text_lower = text.lower()
                            if any(keyword in text_lower for keyword in keywords):
                                content_parts.append(text)
                                content_found = True
                except:
                    pass
            
            # Approach 4: Look for lists that might contain side effects
            if not content_found:
                try:
                    lists = self.driver.find_elements(By.CSS_SELECTOR, "ul, ol")
                    for list_elem in lists:
                        text = list_elem.text.strip()
                        if text and len(text) > 50:
                            text_lower = text.lower()
                            if any(keyword in text_lower for keyword in ['side effect', 'adverse', 'reaction']):
                                content_parts.append("--- Side Effects List ---")
                                content_parts.append(text)
                                content_found = True
                                break
                except:
                    pass
            
            if content_parts:
                content = '\n\n'.join(content_parts)
                content = self.clean_content(content)
                return content
            else:
                return f"No specific side effects content found for {medication}"
                
        except Exception as e:
            return f"Error extracting side effects content: {str(e)}"
    
    def clean_content(self, content):
        """Clean and format the side effects content"""
        if not content:
            return ""
        
        # Remove excessive whitespace and unwanted elements
        lines = content.split('\n')
        cleaned_lines = []
        
        for line in lines:
            line = line.strip()
            if line and not any(skip in line.lower() for skip in [
                'advertisement', 'ads by', 'sponsored', 'cookie', 'privacy',
                'terms of use', 'about us', 'contact us', 'site map',
                'subscribe', 'newsletter', 'email', 'follow us'
            ]):
                cleaned_lines.append(line)
        
        # Join lines and remove excessive blank lines
        content = '\n'.join(cleaned_lines)
        
        # Remove multiple consecutive newlines
        while '\n\n\n' in content:
            content = content.replace('\n\n\n', '\n\n')
        
        return content.strip()
    
    def add_delay(self):
        """Add random delay between requests"""
        delay = random.uniform(5, 9)  # Longer delays to avoid detection
        print(f"  ⏰ Waiting {delay:.1f} seconds before next request...")
        time.sleep(delay)
    
    def close(self):
        """Close the browser"""
        self.driver.quit()

def update_excel_with_side_effects(max_medications=None):
    """Update Excel file with side effects for all medications"""
    
    excel_path = '/Users/juanlu/Documents/Wye/scrapper/Analysis/main_diseases_analysis_final.xlsx'
    
    if not os.path.exists(excel_path):
        print(f"❌ Excel file not found: {excel_path}")
        return
    
    # Load workbook
    wb = load_workbook(excel_path)
    
    if "All Unique Medications" not in wb.sheetnames:
        print("❌ 'All Unique Medications' sheet not found")
        return
    
    ws = wb["All Unique Medications"]
    
    # Initialize scraper
    scraper = DrugsScraper(headless=True)  # Set to True for headless mode
    
    try:
        # Get all medications (starting from row 9 where actual data begins)
        medications = []
        for row in range(9, ws.max_row + 1):  # Start from row 9
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and isinstance(cell_value, str) and cell_value.strip():
                medications.append((row, cell_value.strip()))
        
        if max_medications:
            medications = medications[:max_medications]
        
        print(f"📊 Found {len(medications)} medications to process")
        
        # Process each medication
        for i, (row_num, medication) in enumerate(medications, 1):
            print(f"\n{'='*60}")
            print(f"Processing {i}/{len(medications)}: {medication}")
            print(f"{'='*60}")
            
            # Check if already processed
            try:
                current_value = ws.cell(row=row_num, column=2).value
                if current_value and isinstance(current_value, str) and len(current_value) > 100:
                    print(f"  ⏭️  Already processed: {medication}")
                    continue
            except Exception as e:
                print(f"  ⚠️  Error checking existing value: {e}")
            
            # Scrape side effects
            side_effects = scraper.scrape_side_effects(medication)
            
            # Update Excel
            try:
                ws.cell(row=row_num, column=2, value=side_effects)
                print(f"  ✅ Updated Excel for {medication}")
            except Exception as e:
                print(f"  ❌ Error updating Excel for {medication}: {e}")
            
            # Save progress every 5 medications
            if i % 5 == 0:
                try:
                    wb.save(excel_path)
                    print(f"  💾 Progress saved after {i} medications")
                except Exception as e:
                    print(f"  ❌ Error saving progress: {e}")
            
            # Add delay between requests
            if i < len(medications):
                scraper.add_delay()
        
        # Final save
        try:
            wb.save(excel_path)
            print(f"\n✅ Successfully processed {len(medications)} medications!")
            print(f"📄 Updated Excel file: {excel_path}")
        except Exception as e:
            print(f"❌ Error saving final file: {e}")
        
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        try:
            wb.save(excel_path)  # Save progress even on error
        except:
            pass
    
    finally:
        scraper.close()

if __name__ == "__main__":
    print("🚀 Starting Drugs.com Side Effects Scraper - Firefox Version")
    print("="*60)
    print("🦊 Using Firefox browser with 5-9 second delays")
    print("="*60)
    
    # Process ALL medications
    update_excel_with_side_effects()  # No limit - process all
    
    print("\n" + "="*60)
    print("🎉 FULL SCRAPING COMPLETED!")
    print("="*60)
