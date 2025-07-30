import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import re
import requests
from bs4 import BeautifulSoup
import time
from dotenv import load_dotenv
from google.generativeai import GenerativeModel
import google.generativeai as genai

# Load environment variables
load_dotenv('../.env')

# Initialize LLM
api_key = os.getenv('GOOGLE_GEMINI_API_KEY')
if not api_key:
    print("❌ GOOGLE_GEMINI_API_KEY not found in environment variables")
    print("Please set your Google Gemini API key in the .env file")
    exit(1)

genai.configure(api_key=api_key)
model = GenerativeModel('gemini-1.5-flash')

def extract_tests_and_treatments():
    """
    Extract all unique tests and treatments from the top 10 diseases
    and create a comprehensive Excel file with detailed information
    """
    
    # Target diseases to extract (same as main_diseases_analyzer_final.py)
    target_diseases = [
        'Heart disease',
        'Chronic kidney disease',
        'COPD',
        'Pneumonia',
        'Stroke',
        'Dementia',
        'Depression (major depressive disorder)',
        'High cholesterol',
        'Obesity', 
        'Arthritis'
    ]
    
    # Read the CSV file
    csv_path = '/Users/juanlu/Documents/Wye/scrapper/CSV/final_diseases_complete.csv'
    df = pd.read_csv(csv_path)
    
    print(f"📊 Processing {len(target_diseases)} target diseases...")
    print(f"📄 Source data: {len(df)} total diseases in CSV")
    
    # Collect all tests and treatments with their disease associations
    all_tests = {}  # {test_name: [list_of_diseases]}
    all_treatments = {}  # {treatment_name: [list_of_diseases]}
    
    processed_diseases = []
    
    for disease in target_diseases:
        print(f"\n🔍 Processing: {disease}")
        
        # Find matching rows with specific matching (same logic as main analyzer)
        if disease == 'Heart disease':
            disease_data = df[df['Disease_Name_English'].str.contains('^Heart disease$', case=False, na=False, regex=True)]
        elif disease == 'Obesity':
            disease_data = df[df['Disease_Name_English'].str.contains('^Obesity$', case=False, na=False, regex=True)]
        elif disease == 'Stroke':
            disease_data = df[df['Disease_Name_English'].str.contains('^Stroke$', case=False, na=False, regex=True)]
        else:
            disease_data = df[df['Disease_Name_English'].str.contains(f'^{disease}$', case=False, na=False, regex=True)]
        
        if disease_data.empty:
            print(f"  ⚠️ No exact match found for {disease}, trying partial match...")
            disease_data = df[df['Disease_Name_English'].str.contains(disease, case=False, na=False, regex=False)]
            
        if disease_data.empty:
            print(f"  ❌ No data found for {disease}")
            continue
            
        # Get the first match
        disease_row = disease_data.iloc[0]
        disease_name = disease_row['Disease_Name_English']
        disease_spanish = disease_row['Disease_Name_Spanish']
        
        processed_diseases.append({
            'original': disease,
            'matched': disease_name,
            'spanish': disease_spanish
        })
        
        print(f"  ✅ Found: {disease_name}")
        
        # Extract Tests
        tests_raw = disease_row['Tests'] if pd.notna(disease_row['Tests']) else ''
        if tests_raw:
            # Split by common separators and clean
            test_items = split_medical_items(tests_raw)
            print(f"  📋 Found {len(test_items)} tests")
            
            for test in test_items:
                test = clean_item_name(test)
                if test and len(test) > 2:  # Only meaningful test names
                    if test not in all_tests:
                        all_tests[test] = []
                    all_tests[test].append(disease_name)
        
        # Extract Treatments
        treatments_raw = disease_row['Treatments'] if pd.notna(disease_row['Treatments']) else ''
        if treatments_raw:
            # Split by common separators and clean
            treatment_items = split_medical_items(treatments_raw)
            print(f"  💊 Found {len(treatment_items)} treatments")
            
            for treatment in treatment_items:
                treatment = clean_item_name(treatment)
                if treatment and len(treatment) > 2:  # Only meaningful treatment names
                    if treatment not in all_treatments:
                        all_treatments[treatment] = []
                    all_treatments[treatment].append(disease_name)
    
    print(f"\n📊 EXTRACTION SUMMARY:")
    print(f"✅ Processed diseases: {len(processed_diseases)}")
    print(f"🧪 Unique tests found: {len(all_tests)}")
    print(f"💊 Unique treatments found: {len(all_treatments)}")
    
    # Enhance with Mayo Clinic data
    enhanced_tests = enhance_items_with_mayo_clinic(all_tests, "test")
    enhanced_treatments = enhance_items_with_mayo_clinic(all_treatments, "treatment")
    
    # Create Excel workbook
    wb = Workbook()
    
    # Create Tests sheet
    tests_ws = wb.active
    tests_ws.title = "Tests"
    create_enhanced_tests_sheet(tests_ws, enhanced_tests)
    
    # Create Treatments sheet
    treatments_ws = wb.create_sheet(title="Treatments")
    create_enhanced_treatments_sheet(treatments_ws, enhanced_treatments)
    
    # Create Summary sheet
    summary_ws = wb.create_sheet(title="Summary")
    create_enhanced_summary_sheet(summary_ws, processed_diseases, enhanced_tests, enhanced_treatments)
    
    # Save the workbook
    output_path = '/Users/juanlu/Documents/Wye/scrapper/Analysis/tests_treatments_analysis.xlsx'
    wb.save(output_path)
    print(f"\n💾 Analysis saved to: {output_path}")
    
    return output_path

def split_medical_items(text):
    """Split medical text into individual items using various separators"""
    if not text or pd.isna(text):
        return []
    
    # Skip very long descriptive text (likely descriptions, not item lists)
    if len(text) > 300:
        return []
    
    # Special handling for common compound medical terms that should NOT be split
    compound_terms = [
        'heat and cold therapy',
        'heat/cold therapy', 
        'cold and heat therapy',
        'physical and occupational therapy',
        'speech and language therapy'
    ]
    
    # Check if this text contains compound terms that shouldn't be split
    text_lower = text.lower().strip()
    is_single_compound_term = False
    
    # Only treat as single compound if the ENTIRE text is just the compound term
    for compound in compound_terms:
        if text_lower == compound or text_lower == compound.strip():
            is_single_compound_term = True
            break
    
    if is_single_compound_term:
        # If it's a single compound term, return it as a single item
        cleaned = clean_item_name(text)
        return [cleaned] if cleaned and len(cleaned) > 3 else []
    
    # Handle parentheses carefully AND preserve compound terms
    # First split by semicolons while respecting parentheses
    items = []
    current_item = ""
    paren_depth = 0
    
    i = 0
    while i < len(text):
        char = text[i]
        
        if char == '(':
            paren_depth += 1
            current_item += char
        elif char == ')':
            paren_depth -= 1
            current_item += char
        elif char == ';' and paren_depth == 0:
            # Only split on semicolon if we're not inside parentheses
            if current_item.strip():
                items.append(current_item.strip())
            current_item = ""
        else:
            current_item += char
        
        i += 1
    
    # Add the last item
    if current_item.strip():
        items.append(current_item.strip())
    
    # Now protect compound terms within each item
    protected_items = []
    for item in items:
        item_lower = item.lower().strip()
        
        # Check if this individual item is a compound term that should stay whole
        is_protected = any(compound in item_lower for compound in compound_terms)
        
        if is_protected:
            # Keep the whole item as-is
            protected_items.append(item)
        else:
            # This item can be further split if needed
            protected_items.append(item)
    
    items = protected_items
    
    # If no semicolons were found, try other separators (but more carefully)
    if len(items) == 1:
        original_text = items[0]
        
        # Try splitting by other separators only if no parentheses
        if '(' not in original_text or ')' not in original_text:
            separators = ['\n', ' and ', ' or ', ' / ', ' | ', ',']
            
            for separator in separators:
                if separator in original_text:
                    # Special check: don't split if it would break compound terms
                    potential_items = original_text.split(separator)
                    
                    # Check if any item contains compound terms
                    safe_to_split = True
                    for item in potential_items:
                        item_lower = item.lower().strip()
                        for compound in compound_terms:
                            if compound in item_lower and item_lower != compound:
                                # This item contains part of a compound term
                                safe_to_split = False
                                break
                        if not safe_to_split:
                            break
                    
                    if safe_to_split:
                        items = potential_items
                        break
    
    # Clean and filter items
    cleaned_items = []
    for item in items:
        cleaned = clean_item_name(item)
        if cleaned and len(cleaned) > 3:  # Only meaningful items
            cleaned_items.append(cleaned)
    
    # Remove duplicates while preserving order
    unique_items = []
    seen = set()
    for item in cleaned_items:
        if item.lower() not in seen:
            unique_items.append(item)
            seen.add(item.lower())
    
    return unique_items

def clean_item_name(item):
    """Clean and standardize item names"""
    if not item or pd.isna(item):
        return ""
    
    # Convert to string and strip
    item = str(item).strip()
    
    # Remove common prefixes/suffixes
    prefixes_to_remove = [
        'test:', 'tests:', 'testing:', 'treatment:', 'treatments:',
        'procedure:', 'procedures:', 'therapy:', 'therapies:',
        'including:', 'such as:', 'like:', 'for example:',
        '- ', '• ', '◦ ', '· '
    ]
    
    for prefix in prefixes_to_remove:
        if item.lower().startswith(prefix):
            item = item[len(prefix):].strip()
    
    # Remove numbers at the beginning
    item = re.sub(r'^\d+\.?\s*', '', item)
    
    # Remove extra whitespace
    item = re.sub(r'\s+', ' ', item).strip()
    
    # Remove parenthetical explanations that are too long
    if '(' in item and ')' in item:
        before_paren = item.split('(')[0].strip()
        if len(before_paren) > 5:  # Keep the part before parentheses if meaningful
            item = before_paren
    
    # Capitalize first letter
    if item:
        item = item[0].upper() + item[1:]
    
    return item

def create_enhanced_tests_sheet(ws, enhanced_tests):
    """Create the Tests sheet with enhanced Mayo Clinic data"""
    
    # Header styling
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
    subheader_font = Font(bold=True, size=12, color="FFFFFF")
    subheader_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    
    # Border styling
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Sheet title
    ws['A1'] = 'ENHANCED DIAGNOSTIC TESTS FROM TOP 10 DISEASES'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:G1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    # Information section
    ws['A3'] = 'INFORMATION'
    ws['A3'].font = subheader_font
    ws['A3'].fill = subheader_fill
    ws.merge_cells('A3:G3')
    ws['A3'].alignment = Alignment(horizontal='center')
    
    ws['A4'] = 'Purpose:'
    ws['B4'] = 'Comprehensive list of diagnostic tests enhanced with Mayo Clinic data'
    ws['A5'] = 'Source:'
    ws['B5'] = 'final_diseases_complete.csv + Mayo Clinic + LLM enhancement'
    ws['A6'] = 'Total Tests:'
    ws['B6'] = len(enhanced_tests)
    
    # Style the info cells
    for row in [4, 5, 6]:
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="F0F8F0", end_color="F0F8F0", fill_type="solid")
    
    # Create the table headers
    header_row = 8
    ws[f'A{header_row}'] = 'TEST NAME'
    ws[f'B{header_row}'] = 'SPANISH NAME'
    ws[f'C{header_row}'] = 'DESCRIPTION'
    ws[f'D{header_row}'] = 'BACKGROUND INFORMATION'
    ws[f'E{header_row}'] = 'MAIN DISEASES/CONDITIONS'
    ws[f'F{header_row}'] = 'COUNT'
    ws[f'G{header_row}'] = 'MAYO CLINIC URL'
    
    # Style header row
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        cell = ws[f'{col}{header_row}']
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Sort tests alphabetically
    sorted_tests = sorted(enhanced_tests.items(), key=lambda x: x[0].lower())
    
    # Add all tests
    for i, (test_name, test_info) in enumerate(sorted_tests):
        row_num = header_row + 1 + i
        
        ws[f'A{row_num}'] = test_name
        ws[f'B{row_num}'] = test_info['spanish_name']
        ws[f'C{row_num}'] = test_info['description']
        ws[f'D{row_num}'] = test_info['background']
        ws[f'E{row_num}'] = test_info['main_diseases']
        ws[f'F{row_num}'] = len(test_info['diseases'])
        ws[f'G{row_num}'] = test_info['mayo_url'] if test_info['mayo_url'] else 'Not found'
        
        # Add borders and formatting
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            cell = ws[f'{col}{row_num}']
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # Alternate row colors
        if i % 2 == 0:
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                ws[f'{col}{row_num}'].fill = PatternFill(start_color="F8FFF8", end_color="F8FFF8", fill_type="solid")
    
    # Set column widths
    ws.column_dimensions['A'].width = 25  # Test name
    ws.column_dimensions['B'].width = 25  # Spanish name
    ws.column_dimensions['C'].width = 40  # Description
    ws.column_dimensions['D'].width = 50  # Background
    ws.column_dimensions['E'].width = 30  # Main diseases
    ws.column_dimensions['F'].width = 10  # Count
    ws.column_dimensions['G'].width = 25  # Mayo URL
    
    print(f"✅ Created enhanced Tests sheet with {len(enhanced_tests)} tests")

def create_enhanced_treatments_sheet(ws, enhanced_treatments):
    """Create the Treatments sheet with enhanced Mayo Clinic data"""
    
    # Header styling
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="4169E1", end_color="4169E1", fill_type="solid")
    subheader_font = Font(bold=True, size=12, color="FFFFFF")
    subheader_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    
    # Border styling
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Sheet title
    ws['A1'] = 'ENHANCED TREATMENTS FROM TOP 10 DISEASES'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:G1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    # Information section
    ws['A3'] = 'INFORMATION'
    ws['A3'].font = subheader_font
    ws['A3'].fill = subheader_fill
    ws.merge_cells('A3:G3')
    ws['A3'].alignment = Alignment(horizontal='center')
    
    ws['A4'] = 'Purpose:'
    ws['B4'] = 'Comprehensive list of treatments enhanced with Mayo Clinic data'
    ws['A5'] = 'Source:'
    ws['B5'] = 'final_diseases_complete.csv + Mayo Clinic + LLM enhancement'
    ws['A6'] = 'Total Treatments:'
    ws['B6'] = len(enhanced_treatments)
    
    # Style the info cells
    for row in [4, 5, 6]:
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
    
    # Create the table headers
    header_row = 8
    ws[f'A{header_row}'] = 'TREATMENT NAME'
    ws[f'B{header_row}'] = 'SPANISH NAME'
    ws[f'C{header_row}'] = 'DESCRIPTION'
    ws[f'D{header_row}'] = 'BACKGROUND INFORMATION'
    ws[f'E{header_row}'] = 'MAIN DISEASES/CONDITIONS'
    ws[f'F{header_row}'] = 'COUNT'
    ws[f'G{header_row}'] = 'MAYO CLINIC URL'
    
    # Style header row
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        cell = ws[f'{col}{header_row}']
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4169E1", end_color="4169E1", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Sort treatments alphabetically
    sorted_treatments = sorted(enhanced_treatments.items(), key=lambda x: x[0].lower())
    
    # Add all treatments
    for i, (treatment_name, treatment_info) in enumerate(sorted_treatments):
        row_num = header_row + 1 + i
        
        ws[f'A{row_num}'] = treatment_name
        ws[f'B{row_num}'] = treatment_info['spanish_name']
        ws[f'C{row_num}'] = treatment_info['description']
        ws[f'D{row_num}'] = treatment_info['background']
        ws[f'E{row_num}'] = treatment_info['main_diseases']
        ws[f'F{row_num}'] = len(treatment_info['diseases'])
        ws[f'G{row_num}'] = treatment_info['mayo_url'] if treatment_info['mayo_url'] else 'Not found'
        
        # Add borders and formatting
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            cell = ws[f'{col}{row_num}']
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # Alternate row colors
        if i % 2 == 0:
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                ws[f'{col}{row_num}'].fill = PatternFill(start_color="F8F8FF", end_color="F8F8FF", fill_type="solid")
    
    # Set column widths
    ws.column_dimensions['A'].width = 25  # Treatment name
    ws.column_dimensions['B'].width = 25  # Spanish name
    ws.column_dimensions['C'].width = 40  # Description
    ws.column_dimensions['D'].width = 50  # Background
    ws.column_dimensions['E'].width = 30  # Main diseases
    ws.column_dimensions['F'].width = 10  # Count
    ws.column_dimensions['G'].width = 25  # Mayo URL
    
    print(f"✅ Created enhanced Treatments sheet with {len(enhanced_treatments)} treatments")

def create_enhanced_summary_sheet(ws, processed_diseases, enhanced_tests, enhanced_treatments):
    """Create a summary sheet with enhanced overview information"""
    
    # Header styling
    header_font = Font(bold=True, size=16, color="FFFFFF")
    header_fill = PatternFill(start_color="8B008B", end_color="8B008B", fill_type="solid")
    subheader_font = Font(bold=True, size=12, color="FFFFFF")
    subheader_fill = PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid")
    
    # Title
    ws['A1'] = 'ENHANCED TESTS & TREATMENTS ANALYSIS SUMMARY'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Analysis info
    ws['A3'] = 'Analysis Date:'
    ws['B3'] = '2025-07-15'
    ws['A4'] = 'Source Data:'
    ws['B4'] = 'final_diseases_complete.csv + Mayo Clinic + LLM Enhancement'
    ws['A5'] = 'Diseases Processed:'
    ws['B5'] = len(processed_diseases)
    ws['A6'] = 'Total Tests:'
    ws['B6'] = len(enhanced_tests)
    ws['A7'] = 'Total Treatments:'
    ws['B7'] = len(enhanced_treatments)
    
    # Count Mayo Clinic matches
    tests_with_mayo = sum(1 for test_info in enhanced_tests.values() if test_info['mayo_url'])
    treatments_with_mayo = sum(1 for treatment_info in enhanced_treatments.values() if treatment_info['mayo_url'])
    
    ws['A8'] = 'Tests with Mayo Clinic Data:'
    if len(enhanced_tests) > 0:
        ws['B8'] = f"{tests_with_mayo} / {len(enhanced_tests)} ({tests_with_mayo/len(enhanced_tests)*100:.1f}%)"
    else:
        ws['B8'] = "0 / 0 (0%)"
    
    ws['A9'] = 'Treatments with Mayo Clinic Data:'
    if len(enhanced_treatments) > 0:
        ws['B9'] = f"{treatments_with_mayo} / {len(enhanced_treatments)} ({treatments_with_mayo/len(enhanced_treatments)*100:.1f}%)"
    else:
        ws['B9'] = "0 / 0 (0%)"
    
    # Style info cells
    for row in [3, 4, 5, 6, 7, 8, 9]:
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="F5F0FF", end_color="F5F0FF", fill_type="solid")
    
    # Processed diseases list
    ws['A11'] = 'PROCESSED DISEASES'
    ws['A11'].font = subheader_font
    ws['A11'].fill = subheader_fill
    ws.merge_cells('A11:E11')
    ws['A11'].alignment = Alignment(horizontal='center')
    
    ws['A12'] = 'Original Name'
    ws['B12'] = 'Matched Name'
    ws['C12'] = 'Spanish Name'
    ws['D12'] = 'Tests Found'
    ws['E12'] = 'Treatments Found'
    
    # Header row style
    for col in ['A', 'B', 'C', 'D', 'E']:
        cell = ws[f'{col}12']
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="8B008B", end_color="8B008B", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Add processed diseases
    for i, disease_info in enumerate(processed_diseases):
        row_num = 13 + i
        
        # Count tests and treatments for this disease
        tests_count = sum(1 for test_info in enhanced_tests.values() 
                         if disease_info['matched'] in test_info['diseases'])
        treatments_count = sum(1 for treatment_info in enhanced_treatments.values() 
                             if disease_info['matched'] in treatment_info['diseases'])
        
        ws[f'A{row_num}'] = disease_info['original']
        ws[f'B{row_num}'] = disease_info['matched']
        ws[f'C{row_num}'] = disease_info['spanish']
        ws[f'D{row_num}'] = tests_count
        ws[f'E{row_num}'] = treatments_count
        
        # Color the row green for successful matches
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{row_num}'].fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    print(f"✅ Created enhanced Summary sheet")

def create_tests_sheet(ws, all_tests):
    """Create the Tests sheet with all unique tests"""
    
    # Header styling
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
    subheader_font = Font(bold=True, size=12, color="FFFFFF")
    subheader_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    
    # Border styling
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Sheet title
    ws['A1'] = 'DIAGNOSTIC TESTS FROM TOP 10 DISEASES'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    # Information section
    ws['A3'] = 'INFORMATION'
    ws['A3'].font = subheader_font
    ws['A3'].fill = subheader_fill
    ws.merge_cells('A3:F3')
    ws['A3'].alignment = Alignment(horizontal='center')
    
    ws['A4'] = 'Purpose:'
    ws['B4'] = 'Comprehensive list of all diagnostic tests used across top 10 diseases'
    ws['A5'] = 'Source:'
    ws['B5'] = 'final_diseases_complete.csv (Mayo Clinic data)'
    ws['A6'] = 'Total Tests:'
    ws['B6'] = len(all_tests)
    
    # Style the info cells
    for row in [4, 5, 6]:
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="F0F8F0", end_color="F0F8F0", fill_type="solid")
    
    # Create the table headers
    header_row = 8
    ws[f'A{header_row}'] = 'TEST NAME'
    ws[f'B{header_row}'] = 'SPANISH NAME'
    ws[f'C{header_row}'] = 'DESCRIPTION'
    ws[f'D{header_row}'] = 'BACKGROUND INFORMATION'
    ws[f'E{header_row}'] = 'DISEASES ASSOCIATED'
    ws[f'F{header_row}'] = 'COUNT'
    
    # Style header row
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        cell = ws[f'{col}{header_row}']
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Sort tests alphabetically
    sorted_tests = sorted(all_tests.items(), key=lambda x: x[0].lower())
    
    # Add all tests
    for i, (test_name, diseases) in enumerate(sorted_tests):
        row_num = header_row + 1 + i
        
        # Combine diseases into a single string
        diseases_str = '; '.join(sorted(set(diseases)))
        
        ws[f'A{row_num}'] = test_name
        ws[f'B{row_num}'] = ''  # To be filled with Spanish name
        ws[f'C{row_num}'] = ''  # To be filled with description
        ws[f'D{row_num}'] = ''  # To be filled with background info
        ws[f'E{row_num}'] = diseases_str
        ws[f'F{row_num}'] = len(diseases)
        
        # Add borders and formatting
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            cell = ws[f'{col}{row_num}']
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Alternate row colors
        if i % 2 == 0:
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                ws[f'{col}{row_num}'].fill = PatternFill(start_color="F8FFF8", end_color="F8FFF8", fill_type="solid")
    
    # Set column widths
    ws.column_dimensions['A'].width = 40  # Test Name
    ws.column_dimensions['B'].width = 40  # Spanish Name
    ws.column_dimensions['C'].width = 50  # Description
    ws.column_dimensions['D'].width = 50  # Background Information
    ws.column_dimensions['E'].width = 60  # Diseases Associated
    ws.column_dimensions['F'].width = 10  # Count
    
    print(f"✅ Created Tests sheet with {len(sorted_tests)} unique tests")

def create_treatments_sheet(ws, all_treatments):
    """Create the Treatments sheet with all unique treatments"""
    
    # Header styling
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="4169E1", end_color="4169E1", fill_type="solid")
    subheader_font = Font(bold=True, size=12, color="FFFFFF")
    subheader_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    
    # Border styling
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Sheet title
    ws['A1'] = 'TREATMENTS FROM TOP 10 DISEASES'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    # Information section
    ws['A3'] = 'INFORMATION'
    ws['A3'].font = subheader_font
    ws['A3'].fill = subheader_fill
    ws.merge_cells('A3:F3')
    ws['A3'].alignment = Alignment(horizontal='center')
    
    ws['A4'] = 'Purpose:'
    ws['B4'] = 'Comprehensive list of all treatments used across top 10 diseases'
    ws['A5'] = 'Source:'
    ws['B5'] = 'final_diseases_complete.csv (Mayo Clinic data)'
    ws['A6'] = 'Total Treatments:'
    ws['B6'] = len(all_treatments)
    
    # Style the info cells
    for row in [4, 5, 6]:
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
    
    # Create the table headers
    header_row = 8
    ws[f'A{header_row}'] = 'TREATMENT NAME'
    ws[f'B{header_row}'] = 'SPANISH NAME'
    ws[f'C{header_row}'] = 'DESCRIPTION'
    ws[f'D{header_row}'] = 'BACKGROUND INFORMATION'
    ws[f'E{header_row}'] = 'DISEASES ASSOCIATED'
    ws[f'F{header_row}'] = 'COUNT'
    
    # Style header row
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        cell = ws[f'{col}{header_row}']
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4169E1", end_color="4169E1", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Sort treatments alphabetically
    sorted_treatments = sorted(all_treatments.items(), key=lambda x: x[0].lower())
    
    # Add all treatments
    for i, (treatment_name, diseases) in enumerate(sorted_treatments):
        row_num = header_row + 1 + i
        
        # Combine diseases into a single string
        diseases_str = '; '.join(sorted(set(diseases)))
        
        ws[f'A{row_num}'] = treatment_name
        ws[f'B{row_num}'] = ''  # To be filled with Spanish name
        ws[f'C{row_num}'] = ''  # To be filled with description
        ws[f'D{row_num}'] = ''  # To be filled with background info
        ws[f'E{row_num}'] = diseases_str
        ws[f'F{row_num}'] = len(diseases)
        
        # Add borders and formatting
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            cell = ws[f'{col}{row_num}']
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Alternate row colors
        if i % 2 == 0:
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                ws[f'{col}{row_num}'].fill = PatternFill(start_color="F8F8FF", end_color="F8F8FF", fill_type="solid")
    
    # Set column widths
    ws.column_dimensions['A'].width = 40  # Treatment Name
    ws.column_dimensions['B'].width = 40  # Spanish Name
    ws.column_dimensions['C'].width = 50  # Description
    ws.column_dimensions['D'].width = 50  # Background Information
    ws.column_dimensions['E'].width = 60  # Diseases Associated
    ws.column_dimensions['F'].width = 10  # Count
    
    print(f"✅ Created Treatments sheet with {len(sorted_treatments)} unique treatments")

def create_summary_sheet(ws, processed_diseases, all_tests, all_treatments):
    """Create a summary sheet with overview information"""
    
    # Header styling
    header_font = Font(bold=True, size=16, color="FFFFFF")
    header_fill = PatternFill(start_color="8B008B", end_color="8B008B", fill_type="solid")
    subheader_font = Font(bold=True, size=12, color="FFFFFF")
    subheader_fill = PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid")
    
    # Title
    ws['A1'] = 'TESTS & TREATMENTS ANALYSIS SUMMARY'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Analysis info
    ws['A3'] = 'Analysis Date:'
    ws['B3'] = '2025-07-14'
    ws['A4'] = 'Source Data:'
    ws['B4'] = 'final_diseases_complete.csv'
    ws['A5'] = 'Diseases Processed:'
    ws['B5'] = len(processed_diseases)
    ws['A6'] = 'Total Tests:'
    ws['B6'] = len(all_tests)
    ws['A7'] = 'Total Treatments:'
    ws['B7'] = len(all_treatments)
    
    # Style info cells
    for row in [3, 4, 5, 6, 7]:
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="F5F0FF", end_color="F5F0FF", fill_type="solid")
    
    # Processed diseases list
    ws['A9'] = 'PROCESSED DISEASES'
    ws['A9'].font = subheader_font
    ws['A9'].fill = subheader_fill
    ws.merge_cells('A9:D9')
    ws['A9'].alignment = Alignment(horizontal='center')
    
    ws['A10'] = 'Original Name'
    ws['B10'] = 'Matched Name'
    ws['C10'] = 'Spanish Name'
    ws['D10'] = 'Status'
    
    # Style header row
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}10'].font = Font(bold=True)
        ws[f'{col}10'].fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    
    # Add processed diseases
    for i, disease_info in enumerate(processed_diseases, 11):
        ws[f'A{i}'] = disease_info['original']
        ws[f'B{i}'] = disease_info['matched']
        ws[f'C{i}'] = disease_info['spanish']
        ws[f'D{i}'] = '✅ Found'
        
        # Color the row green for successful matches
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{i}'].fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
    
    # Statistics
    stats_row = len(processed_diseases) + 13
    ws[f'A{stats_row}'] = 'STATISTICS'
    ws[f'A{stats_row}'].font = Font(bold=True, size=12)
    ws[f'A{stats_row+1}'] = f'Success Rate: {len(processed_diseases)}/10 diseases (100%)'
    ws[f'A{stats_row+2}'] = f'Average Tests per Disease: {len(all_tests)/len(processed_diseases):.1f}'
    ws[f'A{stats_row+3}'] = f'Average Treatments per Disease: {len(all_treatments)/len(processed_diseases):.1f}'
    
    # Top tests and treatments
    top_tests = sorted(all_tests.items(), key=lambda x: len(x[1]), reverse=True)[:5]
    top_treatments = sorted(all_treatments.items(), key=lambda x: len(x[1]), reverse=True)[:5]
    
    ws[f'A{stats_row+5}'] = 'TOP 5 MOST COMMON TESTS:'
    ws[f'A{stats_row+5}'].font = Font(bold=True)
    for i, (test, diseases) in enumerate(top_tests, stats_row+6):
        ws[f'A{i}'] = f"{test} ({len(diseases)} diseases)"
    
    ws[f'A{stats_row+12}'] = 'TOP 5 MOST COMMON TREATMENTS:'
    ws[f'A{stats_row+12}'].font = Font(bold=True)
    for i, (treatment, diseases) in enumerate(top_treatments, stats_row+13):
        ws[f'A{i}'] = f"{treatment} ({len(diseases)} diseases)"
    
    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 20
    
    print(f"✅ Created Summary sheet")

def search_mayo_clinic_direct(test_name):
    """Search for a test/treatment using common Mayo Clinic URL patterns"""
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.5',
        'Connection': 'keep-alive',
    }
    
    # Clean up test name for URL
    test_name_clean = test_name.lower().strip()
    url_name = test_name_clean.replace(' ', '-').replace('(', '').replace(')', '').replace(',', '').replace('/', '-')
    
    # Common Mayo Clinic procedures that we know exist
    known_procedures = {
        'blood test': 'complete-blood-count',
        'blood tests': 'complete-blood-count',
        'urine test': 'urinalysis',
        'urine tests': 'urinalysis',
        'ct scan': 'ct-scan',
        'computerized tomography': 'ct-scan',
        'mri': 'mri',
        'magnetic resonance imaging': 'mri',
        'x-ray': 'x-ray',
        'chest x': 'chest-x-rays',
        'ultrasound': 'ultrasound',
        'echocardiogram': 'echocardiogram',
        'electrocardiogram': 'ekg',
        'ekg': 'ekg',
        'ecg': 'ekg',
        'colonoscopy': 'colonoscopy',
        'biopsy': 'biopsy',
        'kidney biopsy': 'kidney-biopsy',
        'physical therapy': 'physical-therapy',
        'surgery': 'robotic-surgery',
        'chemotherapy': 'chemotherapy',
        'radiation therapy': 'radiation-therapy',
        'dialysis': 'hemodialysis',
        'hemodialysis': 'hemodialysis',
        'peritoneal dialysis': 'peritoneal-dialysis',
        'kidney transplant': 'kidney-transplant',
        'cholesterol test': 'cholesterol-test',
        'spirometry': 'spirometry',
        'stress test': 'stress-test',
        'exercise stress test': 'stress-test',
        'arthroscopy': 'arthroscopy',
        'complete blood count': 'complete-blood-count',
        'pet scan': 'pet-scan',
        'carotid ultrasound': 'carotid-ultrasound',
        'physical exam': 'physical-exam',
        'physical examination': 'physical-exam'
    }
    
    # Check if we have a known procedure
    procedure_url = None
    for known_name, url_suffix in known_procedures.items():
        if known_name in test_name_clean or test_name_clean in known_name:
            procedure_url = f"https://www.mayoclinic.org/tests-procedures/{url_suffix}/about/pac-20384919"
            break
    
    if not procedure_url:
        # Try direct URL pattern
        procedure_url = f"https://www.mayoclinic.org/tests-procedures/{url_name}/about/pac-20384919"
    
    try:
        response = requests.get(procedure_url, headers=headers, timeout=10)
        if response.status_code == 200:
            return procedure_url, test_name
        else:
            return None, None
    except:
        return None, None

def scrape_mayo_clinic_procedure(url, procedure_name):
    """Scrape detailed information from a Mayo Clinic procedure page"""
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.5',
        'Connection': 'keep-alive',
    }
    
    try:
        response = requests.get(url, headers=headers, timeout=15)
        response.raise_for_status()
        soup = BeautifulSoup(response.content, 'html.parser')
        
        # Extract main content
        main_content = soup.find('div', class_='content') or soup.find('main') or soup.find('article')
        if not main_content:
            main_content = soup.find('body')
        
        if main_content:
            # Get all text content
            content_text = main_content.get_text(separator=' ', strip=True)
            # Clean up the text
            content_text = re.sub(r'\s+', ' ', content_text)
            
            return content_text[:6000]  # Limit for LLM processing
            
    except Exception as e:
        print(f"   ❌ Error scraping {url}: {e}")
        
    return None

def extract_procedure_info_with_llm(content_text, procedure_name):
    """Use LLM to extract Spanish name, description, and background info"""
    if not content_text:
        return {
            'spanish_name': 'Información no encontrada',
            'description': 'Information not found',
            'background': 'Information not found',
            'main_diseases': 'Information not found'
        }
    
    try:
        prompt = f"""
You are a medical information extraction expert. Extract information about this medical test/procedure: "{procedure_name}"

CONTENT FROM MAYO CLINIC:
{content_text}

Please extract the following information in this EXACT format:

SPANISH_NAME: [The Spanish medical translation of "{procedure_name}"]
DESCRIPTION: [A concise 2-3 sentence description of what this test/procedure is and what it does]
BACKGROUND: [Background information about when this test is used, why it's performed, what conditions it helps diagnose or treat]
MAIN_DISEASES: [List the 3-5 main diseases/conditions that this test/procedure is most commonly used for, separated by semicolons]

Guidelines:
- Be concise but informative
- Use medical terminology appropriately
- If specific information is not in the content, write "Information not found"
- For SPANISH_NAME: provide accurate medical Spanish translation
- For DESCRIPTION: focus on what the test/procedure is and does
- For BACKGROUND: explain when and why it's used, what it helps with
- For MAIN_DISEASES: list the primary medical conditions this test/procedure is used for (e.g., "Heart disease; Diabetes; Hypertension")
"""

        result = model.generate_content(prompt)
        response = result.text  # Use .text instead of .response.text()
        
        # Parse the LLM response
        extracted_info = {
            'spanish_name': 'Información no encontrada',
            'description': 'Information not found',
            'background': 'Information not found',
            'main_diseases': 'Information not found'
        }
        
        lines = response.split('\n')
        for line in lines:
            line = line.strip()
            if line.startswith('SPANISH_NAME:'):
                extracted_info['spanish_name'] = line.replace('SPANISH_NAME:', '').strip()
            elif line.startswith('DESCRIPTION:'):
                extracted_info['description'] = line.replace('DESCRIPTION:', '').strip()
            elif line.startswith('BACKGROUND:'):
                extracted_info['background'] = line.replace('BACKGROUND:', '').strip()
            elif line.startswith('MAIN_DISEASES:'):
                extracted_info['main_diseases'] = line.replace('MAIN_DISEASES:', '').strip()
        
        return extracted_info
        
    except Exception as e:
        print(f"   ❌ LLM extraction failed: {e}")
        return {
            'spanish_name': 'Error en extracción',
            'description': 'Extraction error',
            'background': 'Extraction error',
            'main_diseases': 'Extraction error'
        }

def enhance_items_with_mayo_clinic(items_dict, item_type="test"):
    """Enhance test/treatment information with Mayo Clinic data and LLM"""
    print(f"\n🔍 Enhancing {len(items_dict)} {item_type}s with Mayo Clinic data...")
    
    enhanced_items = {}
    
    for i, (item_name, diseases) in enumerate(items_dict.items(), 1):
        print(f"\n[{i}/{len(items_dict)}] Processing: {item_name}")
        
        # Search for the item on Mayo Clinic
        mayo_url, mayo_title = search_mayo_clinic_direct(item_name)
        
        if mayo_url:
            print(f"   ✅ Found Mayo Clinic page")
            
            # Scrape the page content
            content = scrape_mayo_clinic_procedure(mayo_url, item_name)
            
            if content:
                print(f"   🤖 Using LLM to extract information...")
                # Extract information using LLM
                llm_info = extract_procedure_info_with_llm(content, item_name)
                
                enhanced_items[item_name] = {
                    'diseases': diseases,
                    'spanish_name': llm_info['spanish_name'],
                    'description': llm_info['description'],
                    'background': llm_info['background'],
                    'main_diseases': llm_info['main_diseases'],
                    'mayo_url': mayo_url,
                    'mayo_title': mayo_title
                }
                
                print(f"   ✅ Information extracted successfully")
            else:
                print(f"   ⚠️ Could not scrape content from {mayo_url}")
                enhanced_items[item_name] = {
                    'diseases': diseases,
                    'spanish_name': 'No se pudo obtener información',
                    'description': 'Could not retrieve information',
                    'background': 'Could not retrieve information',
                    'main_diseases': 'Could not retrieve information',
                    'mayo_url': mayo_url,
                    'mayo_title': mayo_title
                }
        else:
            print(f"   ❌ Not found on Mayo Clinic")
            enhanced_items[item_name] = {
                'diseases': diseases,
                'spanish_name': 'No encontrado en Mayo Clinic',
                'description': 'Not found on Mayo Clinic',
                'background': 'Not found on Mayo Clinic',
                'main_diseases': 'Not found on Mayo Clinic',
                'mayo_url': None,
                'mayo_title': None
            }
        
        # Add delay to be respectful to the server
        time.sleep(1)
    
    return enhanced_items

def load_csv_file():
    """Load and return the CSV file as a pandas DataFrame"""
    csv_path = "../CSV/final_diseases_complete.csv"
    try:
        df = pd.read_csv(csv_path)
        print(f"✅ Successfully loaded CSV file: {csv_path}")
        print(f"📊 Data shape: {df.shape}")
        return df
    except FileNotFoundError:
        print(f"❌ CSV file not found: {csv_path}")
        return None
    except Exception as e:
        print(f"❌ Error loading CSV file: {e}")
        return None

def get_unique_diseases(df):
    """Get unique diseases from the CSV with their counts"""
    if 'Disease_Name_English' not in df.columns:
        print("❌ 'Disease_Name_English' column not found in CSV file")
        return {}
    
    # Count occurrences of each disease
    disease_counts = df['Disease_Name_English'].value_counts()
    
    # Convert to dictionary
    all_diseases = disease_counts.to_dict()
    
    print(f"📊 Found {len(all_diseases)} unique diseases")
    return all_diseases

def get_top_diseases(all_diseases, top_n=10):
    """Get the top N most common diseases"""
    
    # Sort diseases by count (descending) and take top N
    top_diseases = dict(sorted(all_diseases.items(), key=lambda x: x[1], reverse=True)[:top_n])
    
    print(f"🔝 Top {top_n} diseases:")
    for i, (disease, count) in enumerate(top_diseases.items(), 1):
        print(f"   {i}. {disease}: {count} records")
    
    return top_diseases

def process_diseases(df, top_diseases):
    """Process the top diseases and return their information"""
    processed_diseases = []
    
    for disease_name, count in top_diseases.items():
        # Get Spanish name if available
        spanish_name = disease_name
        try:
            spanish_row = df[df['Disease_Name_English'] == disease_name]['Disease_Name_Spanish'].iloc[0]
            if not pd.isna(spanish_row) and spanish_row != '':
                spanish_name = spanish_row
        except:
            pass
        
        disease_info = {
            'original': disease_name,
            'matched': disease_name,
            'spanish': spanish_name,
            'count': count
        }
        processed_diseases.append(disease_info)
    
    return processed_diseases

def clean_test_treatment_name(item):
    """Clean and normalize test/treatment names"""
    if not item or pd.isna(item):
        return ""
    
    # Convert to string and strip whitespace
    item = str(item).strip()
    
    # Remove common prefixes
    prefixes_to_remove = [
        'test for ', 'testing for ', 'test of ', 'tests for ',
        'treatment for ', 'treatment of ', 'therapy for ', 'therapy of ',
        'medication for ', 'medications for ', 'drug for ', 'drugs for '
    ]
    
    for prefix in prefixes_to_remove:
        if item.lower().startswith(prefix):
            item = item[len(prefix):].strip()
    
    # Remove numbers at the beginning
    item = re.sub(r'^\d+\.?\s*', '', item)
    
    # Remove extra whitespace
    item = re.sub(r'\s+', ' ', item).strip()
    
    # Remove parenthetical explanations that are too long
    if '(' in item and ')' in item:
        before_paren = item.split('(')[0].strip()
        if len(before_paren) > 5:  # Keep the part before parentheses if meaningful
            item = before_paren
    
    # Capitalize first letter
    if item:
        item = item[0].upper() + item[1:]
    
    return item

def extract_tests_and_treatments(df, processed_diseases, column_type):
    """Extract tests or treatments from the processed diseases"""
    all_items = {}
    
    # Get the column name
    if column_type == 'tests':
        column_name = 'Tests'
    else:
        column_name = 'Treatments'
    
    if column_name not in df.columns:
        print(f"❌ '{column_name}' column not found in CSV file")
        return all_items
    
    # Get diseases we're interested in
    disease_names = [disease['matched'] for disease in processed_diseases]
    
    # Filter dataframe to only include our diseases
    filtered_df = df[df['Disease_Name_English'].isin(disease_names)]
    
    print(f"📊 Processing {column_type} from {len(filtered_df)} records...")
    
    # Process each row
    for _, row in filtered_df.iterrows():
        disease = row['Disease_Name_English']
        items_text = row[column_name]
        
        if pd.isna(items_text) or items_text == '':
            continue
        
        # Split the items (assuming they're separated by commas or semicolons)
        items = []
        if ';' in items_text:
            items = [item.strip() for item in items_text.split(';')]
        else:
            items = [item.strip() for item in items_text.split(',')]
        
        # Clean and process each item
        for item in items:
            if item:
                # Clean the item
                clean_item = clean_test_treatment_name(item)
                
                if clean_item and len(clean_item) > 2:  # Only keep meaningful items
                    if clean_item not in all_items:
                        all_items[clean_item] = []
                    all_items[clean_item].append(disease)
    
    # Remove duplicates in disease lists
    for item_name in all_items:
        all_items[item_name] = list(set(all_items[item_name]))
    
    print(f"✅ Found {len(all_items)} unique {column_type}")
    return all_items

def main():
    """Main function to process diseases and generate enhanced Excel analysis."""
    
    # Load the CSV file
    df = load_csv_file()
    if df is None:
        return
    
    # Get unique diseases from the CSV
    all_diseases = get_unique_diseases(df)
    
    # Filter to top 10 most common diseases
    top_diseases = get_top_diseases(all_diseases)
    
    # Process diseases
    processed_diseases = process_diseases(df, top_diseases)
    
    # Extract tests and treatments
    all_tests = extract_tests_and_treatments(df, processed_diseases, 'tests')
    all_treatments = extract_tests_and_treatments(df, processed_diseases, 'treatments')
    
    # Enhance with Mayo Clinic data
    print("\n🔍 Enhancing tests with Mayo Clinic data...")
    enhanced_tests = enhance_items_with_mayo_clinic(all_tests, 'test')
    
    print("\n🔍 Enhancing treatments with Mayo Clinic data...")
    enhanced_treatments = enhance_items_with_mayo_clinic(all_treatments, 'treatment')
    
    # Create Excel workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create sheets
    summary_sheet = wb.create_sheet(title="Summary")
    tests_sheet = wb.create_sheet(title="Diagnostic Tests")
    treatments_sheet = wb.create_sheet(title="Treatments")
    
    # Fill sheets with enhanced data
    create_enhanced_summary_sheet(summary_sheet, processed_diseases, enhanced_tests, enhanced_treatments)
    create_enhanced_tests_sheet(tests_sheet, enhanced_tests)
    create_enhanced_treatments_sheet(treatments_sheet, enhanced_treatments)
    
    # Save the Excel file
    output_file = "../Analysis/tests_treatments_enhanced_analysis.xlsx"
    wb.save(output_file)
    
    print(f"\n✅ Enhanced Excel file created successfully: {output_file}")
    print(f"📊 Summary:")
    print(f"   - Diseases processed: {len(processed_diseases)}")
    print(f"   - Tests found: {len(enhanced_tests)}")
    print(f"   - Treatments found: {len(enhanced_treatments)}")
    
    # Display Mayo Clinic enhancement stats
    tests_with_mayo = sum(1 for test_info in enhanced_tests.values() if test_info['mayo_url'])
    treatments_with_mayo = sum(1 for treatment_info in enhanced_treatments.values() if treatment_info['mayo_url'])
    
    if len(enhanced_tests) > 0:
        print(f"   - Tests with Mayo Clinic data: {tests_with_mayo}/{len(enhanced_tests)} ({tests_with_mayo/len(enhanced_tests)*100:.1f}%)")
    else:
        print(f"   - Tests with Mayo Clinic data: 0/0 (0%)")
    
    if len(enhanced_treatments) > 0:
        print(f"   - Treatments with Mayo Clinic data: {treatments_with_mayo}/{len(enhanced_treatments)} ({treatments_with_mayo/len(enhanced_treatments)*100:.1f}%)")
    else:
        print(f"   - Treatments with Mayo Clinic data: 0/0 (0%)")
    
    # Show some examples of enhanced items
    print("\n📝 Examples of enhanced items:")
    
    # Show first test with Mayo Clinic data
    enhanced_test = next(((name, info) for name, info in enhanced_tests.items() if info['mayo_url']), None)
    if enhanced_test:
        test_name, test_info = enhanced_test
        print(f"\n   Test: {test_name}")
        print(f"   Spanish: {test_info['spanish_name']}")
        print(f"   Description: {test_info['description'][:100]}...")
        print(f"   Mayo URL: {test_info['mayo_url']}")
    
    # Show first treatment with Mayo Clinic data
    enhanced_treatment = next(((name, info) for name, info in enhanced_treatments.items() if info['mayo_url']), None)
    if enhanced_treatment:
        treatment_name, treatment_info = enhanced_treatment
        print(f"\n   Treatment: {treatment_name}")
        print(f"   Spanish: {treatment_info['spanish_name']}")
        print(f"   Description: {treatment_info['description'][:100]}...")
        print(f"   Mayo URL: {treatment_info['mayo_url']}")

if __name__ == "__main__":
    main()
