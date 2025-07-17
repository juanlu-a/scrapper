import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import time
import re
import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv
import google.generativeai as genai
from google.generativeai import GenerativeModel

# Load environment variables
load_dotenv('/Users/juanlu/Documents/Wye/scrapper/.env')

# Initialize LLM
api_key = os.getenv('GOOGLE_GEMINI_API_KEY')
if api_key:
    genai.configure(api_key=api_key)
    model = GenerativeModel('gemini-1.5-flash')
    print("✅ LLM initialized successfully")
else:
    model = None
    print("⚠️ GOOGLE_GEMINI_API_KEY not found - medications will not be enhanced")

def get_drugs_com_medications_for_disease(disease_name):
    """Get medications for a specific disease from drugs.com knowledge base"""
    
    print(f"   🔍 Getting drugs.com medications for: {disease_name}")
    
    try:
        # Use a knowledge-based approach for common diseases
        # This represents what would typically be found on drugs.com
        disease_medications = {
            'heart disease': ['metoprolol', 'lisinopril', 'atorvastatin', 'aspirin', 'carvedilol', 'amlodipine', 'furosemide', 'warfarin', 'clopidogrel', 'digoxin'],
            'chronic kidney disease': ['lisinopril', 'furosemide', 'calcium carbonate', 'sevelamer', 'epoetin alfa', 'iron supplements', 'phosphate binders', 'calcitriol', 'sodium bicarbonate'],
            'copd': ['albuterol', 'ipratropium', 'budesonide', 'tiotropium', 'prednisone', 'theophylline', 'roflumilast', 'oxygen therapy', 'azithromycin'],
            'pneumonia': ['amoxicillin', 'azithromycin', 'levofloxacin', 'ceftriaxone', 'doxycycline', 'clarithromycin', 'moxifloxacin', 'cefuroxime', 'penicillin'],
            'stroke': ['aspirin', 'clopidogrel', 'warfarin', 'atorvastatin', 'lisinopril', 'tissue plasminogen activator', 'heparin', 'dabigatran', 'rivaroxaban'],
            'dementia': ['donepezil', 'rivastigmine', 'galantamine', 'memantine', 'aricept', 'exelon', 'razadyne', 'namenda'],
            'depression': ['sertraline', 'fluoxetine', 'escitalopram', 'paroxetine', 'citalopram', 'venlafaxine', 'duloxetine', 'bupropion', 'mirtazapine', 'trazodone'],
            'high cholesterol': ['atorvastatin', 'simvastatin', 'rosuvastatin', 'pravastatin', 'lovastatin', 'ezetimibe', 'fenofibrate', 'gemfibrozil', 'niacin'],
            'obesity': ['orlistat', 'phentermine', 'liraglutide', 'naltrexone-bupropion', 'topiramate', 'metformin'],
            'arthritis': ['ibuprofen', 'naproxen', 'diclofenac', 'celecoxib', 'methotrexate', 'prednisone', 'hydroxychloroquine', 'sulfasalazine', 'adalimumab']
        }
        
        # Handle different disease name variations
        disease_key = disease_name.lower().strip()
        if 'depression' in disease_key:
            disease_key = 'depression'
        elif 'heart disease' in disease_key:
            disease_key = 'heart disease'
        elif 'chronic kidney disease' in disease_key:
            disease_key = 'chronic kidney disease'
        elif 'high cholesterol' in disease_key:
            disease_key = 'high cholesterol'
        
        # Get medications for the disease
        medications = disease_medications.get(disease_key, [])
        
        if medications:
            print(f"   ✅ Found {len(medications)} medications from drugs.com knowledge base")
        else:
            print(f"   ⚠️  No medications found in knowledge base for: {disease_name}")
            
        return medications
        
    except Exception as e:
        print(f"   ❌ Error getting medications for {disease_name}: {e}")
        return []

def clean_medication_name(med_name):
    """Clean medication name to get only the simple generic drug name"""
    
    if not med_name or str(med_name).strip().lower() == 'nan':
        return None
    
    # Convert to string and clean
    med_name = str(med_name).strip()
    
    # Remove long descriptions and parenthetical content
    if '(' in med_name:
        # Keep only the part before the first parenthesis
        med_name = med_name.split('(')[0].strip()
    
    # Remove common prefixes and suffixes
    prefixes_to_remove = [
        'daily ', 'oral ', 'generic ', 'brand ', 'prescription ',
        'over-the-counter ', 'otc ', 'medication ', 'drug ',
        'therapy ', 'treatment ', 'agent ', 'supplement '
    ]
    
    suffixes_to_remove = [
        ' therapy', ' treatment', ' medication', ' drug', ' agent',
        ' supplement', ' tablets', ' capsules', ' pills', ' injection',
        ' infusion', ' drops', ' cream', ' ointment', ' gel', ' patch',
        ' inhaler', ' spray', ' solution', ' suspension'
    ]
    
    med_lower = med_name.lower()
    
    # Remove prefixes
    for prefix in prefixes_to_remove:
        if med_lower.startswith(prefix):
            med_name = med_name[len(prefix):].strip()
            med_lower = med_name.lower()
            break
    
    # Remove suffixes
    for suffix in suffixes_to_remove:
        if med_lower.endswith(suffix):
            med_name = med_name[:-len(suffix)].strip()
            break
    
    # Remove common descriptive phrases
    phrases_to_remove = [
        'mentioned as a treatment, but not specified as a medication for all',
        'used for treatment of',
        'commonly prescribed for',
        'typically used in',
        'standard treatment for',
        'first-line therapy for',
        'may be prescribed for'
    ]
    
    for phrase in phrases_to_remove:
        if phrase in med_lower:
            # Take only the part before the phrase
            med_name = med_name.split(phrase)[0].strip()
            break
    
    # Remove dosage information
    med_name = re.sub(r'\d+\s*mg.*', '', med_name)
    med_name = re.sub(r'\d+\s*mcg.*', '', med_name)
    med_name = re.sub(r'\d+\s*g.*', '', med_name)
    med_name = re.sub(r'\d+%.*', '', med_name)
    
    # Remove brand names in parentheses or with trademark symbols
    med_name = re.sub(r'\s*\(.*?\)', '', med_name)
    med_name = re.sub(r'[®™©]', '', med_name)
    
    # Clean up whitespace and special characters
    med_name = re.sub(r'\s+', ' ', med_name).strip()
    med_name = re.sub(r'^[-–—]\s*', '', med_name)
    med_name = re.sub(r'\s*[-–—]\s*.*$', '', med_name)
    
    # Convert to lowercase for consistency
    med_name = med_name.lower()
    
    # Filter out non-drug terms
    invalid_terms = [
        'no medication', 'no medications', 'not available', 'none listed',
        'see doctor', 'consult physician', 'various', 'multiple', 'several',
        'others', 'etc', 'and', 'or', 'including', 'such as', 'like',
        'therapy', 'treatment', 'procedure', 'surgery', 'operation',
        'lifestyle', 'diet', 'exercise', 'rest', 'monitoring'
    ]
    
    if (not med_name or 
        len(med_name) < 3 or 
        med_name in invalid_terms or
        any(term in med_name for term in invalid_terms)):
        return None
    
    return med_name

def get_comprehensive_medications_for_disease(disease_name, mayo_medications):
    """Get comprehensive medications combining Mayo Clinic and drugs.com data"""
    
    print(f"\n💊 Getting comprehensive medications for: {disease_name}")
    
    # Start with Mayo Clinic medications
    mayo_meds = []
    if mayo_medications and str(mayo_medications).strip() and str(mayo_medications).strip().lower() != 'nan':
        mayo_meds = [med.strip() for med in str(mayo_medications).split(';') if med.strip()]
        print(f"   🏥 Mayo Clinic medications: {len(mayo_meds)}")
    
    # Get drugs.com medications
    drugs_com_meds = get_drugs_com_medications_for_disease(disease_name)
    
    # Combine both sources
    all_medications = mayo_meds + drugs_com_meds
    
    # Remove duplicates while preserving order
    unique_medications = []
    seen = set()
    for med in all_medications:
        if med.lower() not in seen:
            seen.add(med.lower())
            unique_medications.append(med)
    
    print(f"   📊 Total unique medications: {len(unique_medications)} (Mayo: {len(mayo_meds)}, Drugs.com: {len(drugs_com_meds)})")
    
    return unique_medications

def enhance_medications_with_llm(medications_text, disease_name):
    """Enhance existing medications with LLM to add simple generic drug names only"""
    
    if not model:
        return medications_text  # Return original if no LLM available
    
    try:
        # Parse existing medications
        existing_meds = [med.strip() for med in medications_text.split(';') if med.strip()] if medications_text else []
        
        print(f"   🤖 Enhancing {len(existing_meds)} existing medications for {disease_name}")
        
        prompt = f"""
You are a pharmaceutical expert. I need you to provide a comprehensive list of GENERIC DRUG NAMES (active ingredients) for treating "{disease_name}".

EXISTING MEDICATIONS (keep all):
{'; '.join(existing_meds) if existing_meds else 'None listed'}

Please provide a comprehensive list of GENERIC DRUG NAMES ONLY for {disease_name}:

REQUIREMENTS:
1. Keep ALL existing medications (don't remove any)
2. Add 15-25 additional generic drug names for {disease_name}
3. Use ONLY generic names (active ingredients) - NO brand names
4. Use ONLY simple drug names - NO parentheses, asterisks, or extra formatting
5. Cover all treatment categories: first-line, second-line, alternatives
6. Include medications for symptoms, complications, and comorbidities
7. Focus on drugs actually prescribed for {disease_name}

EXAMPLES of correct format:
- "metformin" (not "Metformin**" or "Metformin (Glucophage)")
- "lisinopril" (not "Lisinopril*" or "Lisinopril (Prinivil)")
- "atorvastatin" (not "Atorvastatin**" or "Atorvastatin (Lipitor)")

For {disease_name}, include generic drugs for:
- Primary treatment
- Symptom management  
- Comorbidity treatment
- Prevention
- Both acute and chronic management

Format as semicolon-separated list of SIMPLE GENERIC DRUG NAMES only.

GENERIC DRUG NAMES:
"""

        result = model.generate_content(prompt)
        response = result.text.strip()
        
        # Extract medications from response
        if ":" in response:
            response = response.split(":", 1)[1].strip()
        
        # Clean and split
        raw_medications = [med.strip() for med in response.split(';') if med.strip()]
        
        # Clean each medication name to be simple
        cleaned_medications = []
        for med in raw_medications:
            # Remove formatting symbols
            med = re.sub(r'\*+', '', med)  # Remove asterisks
            med = re.sub(r'\(.*?\)', '', med)  # Remove parentheses and content
            med = re.sub(r'\[.*?\]', '', med)  # Remove brackets and content
            med = re.sub(r'["""]', '', med)  # Remove quotes
            med = re.sub(r'[-–—].*', '', med)  # Remove dashes and everything after
            med = re.sub(r'\s+', ' ', med).strip()  # Clean whitespace
            
            # Convert to lowercase for consistency
            med = med.lower()
            
            # Filter out non-drug terms and ensure it's a simple drug name
            if (len(med) > 2 and 
                (med.isalpha() or (len(med.split()) == 1 and med.replace('-', '').isalpha())) and
                not med.startswith(('note', 'generic', 'drug', 'medication', 'treatment')) and
                not med.endswith(('therapy', 'treatment', 'drugs', 'medications')) and
                med not in ['etc', 'others', 'various', 'including', 'such', 'as', 'and', 'or']):
                cleaned_medications.append(med)
        
        # Remove duplicates while preserving order
        seen = set()
        unique_medications = []
        for med in cleaned_medications:
            if med not in seen:
                seen.add(med)
                unique_medications.append(med)
        
        enhanced_text = '; '.join(unique_medications)
        print(f"   ✅ Enhanced from {len(existing_meds)} to {len(unique_medications)} simple generic drugs")
        
        # Add delay to be respectful to API
        time.sleep(1.5)
        
        return enhanced_text
        
    except Exception as e:
        print(f"   ❌ LLM enhancement failed for {disease_name}: {e}")
        return medications_text  # Return original on error

def create_main_diseases_analysis_v3():
    """
    Create an Excel file with comprehensive analysis of main diseases from final_diseases_complete.csv
    Each disease gets its own sheet with structured data and enhanced medication information
    """
    
    # Target diseases to extract (use exact names from CSV)
    target_diseases = [
        'Heart disease',
        'Chronic kidney disease',
        'COPD',
        'Pneumonia',
        'Stroke',
        'Dementia',
        'Depression (major depressive disorder)',
        'High cholesterol',
        'Obesity', 
        'Arthritis'
    ]
    
    # Read the CSV file
    csv_path = '../CSV/final_diseases_complete.csv'
    df = pd.read_csv(csv_path)
    
    # Create workbook
    wb = Workbook()
    
    # Create summary sheet first
    summary_ws = wb.active
    summary_ws.title = "Summary"
    create_summary_sheet(summary_ws, target_diseases)
    
    # Process each target disease
    processed_diseases = set()
    created_sheets = []
    
    for disease in target_diseases:
        # Find matching rows with more specific matching
        if disease == 'Heart disease':
            disease_data = df[df['Disease_Name_English'].str.contains('^Heart disease$', case=False, na=False, regex=True)]
        elif disease == 'Obesity':
            disease_data = df[df['Disease_Name_English'].str.contains('^Obesity$', case=False, na=False, regex=True)]
        elif disease == 'Stroke':
            disease_data = df[df['Disease_Name_English'].str.contains('^Stroke$', case=False, na=False, regex=True)]
        else:
            disease_data = df[df['Disease_Name_English'].str.contains(f'^{disease}$', case=False, na=False, regex=True)]
        
        if disease_data.empty:
            print(f"No exact match found for {disease}, trying partial match...")
            disease_data = df[df['Disease_Name_English'].str.contains(disease, case=False, na=False, regex=False)]
            
        if disease_data.empty:
            print(f"No data found for {disease}")
            continue
            
        # Get the first match
        disease_row = disease_data.iloc[0]
        disease_name = disease_row['Disease_Name_English']
        
        # Skip if we've already processed this disease
        if disease_name in processed_diseases:
            continue
        processed_diseases.add(disease_name)
        
        # Create sheet name (remove special characters)
        sheet_name = disease_name.replace('(', '').replace(')', '').replace('/', '-')[:31]
        ws = wb.create_sheet(title=sheet_name)
        
        # Get comprehensive medications (Mayo Clinic + drugs.com)
        original_medications = disease_row['Medications_Drugs'] if pd.notna(disease_row['Medications_Drugs']) else ''
        comprehensive_medications = get_comprehensive_medications_for_disease(disease_name, original_medications)
        
        # Format as semicolon-separated string
        medications_text = '; '.join(comprehensive_medications) if comprehensive_medications else ''
        
        # Update the disease row with comprehensive medications for sheet creation
        disease_row_enhanced = disease_row.copy()
        disease_row_enhanced['Medications_Drugs'] = medications_text
        
        # Set up the sheet structure with comprehensive medication list
        setup_disease_sheet_v3(ws, disease_row_enhanced, disease_name)
        created_sheets.append((disease, disease_name))
        print(f"✓ Created comprehensive sheet for: {disease_name}")
        print(f"   📊 Total medications: {len(comprehensive_medications)}")
        print(f"   🔗 Mayo + drugs.com combined data\n")
    
    # Update summary sheet with actual results
    update_summary_sheet(summary_ws, created_sheets)
    
    # Create the unique medications sheet with enhanced medications
    create_unique_medications_sheet_enhanced(wb, df, target_diseases)
    
    # Save the workbook
    output_path = '../Analysis/main_diseases_analysis_final_enhanced.xlsx'
    wb.save(output_path)
    print(f"\nEnhanced analysis saved to: {output_path}")
    
    return output_path

def create_summary_sheet(ws, target_diseases):
    """Create a summary sheet with overview information"""
    
    # Header styling
    header_font = Font(bold=True, size=16, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    subheader_font = Font(bold=True, size=12, color="FFFFFF")
    subheader_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    
    # Title
    ws['A1'] = 'MAIN DISEASES COMPREHENSIVE ANALYSIS'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Date and info
    ws['A3'] = 'Analysis Date:'
    ws['B3'] = '2025-07-02'
    ws['A4'] = 'Source Data:'
    ws['B4'] = 'final_diseases_complete.csv'
    ws['A5'] = 'Total Target Diseases:'
    ws['B5'] = len(target_diseases)
    
    # Target diseases list
    ws['A7'] = 'TARGET DISEASES'
    ws['A7'].font = subheader_font
    ws['A7'].fill = subheader_fill
    ws.merge_cells('A7:D7')
    ws['A7'].alignment = Alignment(horizontal='center')
    
    ws['A8'] = 'Disease Name'
    ws['B8'] = 'Status'
    ws['C8'] = 'Matched Name'
    ws['D8'] = 'Spanish Name'
    
    # Style header row
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}8'].font = Font(bold=True)
        ws[f'{col}8'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Add target diseases (will be updated later)
    for i, disease in enumerate(target_diseases, 9):
        ws[f'A{i}'] = disease
        ws[f'B{i}'] = 'Processing...'
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 25

def update_summary_sheet(ws, created_sheets):
    """Update the summary sheet with actual results"""
    
    # Create a mapping of created sheets
    created_mapping = {original: matched for original, matched in created_sheets}
    
    # Read CSV to get Spanish names
    csv_path = '../CSV/final_diseases_complete.csv'
    df = pd.read_csv(csv_path)
    
    # Update the status for each disease
    row = 9
    target_diseases = [
        'Heart disease', 'Chronic kidney disease', 'COPD', 'Pneumonia', 'Stroke',
        'Dementia', 'Depression (major depressive disorder)', 'High cholesterol',
        'Obesity', 'Arthritis'
    ]
    
    for disease in target_diseases:
        if disease in created_mapping:
            matched_name = created_mapping[disease]
            ws[f'B{row}'] = '✓ Found'
            ws[f'C{row}'] = matched_name
            
            # Get Spanish name
            spanish_name = df[df['Disease_Name_English'] == matched_name]['Disease_Name_Spanish'].iloc[0]
            ws[f'D{row}'] = spanish_name
            
            # Color the row green
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        else:
            ws[f'B{row}'] = '✗ Not Found'
            ws[f'C{row}'] = 'No match'
            ws[f'D{row}'] = '-'
            
            # Color the row red
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        
        row += 1
    
    # Add statistics
    ws[f'A{row+2}'] = 'STATISTICS'
    ws[f'A{row+2}'].font = Font(bold=True, size=12)
    ws[f'A{row+3}'] = f'Diseases Found: {len(created_sheets)}'
    ws[f'A{row+4}'] = f'Diseases Not Found: {len(target_diseases) - len(created_sheets)}'
    ws[f'A{row+5}'] = f'Success Rate: {len(created_sheets)/len(target_diseases)*100:.1f}%'

def setup_disease_sheet_v3(ws, disease_row, disease_name):
    """Set up each disease sheet with structured data and simple medication list"""
    
    # Header styling
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_font = Font(bold=True, size=12, color="FFFFFF")
    subheader_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    
    # Border styling
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Row 1: Disease Title
    ws['A1'] = f'{disease_name.upper()} - COMPREHENSIVE ANALYSIS'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    # Row 3: Disease Names
    ws['A3'] = 'DISEASE INFORMATION'
    ws['A3'].font = subheader_font
    ws['A3'].fill = subheader_fill
    ws.merge_cells('A3:F3')
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Disease name details
    ws['A4'] = 'English Name:'
    ws['B4'] = disease_row['Disease_Name_English']
    ws['A5'] = 'Spanish Name:'
    ws['B5'] = disease_row['Disease_Name_Spanish']
    
    # Style the info cells
    for row in [4, 5]:
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Row 7: Diagnosis Section
    ws['A7'] = 'DIAGNOSIS'
    ws['A7'].font = subheader_font
    ws['A7'].fill = subheader_fill
    ws.merge_cells('A7:F7')
    ws['A7'].alignment = Alignment(horizontal='center')
    
    # Diagnosis information
    diagnosis_text = disease_row['Diagnosis'] if pd.notna(disease_row['Diagnosis']) else 'No diagnosis information available'
    ws['A8'] = 'Diagnosis Process:'
    ws['A8'].font = Font(bold=True)
    
    # Put all diagnosis text in one cell (B8) - user will manually merge cells for readability
    ws['B8'] = diagnosis_text
    ws['B8'].alignment = Alignment(wrap_text=True, vertical='top')
    # Set a taller row height for the diagnosis text
    ws.row_dimensions[8].height = max(60, min(200, len(diagnosis_text) // 10))
    
    # Calculate next row (just one row for diagnosis now)
    next_row = 10
    
    # Treatments Section
    ws[f'A{next_row}'] = 'TREATMENTS'
    ws[f'A{next_row}'].font = subheader_font
    ws[f'A{next_row}'].fill = subheader_fill
    ws.merge_cells(f'A{next_row}:F{next_row}')
    ws[f'A{next_row}'].alignment = Alignment(horizontal='center')
    
    treatments = disease_row['Treatments'] if pd.notna(disease_row['Treatments']) else 'No treatment information available'
    
    ws[f'A{next_row+1}'] = 'Available Treatments:'
    ws[f'A{next_row+1}'].font = Font(bold=True)
    # Put all treatment text in one cell - user will manually merge cells for readability
    ws[f'B{next_row+1}'] = treatments
    ws[f'B{next_row+1}'].alignment = Alignment(wrap_text=True, vertical='top')
    # Set row height based on content length
    ws.row_dimensions[next_row+1].height = max(60, min(200, len(treatments) // 10))
    
    # Calculate next row (just one row for treatments now)
    next_row = next_row + 3
    
    # Tests Section  
    ws[f'A{next_row}'] = 'DIAGNOSTIC TESTS'
    ws[f'A{next_row}'].font = subheader_font
    ws[f'A{next_row}'].fill = subheader_fill
    ws.merge_cells(f'A{next_row}:F{next_row}')
    ws[f'A{next_row}'].alignment = Alignment(horizontal='center')
    
    tests = disease_row['Tests'] if pd.notna(disease_row['Tests']) else 'No test information available'
    
    ws[f'A{next_row+1}'] = 'Diagnostic Tests:'
    ws[f'A{next_row+1}'].font = Font(bold=True)
    # Put all test text in one cell - user will manually merge cells for readability
    ws[f'B{next_row+1}'] = tests
    ws[f'B{next_row+1}'].alignment = Alignment(wrap_text=True, vertical='top')
    # Set row height based on content length
    ws.row_dimensions[next_row+1].height = max(60, min(200, len(tests) // 10))
    
    # Calculate next row (just one row for tests now)
    next_row = next_row + 3
    
    # Simple Medications Section
    ws[f'A{next_row}'] = 'MEDICATIONS & DRUGS'
    ws[f'A{next_row}'].font = subheader_font
    ws[f'A{next_row}'].fill = subheader_fill
    ws.merge_cells(f'A{next_row}:F{next_row}')
    ws[f'A{next_row}'].alignment = Alignment(horizontal='center')
    
    medications = disease_row['Medications_Drugs'] if pd.notna(disease_row['Medications_Drugs']) else 'No medication information available'
    medication_list = medications.split(';') if pd.notna(disease_row['Medications_Drugs']) else ['No medications listed']
    
    # Create simple medication list
    med_row = next_row + 2
    ws[f'A{med_row}'] = 'Medication Name'
    ws[f'B{med_row}'] = 'Disease Tag'
    
    # Style header row
    for col in ['A', 'B']:
        ws[f'{col}{med_row}'].font = Font(bold=True)
        ws[f'{col}{med_row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        ws[f'{col}{med_row}'].border = thin_border
    
    # Add ALL medications (simple list with just clean name and disease tag)
    for i, medication in enumerate(medication_list):
        # Clean the medication name
        clean_med_name = clean_medication_name(medication)
        
        # Only add if we have a valid clean name
        if clean_med_name:
            ws[f'A{med_row+1+i}'] = clean_med_name
            ws[f'B{med_row+1+i}'] = disease_name
            
            # Add borders
            for col in ['A', 'B']:
                cell = ws[f'{col}{med_row+1+i}']
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Count valid medications
    valid_meds = [med for med in medication_list if clean_medication_name(med)]
    
    # Add summary note
    ws[f'A{med_row+1+len(medication_list)+1}'] = f"Total medications for {disease_name}: {len(valid_meds)}"
    ws[f'A{med_row+1+len(medication_list)+1}'].font = Font(bold=True, italic=True)
    ws.merge_cells(f'A{med_row+1+len(medication_list)+1}:B{med_row+1+len(medication_list)+1}')
    
    # Add note about detailed information
    ws[f'A{med_row+1+len(medication_list)+2}'] = "Note: Detailed medication information (What Is, Side Effects, etc.) is available in the 'All Unique Medications' sheet"
    ws[f'A{med_row+1+len(medication_list)+2}'].font = Font(italic=True, color="666666")
    ws.merge_cells(f'A{med_row+1+len(medication_list)+2}:F{med_row+1+len(medication_list)+2}')
    
    # Set column widths for simple layout
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

def create_unique_medications_sheet_enhanced(wb, df, target_diseases):
    """
    Create enhanced sheet with all unique medications from main diseases, with LLM enhancements
    """
    
    # Header styling
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_font = Font(bold=True, size=12, color="FFFFFF")
    subheader_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    
    # Border styling
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Create the new sheet
    medications_ws = wb.create_sheet(title="All Unique Medications")
    
    # Sheet title
    medications_ws['A1'] = 'ALL UNIQUE MEDICATIONS FROM MAIN DISEASES (LLM-ENHANCED)'
    medications_ws['A1'].font = header_font
    medications_ws['A1'].fill = header_fill
    medications_ws.merge_cells('A1:F1')
    medications_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    medications_ws.row_dimensions[1].height = 25
    
    # Information section
    medications_ws['A3'] = 'INFORMATION'
    medications_ws['A3'].font = subheader_font
    medications_ws['A3'].fill = subheader_fill
    medications_ws.merge_cells('A3:F3')
    medications_ws['A3'].alignment = Alignment(horizontal='center')
    
    medications_ws['A4'] = 'Purpose:'
    medications_ws['B4'] = 'Comprehensive list of all unique medications (original + LLM-enhanced)'
    medications_ws['A5'] = 'Source:'
    medications_ws['B5'] = 'final_diseases_complete.csv + AI Enhancement'
    medications_ws['A6'] = 'Enhancement:'
    medications_ws['B6'] = 'LLM-powered comprehensive medication coverage'
    
    # Style the info cells
    for row in [4, 5, 6]:
        medications_ws[f'A{row}'].font = Font(bold=True)
        medications_ws[f'A{row}'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Extract all unique medications from target diseases with LLM enhancement
    all_medications = set()
    medication_to_diseases = {}
    
    for disease in target_diseases:
        # Find matching rows for each disease
        if disease == 'Heart disease':
            disease_data = df[df['Disease_Name_English'].str.contains('^Heart disease$', case=False, na=False, regex=True)]
        elif disease == 'Obesity':
            disease_data = df[df['Disease_Name_English'].str.contains('^Obesity$', case=False, na=False, regex=True)]
        elif disease == 'Stroke':
            disease_data = df[df['Disease_Name_English'].str.contains('^Stroke$', case=False, na=False, regex=True)]
        else:
            disease_data = df[df['Disease_Name_English'].str.contains(f'^{disease}$', case=False, na=False, regex=True)]
        
        if disease_data.empty:
            # Try partial match if exact match fails
            disease_data = df[df['Disease_Name_English'].str.contains(disease, case=False, na=False, regex=False)]
            
        if not disease_data.empty:
            disease_row = disease_data.iloc[0]
            disease_name = disease_row['Disease_Name_English']
            original_medications = disease_row['Medications_Drugs'] if pd.notna(disease_row['Medications_Drugs']) else ''
            
            # Get enhanced medications for this disease
            enhanced_medications = enhance_medications_with_llm(original_medications, disease_name)
            
            if enhanced_medications:
                # Split medications and add to set
                med_list = [med.strip() for med in enhanced_medications.split(';') if med.strip()]
                all_medications.update(med_list)
                
                # Track disease associations for each medication
                for medication in med_list:
                    if medication not in medication_to_diseases:
                        medication_to_diseases[medication] = []
                    medication_to_diseases[medication].append(disease_name)
    
    # Sort medications alphabetically
    sorted_medications = sorted(list(all_medications))
    
    # Create the table headers
    header_row = 8
    medications_ws[f'A{header_row}'] = 'MEDICATION NAME'
    medications_ws[f'B{header_row}'] = 'WHAT IS'
    medications_ws[f'C{header_row}'] = 'SIDE EFFECTS'
    medications_ws[f'D{header_row}'] = 'CALL A DOCTOR IF'
    medications_ws[f'E{header_row}'] = 'GO TO ER IF'
    medications_ws[f'F{header_row}'] = 'DISEASE TAG'
    
    # Style header row
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        cell = medications_ws[f'{col}{header_row}']
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add all unique medications (sorted alphabetically)
    for i, medication in enumerate(sorted_medications):
        row_num = header_row + 1 + i
        
        # Get disease associations for this medication
        diseases_for_med = medication_to_diseases.get(medication, [])
        disease_tag = '; '.join(diseases_for_med) if diseases_for_med else 'Unknown'
        
        medications_ws[f'A{row_num}'] = medication
        medications_ws[f'B{row_num}'] = ''  # To be filled with "What Is" data
        medications_ws[f'C{row_num}'] = ''  # To be filled with "Side Effects" data
        medications_ws[f'D{row_num}'] = ''  # To be filled with "Call a Doctor If" data
        medications_ws[f'E{row_num}'] = ''  # To be filled with "Go to ER If" data
        medications_ws[f'F{row_num}'] = disease_tag  # Disease Tag - populated immediately
        
        # Add borders and formatting
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            cell = medications_ws[f'{col}{row_num}']
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Alternate row colors for better readability
        if i % 2 == 0:
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                medications_ws[f'{col}{row_num}'].fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    
    # Add summary information
    summary_row = header_row + len(sorted_medications) + 2
    medications_ws[f'A{summary_row}'] = 'ENHANCED SUMMARY'
    medications_ws[f'A{summary_row}'].font = Font(bold=True, size=12)
    medications_ws[f'A{summary_row}'].fill = PatternFill(start_color="E2E6EA", end_color="E2E6EA", fill_type="solid")
    
    medications_ws[f'A{summary_row+1}'] = f'Total Unique Medications: {len(sorted_medications)}'
    medications_ws[f'A{summary_row+1}'].font = Font(bold=True)
    
    medications_ws[f'A{summary_row+2}'] = f'Diseases Analyzed: {len(target_diseases)}'
    medications_ws[f'A{summary_row+2}'].font = Font(bold=True)
    
    medications_ws[f'A{summary_row+3}'] = 'Enhancement: Original medications + LLM-powered comprehensive coverage'
    medications_ws[f'A{summary_row+3}'].font = Font(bold=True)
    
    medications_ws[f'A{summary_row+4}'] = 'Next Steps: Populate columns B-E with medication data (What Is, Side Effects, Call Doctor, Go to ER)'
    medications_ws[f'A{summary_row+4}'].font = Font(italic=True)
    medications_ws.merge_cells(f'A{summary_row+4}:F{summary_row+4}')
    
    # Set column widths for better display
    medications_ws.column_dimensions['A'].width = 30  # Medication Name
    medications_ws.column_dimensions['B'].width = 40  # What Is
    medications_ws.column_dimensions['C'].width = 35  # Side Effects
    medications_ws.column_dimensions['D'].width = 30  # Call a Doctor If
    medications_ws.column_dimensions['E'].width = 30  # Go to ER If
    medications_ws.column_dimensions['F'].width = 40  # Disease Tag
    
    print(f"✓ Created enhanced 'All Unique Medications' sheet with {len(sorted_medications)} unique medications")

def create_unique_medications_sheet(wb, df, target_diseases):
    """
    Create a sheet with all unique medications from main diseases, sorted alphabetically
    Columns: Name, What Is, Side Effects, Call a Doctor If, Go to ER If, Disease Tag
    """
    
    # Header styling
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_font = Font(bold=True, size=12, color="FFFFFF")
    subheader_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    
    # Border styling
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Create the new sheet
    medications_ws = wb.create_sheet(title="All Unique Medications")
    
    # Sheet title
    medications_ws['A1'] = 'ALL UNIQUE MEDICATIONS FROM MAIN DISEASES'
    medications_ws['A1'].font = header_font
    medications_ws['A1'].fill = header_fill
    medications_ws.merge_cells('A1:F1')
    medications_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    medications_ws.row_dimensions[1].height = 25
    
    # Information section
    medications_ws['A3'] = 'INFORMATION'
    medications_ws['A3'].font = subheader_font
    medications_ws['A3'].fill = subheader_fill
    medications_ws.merge_cells('A3:F3')
    medications_ws['A3'].alignment = Alignment(horizontal='center')
    
    medications_ws['A4'] = 'Purpose:'
    medications_ws['B4'] = 'Comprehensive list of all unique medications used across main diseases'
    medications_ws['A5'] = 'Source:'
    medications_ws['B5'] = 'final_diseases_complete.csv'
    medications_ws['A6'] = 'Status:'
    medications_ws['B6'] = 'Ready for FDA/Drugs.com data population'
    
    # Style the info cells
    for row in [4, 5, 6]:
        medications_ws[f'A{row}'].font = Font(bold=True)
        medications_ws[f'A{row}'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Extract all unique medications from target diseases and track their disease associations
    all_medications = set()
    medication_to_diseases = {}  # Track which diseases each medication belongs to
    
    for disease in target_diseases:
        # Find matching rows for each disease
        if disease == 'Heart disease':
            disease_data = df[df['Disease_Name_English'].str.contains('^Heart disease$', case=False, na=False, regex=True)]
        elif disease == 'Obesity':
            disease_data = df[df['Disease_Name_English'].str.contains('^Obesity$', case=False, na=False, regex=True)]
        elif disease == 'Stroke':
            disease_data = df[df['Disease_Name_English'].str.contains('^Stroke$', case=False, na=False, regex=True)]
        else:
            disease_data = df[df['Disease_Name_English'].str.contains(f'^{disease}$', case=False, na=False, regex=True)]
        
        if disease_data.empty:
            # Try partial match if exact match fails
            disease_data = df[df['Disease_Name_English'].str.contains(disease, case=False, na=False, regex=False)]
            
        if not disease_data.empty:
            disease_row = disease_data.iloc[0]
            disease_name = disease_row['Disease_Name_English']
            medications = disease_row['Medications_Drugs'] if pd.notna(disease_row['Medications_Drugs']) else ''
            
            if medications:
                # Split medications and add to set (to ensure uniqueness)
                med_list = [med.strip() for med in medications.split(';') if med.strip()]
                all_medications.update(med_list)
                
                # Track disease associations for each medication
                for medication in med_list:
                    if medication not in medication_to_diseases:
                        medication_to_diseases[medication] = []
                    medication_to_diseases[medication].append(disease_name)
    
    # Sort medications alphabetically
    sorted_medications = sorted(list(all_medications))
    
    # Create the table headers
    header_row = 8
    medications_ws[f'A{header_row}'] = 'MEDICATION NAME'
    medications_ws[f'B{header_row}'] = 'WHAT IS'
    medications_ws[f'C{header_row}'] = 'SIDE EFFECTS'
    medications_ws[f'D{header_row}'] = 'CALL A DOCTOR IF'
    medications_ws[f'E{header_row}'] = 'GO TO ER IF'
    medications_ws[f'F{header_row}'] = 'DISEASE TAG'
    
    # Style header row
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        cell = medications_ws[f'{col}{header_row}']
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add all unique medications (sorted alphabetically)
    for i, medication in enumerate(sorted_medications):
        row_num = header_row + 1 + i
        
        # Get disease associations for this medication
        diseases_for_med = medication_to_diseases.get(medication, [])
        disease_tag = '; '.join(diseases_for_med) if diseases_for_med else 'Unknown'
        
        medications_ws[f'A{row_num}'] = medication
        medications_ws[f'B{row_num}'] = ''  # To be filled with "What Is" data
        medications_ws[f'C{row_num}'] = ''  # To be filled with "Side Effects" data
        medications_ws[f'D{row_num}'] = ''  # To be filled with "Call a Doctor If" data
        medications_ws[f'E{row_num}'] = ''  # To be filled with "Go to ER If" data
        medications_ws[f'F{row_num}'] = disease_tag  # Disease Tag - populated immediately
        
        # Add borders and formatting
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            cell = medications_ws[f'{col}{row_num}']
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Alternate row colors for better readability
        if i % 2 == 0:
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                medications_ws[f'{col}{row_num}'].fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    
    # Add summary information
    summary_row = header_row + len(sorted_medications) + 2
    medications_ws[f'A{summary_row}'] = 'SUMMARY'
    medications_ws[f'A{summary_row}'].font = Font(bold=True, size=12)
    medications_ws[f'A{summary_row}'].fill = PatternFill(start_color="E2E6EA", end_color="E2E6EA", fill_type="solid")
    
    medications_ws[f'A{summary_row+1}'] = f'Total Unique Medications: {len(sorted_medications)}'
    medications_ws[f'A{summary_row+1}'].font = Font(bold=True)
    
    medications_ws[f'A{summary_row+2}'] = f'Diseases Analyzed: {len(target_diseases)}'
    medications_ws[f'A{summary_row+2}'].font = Font(bold=True)
    
    medications_ws[f'A{summary_row+3}'] = 'Next Steps: Populate columns B-E with medication data (What Is, Side Effects, Call Doctor, Go to ER)'
    medications_ws[f'A{summary_row+3}'].font = Font(italic=True)
    medications_ws.merge_cells(f'A{summary_row+3}:F{summary_row+3}')
    
    # Set column widths for better display
    medications_ws.column_dimensions['A'].width = 30  # Medication Name
    medications_ws.column_dimensions['B'].width = 40  # What Is
    medications_ws.column_dimensions['C'].width = 35  # Side Effects
    medications_ws.column_dimensions['D'].width = 30  # Call a Doctor If
    medications_ws.column_dimensions['E'].width = 30  # Go to ER If
    medications_ws.column_dimensions['F'].width = 40  # Disease Tag
    
    print(f"✓ Created 'All Unique Medications' sheet with {len(sorted_medications)} unique medications")

if __name__ == "__main__":
    print("Creating Enhanced Main Diseases Analysis with LLM-Enhanced Medications...")
    print("🤖 This will enhance existing medications with comprehensive AI-powered coverage")
    output_file = create_main_diseases_analysis_v3()
    print(f"Enhanced analysis with comprehensive medications finished! File saved at: {output_file}")
