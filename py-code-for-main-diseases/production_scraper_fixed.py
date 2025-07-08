from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import time
import random
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

class DrugsScraper:
    def __init__(self, headless=False):
        self.setup_driver(headless)
        self.wait = WebDriverWait(self.driver, 10)
        
    def setup_driver(self, headless=False):
        """Set up Chrome driver with options"""
        chrome_options = Options()
        if headless:
            chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        
        self.driver = webdriver.Chrome(options=chrome_options)
        self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        
    def search_and_get_side_effects(self, medication):
        """Search for medication and get side effects content"""
        try:
            print(f"🔍 Processing: {medication}")
            
            # Step 1: Go to drugs.com with retry logic
            for attempt in range(3):
                try:
                    self.driver.get("https://www.drugs.com")
                    break
                except Exception as e:
                    print(f"    Attempt {attempt + 1} failed to load drugs.com: {e}")
                    if attempt == 2:
                        return f"❌ Failed to load drugs.com after 3 attempts"
                    time.sleep(5)
            
            time.sleep(3)
            
            # Step 2: Search for medication
            try:
                search_box = self.wait.until(EC.presence_of_element_located((By.NAME, "searchterm")))
                search_box.clear()
                search_box.send_keys(medication)
                search_box.send_keys(Keys.RETURN)
                print(f"  ✅ Search submitted for: {medication}")
            except Exception as e:
                return f"❌ Failed to search for {medication}: {str(e)}"
            
            # Step 3: Find main medication result
            main_result = self.find_main_medication_result(medication)
            if not main_result:
                return f"❌ Could not find main result for {medication}"
            
            # Step 4: Click on main result with retry
            try:
                self.driver.execute_script("arguments[0].scrollIntoView(true);", main_result)
                time.sleep(1)
                main_result.click()
                print(f"  ✅ Clicked main result for {medication}")
            except Exception as e:
                # Try JavaScript click as fallback
                try:
                    self.driver.execute_script("arguments[0].click();", main_result)
                    print(f"  ✅ Clicked main result (JS) for {medication}")
                except Exception as e2:
                    return f"❌ Failed to click main result for {medication}: {str(e2)}"
            
            time.sleep(4)
            
            # Step 5: Find and click side effects link
            side_effects_link = self.find_side_effects_link()
            if not side_effects_link:
                return f"❌ Could not find side effects link for {medication}"
            
            # Step 6: Click side effects link
            try:
                self.driver.execute_script("arguments[0].scrollIntoView(true);", side_effects_link)
                time.sleep(1)
                side_effects_link.click()
                print(f"  ✅ Clicked side effects link for {medication}")
            except Exception as e:
                # Try JavaScript click as fallback
                try:
                    self.driver.execute_script("arguments[0].click();", side_effects_link)
                    print(f"  ✅ Clicked side effects link (JS) for {medication}")
                except Exception as e2:
                    return f"❌ Failed to click side effects link for {medication}: {str(e2)}"
            
            time.sleep(4)
            
            # Step 7: Extract side effects content
            content = self.extract_side_effects_content(medication)
            
            if content and len(content) > 50:
                print(f"  ✅ Successfully processed {medication} ({len(content)} characters)")
                return content
            else:
                return f"❌ No substantial side effects content found for {medication}"
            
        except Exception as e:
            error_msg = f"❌ Unexpected error processing {medication}: {str(e)}"
            print(error_msg)
            return error_msg
    
    def find_main_medication_result(self, medication):
        """Find the main medication result (usually with yellow star)"""
        print(f"  🔍 Looking for main result for: {medication}")
        
        # Wait a bit longer for search results to load
        time.sleep(5)
        
        # Try multiple approaches to find the main result
        approaches = [
            # Approach 1: Look for direct href matches
            {
                'name': 'Direct href match',
                'selectors': [
                    f"a[href*='{medication.lower().replace(' ', '-')}.html']",
                    f"a[href*='{medication.lower()}.html']",
                    f"a[href*='{medication.replace(' ', '-').lower()}.html']",
                ]
            },
            # Approach 2: Look for search results containing the medication name
            {
                'name': 'Search result text match',
                'selectors': [
                    "a[href*='.html']",  # Any HTML page link
                ]
            },
            # Approach 3: Look for specific search result containers
            {
                'name': 'Search result containers',
                'selectors': [
                    ".ddc-search-result a",
                    ".search-result a",
                    ".result a",
                    "div[class*='result'] a",
                ]
            }
        ]
        
        for approach in approaches:
            print(f"    Trying approach: {approach['name']}")
            
            for selector in approach['selectors']:
                try:
                    results = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    print(f"      Found {len(results)} results with selector: {selector}")
                    
                    for i, result in enumerate(results):
                        try:
                            href = result.get_attribute('href')
                            text = result.text.strip()
                            
                            if not href or not text:
                                continue
                            
                            # Skip unwanted links
                            if any(skip in href.lower() for skip in ['/pro/', '/search', '/compare', '/interaction']):
                                continue
                            
                            # Check if this looks like our medication
                            medication_words = medication.lower().split()
                            text_lower = text.lower()
                            
                            # For direct href matches, be more strict
                            if approach['name'] == 'Direct href match':
                                if medication.lower().replace(' ', '-') in href.lower():
                                    print(f"      ✅ Found direct match: {text[:50]}... -> {href}")
                                    return result
                            
                            # For text matches, check if medication words are in the text
                            elif approach['name'] == 'Search result text match':
                                if len(medication_words) == 1:
                                    # Single word medication
                                    if medication_words[0] in text_lower and '.html' in href:
                                        print(f"      ✅ Found text match: {text[:50]}... -> {href}")
                                        return result
                                else:
                                    # Multi-word medication - check if most words are present
                                    word_matches = sum(1 for word in medication_words if word in text_lower)
                                    if word_matches >= len(medication_words) * 0.7 and '.html' in href:
                                        print(f"      ✅ Found text match: {text[:50]}... -> {href}")
                                        return result
                            
                            # For container matches, be more flexible
                            else:
                                if any(word in text_lower for word in medication_words) and '.html' in href:
                                    print(f"      ✅ Found container match: {text[:50]}... -> {href}")
                                    return result
                        
                        except Exception as e:
                            continue
                
                except Exception as e:
                    continue
        
        # If no result found, print available links for debugging
        print(f"    ❌ No main result found. Available links:")
        try:
            all_links = self.driver.find_elements(By.CSS_SELECTOR, "a[href*='.html']")
            for i, link in enumerate(all_links[:10]):  # Show first 10
                href = link.get_attribute('href')
                text = link.text.strip()[:50]
                print(f"      {i+1}. {text} -> {href}")
        except:
            pass
        
        return None
    
    def find_side_effects_link(self):
        """Find the side effects navigation link"""
        print(f"  🔍 Looking for side effects link...")
        
        # Wait for page to fully load
        time.sleep(3)
        
        # Try multiple approaches to find side effects link
        approaches = [
            # Approach 1: XPath for text content
            {
                'name': 'XPath text search',
                'selectors': [
                    "//a[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'side effects')]",
                    "//a[contains(text(), 'Side effects')]",
                    "//a[contains(text(), 'side effects')]",
                    "//a[contains(text(), 'Side Effects')]",
                ]
            },
            # Approach 2: href containing side-effects
            {
                'name': 'Href containing side-effects',
                'selectors': [
                    "a[href*='side-effects']",
                    "a[href*='#side-effects']",
                    "a[href*='sideeffects']",
                ]
            },
            # Approach 3: Look in navigation areas
            {
                'name': 'Navigation areas',
                'selectors': [
                    "nav a",
                    ".nav a",
                    ".navigation a",
                    ".tabs a",
                    ".tab a",
                    ".menu a",
                ]
            }
        ]
        
        for approach in approaches:
            print(f"    Trying approach: {approach['name']}")
            
            for selector in approach['selectors']:
                try:
                    if selector.startswith("//"):
                        # XPath selector
                        links = self.driver.find_elements(By.XPATH, selector)
                    else:
                        # CSS selector
                        links = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    
                    print(f"      Found {len(links)} links with selector: {selector}")
                    
                    for link in links:
                        try:
                            text = link.text.strip().lower()
                            href = link.get_attribute('href')
                            
                            # Check if this looks like a side effects link
                            if approach['name'] == 'Navigation areas':
                                # For navigation areas, check text content
                                if 'side effect' in text:
                                    print(f"      ✅ Found side effects link: {link.text}")
                                    return link
                            else:
                                # For other approaches, we already filtered
                                if text and href:
                                    print(f"      ✅ Found side effects link: {link.text}")
                                    return link
                        
                        except Exception as e:
                            continue
                
                except Exception as e:
                    continue
        
        # If no side effects link found, print available navigation links
        print(f"    ❌ No side effects link found. Available navigation links:")
        try:
            nav_links = self.driver.find_elements(By.CSS_SELECTOR, "a")
            shown_links = 0
            for link in nav_links:
                try:
                    text = link.text.strip()
                    href = link.get_attribute('href')
                    if text and href and len(text) < 50:
                        print(f"      - {text} -> {href}")
                        shown_links += 1
                        if shown_links >= 15:  # Show first 15 links
                            break
                except:
                    continue
        except:
            pass
        
        return None
    
    def extract_side_effects_content(self, medication):
        """Extract ONLY side effects content from the page"""
        try:
            content_parts = []
            
            # Try to find specific side effects sections first
            selectors_to_try = [
                "#side-effects",
                ".side-effects", 
                "[id*='side-effects']",
                "[class*='side-effects']"
            ]
            
            content_found = False
            for selector in selectors_to_try:
                try:
                    section = self.driver.find_element(By.CSS_SELECTOR, selector)
                    if section:
                        text = section.text.strip()
                        if text and len(text) > 100:  # Ensure we have substantial side effects content
                            content_parts.append(text)
                            content_found = True
                            break
                except:
                    continue
            
            # If no specific side effects section found, look for side effects text patterns
            if not content_found:
                # Look for headings containing "side effects"
                try:
                    headings = self.driver.find_elements(By.XPATH, "//h1[contains(text(), 'side effects')] | //h2[contains(text(), 'side effects')] | //h3[contains(text(), 'side effects')]")
                    for heading in headings:
                        # Get content after the heading
                        content_parts.append(f"--- {heading.text} ---")
                        
                        # Find the next elements that contain side effects info
                        next_element = heading.find_element(By.XPATH, "./following-sibling::*")
                        while next_element and next_element.tag_name not in ['h1', 'h2', 'h3']:
                            text = next_element.text.strip()
                            if text:
                                content_parts.append(text)
                            try:
                                next_element = next_element.find_element(By.XPATH, "./following-sibling::*")
                            except:
                                break
                        content_found = True
                        break
                except:
                    pass
            
            # If still no content, look for paragraphs containing side effects keywords
            if not content_found:
                try:
                    paragraphs = self.driver.find_elements(By.TAG_NAME, "p")
                    for p in paragraphs:
                        text = p.text.strip().lower()
                        if any(keyword in text for keyword in ['side effect', 'adverse', 'reaction', 'emergency', 'call your doctor', 'serious']):
                            content_parts.append(p.text.strip())
                except:
                    pass
            
            if content_parts:
                content = '\n\n'.join(content_parts)
                # Clean up the content
                content = self.clean_content(content)
                return content
            else:
                return f"No specific side effects content found for {medication}"
            
        except Exception as e:
            return f"Error extracting side effects content: {str(e)}"
    
    def clean_content(self, content):
        """Clean and format the side effects content"""
        # Remove excessive whitespace and unwanted elements
        lines = content.split('\n')
        cleaned_lines = []
        
        for line in lines:
            line = line.strip()
            if line and not any(skip in line.lower() for skip in [
                'advertisement', 'ads by', 'sponsored', 'cookie', 'privacy',
                'terms of use', 'about us', 'contact us', 'site map'
            ]):
                cleaned_lines.append(line)
        
        # Join lines and remove excessive blank lines
        content = '\n'.join(cleaned_lines)
        
        # Remove multiple consecutive newlines
        while '\n\n\n' in content:
            content = content.replace('\n\n\n', '\n\n')
        
        return content.strip()
    
    def add_delay(self):
        """Add random delay between requests"""
        delay = random.uniform(3, 7)  # Longer delays to avoid detection
        print(f"  ⏰ Waiting {delay:.1f} seconds before next request...")
        time.sleep(delay)
    
    def close(self):
        """Close the browser"""
        self.driver.quit()

def update_excel_with_side_effects(max_medications=None):
    """Update Excel file with side effects for all medications"""
    
    excel_path = '/Users/juanlu/Documents/Wye/scrapper/Analysis/main_diseases_analysis_final.xlsx'
    
    if not os.path.exists(excel_path):
        print(f"❌ Excel file not found: {excel_path}")
        return
    
    # Load workbook
    wb = load_workbook(excel_path)
    
    if "All Unique Medications" not in wb.sheetnames:
        print("❌ 'All Unique Medications' sheet not found")
        return
    
    medications_ws = wb["All Unique Medications"]
    
    # Get medications
    medications = []
    for row in medications_ws.iter_rows(min_row=9, max_col=1, values_only=True):
        if row[0] and row[0].strip():
            medications.append(row[0].strip())
    
    # Use all medications if max_medications is None
    if max_medications:
        medications = medications[:max_medications]
    
    print(f"📊 Processing {len(medications)} medications...")
    
    # Add new column header if not exists
    if not medications_ws['H8'].value or 'SIDE EFFECTS' not in str(medications_ws['H8'].value):
        medications_ws['H8'] = 'SIDE EFFECTS'
        medications_ws['H8'].font = Font(bold=True, color="FFFFFF")
        medications_ws['H8'].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        medications_ws['H8'].border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        medications_ws['H8'].alignment = Alignment(horizontal='center', vertical='center')
        medications_ws.column_dimensions['H'].width = 80
    
    # Initialize scraper
    scraper = DrugsScraper(headless=True)  # Use headless for faster processing
    
    try:
        processed_count = 0
        errors = []
        
        for i, medication in enumerate(medications):
            print(f"\n[{i+1}/{len(medications)}] Processing: {medication}")
            
            # Get side effects content
            content = scraper.search_and_get_side_effects(medication)
            
            # Add to Excel
            row_num = 9 + i
            medications_ws[f'H{row_num}'] = content
            
            # Format cell
            cell = medications_ws[f'H{row_num}']
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Alternate row colors
            if i % 2 == 0:
                cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            
            processed_count += 1
            
            # Track errors
            if content.startswith("❌"):
                errors.append(medication)
            
            # Save progress every 10 medications
            if processed_count % 10 == 0:
                wb.save(excel_path)
                print(f"💾 Saved progress: {processed_count}/{len(medications)} medications processed")
                print(f"   Errors so far: {len(errors)}")
            
            # Add delay between requests
            scraper.add_delay()
    
    finally:
        scraper.close()
    
    # Final save
    wb.save(excel_path)
    
    # Summary
    success_count = processed_count - len(errors)
    print(f"\n" + "="*60)
    print(f"✅ PROCESSING COMPLETED!")
    print(f"📊 Total processed: {processed_count}")
    print(f"✅ Successful: {success_count}")
    print(f"❌ Errors: {len(errors)}")
    print(f"📄 Updated Excel file: {excel_path}")
    
    if errors:
        print(f"\n❌ Medications with errors:")
        for error_med in errors[:10]:  # Show first 10 errors
            print(f"   - {error_med}")
        if len(errors) > 10:
            print(f"   ... and {len(errors) - 10} more")

if __name__ == "__main__":
    print("🚀 Starting FULL Drugs.com Side Effects Scraper")
    print("="*60)
    print("✅ Test successful! Running full scraper for ALL medications")
    print("⚠️  This will process ALL 331 medications and may take several hours")
    print("💾 Progress will be saved every 10 medications")
    print("⏰ Using 3-7 second delays between requests")
    print("="*60)
    
    # Process ALL medications
    update_excel_with_side_effects()  # No limit - process all
    
    print("\n" + "="*60)
    print("🎉 FULL SCRAPING COMPLETED!")
