from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import time
import random
import re
from openpyxl import load_workbook
import os
import shutil
import google.generativeai as genai
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv('/Users/juanlu/Documents/Wye/scrapper/.env')

class DrugsScraper:
    def __init__(self, headless=False):
        self.headless = headless
        self.driver = None
        self.wait = None
        
        # Configure Google Generative AI
        api_key = os.getenv('GOOGLE_GEMINI_API_KEY')
        if not api_key:
            raise ValueError("GOOGLE_GEMINI_API_KEY not found in environment variables. Please check your .env file.")
        
        genai.configure(api_key=api_key)
        self.model = genai.GenerativeModel("gemini-1.5-flash")
        
        self.init_driver()
        
    def init_driver(self):
        """Initialize or reinitialize the driver"""
        if self.driver:
            try:
                self.driver.quit()
            except:
                pass
        
        self.driver = self.setup_driver(self.headless)
        self.wait = WebDriverWait(self.driver, 10)
        
    def setup_driver(self, headless=False):
        """Set up Chrome driver with options"""
        chrome_options = Options()
        if headless:
            chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_argument("--disable-web-security")
        chrome_options.add_argument("--disable-features=VizDisplayCompositor")
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        
        # Add user agent to avoid detection
        chrome_options.add_argument("--user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
        
        try:
            driver = webdriver.Chrome(options=chrome_options)
            driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            
            # Ensure we start with a valid page
            driver.get("https://www.drugs.com")
            time.sleep(2)
            
            return driver
        except Exception as e:
            print(f"❌ Error setting up Chrome driver: {e}")
            raise e

    def ensure_valid_page(self):
        """Ensure we're on a valid drugs.com page"""
        try:
            current_url = self.driver.current_url
            
            # If we're on data:, or invalid page, go to drugs.com
            if current_url.startswith("data:") or "drugs.com" not in current_url or current_url == "about:blank":
                print(f"  🔄 Invalid page detected ({current_url}), navigating to drugs.com...")
                self.driver.get("https://www.drugs.com")
                time.sleep(3)
                
                # Verify the navigation worked
                new_url = self.driver.current_url
                if new_url.startswith("data:") or "drugs.com" not in new_url:
                    raise Exception(f"Failed to navigate to valid page, still on: {new_url}")
                
                print(f"  ✅ Successfully navigated to: {new_url}")
                
        except Exception as e:
            print(f"  ❌ Error ensuring valid page: {e}")
            raise e
    
    def check_connection(self):
        """Check if the driver connection is still alive"""
        try:
            # Try to get current URL - if this fails, connection is lost
            current_url = self.driver.current_url
            
            # Check if we're on a valid page (not data:, or empty)
            if current_url.startswith("data:") or not current_url or current_url == "about:blank":
                print(f"  ⚠️  Invalid URL detected: {current_url}")
                return False
                
            return True
        except Exception as e:
            print(f"  ⚠️  Connection lost: {e}")
            return False
    
    def close_modal_popups(self):
        """Close any modal popups that might be blocking the page with timeout protection"""
        try:
            # Quick and aggressive approach: send multiple escape keys
            from selenium.webdriver.common.keys import Keys
            body = self.driver.find_element(By.TAG_NAME, "body")
            
            # Send multiple escape keys to be extra sure
            body.send_keys(Keys.ESCAPE)
            time.sleep(0.1)
            body.send_keys(Keys.ESCAPE)
            time.sleep(0.1)
            body.send_keys(Keys.ESCAPE)  # Third escape for persistent popups
            print(f"  ✅ Sent multiple Escape keys to close modals")
            time.sleep(0.3)  # Very short wait
            
            # Try to click any obvious close buttons quickly
            quick_selectors = [
                "button[aria-label*='close']",
                "button[aria-label*='Close']", 
                ".close-button",
                "button.close",
                "[data-dismiss='modal']",
                ".modal-close",
                ".popup-close"
            ]
            
            for selector in quick_selectors[:4]:  # Try first 4 selectors
                try:
                    buttons = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    for button in buttons[:1]:  # Only try first button
                        try:
                            if button.is_displayed():
                                button.click()
                                print(f"  ✅ Closed modal with {selector}")
                                time.sleep(0.2)
                                break
                        except:
                            continue
                    if buttons:  # If we found buttons, stop trying other selectors
                        break
                except:
                    continue
                    
        except Exception as e:
            # Don't let modal closing errors stop the process
            pass
    
    def extract_what_is_info(self, medication):
        """Extract 'What Is' information from the main medication page"""
        try:
            print(f"    📋 Extracting 'What Is' information for {medication}...")
            
            # Close any popups first
            self.close_modal_popups()
            
            what_is_content = ""
            
            # Strategy 1: Look for "What is [drug_name]?" section specifically
            try:
                # Try to find headings that match "What is [drug_name]?"
                headings_to_try = [
                    f"What is {medication}?",
                    f"What is {medication.lower()}?",
                    f"What is {medication.capitalize()}?",
                    "What is this medicine?",
                    "What is this medication?",
                    "What is this drug?"
                ]
                
                for heading_text in headings_to_try:
                    try:
                        # Try XPath to find heading with specific text
                        xpath = f"//h1[contains(text(), '{heading_text}')] | //h2[contains(text(), '{heading_text}')] | //h3[contains(text(), '{heading_text}')]"
                        headings = self.driver.find_elements(By.XPATH, xpath)
                        
                        if headings:
                            print(f"      ✅ Found 'What is' heading: {heading_text}")
                            
                            # Get the next sibling elements (paragraphs) after this heading
                            heading = headings[0]
                            following_elements = heading.find_elements(By.XPATH, "./following-sibling::p | ./following-sibling::div//p")
                            
                            for elem in following_elements[:3]:  # Get first 3 paragraphs
                                text = elem.text.strip()
                                if len(text) > 30:  # Only meaningful paragraphs
                                    what_is_content += text + " "
                                    
                            if what_is_content:
                                break
                                
                    except Exception:
                        continue
                        
            except Exception as e:
                print(f"      ⚠️ Error in strategy 1: {e}")
            
            # Strategy 2: Look for drug overview/description sections
            if not what_is_content:
                try:
                    print(f"      🔍 Strategy 2: Looking for overview sections...")
                    
                    overview_selectors = [
                        ".drug-overview",
                        ".drug-description", 
                        ".medication-overview",
                        ".drug-summary",
                        "[class*='overview']",
                        "[class*='description']",
                        "[class*='summary']",
                        ".content-body p:first-of-type",
                        ".drug-info p:first-of-type",
                        "main p:first-of-type"
                    ]
                    
                    for selector in overview_selectors:
                        try:
                            elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                            
                            for element in elements[:2]:  # Only check first 2 elements
                                text = element.text.strip()
                                
                                # Check if this looks like a drug description
                                if (len(text) > 50 and 
                                    any(keyword in text.lower() for keyword in [
                                        'used to treat', 'used for', 'medication', 'drug', 'prescribed',
                                        'belongs to', 'class of', 'works by', 'helps', 'treats',
                                        medication.lower().split()[0] if medication else ''
                                    ])):
                                    
                                    what_is_content += text + " "
                                    print(f"      ✅ Found description from {selector}")
                                    break
                                    
                            if what_is_content:
                                break
                                
                        except Exception:
                            continue
                            
                except Exception as e:
                    print(f"      ⚠️ Error in strategy 2: {e}")
            
            # Strategy 3: Look for first meaningful paragraphs on the page
            if not what_is_content:
                try:
                    print(f"      🔍 Strategy 3: Getting first meaningful paragraphs...")
                    
                    # Get all paragraphs and find the first ones that seem descriptive
                    paragraphs = self.driver.find_elements(By.CSS_SELECTOR, "p")
                    
                    for paragraph in paragraphs[:10]:  # Check first 10 paragraphs
                        try:
                            text = paragraph.text.strip()
                            
                            # Skip navigation, footer, and other non-content paragraphs
                            if (len(text) > 40 and 
                                not any(skip_word in text.lower() for skip_word in [
                                    'cookie', 'privacy', 'navigation', 'menu', 'search',
                                    'copyright', 'terms', 'conditions', 'policy'
                                ]) and
                                any(drug_keyword in text.lower() for drug_keyword in [
                                    'medication', 'drug', 'medicine', 'treatment', 'prescribed',
                                    'used to', 'treats', 'helps', medication.lower().split()[0] if medication else ''
                                ])):
                                
                                what_is_content += text + " "
                                print(f"      ✅ Added paragraph: {text[:50]}...")
                                
                                # If we have enough content, stop
                                if len(what_is_content) > 200:
                                    break
                                    
                        except Exception:
                            continue
                            
                except Exception as e:
                    print(f"      ⚠️ Error in strategy 3: {e}")
            
            # Clean and format the content
            if what_is_content:
                # Remove extra whitespace and limit length
                what_is_content = re.sub(r'\s+', ' ', what_is_content).strip()
                
                # Limit to a reasonable length for Excel
                if len(what_is_content) > 500:
                    # Try to cut at sentence boundary
                    sentences = what_is_content.split('. ')
                    truncated = ""
                    for sentence in sentences:
                        if len(truncated + sentence + '. ') <= 500:
                            truncated += sentence + '. '
                        else:
                            break
                    what_is_content = truncated.strip() if truncated else what_is_content[:500] + "..."
                
                print(f"    ✅ Extracted 'What Is' info: {len(what_is_content)} characters")
                return what_is_content
            
            # Fallback: try to get any descriptive text from the page
            try:
                print(f"      🔍 Fallback: Looking for any descriptive content...")
                body_text = self.driver.find_element(By.TAG_NAME, "body").text
                lines = body_text.split('\n')
                
                for line in lines[:30]:  # Check first 30 lines
                    line = line.strip()
                    if (len(line) > 50 and 
                        any(keyword in line.lower() for keyword in [
                            'used to treat', 'used for', 'medication', 'drug', 'prescribed'
                        ])):
                        return line[:500]
                        
            except Exception:
                pass
            
            print(f"    ⚠️ Could not extract 'What Is' information for {medication}")
            return f"Description not available for {medication}"
            
        except Exception as e:
            print(f"    ❌ Error extracting 'What Is' info for {medication}: {e}")
            return f"Error extracting description for {medication}"
    
    def search_and_get_side_effects(self, medication):
        """Search for medication and get side effects content with LLM processing"""
        try:
            print(f"🔍 Processing: {medication}")
            
            # Ensure we start with a valid page
            self.ensure_valid_page()
            
            # Close any modals that might be open
            self.close_modal_popups()
            
            # Step 1: Go to drugs.com
            self.driver.get("https://www.drugs.com")
            time.sleep(2)
            self.close_modal_popups()
            
            # Step 2: Search for medication
            search_box = self.wait.until(EC.presence_of_element_located((By.NAME, "searchterm")))
            search_box.clear()
            search_box.send_keys(medication)
            search_box.send_keys(Keys.RETURN)
            print(f"  ✅ Search submitted for: {medication}")
            
            # Step 3: Find main medication result
            self.close_modal_popups()  # Close popups before searching
            main_result = self.find_main_medication_result(medication)
            if not main_result:
                return f"❌ Could not find main result for {medication}"
            
            # Step 4: Click on main result
            try:
                self.close_modal_popups()
                main_result.click()
                print(f"  ✅ Clicked main result for {medication}")
                time.sleep(1)
                self.close_modal_popups()
            except Exception as e:
                try:
                    self.driver.execute_script("arguments[0].click();", main_result)
                    print(f"  ✅ Clicked main result (JS) for {medication}")
                    time.sleep(1)
                    self.close_modal_popups()
                except Exception as e2:
                    return f"❌ Failed to click main result for {medication}: {str(e2)}"
            
            # Step 4.5: Extract "What Is" information from main page before going to side effects
            what_is_info = self.extract_what_is_info(medication)
            
            # Step 5: Find and click side effects link
            self.close_modal_popups()  # Close popups before searching for side effects link
            side_effects_link = self.find_side_effects_link()
            if not side_effects_link:
                return f"❌ Could not find side effects link for {medication}"
            
            # Step 6: Click side effects link
            try:
                self.close_modal_popups()  # Close popups before clicking
                side_effects_link.click()
                print(f"  ✅ Clicked side effects link for {medication}")
                time.sleep(1)
                self.close_modal_popups()  # Close popups immediately after click
                time.sleep(1)
                self.close_modal_popups()  # Close popups again to be extra sure
            except Exception as e:
                try:
                    self.driver.execute_script("arguments[0].click();", side_effects_link)
                    print(f"  ✅ Clicked side effects link (JS) for {medication}")
                    time.sleep(1)
                    self.close_modal_popups()  # Close popups immediately after JS click
                    time.sleep(1)
                    self.close_modal_popups()  # Close popups again to be extra sure
                except Exception as e2:
                    return f"❌ Failed to click side effects link for {medication}: {str(e2)}"
            
            # Step 7: Extract comprehensive side effects content with timeout protection
            print(f"  📝 Extracting comprehensive side effects content...")
            
            # Aggressive popup closing after page load
            self.close_modal_popups()  # First close
            time.sleep(0.5)
            self.close_modal_popups()  # Second close
            time.sleep(0.5)  
            self.close_modal_popups()  # Third close to be extra sure
            
            comprehensive_content = self.extract_comprehensive_side_effects(medication)
            
            # Quick sanity check
            if not comprehensive_content or len(comprehensive_content) < 50:
                print(f"  ⚠️ Extraction returned minimal content, attempting quick recovery...")
                comprehensive_content = self.extract_comprehensive_side_effects_quick(medication)
            
            # Step 8: Process with LLM to categorize information
            print(f"  🤖 Processing content with LLM...")
            categorized_data = self.process_content_with_llm(medication, comprehensive_content, what_is_info)
            
            print(f"  ✅ Successfully processed {medication}")
            return categorized_data
            
        except Exception as e:
            error_msg = f"❌ Unexpected error processing {medication}: {str(e)}"
            print(error_msg)
            try:
                print("  🔄 Attempting to recover from error...")
                self.init_driver()
                time.sleep(3)
            except:
                pass
            return error_msg
    
    def find_main_medication_result(self, medication):
        """Find the main medication result"""
        print(f"  🔍 Looking for main result for: {medication}")
        time.sleep(3)
        
        # Close popups before searching
        self.close_modal_popups()
        
        if not self.check_connection():
            print("  🔄 Reconnecting before searching for results...")
            self.init_driver()
            time.sleep(3)
            return None
        
        # Try direct href matches first
        direct_selectors = [
            f"a[href*='{medication.lower().replace(' ', '-')}.html']",
            f"a[href*='{medication.lower()}.html']",
            f"a[href*='{medication.replace(' ', '-').lower()}.html']",
        ]
        
        for selector in direct_selectors:
            try:
                self.close_modal_popups()  # Close popups before each search attempt
                results = self.driver.find_elements(By.CSS_SELECTOR, selector)
                if results:
                    print(f"      ✅ Found direct match: {results[0].text[:50]}...")
                    return results[0]
            except:
                continue
        
        # Try text-based search
        try:
            self.close_modal_popups()  # Close popups before text search
            all_links = self.driver.find_elements(By.CSS_SELECTOR, "a[href*='.html']")
            medication_words = medication.lower().split()
            
            for link in all_links:
                try:
                    text = link.text.strip().lower()
                    href = link.get_attribute('href')
                    
                    if not text or not href:
                        continue
                    
                    # Skip unwanted links
                    if any(skip in href.lower() for skip in ['/pro/', '/search', '/compare', '/interaction']):
                        continue
                    
                    # Check if medication words are in the text
                    if len(medication_words) == 1:
                        if medication_words[0] in text and '.html' in href:
                            print(f"      ✅ Found text match: {link.text[:50]}...")
                            return link
                    else:
                        word_matches = sum(1 for word in medication_words if word in text)
                        if word_matches >= len(medication_words) * 0.7 and '.html' in href:
                            print(f"      ✅ Found text match: {link.text[:50]}...")
                            return link
                except:
                    continue
        except:
            pass
        
        print(f"    ❌ No main result found for {medication}")
        return None
    
    def find_side_effects_link(self):
        """Find the side effects navigation link"""
        print(f"  🔍 Looking for side effects link...")
        time.sleep(2)
        
        # Close popups before searching
        self.close_modal_popups()
        
        if not self.check_connection():
            print("  🔄 Reconnecting before searching for side effects link...")
            self.init_driver()
            time.sleep(3)
            return None
        
        # Try XPath text search first
        xpath_selectors = [
            "//a[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'side effects')]",
            "//a[contains(text(), 'Side effects')]",
            "//a[contains(text(), 'side effects')]",
        ]
        
        for selector in xpath_selectors:
            try:
                self.close_modal_popups()  # Close popups before each search attempt
                links = self.driver.find_elements(By.XPATH, selector)
                if links:
                    print(f"      ✅ Found side effects link: {links[0].text}")
                    return links[0]
            except:
                continue
        
        # Try href-based search
        href_selectors = [
            "a[href*='side-effects']",
            "a[href*='#side-effects']",
            "a[href*='sideeffects']",
        ]
        
        for selector in href_selectors:
            try:
                self.close_modal_popups()  # Close popups before each search attempt
                links = self.driver.find_elements(By.CSS_SELECTOR, selector)
                if links:
                    print(f"      ✅ Found side effects link: {links[0].text}")
                    return links[0]
            except:
                continue
        
        print(f"    ❌ No side effects link found")
        return None
    
    def extract_comprehensive_side_effects(self, medication):
        """Extract ALL side effects content comprehensively with timeout protection"""
        try:
            print(f"    🔍 Starting comprehensive extraction for {medication}")
            
            # Quick connection check
            if not self.check_connection():
                return f"❌ Connection lost during content extraction for {medication}"
            
            # Close popups at start of extraction
            self.close_modal_popups()
            time.sleep(2)  # Reduced wait time
            
            all_content = []
            
            # Strategy 1: Get ALL text from the entire page and filter (with timeout protection)
            try:
                print(f"      📄 Strategy 1: Full page text extraction...")
                self.close_modal_popups()  # Close popups before extraction
                
                body_element = self.driver.find_element(By.TAG_NAME, "body")
                full_page_text = body_element.text
                
                # Quick check if we got content
                if len(full_page_text) < 100:
                    print(f"      ⚠️ Page text too short ({len(full_page_text)} chars), might be loading issue")
                    self.close_modal_popups()  # Close popups before retry
                    time.sleep(3)
                    body_element = self.driver.find_element(By.TAG_NAME, "body")
                    full_page_text = body_element.text
                
                paragraphs = full_page_text.split('\n')
                relevant_content = []
                
                side_effects_keywords = [
                    'side effect', 'adverse', 'reaction', 'warning', 'precaution',
                    'common side effects', 'serious side effects', 'rare side effects',
                    'call your doctor', 'emergency', 'seek immediate help',
                    'stop taking', 'discontinue', 'allergic', 'overdose',
                    'nausea', 'vomiting', 'diarrhea', 'headache', 'dizziness',
                    'rash', 'fever', 'breathing', 'chest pain', 'swelling',
                    'mood changes', 'depression', 'anxiety', 'suicidal',
                    'liver problems', 'kidney problems', 'heart problems',
                    'bleeding', 'bruising', 'infection', 'seizure'
                ]
                
                # Process paragraphs with limit to prevent hanging
                for i, paragraph in enumerate(paragraphs[:500]):  # Limit processing
                    if i % 50 == 0:  # Close popups every 50 paragraphs and check connection
                        self.close_modal_popups()
                        if not self.check_connection():
                            print(f"      ⚠️ Connection lost during paragraph processing")
                            break
                            
                    paragraph = paragraph.strip()
                    if len(paragraph) > 20:
                        if any(keyword in paragraph.lower() for keyword in side_effects_keywords):
                            relevant_content.append(paragraph)
                
                if relevant_content:
                    all_content.extend(relevant_content)
                    print(f"      ✅ Found {len(relevant_content)} relevant paragraphs from full page")
                
            except Exception as e:
                print(f"      ⚠️ Error extracting from full page: {e}")
            
            # Strategy 2: Look for specific sections and their content (with timeout protection)
            print(f"      🎯 Strategy 2: Section-based extraction...")
            self.close_modal_popups()  # Close popups before section extraction
            
            section_selectors = [
                "#side-effects",
                ".side-effects",
                "[class*='side-effect']",
                "[id*='side-effect']",
                ".adverse-reactions",
                "[class*='adverse']",
                "[class*='warning']",
                "[class*='precaution']"
            ]
            
            for i, selector in enumerate(section_selectors):
                try:
                    # Close popups before each section search
                    if i % 2 == 0:  # Every 2 selectors
                        self.close_modal_popups()
                        
                    # Quick connection check
                    if not self.check_connection():
                        print(f"      ⚠️ Connection lost during section processing")
                        break
                        
                    sections = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    for section in sections:
                        text = section.text.strip()
                        if text:
                            all_content.append(f"=== SECTION: {selector} ===")
                            all_content.append(text)
                            print(f"      ✅ Found section with {len(text)} characters: {selector}")
                            
                            # Special handling for #side-effects - get sibling content
                            if selector == "#side-effects":
                                try:
                                    # Get following siblings that contain the actual side effects content
                                    siblings = section.find_elements(By.XPATH, "./following-sibling::*")
                                    print(f"      🔍 Found {len(siblings)} siblings after #side-effects")
                                    
                                    for i, sibling in enumerate(siblings[:5]):  # Get first 5 siblings
                                        sibling_text = sibling.text.strip()
                                        if len(sibling_text) > 20:
                                            all_content.append(f"=== SIDE EFFECTS CONTENT {i+1} ===")
                                            all_content.append(sibling_text)
                                            print(f"      ✅ Added sibling {i+1} with {len(sibling_text)} characters")
                                        
                                except Exception as e:
                                    print(f"      ⚠️ Error getting side effects siblings: {e}")
                except:
                    continue
            
            # Combine all content
            if all_content:
                combined_content = '\n\n'.join(all_content)
                print(f"    ✅ Successfully extracted {len(combined_content)} characters of comprehensive content")
                return combined_content
            else:
                print(f"    ⚠️ No side effects content found")
                return f"No side effects content found for {medication}"
            
        except Exception as e:
            print(f"    ❌ Exception during comprehensive extraction: {e}")
            return f"Error extracting side effects: {str(e)}"
    
    def extract_comprehensive_side_effects_quick(self, medication):
        """Quick fallback extraction method with minimal processing"""
        try:
            print(f"    🚀 Quick extraction fallback for {medication}")
            
            # Simple strategy: just get the side effects section content
            try:
                side_effects_elem = self.driver.find_element(By.CSS_SELECTOR, "#side-effects")
                # Get parent element which likely contains the content
                parent = side_effects_elem.find_element(By.XPATH, "..")
                content = parent.text.strip()
                
                if len(content) > 100:
                    print(f"    ✅ Quick extraction got {len(content)} characters")
                    return content
                    
            except Exception as e:
                print(f"    ⚠️ Quick extraction fallback failed: {e}")
            
            # Ultimate fallback: get visible text from body
            try:
                body = self.driver.find_element(By.TAG_NAME, "body")
                text = body.text
                if "side effect" in text.lower():
                    return text[:5000]  # Limit to first 5000 characters
            except:
                pass
                
            return f"Quick extraction failed for {medication}"
            
        except Exception as e:
            return f"Quick extraction error for {medication}: {str(e)}"

    def process_content_with_llm(self, medication, comprehensive_content, what_is_info):
        """Use LLM to categorize comprehensive content into structured columns including What Is information"""
        try:
            print(f"    🤖 Processing content with LLM for {medication}...")
            
            if not comprehensive_content or comprehensive_content.startswith("❌") or comprehensive_content.startswith("No side effects"):
                return {
                    'what_is': what_is_info if what_is_info else f"No description available for {medication}",
                    'side_effects': f"No side effects information found for {medication}",
                    'call_doctor': f"No doctor guidance found for {medication}",
                    'go_to_er': f"No emergency guidance found for {medication}"
                }
            
            # Create a comprehensive prompt for the LLM
            prompt = f"""
You are a medical information expert. Please analyze the following information for the medication "{medication}" and categorize it into four specific columns:

WHAT IS INFORMATION:
{what_is_info}

RAW SIDE EFFECTS TEXT:
{comprehensive_content}

Please extract and organize this information into exactly four categories:

1. WHAT IS: Provide a clear, concise description of what this medication is and what it's used for. Use the "What Is Information" provided above, but feel free to refine and improve it for clarity.

2. SIDE EFFECTS: List all the side effects mentioned (common, uncommon, serious, mild, etc.). Include symptoms, reactions, and any physical or mental effects.

3. CALL A DOCTOR IF: Extract information about when patients should contact their doctor. This includes warnings, concerning symptoms, or situations requiring medical consultation.

4. GO TO ER IF: Extract emergency situations, severe reactions, or life-threatening symptoms that require immediate emergency medical attention.

Format your response as:

WHAT IS:
[Clear description of the medication and its uses]

SIDE EFFECTS:
[List all side effects here]

CALL A DOCTOR IF:
[List situations requiring doctor consultation]

GO TO ER IF:
[List emergency situations]

Important: 
- Be comprehensive and include ALL relevant information from the text
- Use clear, readable language
- Separate different symptoms with bullet points or clear formatting
- If no information is available for a category, write "No specific information provided"
"""

            # Generate response from LLM
            response = self.model.generate_content(prompt)
            llm_response = response.text
            
            print(f"    ✅ LLM processing completed ({len(llm_response)} characters)")
            
            # Parse the LLM response into structured data
            parsed_data = self.parse_llm_response(llm_response)
            
            return parsed_data
            
        except Exception as e:
            print(f"    ❌ Error processing with LLM: {e}")
            return {
                'what_is': what_is_info if what_is_info else f"Error getting description for {medication}: {str(e)}",
                'side_effects': f"Error processing side effects for {medication}: {str(e)}",
                'call_doctor': f"Error processing doctor guidance for {medication}: {str(e)}",
                'go_to_er': f"Error processing emergency guidance for {medication}: {str(e)}"
            }
    
    def parse_llm_response(self, llm_response):
        """Parse the LLM response into structured categories including What Is"""
        try:
            # Initialize default values
            what_is = "No specific information provided"
            side_effects = "No specific information provided"
            call_doctor = "No specific information provided"
            go_to_er = "No specific information provided"
            
            # Split response by sections
            sections = llm_response.split('\n')
            current_section = None
            current_content = []
            
            for i, line in enumerate(sections):
                line = line.strip()
                
                # Handle different formats: "WHAT IS:", "**WHAT IS:**", etc.
                line_upper = line.upper().replace('*', '').replace(':', '').strip()
                
                # Make section detection more strict - must START with the section name
                if line_upper == 'WHAT IS' or line_upper.startswith('WHAT IS'):
                    if current_section and current_content:
                        content = '\n'.join(current_content).strip()
                        if current_section == 'what_is':
                            what_is = content
                        elif current_section == 'side_effects':
                            side_effects = content
                        elif current_section == 'call_doctor':
                            call_doctor = content
                        elif current_section == 'go_to_er':
                            go_to_er = content
                    
                    current_section = 'what_is'
                    current_content = []
                    # Add any content after the header
                    if ':' in line:
                        after_colon = line.split(':', 1)[1].strip()
                        if after_colon and not after_colon.startswith('*'):
                            current_content.append(after_colon)
                            
                elif line_upper == 'SIDE EFFECTS' or line_upper.startswith('SIDE EFFECTS'):
                    if current_section and current_content:
                        content = '\n'.join(current_content).strip()
                        if current_section == 'what_is':
                            what_is = content
                        elif current_section == 'side_effects':
                            side_effects = content
                        elif current_section == 'call_doctor':
                            call_doctor = content
                        elif current_section == 'go_to_er':
                            go_to_er = content
                    
                    current_section = 'side_effects'
                    current_content = []
                    # Add any content after the header
                    if ':' in line:
                        after_colon = line.split(':', 1)[1].strip()
                        if after_colon and not after_colon.startswith('*'):
                            current_content.append(after_colon)
                            
                elif line_upper.startswith('CALL A DOCTOR IF') or line_upper.startswith('CALL DOCTOR IF'):
                    if current_section and current_content:
                        content = '\n'.join(current_content).strip()
                        if current_section == 'what_is':
                            what_is = content
                        elif current_section == 'side_effects':
                            side_effects = content
                        elif current_section == 'call_doctor':
                            call_doctor = content
                        elif current_section == 'go_to_er':
                            go_to_er = content
                    
                    current_section = 'call_doctor'
                    current_content = []
                    # Add any content after the header
                    if ':' in line:
                        after_colon = line.split(':', 1)[1].strip()
                        if after_colon and not after_colon.startswith('*'):
                            current_content.append(after_colon)
                            
                elif line_upper.startswith('GO TO ER IF') or line_upper.startswith('EMERGENCY'):
                    if current_section and current_content:
                        content = '\n'.join(current_content).strip()
                        if current_section == 'what_is':
                            what_is = content
                        elif current_section == 'side_effects':
                            side_effects = content
                        elif current_section == 'call_doctor':
                            call_doctor = content
                        elif current_section == 'go_to_er':
                            go_to_er = content
                    
                    current_section = 'go_to_er'
                    current_content = []
                    # Add any content after the header
                    if ':' in line:
                        after_colon = line.split(':', 1)[1].strip()
                        if after_colon and not after_colon.startswith('*'):
                            current_content.append(after_colon)
                            
                elif line and current_section:
                    # Collect content within the current section
                    if line.strip() and line.strip() != '*':
                        current_content.append(line)
            
            # Save the last section
            if current_section and current_content:
                content = '\n'.join(current_content).strip()
                if current_section == 'what_is':
                    what_is = content
                elif current_section == 'side_effects':
                    side_effects = content
                elif current_section == 'call_doctor':
                    call_doctor = content
                elif current_section == 'go_to_er':
                    go_to_er = content
            
            return {
                'what_is': what_is if what_is != "No specific information provided" else "No specific information provided",
                'side_effects': side_effects if side_effects != "No specific information provided" else "No specific information provided",
                'call_doctor': call_doctor if call_doctor != "No specific information provided" else "No specific information provided",
                'go_to_er': go_to_er if go_to_er != "No specific information provided" else "No specific information provided"
            }
            
        except Exception as e:
            print(f"    ❌ Error parsing LLM response: {e}")
            return {
                'what_is': f"Error parsing response: {str(e)}",
                'side_effects': f"Error parsing response: {str(e)}",
                'call_doctor': f"Error parsing response: {str(e)}",
                'go_to_er': f"Error parsing response: {str(e)}"
            }
    
    def add_delay(self):
        """Add random delay between requests"""
        delay = random.uniform(1.5, 2.5)
        print(f"  ⏰ Waiting {delay:.1f} seconds before next request...")
        time.sleep(delay)
    
    def close(self):
        """Close the browser"""
        if self.driver:
            self.driver.quit()

def sanitize_text_for_excel(text):
    """Sanitize text to prevent Excel corruption"""
    if not text:
        return "No content"
    
    try:
        text = str(text)
        
        # Remove problematic characters
        text = re.sub(r'[^\x20-\x7E\n\r\t]', ' ', text)  # Keep only printable ASCII + newlines/tabs
        text = re.sub(r'\s+', ' ', text)  # Replace multiple spaces with single space
        
        # Limit length to prevent Excel issues
        if len(text) > 2000:
            text = text[:2000] + "... [truncated]"
        
        if not text.strip():
            return "Content processed but empty"
            
        return text.strip()
        
    except Exception as e:
        return f"Error processing content: {str(e)[:50]}"

def update_excel_with_side_effects(max_medications=None, start_from=0):
    """Update Excel file with side effects for all medications using LLM categorization"""
    
    excel_path = '/Users/juanlu/Documents/Wye/scrapper/Analysis/main_diseases_analysis_final.xlsx'
    
    if not os.path.exists(excel_path):
        print(f"❌ Excel file not found: {excel_path}")
        return
    
    # Load workbook
    wb = load_workbook(excel_path)
    
    if "All Unique Medications" not in wb.sheetnames:
        print("❌ 'All Unique Medications' sheet not found")
        return
    
    medications_ws = wb["All Unique Medications"]
    
    # Get medications
    medications = []
    for row in medications_ws.iter_rows(min_row=9, max_col=1, values_only=True):
        if row[0] and row[0].strip():
            medications.append(row[0].strip())
    
    # Check what's already been processed (check column B for existing data)
    processed_count = 0
    for i in range(9, 9 + len(medications)):
        if medications_ws[f'B{i}'].value:
            processed_count += 1
        else:
            break
    
    print(f"📊 Found {len(medications)} total medications")
    print(f"✅ Already processed: {processed_count} medications")
    
    # Start from where we left off, or from start_from parameter
    start_index = max(processed_count, start_from)
    remaining_medications = medications[start_index:]
    
    if not remaining_medications:
        print("🎉 All medications have already been processed!")
        return
    
    # Use all remaining medications if max_medications is None
    if max_medications:
        remaining_medications = remaining_medications[:max_medications]
    
    print(f"📊 Processing {len(remaining_medications)} remaining medications (starting from #{start_index + 1})...")
    
    # Update column headers for the new structure
    medications_ws['B8'] = 'WHAT IS'
    medications_ws['C8'] = 'SIDE EFFECTS'
    medications_ws['D8'] = 'CALL A DOCTOR IF'
    medications_ws['E8'] = 'GO TO ER IF'
    
    # Initialize scraper
    scraper = DrugsScraper(headless=False)
    
    try:
        current_processed = processed_count
        errors = []
        
        for i, medication in enumerate(remaining_medications):
            medication_index = start_index + i
            print(f"\n[{medication_index + 1}/{len(medications)}] Processing: {medication}")
            
            # Close popups at start of each medication processing
            try:
                scraper.close_modal_popups()
            except:
                pass
            
            # Check if scraper connection is still alive
            if not scraper.check_connection():
                print("  🔄 Reconnecting scraper...")
                scraper.init_driver()
                time.sleep(5)
            
            # Get structured side effects data with LLM processing
            max_retries = 3
            categorized_data = None
            
            for attempt in range(max_retries):
                try:
                    print(f"  🔄 Attempt {attempt + 1} of {max_retries}")
                    
                    # Close popups before each attempt
                    try:
                        scraper.close_modal_popups()
                    except:
                        pass
                    
                    start_time = time.time()
                    
                    categorized_data = scraper.search_and_get_side_effects(medication)
                    
                    elapsed = time.time() - start_time
                    if elapsed > 300:
                        print(f"  ⏰ Process took {elapsed:.1f} seconds (longer than expected)")
                    
                    break
                        
                except Exception as e:
                    print(f"  ⚠️  Attempt {attempt + 1} failed: {e}")
                    if attempt < max_retries - 1:
                        print("  🔄 Reinitializing scraper and retrying...")
                        try:
                            scraper.close_modal_popups()
                        except:
                            pass
                        scraper.init_driver()
                        time.sleep(10)
                    else:
                        categorized_data = {
                            'side_effects': f"❌ Failed to process {medication} after {max_retries} attempts",
                            'call_doctor': f"❌ Failed to process {medication} after {max_retries} attempts", 
                            'go_to_er': f"❌ Failed to process {medication} after {max_retries} attempts"
                        }
            
            # Add structured data to Excel columns B, C, D
            row_num = 9 + medication_index
            
            try:
                if isinstance(categorized_data, dict):
                    # LLM processing succeeded - save structured data
                    what_is = sanitize_text_for_excel(categorized_data.get('what_is', ''))
                    side_effects = sanitize_text_for_excel(categorized_data.get('side_effects', ''))
                    call_doctor = sanitize_text_for_excel(categorized_data.get('call_doctor', ''))
                    go_to_er = sanitize_text_for_excel(categorized_data.get('go_to_er', ''))
                    
                    medications_ws[f'B{row_num}'] = what_is
                    medications_ws[f'C{row_num}'] = side_effects
                    medications_ws[f'D{row_num}'] = call_doctor
                    medications_ws[f'E{row_num}'] = go_to_er
                    
                    print(f"  ✅ Saved structured data for {medication}")
                    print(f"    - What Is: {len(what_is)} chars")
                    print(f"    - Side Effects: {len(side_effects)} chars")
                    print(f"    - Call Doctor: {len(call_doctor)} chars")
                    print(f"    - Go to ER: {len(go_to_er)} chars")
                    
                else:
                    # LLM processing failed - save error message
                    error_msg = sanitize_text_for_excel(str(categorized_data) if categorized_data else f"❌ Failed to process {medication}")
                    
                    medications_ws[f'B{row_num}'] = error_msg
                    medications_ws[f'C{row_num}'] = "Processing failed"
                    medications_ws[f'D{row_num}'] = "Processing failed"
                    medications_ws[f'E{row_num}'] = "Processing failed"
                    
                    print(f"  ❌ Saved error data for {medication}")
                    
            except Exception as write_error:
                print(f"  ⚠️  Error writing to Excel: {write_error}")
                try:
                    error_msg = f"Error processing {medication}"
                    medications_ws[f'B{row_num}'] = error_msg
                    medications_ws[f'C{row_num}'] = error_msg
                    medications_ws[f'D{row_num}'] = error_msg
                    medications_ws[f'E{row_num}'] = error_msg
                except Exception as fallback_error:
                    print(f"  ❌ Fatal Excel write error: {fallback_error}")
                    continue
            
            current_processed += 1
            
            # Track errors
            if isinstance(categorized_data, str) and categorized_data.startswith("❌"):
                errors.append(medication)
            
            # Save progress every 5 medications
            if current_processed % 5 == 0:
                try:
                    wb.save(excel_path)
                    print(f"💾 Progress saved: {current_processed}/{len(medications)} medications processed")
                    print(f"   Errors so far: {len(errors)}")
                except Exception as save_error:
                    print(f"  ⚠️  Error saving progress: {save_error}")
            
            # Add delay between requests
            scraper.add_delay()
    
    finally:
        scraper.close()
    
    # Final save
    try:
        wb.save(excel_path)
        print(f"💾 Final save completed")
    except Exception as save_error:
        print(f"❌ Error in final save: {save_error}")
    
    # Summary
    success_count = current_processed - len(errors)
    print(f"\n" + "="*60)
    print(f"✅ PROCESSING COMPLETED!")
    print(f"📊 Total processed: {current_processed}")
    print(f"✅ Successful: {success_count}")
    print(f"❌ Errors: {len(errors)}")
    print(f"📄 Updated Excel file: {excel_path}")
    
    if errors:
        print(f"\n❌ Medications with errors:")
        for error_med in errors[:10]:
            print(f"   - {error_med}")
        if len(errors) > 10:
            print(f"   ... and {len(errors) - 10} more")

if __name__ == "__main__":
    print("🚀 Starting Enhanced LLM-Powered Medication Data Scraper")
    print("="*60)
    print("🔧 Enhanced features:")
    print("   - Direct LLM processing of medication information and side effects")
    print("   - Structured categorization into 4 columns:")
    print("     * WHAT IS (medication description)")
    print("     * SIDE EFFECTS")
    print("     * CALL A DOCTOR IF") 
    print("     * GO TO ER IF")
    print("   - Automatic reconnection after computer sleep")
    print("   - Connection health monitoring")
    print("   - Robust error recovery")
    print("   - Multiple retry attempts")
    print("   - Modal/popup handling")
    print("   - Resume from where it left off")
    print("="*60)
    
    # Run for 1 medication with DEBUG
    print("🧪 Running DEBUG mode with 1 medication...")
    update_excel_with_side_effects()  # Process ALL medications
    
    print("\n" + "="*60)
    print("🎉 DEBUG RUN COMPLETED! Check the output above for debug info.")
    print("💡 If the test looks good, remove the max_medications parameter")
    print("   to process ALL medications.")
