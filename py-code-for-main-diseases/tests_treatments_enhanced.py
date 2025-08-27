#!/usr/bin/env python3

import pandas as pd
from openpyxl import load_workbook
import os

def recreate_clean_excel():
    print("🚀 Recreating clean Excel...")
    
    original_file = "tests_treatments_enhanced_analysis.xlsx"
    
    if not os.path.exists(original_file):
        print(f"❌ File not found: {original_file}")
        return
    
    try:
        wb_original = load_workbook(original_file)
        wb_copy = load_workbook(original_file)
        
        if 'Tests' in wb_copy.sheetnames:
            print("🔬 Completing Tests...")
            complete_tests_sheet(wb_copy['Tests'])
        
        if 'Treatments' in wb_copy.sheetnames:
            print("💊 Completing Treatments...")
            complete_treatments_sheet(wb_copy['Treatments'])
        
        output_file = "../Analysis/tests_treatments_enhanced_analysis_COMPLETED.xlsx"
        wb_copy.save(output_file)
        print(f"✅ Completed file saved as: {output_file}")
        
        wb_original.close()
        wb_copy.close()
        
    except Exception as e:
        print(f"❌ Error: {e}")

def complete_tests_sheet(ws):
    tests_info = {
        'Blood test': {
            'description': 'Blood tests are laboratory examinations of blood samples to evaluate overall health, detect diseases, and monitor treatment effectiveness. They can measure various components like red blood cells, white blood cells, platelets, and chemical substances.',
            'background': 'Doctors order blood tests to screen for conditions, diagnose problems, monitor chronic diseases, and check how well treatments are working. Results help identify infections, anemia, diabetes, heart disease, and many other health issues.',
            'spanish_name': 'Análisis de sangre'
        },
        'Blood tests': {
            'description': 'Blood tests are laboratory examinations of blood samples to evaluate overall health, detect diseases, and monitor treatment effectiveness. They can measure various components like red blood cells, white blood cells, platelets, and chemical substances.',
            'background': 'Doctors order blood tests to screen for conditions, diagnose problems, monitor chronic diseases, and check how well treatments are working. Results help identify infections, anemia, diabetes, heart disease, and many other health issues.',
            'spanish_name': 'Análisis de sangre'
        },
        'PET scan': {
            'description': 'A PET scan (positron emission tomography) is a nuclear medicine imaging test that uses a radioactive tracer to show how tissues and organs are functioning. It can detect cancer, heart disease, and brain disorders by showing metabolic activity.',
            'background': 'PET scans work by injecting a small amount of radioactive material that collects in areas of high chemical activity, which often indicates disease. The scan creates detailed images showing how your tissues and organs are working, not just their structure.',
            'spanish_name': 'Tomografía por emisión de positrones (TEP)'
        },
        'X-ray': {
            'description': 'An X-ray is a type of picture taken of the inside of your body using special rays.',
            'background': 'X-rays use electromagnetic radiation to create images of bones and soft tissues inside the body. They are commonly used to diagnose fractures, lung conditions, and other internal problems. The images help doctors see the structure and condition of internal organs and bones.'
        }
    }
    
    for row in range(7, ws.max_row + 1):
        test_name = ws.cell(row=row, column=1).value
        if not test_name or test_name == 'TEST NAME':
            continue
        
        if test_name in tests_info:
            current_spanish = ws.cell(row=row, column=2).value
            current_desc = ws.cell(row=row, column=3).value
            current_bg = ws.cell(row=row, column=4).value
            
            if (pd.isna(current_spanish) or 
                str(current_spanish).strip() == "" or
                'Information not found' in str(current_spanish) or 
                'No encontrado' in str(current_spanish) or
                'No encontrado en Mayo Clinic' in str(current_spanish) or
                'Información no encontrada' in str(current_spanish)):
                if 'spanish_name' in tests_info[test_name]:
                    ws.cell(row=row, column=2, value=tests_info[test_name]['spanish_name'])
                    print(f"  ✅ {test_name}: Spanish name completed")
            
            if (pd.isna(current_desc) or 
                str(current_desc).strip() == "" or
                'Information not found' in str(current_desc) or 
                'Not found on Mayo Clinic' in str(current_desc) or
                'No encontrado en Mayo Clinic' in str(current_desc)):
                ws.cell(row=row, column=3, value=tests_info[test_name]['description'])
                print(f"  ✅ {test_name}: Description completed")
            
            if (pd.isna(current_bg) or 
                str(current_bg).strip() == "" or
                'Information not found' in str(current_bg) or 
                'Not found on Mayo Clinic' in str(current_bg) or
                'No encontrado en Mayo Clinic' in str(current_bg)):
                ws.cell(row=row, column=4, value=tests_info[test_name]['background'])
                print(f"  ✅ {test_name}: Background information completed")

def complete_treatments_sheet(ws):
    treatments_info = {
        'Assistive devices': {
            'description': 'Assistive devices are tools, equipment, or products that help people with disabilities or medical conditions perform daily activities more easily and safely. They can include mobility aids, communication devices, and adaptive equipment.',
            'background': 'These devices help improve independence, safety, and quality of life for people with various conditions. They can range from simple tools like grab bars to complex technology like voice-activated computers.'
        },
        'Chondroitin': {
            'description': 'Chondroitin is a natural substance found in cartilage around joints. It is commonly used as a dietary supplement to help with osteoarthritis pain and joint health.',
            'background': 'Chondroitin works by helping to maintain the structure of cartilage and may reduce inflammation in joints. It is often taken together with glucosamine for joint health support.'
        },
        'Donanemab-azbt': {
            'description': 'Donanemab is a monoclonal antibody treatment for Alzheimer\'s disease that targets amyloid plaques in the brain. It works by helping the body remove these harmful protein deposits.',
            'background': 'This treatment is designed to slow the progression of Alzheimer\'s disease by reducing amyloid plaque buildup. It represents a new approach to treating the underlying causes of the disease.'
        },
        'Environmental modifications': {
            'description': 'Environmental modifications involve changing the physical environment to make it safer and more accessible for people with medical conditions or disabilities.',
            'background': 'These changes can include installing ramps, removing trip hazards, improving lighting, and adapting living spaces to meet specific medical needs and improve safety.'
        },
        'Exercise': {
            'description': 'Exercise is physical activity that improves health, fitness, and well-being. It includes activities like walking, swimming, strength training, and flexibility exercises.',
            'background': 'Regular exercise helps prevent and manage many health conditions, improves cardiovascular health, strengthens muscles and bones, and enhances mental health and mood.'
        },
        'Lecanemab-irmb': {
            'description': 'Lecanemab is a monoclonal antibody treatment for Alzheimer\'s disease that targets amyloid beta proteins in the brain to slow disease progression.',
            'background': 'This treatment works by binding to amyloid plaques and helping the immune system remove them, potentially slowing cognitive decline in early-stage Alzheimer\'s disease.'
        },
        'Managing complications': {
            'description': 'Managing complications involves identifying, treating, and preventing secondary health problems that can arise from primary medical conditions or treatments.',
            'background': 'This approach focuses on early detection and intervention to minimize the impact of complications on overall health and treatment outcomes.'
        },
        'Medication': {
            'description': 'Medication refers to drugs or pharmaceutical substances used to treat, cure, or prevent diseases and medical conditions.',
            'background': 'Medications work in various ways to target specific disease processes, relieve symptoms, or support bodily functions. They can be taken orally, injected, or applied topically.'
        },
        'Medications': {
            'description': 'Medications are pharmaceutical drugs used to treat various medical conditions, manage symptoms, and support health. They can include prescription drugs, over-the-counter medicines, and supplements.',
            'background': 'Different types of medications work through various mechanisms to address specific health issues, from antibiotics that fight infections to pain relievers that reduce discomfort.'
        },
        'Simpler tasks and structured routines': {
            'description': 'Simpler tasks and structured routines involve breaking down complex activities into manageable steps and establishing consistent daily patterns to help people with cognitive or physical challenges.',
            'background': 'This approach helps improve function, reduce confusion, and increase independence for people with various conditions by creating predictable and manageable daily activities.'
        },
        'Surgery': {
            'description': 'Surgery is a medical procedure that involves cutting into the body to repair, remove, or replace damaged or diseased tissue, organs, or structures.',
            'background': 'Surgical procedures can range from minor outpatient operations to complex major surgeries requiring hospitalization. They are performed to treat injuries, diseases, and congenital conditions.'
        }
    }
    
    for row in range(7, ws.max_row + 1):
        treatment_name = ws.cell(row=row, column=1).value
        if not treatment_name or treatment_name == 'TREATMENT NAME':
            continue
        
        if treatment_name in treatments_info:
            current_desc = ws.cell(row=row, column=3).value
            current_bg = ws.cell(row=row, column=4).value
            
            if (pd.isna(current_desc) or 
                str(current_desc).strip() == "" or
                'Information not found' in str(current_desc) or 
                'Not found on Mayo Clinic' in str(current_desc) or
                'No encontrado en Mayo Clinic' in str(current_desc)):
                ws.cell(row=row, column=3, value=treatments_info[treatment_name]['description'])
                print(f"  ✅ {treatment_name}: Description completed")
            
            if (pd.isna(current_bg) or 
                str(current_bg).strip() == "" or
                'Information not found' in str(current_bg) or 
                'Not found on Mayo Clinic' in str(current_bg) or
                'No encontrado en Mayo Clinic' in str(current_bg)):
                ws.cell(row=row, column=4, value=treatments_info[treatment_name]['background'])
                print(f"  ✅ {treatment_name}: Background information completed")

def main():
    print("🎯 Script to recreate clean Excel from scratch")
    print("=" * 50)
    print("📝 NOTE: Original file will NOT be modified")
    print("🆕 A new file will be created with completed information")
    print("=" * 50)
    recreate_clean_excel()
    print("\n🎉 Process completed!")

if __name__ == "__main__":
    main()

