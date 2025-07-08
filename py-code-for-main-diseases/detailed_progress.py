#!/usr/bin/env python3
"""
Check which medications failed and get a detailed summary
"""
from openpyxl import load_workbook
import time
import os

def detailed_progress():
    excel_path = '/Users/juanlu/Documents/Wye/scrapper/Analysis/main_diseases_analysis_final.xlsx'
    
    if not os.path.exists(excel_path):
        print("❌ Excel file not found")
        return
    
    try:
        wb = load_workbook(excel_path)
        if "All Unique Medications" not in wb.sheetnames:
            print("❌ 'All Unique Medications' sheet not found")
            return
        
        ws = wb["All Unique Medications"]
        
        # Count and analyze all medications
        total_medications = 0
        processed_medications = 0
        successful_medications = 0
        failed_medications = []
        
        for row_num in range(9, 1000):  # Check up to row 1000
            med_name = ws[f'A{row_num}'].value
            side_effects = ws[f'H{row_num}'].value
            
            if not med_name or not med_name.strip():
                break
                
            total_medications += 1
            
            if side_effects:
                processed_medications += 1
                if str(side_effects).startswith("❌"):
                    failed_medications.append(med_name)
                else:
                    successful_medications += 1
        
        print(f"📊 DETAILED PROGRESS REPORT")
        print(f"=" * 50)
        print(f"📝 Total medications: {total_medications}")
        print(f"⚡ Processed: {processed_medications}")
        print(f"✅ Successful: {successful_medications}")
        print(f"❌ Failed: {len(failed_medications)}")
        print(f"📊 Progress: {(processed_medications/total_medications)*100:.1f}%")
        
        if processed_medications > 0:
            print(f"🎯 Success rate: {(successful_medications/processed_medications)*100:.1f}%")
        
        if processed_medications < total_medications:
            print(f"\n⏳ Still processing: {total_medications - processed_medications} medications remaining")
        
        # Show failed medications
        if failed_medications:
            print(f"\n❌ FAILED MEDICATIONS ({len(failed_medications)}):")
            for i, med in enumerate(failed_medications):
                print(f"   {i+1}. {med}")
                if i >= 19:  # Show first 20
                    print(f"   ... and {len(failed_medications) - 20} more")
                    break
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Error checking progress: {e}")

if __name__ == "__main__":
    detailed_progress()
