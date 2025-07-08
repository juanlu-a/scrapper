from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import time
import random
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

class DrugsScraper:
    def __init__(self, headless=False):
        self.driver = self.setup_driver(headless)
        self.wait = WebDriverWait(self.driver, 10)
        
    def setup_driver(self, headless=False):
        """Set up Chrome driver with options"""
        chrome_options = Options()
        if headless:
            chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        
        # User agent to appear more human-like
        chrome_options.add_argument("--user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
        
        driver = webdriver.Chrome(options=chrome_options)
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        
        return driver
    
    def scrape_side_effects(self, medication, max_retries=3):
        """Scrape side effects for a specific medication with enhanced error handling"""
        print(f"\n🔍 Processing: {medication}")
        
        for attempt in range(max_retries):
            try:
                if attempt > 0:
                    print(f"  🔄 Retry attempt {attempt + 1}/{max_retries}")
                
                # Step 1: Navigate to drugs.com
                self.driver.get("https://www.drugs.com")
                time.sleep(3)
                
                # Step 2: Find search box and search
                search_box = self.wait.until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='search'], input[name='q'], input[placeholder*='search'], #livesearch"))
                )
                search_box.clear()
                search_box.send_keys(medication)
                time.sleep(1)
                search_box.send_keys(Keys.RETURN)
                
                print(f"  ✅ Searched for: {medication}")
                time.sleep(4)
                
                # Step 3: Find main result
                main_result = self.find_main_result(medication)
                if not main_result:
                    print(f"  ❌ Could not find main result for {medication}")
                    continue
                
                # Step 4: Click on main result
                try:
                    self.driver.execute_script("arguments[0].scrollIntoView(true);", main_result)
                    time.sleep(1)
                    main_result.click()
                    print(f"  ✅ Clicked main result for {medication}")
                except Exception as e:
                    try:
                        self.driver.execute_script("arguments[0].click();", main_result)
                        print(f"  ✅ Clicked main result (JS) for {medication}")
                    except Exception as e2:
                        print(f"  ❌ Failed to click main result for {medication}: {str(e2)}")
                        continue
                
                time.sleep(4)
                
                # Step 5: Find and click side effects link
                side_effects_link = self.find_side_effects_link()
                if not side_effects_link:
                    print(f"  ❌ Could not find side effects link for {medication}")
                    continue
                
                # Step 6: Click side effects link
                try:
                    self.driver.execute_script("arguments[0].scrollIntoView(true);", side_effects_link)
                    time.sleep(1)
                    side_effects_link.click()
                    print(f"  ✅ Clicked side effects link for {medication}")
                except Exception as e:
                    try:
                        self.driver.execute_script("arguments[0].click();", side_effects_link)
                        print(f"  ✅ Clicked side effects link (JS) for {medication}")
                    except Exception as e2:
                        print(f"  ❌ Failed to click side effects link for {medication}: {str(e2)}")
                        continue
                
                time.sleep(4)
                
                # Step 7: Extract side effects content
                content = self.extract_side_effects_content(medication)
                
                if content and len(content) > 50:
                    print(f"  ✅ Successfully processed {medication} ({len(content)} characters)")
                    return content
                else:
                    print(f"  ❌ No substantial side effects content found for {medication}")
                    if attempt < max_retries - 1:
                        continue
                    else:
                        return f"No detailed side effects information found for {medication}"
                
            except Exception as e:
                error_msg = f"❌ Error processing {medication} (attempt {attempt + 1}): {str(e)}"
                print(error_msg)
                if attempt < max_retries - 1:
                    time.sleep(5)  # Wait before retry
                else:
                    return error_msg
        
        return f"Failed to process {medication} after {max_retries} attempts"
    
    def find_main_result(self, medication):
        """Find the main result for the medication with enhanced matching"""
        print(f"  🔍 Looking for main result for: {medication}")
        
        # Clean medication name for matching
        medication_clean = medication.lower().strip()
        medication_words = [word for word in medication_clean.split() if len(word) > 2]
        
        # Try different selectors to find results
        selectors = [
            "a[href*='.html']",
            ".search-results a",
            ".result a",
            ".drug-results a",
            "a[href*='/drug/']",
            "li a[href*='.html']",
            ".search-result a",
            ".results a"
        ]
        
        for selector in selectors:
            try:
                results = self.driver.find_elements(By.CSS_SELECTOR, selector)
                print(f"    Found {len(results)} results with selector: {selector}")
                
                for result in results:
                    try:
                        href = result.get_attribute('href')
                        if not href or '.html' not in href:
                            continue
                        
                        # Get text content
                        text = result.text.strip().lower()
                        
                        # Check if this looks like a main medication result
                        if not text:
                            continue
                        
                        # Look for star indicators or main drug page indicators
                        parent = result.find_element(By.XPATH, "./..")
                        parent_text = parent.text.lower()
                        
                        # Check for star or main result indicators
                        has_star = '★' in parent_text or 'star' in parent_text
                        
                        # Check text matching
                        if len(medication_words) == 1:
                            # Single word medication
                            if medication_words[0] in text:
                                print(f"      ✅ Found single word match: {text[:50]}...")
                                return result
                        else:
                            # Multi-word medication - check for word matches
                            word_matches = sum(1 for word in medication_words if word in text)
                            if word_matches >= len(medication_words) * 0.7:
                                print(f"      ✅ Found multi-word match: {text[:50]}...")
                                return result
                        
                        # If has star indicator, be more flexible
                        if has_star and any(word in text for word in medication_words):
                            print(f"      ✅ Found star result: {text[:50]}...")
                            return result
                    
                    except Exception as e:
                        continue
                
            except Exception as e:
                continue
        
        return None
    
    def find_side_effects_link(self):
        """Find the side effects navigation link with enhanced detection"""
        print(f"  🔍 Looking for side effects link...")
        
        time.sleep(3)
        
        # Try multiple approaches to find side effects link
        approaches = [
            # Approach 1: Direct text matching
            {
                'name': 'Direct text matching',
                'method': 'xpath',
                'selectors': [
                    "//a[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'side effects')]",
                    "//a[text()='Side effects']",
                    "//a[text()='side effects']",
                    "//a[text()='Side Effects']",
                    "//span[contains(text(), 'side effects')]/parent::a",
                    "//span[contains(text(), 'Side effects')]/parent::a"
                ]
            },
            # Approach 2: href containing side-effects
            {
                'name': 'Href containing side-effects',
                'method': 'css',
                'selectors': [
                    "a[href*='side-effects']",
                    "a[href*='#side-effects']",
                    "a[href*='sideeffects']",
                    "a[href*='adverse-reactions']"
                ]
            },
            # Approach 3: Navigation areas
            {
                'name': 'Navigation areas',
                'method': 'css',
                'selectors': [
                    "nav a",
                    ".nav a",
                    ".navigation a",
                    ".tabs a",
                    ".tab a",
                    ".menu a",
                    ".drug-nav a",
                    ".page-nav a",
                    "ul.nav a",
                    ".nav-tabs a"
                ]
            }
        ]
        
        for approach in approaches:
            print(f"    Trying approach: {approach['name']}")
            
            for selector in approach['selectors']:
                try:
                    if approach['method'] == 'xpath':
                        links = self.driver.find_elements(By.XPATH, selector)
                    else:
                        links = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    
                    print(f"      Found {len(links)} links with selector: {selector}")
                    
                    for link in links:
                        try:
                            text = link.text.strip().lower()
                            href = link.get_attribute('href')
                            
                            # Check if this looks like a side effects link
                            if approach['name'] == 'Navigation areas':
                                # For navigation areas, check text content
                                if 'side effect' in text or 'adverse' in text:
                                    print(f"      ✅ Found side effects link: {link.text}")
                                    return link
                            else:
                                # For other approaches, we already filtered
                                if text and href:
                                    print(f"      ✅ Found side effects link: {link.text}")
                                    return link
                        except Exception as e:
                            continue
                            
                except Exception as e:
                    continue
        
        # If no link found, try to find side effects content on current page
        print(f"    No side effects link found, checking current page...")
        return None
    
    def extract_side_effects_content(self, medication):
        """Extract side effects content from the current page"""
        print(f"  📄 Extracting side effects content for {medication}")
        
        try:
            content_parts = []
            content_found = False
            
            # Wait for page to load
            time.sleep(3)
            
            # Try multiple approaches to find side effects content
            
            # Approach 1: Look for side effects sections
            section_selectors = [
                "#side-effects",
                ".side-effects",
                "[id*='side-effects']",
                "[class*='side-effects']",
                "[id*='adverse']",
                "[class*='adverse']"
            ]
            
            for selector in section_selectors:
                try:
                    sections = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    for section in sections:
                        text = section.text.strip()
                        if text and len(text) > 50:
                            content_parts.append(f"--- Side Effects ---")
                            content_parts.append(text)
                            content_found = True
                            break
                except:
                    continue
            
            # Approach 2: Look for headings containing "side effects"
            if not content_found:
                try:
                    headings = self.driver.find_elements(By.XPATH, 
                        "//h1[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'side effects')] | "
                        "//h2[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'side effects')] | "
                        "//h3[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'side effects')]"
                    )
                    
                    for heading in headings:
                        content_parts.append(f"--- {heading.text} ---")
                        
                        # Find the following content
                        try:
                            next_sibling = heading.find_element(By.XPATH, "./following-sibling::*")
                            count = 0
                            while next_sibling and count < 5:  # Limit to avoid infinite loop
                                tag = next_sibling.tag_name.lower()
                                if tag in ['h1', 'h2', 'h3']:
                                    break
                                    
                                text = next_sibling.text.strip()
                                if text and len(text) > 20:
                                    content_parts.append(text)
                                    content_found = True
                                
                                try:
                                    next_sibling = next_sibling.find_element(By.XPATH, "./following-sibling::*")
                                    count += 1
                                except:
                                    break
                        except:
                            pass
                        
                        if content_found:
                            break
                except:
                    pass
            
            # Approach 3: Look for paragraphs with side effects keywords
            if not content_found:
                try:
                    keywords = ['side effect', 'adverse', 'reaction', 'emergency', 'call your doctor', 'serious', 'common', 'rare']
                    paragraphs = self.driver.find_elements(By.TAG_NAME, "p")
                    
                    for p in paragraphs:
                        text = p.text.strip()
                        if text and len(text) > 30:
                            text_lower = text.lower()
                            if any(keyword in text_lower for keyword in keywords):
                                content_parts.append(text)
                                content_found = True
                except:
                    pass
            
            # Approach 4: Look for lists that might contain side effects
            if not content_found:
                try:
                    lists = self.driver.find_elements(By.CSS_SELECTOR, "ul, ol")
                    for list_elem in lists:
                        text = list_elem.text.strip()
                        if text and len(text) > 50:
                            text_lower = text.lower()
                            if any(keyword in text_lower for keyword in ['side effect', 'adverse', 'reaction']):
                                content_parts.append("--- Side Effects List ---")
                                content_parts.append(text)
                                content_found = True
                                break
                except:
                    pass
            
            if content_parts:
                content = '\n\n'.join(content_parts)
                content = self.clean_content(content)
                return content
            else:
                return f"No specific side effects content found for {medication}"
                
        except Exception as e:
            return f"Error extracting side effects content: {str(e)}"
    
    def clean_content(self, content):
        """Clean and format the side effects content"""
        if not content:
            return ""
        
        # Remove excessive whitespace and unwanted elements
        lines = content.split('\n')
        cleaned_lines = []
        
        for line in lines:
            line = line.strip()
            if line and not any(skip in line.lower() for skip in [
                'advertisement', 'ads by', 'sponsored', 'cookie', 'privacy',
                'terms of use', 'about us', 'contact us', 'site map',
                'subscribe', 'newsletter', 'email', 'follow us'
            ]):
                cleaned_lines.append(line)
        
        # Join lines and remove excessive blank lines
        content = '\n'.join(cleaned_lines)
        
        # Remove multiple consecutive newlines
        while '\n\n\n' in content:
            content = content.replace('\n\n\n', '\n\n')
        
        return content.strip()
    
    def add_delay(self):
        """Add random delay between requests"""
        delay = random.uniform(4, 8)  # Longer delays to avoid detection
        print(f"  ⏰ Waiting {delay:.1f} seconds before next request...")
        time.sleep(delay)
    
    def close(self):
        """Close the browser"""
        self.driver.quit()

def update_excel_with_side_effects(max_medications=None):
    """Update Excel file with side effects for all medications"""
    
    excel_path = '/Users/juanlu/Documents/Wye/scrapper/Analysis/main_diseases_analysis_final.xlsx'
    
    if not os.path.exists(excel_path):
        print(f"❌ Excel file not found: {excel_path}")
        return
    
    # Load workbook
    wb = load_workbook(excel_path)
    
    if "All Unique Medications" not in wb.sheetnames:
        print("❌ 'All Unique Medications' sheet not found")
        return
    
    ws = wb["All Unique Medications"]
    
    # Initialize scraper
    scraper = DrugsScraper(headless=False)  # Set to True for headless mode
    
    try:
        # Get all medications
        medications = []
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0]:  # If there's a medication name
                medications.append(row[0])
        
        if max_medications:
            medications = medications[:max_medications]
        
        print(f"📊 Found {len(medications)} medications to process")
        
        # Process each medication
        for i, medication in enumerate(medications, 1):
            print(f"\n{'='*60}")
            print(f"Processing {i}/{len(medications)}: {medication}")
            print(f"{'='*60}")
            
            # Check if already processed
            current_value = ws.cell(row=i+1, column=2).value
            if current_value and len(str(current_value)) > 100:
                print(f"  ⏭️  Already processed: {medication}")
                continue
            
            # Scrape side effects
            side_effects = scraper.scrape_side_effects(medication)
            
            # Update Excel
            ws.cell(row=i+1, column=2, value=side_effects)
            
            # Save progress every 5 medications
            if i % 5 == 0:
                wb.save(excel_path)
                print(f"  💾 Progress saved after {i} medications")
            
            # Add delay between requests
            if i < len(medications):
                scraper.add_delay()
        
        # Final save
        wb.save(excel_path)
        print(f"\n✅ Successfully processed {len(medications)} medications!")
        print(f"📄 Updated Excel file: {excel_path}")
        
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        wb.save(excel_path)  # Save progress even on error
    
    finally:
        scraper.close()

if __name__ == "__main__":
    print("🚀 Starting Drugs.com Side Effects Scraper")
    print("="*60)
    print("⏰ Using 4-8 second delays between requests")
    print("="*60)
    
    # Process ALL medications
    update_excel_with_side_effects()  # No limit - process all
    
    print("\n" + "="*60)
    print("🎉 FULL SCRAPING COMPLETED!")
    print("="*60)
