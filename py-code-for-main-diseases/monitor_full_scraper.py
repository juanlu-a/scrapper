#!/usr/bin/env python3
"""
Monitor the progress of the full scraper run
"""
from openpyxl import load_workbook
import time
import os
from datetime import datetime

def monitor_progress():
    """Monitor the progress of the scraper"""
    excel_path = '/Users/juanlu/Documents/Wye/scrapper/Analysis/main_diseases_analysis_final.xlsx'
    
    if not os.path.exists(excel_path):
        print("❌ Excel file not found")
        return
    
    try:
        wb = load_workbook(excel_path)
        if "All Unique Medications" not in wb.sheetnames:
            print("❌ 'All Unique Medications' sheet not found")
            return
        
        ws = wb["All Unique Medications"]
        
        # Count total medications
        total_medications = 0
        for row in ws.iter_rows(min_row=9, max_col=1, values_only=True):
            if row[0] and row[0].strip():
                total_medications += 1
        
        # Count processed medications (those with content in column G)
        processed_medications = 0
        successful_medications = 0
        
        for row in ws.iter_rows(min_row=9, max_col=7, values_only=True):
            if row[0] and row[0].strip():  # Has medication name
                if row[6]:  # Has content in column G (Full Information)
                    processed_medications += 1
                    if not str(row[6]).startswith("❌"):
                        successful_medications += 1
        
        print(f"📊 SCRAPER PROGRESS - {datetime.now().strftime('%H:%M:%S')}")
        print(f"=" * 50)
        print(f"📝 Total medications: {total_medications}")
        print(f"⚡ Processed: {processed_medications}")
        print(f"✅ Successful: {successful_medications}")
        print(f"❌ Errors: {processed_medications - successful_medications}")
        print(f"📊 Progress: {(processed_medications/total_medications)*100:.1f}%")
        
        if processed_medications > 0:
            print(f"🎯 Success rate: {(successful_medications/processed_medications)*100:.1f}%")
        
        # Show latest processed medications
        print(f"\n📋 Latest processed medications:")
        count = 0
        for row_num in range(9, 9 + total_medications):
            cell_value = ws[f'G{row_num}'].value
            if cell_value:
                med_name = ws[f'A{row_num}'].value
                status = "✅" if not str(cell_value).startswith("❌") else "❌"
                char_count = len(str(cell_value))
                print(f"   {status} {med_name} ({char_count} chars)")
                count += 1
                if count >= 5:  # Show last 5
                    break
        
        # Estimate time remaining
        if processed_medications > 0:
            avg_time_per_med = 6  # seconds (4 second delay + 2 seconds processing)
            remaining = total_medications - processed_medications
            eta_minutes = (remaining * avg_time_per_med) / 60
            print(f"\n⏱️  Estimated time remaining: {eta_minutes:.1f} minutes")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Error checking progress: {e}")

if __name__ == "__main__":
    monitor_progress()
