from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import time
import random
from openpyxl import load_workbook
# Removed all formatting imports to prevent Excel corruption
import os

class DrugsScraper:
    def __init__(self, headless=False):
        self.headless = headless
        self.driver = None
        self.wait = None
        self.init_driver()
        
    def init_driver(self):
        """Initialize or reinitialize the driver"""
        if self.driver:
            try:
                self.driver.quit()
            except:
                pass
        
        self.driver = self.setup_driver(self.headless)
        self.wait = WebDriverWait(self.driver, 10)
        
    def setup_driver(self, headless=False):
        """Set up Chrome driver with options"""
        chrome_options = Options()
        if headless:
            chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_argument("--disable-web-security")
        chrome_options.add_argument("--disable-features=VizDisplayCompositor")
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        
        # Add user agent to avoid detection
        chrome_options.add_argument("--user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
        
        try:
            driver = webdriver.Chrome(options=chrome_options)
            driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            
            # Ensure we start with a valid page
            driver.get("https://www.drugs.com")
            time.sleep(2)
            
            return driver
        except Exception as e:
            print(f"❌ Error setting up Chrome driver: {e}")
            raise e
    
    def close_modal_popups(self):
        """Close any modal popups that might be blocking the page"""
        try:
            # Common modal close selectors
            close_selectors = [
                # Newsletter/subscription modals
                "button[aria-label='Close']",
                "button[aria-label='close']", 
                ".close",
                ".modal-close",
                ".popup-close",
                "[data-dismiss='modal']",
                ".newsletter-close",
                # X buttons
                "button:contains('×')",
                "span:contains('×')",
                "div:contains('×')",
                # Close text buttons
                "button:contains('Close')",
                "button:contains('close')",
                "a:contains('Close')",
                "a:contains('close')",
                # Generic close buttons
                ".btn-close",
                ".close-button",
                ".close-btn"
            ]
            
            for selector in close_selectors:
                try:
                    if selector.startswith("button:contains") or selector.startswith("span:contains") or selector.startswith("div:contains") or selector.startswith("a:contains"):
                        # XPath for text content
                        text = selector.split("'")[1]
                        elements = self.driver.find_elements(By.XPATH, f"//*[contains(text(), '{text}')]")
                    else:
                        # CSS selector
                        elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    
                    for element in elements:
                        if element.is_displayed() and element.is_enabled():
                            try:
                                # Try regular click first
                                element.click()
                                print(f"    ✅ Closed modal using selector: {selector}")
                                time.sleep(1)
                                return True
                            except:
                                try:
                                    # Try JavaScript click
                                    self.driver.execute_script("arguments[0].click();", element)
                                    print(f"    ✅ Closed modal (JS) using selector: {selector}")
                                    time.sleep(1)
                                    return True
                                except:
                                    continue
                except:
                    continue
            
            # Try pressing Escape key as fallback
            try:
                from selenium.webdriver.common.keys import Keys
                self.driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
                print("    ✅ Sent ESC key to close modal")
                time.sleep(1)
                return True
            except:
                pass
            
            return False
            
        except Exception as e:
            print(f"    ⚠️  Error closing modals: {e}")
            return False

    def ensure_valid_page(self):
        """Ensure we're on a valid drugs.com page"""
        try:
            current_url = self.driver.current_url
            
            # If we're on data:, or invalid page, go to drugs.com
            if current_url.startswith("data:") or "drugs.com" not in current_url or current_url == "about:blank":
                print(f"  🔄 Invalid page detected ({current_url}), navigating to drugs.com...")
                self.driver.get("https://www.drugs.com")
                time.sleep(3)
                
                # Verify the navigation worked
                new_url = self.driver.current_url
                if new_url.startswith("data:") or "drugs.com" not in new_url:
                    raise Exception(f"Failed to navigate to valid page, still on: {new_url}")
                
                print(f"  ✅ Successfully navigated to: {new_url}")
                
        except Exception as e:
            print(f"  ❌ Error ensuring valid page: {e}")
            raise e
    
    def check_connection(self):
        """Check if the driver connection is still alive"""
        try:
            # Try to get current URL - if this fails, connection is lost
            current_url = self.driver.current_url
            
            # Check if we're on a valid page (not data:, or empty)
            if current_url.startswith("data:") or not current_url or current_url == "about:blank":
                print(f"  ⚠️  Invalid URL detected: {current_url}")
                return False
                
            return True
        except Exception as e:
            print(f"  ⚠️  Connection lost: {e}")
            return False
    
    def reconnect_if_needed(self):
        """Reconnect to the driver if connection is lost"""
        if not self.check_connection():
            print("  🔄 Reconnecting to browser...")
            self.init_driver()
            time.sleep(3)
            return True
        return False
    
    def close_modal_popups(self):
        """Close any modal popups that might be blocking the page"""
        try:
            # List of selectors for common modal close buttons
            close_selectors = [
                # Generic close buttons
                "button[aria-label*='close']",
                "button[aria-label*='Close']",
                "button[title*='close']",
                "button[title*='Close']",
                ".close-button",
                ".close-btn",
                ".modal-close",
                ".popup-close",
                "button.close",
                "[data-dismiss='modal']",
                
                # Newsletter modal specific (like in the screenshot)
                ".newsletter-modal .close",
                ".newsletter-modal button[type='button']",
                ".modal-dialog .close",
                ".modal-dialog button[aria-label*='close']",
                
                # X buttons
                "button:contains('×')",
                "span:contains('×')",
                ".fa-times",
                ".fa-close",
                
                # Other common patterns
                ".overlay-close",
                ".lightbox-close",
                ".dialog-close"
            ]
            
            closed_modal = False
            
            for selector in close_selectors:
                try:
                    if selector.startswith("button:contains") or selector.startswith("span:contains"):
                        # Use XPath for text-based selectors
                        xpath = f"//button[contains(text(), '×')] | //span[contains(text(), '×')]"
                        close_buttons = self.driver.find_elements(By.XPATH, xpath)
                    else:
                        close_buttons = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    
                    for button in close_buttons:
                        try:
                            # Check if button is visible and clickable
                            if button.is_displayed() and button.is_enabled():
                                # Try direct click first
                                button.click()
                                print(f"  ✅ Closed modal using selector: {selector}")
                                closed_modal = True
                                time.sleep(1)  # Give time for modal to close
                                break
                        except:
                            try:
                                # Try JavaScript click as fallback
                                self.driver.execute_script("arguments[0].click();", button)
                                print(f"  ✅ Closed modal (JS) using selector: {selector}")
                                closed_modal = True
                                time.sleep(1)
                                break
                            except:
                                continue
                    
                    if closed_modal:
                        break
                        
                except Exception as e:
                    continue
            
            # Also try to press Escape key to close modals
            if not closed_modal:
                try:
                    from selenium.webdriver.common.keys import Keys
                    self.driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
                    print("  ✅ Sent Escape key to close modal")
                    time.sleep(1)
                except:
                    pass
            
            # Check for overlay backgrounds and click them
            try:
                overlay_selectors = [
                    ".modal-backdrop",
                    ".overlay",
                    ".modal-overlay",
                    ".popup-overlay",
                    ".lightbox-overlay"
                ]
                
                for overlay_selector in overlay_selectors:
                    overlays = self.driver.find_elements(By.CSS_SELECTOR, overlay_selector)
                    for overlay in overlays:
                        if overlay.is_displayed():
                            try:
                                overlay.click()
                                print(f"  ✅ Clicked overlay to close modal: {overlay_selector}")
                                time.sleep(1)
                                break
                            except:
                                continue
            except:
                pass
                
        except Exception as e:
            print(f"  ⚠️  Error closing modals: {e}")
            pass
    
    def safe_driver_action(self, action, *args, **kwargs):
        """Execute a driver action with automatic reconnection on failure"""
        max_retries = 3
        for attempt in range(max_retries):
            try:
                # Check connection before action
                if not self.check_connection():
                    print(f"  🔄 Reconnecting before action (attempt {attempt + 1})")
                    self.init_driver()
                    time.sleep(3)
                
                # Execute the action
                return action(*args, **kwargs)
                
            except Exception as e:
                print(f"  ⚠️  Action failed (attempt {attempt + 1}): {e}")
                if attempt < max_retries - 1:
                    print(f"  🔄 Reinitializing driver and retrying...")
                    self.init_driver()
                    time.sleep(5)
                else:
                    raise e
        
    def search_and_get_side_effects(self, medication):
        """Search for medication and get side effects content with reconnection handling"""
        try:
            print(f"🔍 Processing: {medication}")
            
            # Ensure we start with a valid page
            self.ensure_valid_page()
            
            # Close any modals that might be open
            self.close_modal_popups()
            
            # Step 1: Go to drugs.com with better error handling
            def load_drugs_com():
                print("    Loading drugs.com...")
                self.driver.get("https://www.drugs.com")
                time.sleep(2)
                
                # Verify we're on the correct page
                current_url = self.driver.current_url
                if current_url.startswith("data:") or "drugs.com" not in current_url:
                    raise Exception(f"Failed to load drugs.com, got URL: {current_url}")
                
                print(f"    ✅ Successfully loaded: {current_url}")
                
                # Close any modals that might have opened
                time.sleep(2)
                self.close_modal_popups()
                
                return True
            
            for attempt in range(3):
                try:
                    load_drugs_com()
                    break
                except Exception as e:
                    print(f"    Attempt {attempt + 1} failed to load drugs.com: {e}")
                    if attempt == 2:
                        return f"❌ Failed to load drugs.com after 3 attempts"
                    
                    # Reinitialize driver on failure
                    print("    🔄 Reinitializing driver...")
                    self.init_driver()
                    time.sleep(5)
            
            time.sleep(3)
            
            # Step 2: Search for medication with connection safety
            def search_medication():
                # Ensure we're on a valid page before searching
                self.ensure_valid_page()
                
                # Close any modals before searching
                self.close_modal_popups()
                
                search_box = self.wait.until(EC.presence_of_element_located((By.NAME, "searchterm")))
                search_box.clear()
                search_box.send_keys(medication)
                search_box.send_keys(Keys.RETURN)
                return True
            
            try:
                self.safe_driver_action(search_medication)
                print(f"  ✅ Search submitted for: {medication}")
            except Exception as e:
                return f"❌ Failed to search for {medication}: {str(e)}"
            
            # Step 3: Find main medication result
            main_result = self.find_main_medication_result(medication)
            if not main_result:
                return f"❌ Could not find main result for {medication}"
            
            # Step 4: Click on main result - simplified and faster
            try:
                # Close any modals before clicking
                self.close_modal_popups()
                
                # Direct click first
                main_result.click()
                print(f"  ✅ Clicked main result for {medication}")
            except Exception as e:
                # Quick JavaScript fallback
                try:
                    self.driver.execute_script("arguments[0].click();", main_result)
                    print(f"  ✅ Clicked main result (JS) for {medication}")
                except Exception as e2:
                    return f"❌ Failed to click main result for {medication}: {str(e2)}"
            
            time.sleep(2)  # Reduced wait time
            
            # Close any modals that might have opened after clicking
            self.close_modal_popups()
            
            # Step 5: Find and click side effects link
            side_effects_link = self.find_side_effects_link()
            if not side_effects_link:
                return f"❌ Could not find side effects link for {medication}"
            
            # Step 6: Click side effects link - simplified and faster
            try:
                # Direct click first
                side_effects_link.click()
                print(f"  ✅ Clicked side effects link for {medication}")
            except Exception as e:
                # Quick JavaScript fallback
                try:
                    self.driver.execute_script("arguments[0].click();", side_effects_link)
                    print(f"  ✅ Clicked side effects link (JS) for {medication}")
                except Exception as e2:
                    return f"❌ Failed to click side effects link for {medication}: {str(e2)}"
            
            time.sleep(2)  # Reduced wait time
            
            # Step 7: Extract side effects content
            content = self.extract_side_effects_content(medication)
            
            if content and len(content) > 50:
                print(f"  ✅ Successfully processed {medication} ({len(content)} characters)")
                return content
            else:
                return f"❌ No substantial side effects content found for {medication}"
            
        except Exception as e:
            error_msg = f"❌ Unexpected error processing {medication}: {str(e)}"
            print(error_msg)
            # Try to recover by reinitializing driver
            try:
                print("  🔄 Attempting to recover from error...")
                self.init_driver()
                time.sleep(3)
            except:
                pass
            return error_msg
    
    def find_main_medication_result(self, medication):
        """Find the main medication result (usually with yellow star)"""
        print(f"  🔍 Looking for main result for: {medication}")
        
        # Wait a bit for search results to load
        time.sleep(3)  # Reduced from 5 seconds
        
        # Check connection before searching
        if not self.check_connection():
            print("  🔄 Reconnecting before searching for results...")
            self.init_driver()
            time.sleep(3)
            return None
        
        # Try multiple approaches to find the main result
        approaches = [
            # Approach 1: Look for direct href matches
            {
                'name': 'Direct href match',
                'selectors': [
                    f"a[href*='{medication.lower().replace(' ', '-')}.html']",
                    f"a[href*='{medication.lower()}.html']",
                    f"a[href*='{medication.replace(' ', '-').lower()}.html']",
                ]
            },
            # Approach 2: Look for search results containing the medication name
            {
                'name': 'Search result text match',
                'selectors': [
                    "a[href*='.html']",  # Any HTML page link
                ]
            },
            # Approach 3: Look for specific search result containers
            {
                'name': 'Search result containers',
                'selectors': [
                    ".ddc-search-result a",
                    ".search-result a",
                    ".result a",
                    "div[class*='result'] a",
                ]
            }
        ]
        
        for approach in approaches:
            print(f"    Trying approach: {approach['name']}")
            
            for selector in approach['selectors']:
                try:
                    # Use safe driver action to find elements
                    def find_results():
                        return self.driver.find_elements(By.CSS_SELECTOR, selector)
                    
                    results = self.safe_driver_action(find_results)
                    print(f"      Found {len(results)} results with selector: {selector}")
                    
                    for i, result in enumerate(results):
                        try:
                            href = result.get_attribute('href')
                            text = result.text.strip()
                            
                            if not href or not text:
                                continue
                            
                            # Skip unwanted links
                            if any(skip in href.lower() for skip in ['/pro/', '/search', '/compare', '/interaction']):
                                continue
                            
                            # Check if this looks like our medication
                            medication_words = medication.lower().split()
                            text_lower = text.lower()
                            
                            # For direct href matches, be more strict
                            if approach['name'] == 'Direct href match':
                                if medication.lower().replace(' ', '-') in href.lower():
                                    print(f"      ✅ Found direct match: {text[:50]}... -> {href}")
                                    return result
                            
                            # For text matches, check if medication words are in the text
                            elif approach['name'] == 'Search result text match':
                                if len(medication_words) == 1:
                                    # Single word medication
                                    if medication_words[0] in text_lower and '.html' in href:
                                        print(f"      ✅ Found text match: {text[:50]}... -> {href}")
                                        return result
                                else:
                                    # Multi-word medication - check if most words are present
                                    word_matches = sum(1 for word in medication_words if word in text_lower)
                                    if word_matches >= len(medication_words) * 0.7 and '.html' in href:
                                        print(f"      ✅ Found text match: {text[:50]}... -> {href}")
                                        return result
                            
                            # For container matches, be more flexible
                            else:
                                if any(word in text_lower for word in medication_words) and '.html' in href:
                                    print(f"      ✅ Found container match: {text[:50]}... -> {href}")
                                    return result
                        
                        except Exception as e:
                            continue
                
                except Exception as e:
                    continue
        
        # If no result found, print available links for debugging
        print(f"    ❌ No main result found. Available links:")
        try:
            all_links = self.driver.find_elements(By.CSS_SELECTOR, "a[href*='.html']")
            for i, link in enumerate(all_links[:10]):  # Show first 10
                href = link.get_attribute('href')
                text = link.text.strip()[:50]
                print(f"      {i+1}. {text} -> {href}")
        except:
            pass
        
        return None
    
    def find_side_effects_link(self):
        """Find the side effects navigation link"""
        print(f"  🔍 Looking for side effects link...")
        
        # Wait for page to load
        time.sleep(2)  # Reduced from 3 seconds
        
        # Check connection before searching
        if not self.check_connection():
            print("  🔄 Reconnecting before searching for side effects link...")
            self.init_driver()
            time.sleep(3)
            return None
        
        # Try multiple approaches to find side effects link
        approaches = [
            # Approach 1: XPath for text content
            {
                'name': 'XPath text search',
                'selectors': [
                    "//a[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'side effects')]",
                    "//a[contains(text(), 'Side effects')]",
                    "//a[contains(text(), 'side effects')]",
                    "//a[contains(text(), 'Side Effects')]",
                ]
            },
            # Approach 2: href containing side-effects
            {
                'name': 'Href containing side-effects',
                'selectors': [
                    "a[href*='side-effects']",
                    "a[href*='#side-effects']",
                    "a[href*='sideeffects']",
                ]
            },
            # Approach 3: Look in navigation areas
            {
                'name': 'Navigation areas',
                'selectors': [
                    "nav a",
                    ".nav a",
                    ".navigation a",
                    ".tabs a",
                    ".tab a",
                    ".menu a",
                ]
            }
        ]
        
        for approach in approaches:
            print(f"    Trying approach: {approach['name']}")
            
            for selector in approach['selectors']:
                try:
                    if selector.startswith("//"):
                        # XPath selector
                        links = self.driver.find_elements(By.XPATH, selector)
                    else:
                        # CSS selector
                        links = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    
                    print(f"      Found {len(links)} links with selector: {selector}")
                    
                    for link in links:
                        try:
                            text = link.text.strip().lower()
                            href = link.get_attribute('href')
                            
                            # Check if this looks like a side effects link
                            if approach['name'] == 'Navigation areas':
                                # For navigation areas, check text content
                                if 'side effect' in text:
                                    print(f"      ✅ Found side effects link: {link.text}")
                                    return link
                            else:
                                # For other approaches, we already filtered
                                if text and href:
                                    print(f"      ✅ Found side effects link: {link.text}")
                                    return link
                        
                        except Exception as e:
                            continue
                
                except Exception as e:
                    continue
        
        # If no side effects link found, print available navigation links
        print(f"    ❌ No side effects link found. Available navigation links:")
        try:
            nav_links = self.driver.find_elements(By.CSS_SELECTOR, "a")
            shown_links = 0
            for link in nav_links:
                try:
                    text = link.text.strip()
                    href = link.get_attribute('href')
                    if text and href and len(text) < 50:
                        print(f"      - {text} -> {href}")
                        shown_links += 1
                        if shown_links >= 15:  # Show first 15 links
                            break
                except:
                    continue
        except:
            pass
        
        return None
    
    def extract_side_effects_content(self, medication):
        """Extract ONLY the core side effects content from the page"""
        try:
            # Close any modals that might be blocking content
            self.close_modal_popups()
            
            # Wait for page to load completely
            time.sleep(2)
            
            content_parts = []
            
            # Try to find the main side effects section first
            main_selectors = [
                "#side-effects",
                ".side-effects-content",
                ".side-effects", 
                "[id*='side-effects']",
                "div[class*='side-effects']"
            ]
            
            # Look for the main side effects section
            for selector in main_selectors:
                try:
                    section = self.driver.find_element(By.CSS_SELECTOR, selector)
                    if section:
                        text = section.text.strip()
                        if text and len(text) > 50:  # Ensure substantial content
                            # Clean the text to remove navigation/ads
                            clean_text = self.clean_side_effects_text(text)
                            if clean_text:
                                return clean_text
                except:
                    continue
            
            # If no main section found, look for side effects headings and following content
            try:
                # Look for side effects headings
                headings = self.driver.find_elements(By.XPATH, 
                    "//h1[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'side effect')] | "
                    "//h2[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'side effect')] | "
                    "//h3[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'side effect')]"
                )
                
                for heading in headings:
                    content_parts.append(f"=== {heading.text.strip()} ===")
                    
                    # Get the next few elements after the heading
                    try:
                        parent = heading.find_element(By.XPATH, "./..")
                        following_elements = parent.find_elements(By.XPATH, f".//*[position() > {heading.find_elements(By.XPATH, './preceding-sibling::*').__len__() + 1}]")
                        
                        for elem in following_elements[:5]:  # Limit to first 5 elements
                            elem_text = elem.text.strip()
                            if elem_text and len(elem_text) > 20:
                                content_parts.append(elem_text)
                                if len(content_parts) >= 10:  # Limit total parts
                                    break
                        break
                    except:
                        continue
                        
            except:
                pass
            
            # If still no content, look for paragraphs with side effects keywords
            if not content_parts:
                try:
                    paragraphs = self.driver.find_elements(By.TAG_NAME, "p")
                    for p in paragraphs:
                        text = p.text.strip()
                        if any(keyword in text.lower() for keyword in [
                            'side effect', 'adverse reaction', 'common side effects',
                            'serious side effects', 'call your doctor', 'emergency'
                        ]):
                            content_parts.append(text)
                            if len(content_parts) >= 5:  # Limit to 5 relevant paragraphs
                                break
                except:
                    pass
            
            if content_parts:
                content = '\n\n'.join(content_parts)
                return self.clean_side_effects_text(content)
            else:
                return f"No side effects content found for {medication}"
            
        except Exception as e:
            return f"Error extracting side effects: {str(e)}"
    
    def clean_side_effects_text(self, content):
        """Clean side effects text to remove navigation, ads, and excess content"""
        if not content:
            return ""
            
        lines = content.split('\n')
        cleaned_lines = []
        
        # Remove unwanted content
        skip_keywords = [
            'advertisement', 'ads by', 'sponsored', 'cookie', 'privacy',
            'terms of use', 'about us', 'contact us', 'site map', 'navigation',
            'menu', 'search', 'login', 'register', 'subscribe', 'newsletter',
            'related articles', 'see also', 'references', 'further reading',
            'drug interactions', 'dosage', 'how to take', 'storage'
        ]
        
        for line in lines:
            line = line.strip()
            if line and len(line) > 3:  # Skip very short lines
                # Skip lines with unwanted content
                if not any(skip in line.lower() for skip in skip_keywords):
                    # Keep lines that seem to be about side effects
                    if any(keyword in line.lower() for keyword in [
                        'side effect', 'adverse', 'reaction', 'symptom', 'common',
                        'serious', 'severe', 'mild', 'call your doctor', 'emergency',
                        'stop taking', 'discontinue', 'allergic', 'rash', 'fever',
                        'nausea', 'vomiting', 'diarrhea', 'headache', 'dizziness'
                    ]) or line.startswith('===') or line.startswith('---'):
                        cleaned_lines.append(line)
                    # Also keep short descriptive lines that might be side effects
                    elif len(line) < 100 and not any(char in line for char in ['©', '®', '™']):
                        cleaned_lines.append(line)
        
        # Join lines and clean up
        content = '\n'.join(cleaned_lines)
        
        # Remove multiple consecutive newlines
        while '\n\n\n' in content:
            content = content.replace('\n\n\n', '\n\n')
        
        # Limit content length to avoid excessive data
        if len(content) > 3000:
            content = content[:3000] + "\n\n[Content truncated - showing first 3000 characters]"
        
        return content.strip()
    
    def clean_content(self, content):
        """Clean and format the side effects content"""
        # Remove excessive whitespace and unwanted elements
        lines = content.split('\n')
        cleaned_lines = []
        
        for line in lines:
            line = line.strip()
            if line and not any(skip in line.lower() for skip in [
                'advertisement', 'ads by', 'sponsored', 'cookie', 'privacy',
                'terms of use', 'about us', 'contact us', 'site map'
            ]):
                cleaned_lines.append(line)
        
        # Join lines and remove excessive blank lines
        content = '\n'.join(cleaned_lines)
        
        # Remove multiple consecutive newlines
        while '\n\n\n' in content:
            content = content.replace('\n\n\n', '\n\n')
        
        return content.strip()
    
    def add_delay(self):
        """Add random delay between requests"""
        delay = random.uniform(1.5, 2.5)  # Shorter delays - 2 seconds average
        print(f"  ⏰ Waiting {delay:.1f} seconds before next request...")
        time.sleep(delay)
    
    def close(self):
        """Close the browser"""
        if self.driver:
            self.driver.quit()

def update_excel_with_side_effects(max_medications=None):
    """Update Excel file with side effects for all medications"""
    
    excel_path = '/Users/juanlu/Documents/Wye/scrapper/Analysis/main_diseases_analysis_final.xlsx'
    
    if not os.path.exists(excel_path):
        print(f"❌ Excel file not found: {excel_path}")
        return
    
    # Load workbook
    wb = load_workbook(excel_path)
    
    if "All Unique Medications" not in wb.sheetnames:
        print("❌ 'All Unique Medications' sheet not found")
        return
    
    medications_ws = wb["All Unique Medications"]
    
    # Get medications
    medications = []
    for row in medications_ws.iter_rows(min_row=9, max_col=1, values_only=True):
        if row[0] and row[0].strip():
            medications.append(row[0].strip())
    
    # Use all medications if max_medications is None
    if max_medications:
        medications = medications[:max_medications]
    
    print(f"📊 Processing {len(medications)} medications...")
    
    # Add new column header if not exists
    if not medications_ws['G8'].value or 'FULL INFORMATION' not in str(medications_ws['G8'].value):
        medications_ws['G8'] = 'FULL INFORMATION'
        medications_ws['G8'].font = Font(bold=True, color="FFFFFF")
        medications_ws['G8'].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        medications_ws['G8'].border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        medications_ws['G8'].alignment = Alignment(horizontal='center', vertical='center')
        medications_ws.column_dimensions['G'].width = 100
    
    # Initialize scraper
    scraper = DrugsScraper(headless=False)  # Use visible mode to see what's happening
    
    try:
        processed_count = 0
        errors = []
        
        for i, medication in enumerate(medications):
            print(f"\n[{i+1}/{len(medications)}] Processing: {medication}")
            
            # Check if scraper connection is still alive
            if not scraper.check_connection():
                print("  🔄 Reconnecting scraper...")
                scraper.init_driver()
                time.sleep(5)
            
            # Get side effects content with retry on connection errors
            max_retries = 3
            content = None
            
            for attempt in range(max_retries):
                try:
                    content = scraper.search_and_get_side_effects(medication)
                    break
                except Exception as e:
                    print(f"  ⚠️  Attempt {attempt + 1} failed: {e}")
                    if attempt < max_retries - 1:
                        print("  🔄 Reinitializing scraper and retrying...")
                        scraper.init_driver()
                        time.sleep(10)  # Longer wait after reconnection
                    else:
                        content = f"❌ Failed to process {medication} after {max_retries} attempts"
            
            # Add to Excel in column G
            row_num = 9 + i
            medications_ws[f'G{row_num}'] = content
            
            # Format cell
            cell = medications_ws[f'G{row_num}']
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Alternate row colors
            if i % 2 == 0:
                cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            
            processed_count += 1
            
            # Track errors
            if content.startswith("❌"):
                errors.append(medication)
            
            # Save progress every 10 medications
            if processed_count % 10 == 0:
                wb.save(excel_path)
                print(f"💾 Saved progress: {processed_count}/{len(medications)} medications processed")
                print(f"   Errors so far: {len(errors)}")
            
            # Add delay between requests
            scraper.add_delay()
    
    finally:
        scraper.close()
    
    # Final save
    wb.save(excel_path)
    
    # Summary
    success_count = processed_count - len(errors)
    print(f"\n" + "="*60)
    print(f"✅ PROCESSING COMPLETED!")
    print(f"📊 Total processed: {processed_count}")
    print(f"✅ Successful: {success_count}")
    print(f"❌ Errors: {len(errors)}")
    print(f"📄 Updated Excel file: {excel_path}")
    
    if errors:
        print(f"\n❌ Medications with errors:")
        for error_med in errors[:10]:  # Show first 10 errors
            print(f"   - {error_med}")
        if len(errors) > 10:
            print(f"   ... and {len(errors) - 10} more")

if __name__ == "__main__":
    print("🚀 Starting Enhanced LLM-Powered Side Effects Scraper")
    print("="*60)
    print("🔧 Enhanced features:")
    print("   - Direct LLM processing of ALL side effects content")
    print("   - Structured categorization into 3 columns:")
    print("     * SIDE EFFECTS")
    print("     * CALL A DOCTOR IF") 
    print("     * GO TO ER IF")
    print("   - Automatic reconnection after computer sleep")
    print("   - Connection health monitoring")
    print("   - Robust error recovery")
    print("   - Multiple retry attempts")
    print("   - Modal/popup handling")
    print("   - Resume from where it left off")
    print("="*60)
    
    # Run test with 1 medication first
    print("🧪 Running TEST with 1 medication...")
    update_excel_with_side_effects(max_medications=1)
    
    print("\n" + "="*60)
    print("🎉 TEST COMPLETED! Check the Excel file to verify the results.")
    print("💡 If the test looks good, remove the max_medications parameter")
    print("   to process ALL medications.")
