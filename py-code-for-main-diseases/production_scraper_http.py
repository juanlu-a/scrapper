import requests
from bs4 import BeautifulSoup
import time
import random
from openpyxl import load_workbook
import os
import urllib.parse

class DrugsScraper:
    def __init__(self):
        self.session = requests.Session()
        
        # Set up headers to appear more human-like
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.5',
            'Accept-Encoding': 'gzip, deflate',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1'
        })
    
    def scrape_side_effects(self, medication, max_retries=3):
        """Scrape side effects for a specific medication"""
        print(f"\n🔍 Processing: {medication}")
        
        for attempt in range(max_retries):
            try:
                if attempt > 0:
                    print(f"  🔄 Retry attempt {attempt + 1}/{max_retries}")
                
                # Step 1: Search for medication
                search_url = f"https://www.drugs.com/search.php?searchterm={urllib.parse.quote(medication)}"
                print(f"  🔍 Searching: {search_url}")
                
                search_response = self.session.get(search_url)
                if search_response.status_code != 200:
                    print(f"  ❌ Search failed with status {search_response.status_code}")
                    continue
                
                search_soup = BeautifulSoup(search_response.content, 'html.parser')
                
                # Step 2: Find main result
                main_result_url = self.find_main_result(search_soup, medication)
                if not main_result_url:
                    print(f"  ❌ Could not find main result for {medication}")
                    continue
                
                # Step 3: Get the main page
                if not main_result_url.startswith('http'):
                    main_result_url = f"https://www.drugs.com{main_result_url}"
                
                print(f"  ✅ Found main page: {main_result_url}")
                
                main_response = self.session.get(main_result_url)
                if main_response.status_code != 200:
                    print(f"  ❌ Main page failed with status {main_response.status_code}")
                    continue
                
                main_soup = BeautifulSoup(main_response.content, 'html.parser')
                
                # Step 4: Look for side effects link
                side_effects_url = self.find_side_effects_link(main_soup, main_result_url)
                
                if side_effects_url:
                    # Step 5: Get the side effects page
                    if not side_effects_url.startswith('http'):
                        side_effects_url = f"https://www.drugs.com{side_effects_url}"
                    
                    print(f"  ✅ Found side effects page: {side_effects_url}")
                    
                    side_effects_response = self.session.get(side_effects_url)
                    if side_effects_response.status_code != 200:
                        print(f"  ❌ Side effects page failed with status {side_effects_response.status_code}")
                        continue
                    
                    side_effects_soup = BeautifulSoup(side_effects_response.content, 'html.parser')
                    content = self.extract_side_effects_content(side_effects_soup, medication)
                else:
                    print(f"  ℹ️  No side effects link found, checking main page")
                    content = self.extract_side_effects_content(main_soup, medication)
                
                if content and len(content) > 50:
                    print(f"  ✅ Successfully processed {medication} ({len(content)} characters)")
                    return content
                else:
                    print(f"  ❌ No substantial side effects content found for {medication}")
                    if attempt < max_retries - 1:
                        continue
                    else:
                        return f"No detailed side effects information found for {medication}"
                
            except Exception as e:
                error_msg = f"❌ Error processing {medication} (attempt {attempt + 1}): {str(e)}"
                print(error_msg)
                if attempt < max_retries - 1:
                    time.sleep(3)  # Wait before retry
                else:
                    return f"Error: {str(e)}"
        
        return f"Failed to process {medication} after {max_retries} attempts"
    
    def find_main_result(self, soup, medication):
        """Find the main medication result URL"""
        medication_words = [word.lower() for word in medication.split() if len(word) > 2]
        
        # Try different selectors for search results
        selectors = [
            'a[href*=".html"]',
            '.search-results a',
            '.result a',
            '.drug-results a',
            'a[href*="/drug/"]',
            'li a[href*=".html"]',
            '.search-result a',
            '.results a'
        ]
        
        for selector in selectors:
            results = soup.select(selector)
            print(f"    Found {len(results)} results with selector: {selector}")
            
            for result in results:
                try:
                    href = result.get('href')
                    if not href or '.html' not in href:
                        continue
                    
                    text = result.get_text().strip().lower()
                    if not text:
                        continue
                    
                    # Check if this looks like a main medication result
                    if len(medication_words) == 1:
                        # Single word medication
                        if medication_words[0] in text:
                            print(f"      ✅ Found single word match: {text[:50]}...")
                            return href
                    else:
                        # Multi-word medication - check for word matches
                        word_matches = sum(1 for word in medication_words if word in text)
                        if word_matches >= len(medication_words) * 0.6:
                            print(f"      ✅ Found multi-word match: {text[:50]}...")
                            return href
                
                except Exception as e:
                    continue
        
        return None
    
    def find_side_effects_link(self, soup, base_url):
        """Find the side effects link"""
        # Try different approaches to find side effects link
        
        # Approach 1: Look for links with "side effects" in text
        side_effects_links = soup.find_all('a', string=lambda text: text and 'side effects' in text.lower())
        if side_effects_links:
            href = side_effects_links[0].get('href')
            if href:
                return href
        
        # Approach 2: Look for links with "side-effects" in href
        side_effects_links = soup.find_all('a', href=lambda href: href and 'side-effects' in href)
        if side_effects_links:
            return side_effects_links[0].get('href')
        
        # Approach 3: Look in navigation areas
        nav_areas = soup.find_all(['nav', 'div', 'ul'], class_=lambda x: x and ('nav' in x or 'tab' in x or 'menu' in x))
        for nav in nav_areas:
            links = nav.find_all('a')
            for link in links:
                text = link.get_text().strip().lower()
                if 'side effect' in text or 'adverse' in text:
                    href = link.get('href')
                    if href:
                        return href
        
        return None
    
    def extract_side_effects_content(self, soup, medication):
        """Extract side effects content from the page"""
        print(f"  📄 Extracting side effects content for {medication}")
        
        content_parts = []
        
        # Try multiple approaches to find side effects content
        
        # Approach 1: Look for sections with side effects in ID or class
        side_effects_sections = soup.find_all(['div', 'section'], 
                                            id=lambda x: x and 'side-effects' in x)
        side_effects_sections.extend(soup.find_all(['div', 'section'], 
                                                  class_=lambda x: x and any('side-effects' in str(cls) for cls in x)))
        
        for section in side_effects_sections:
            text = section.get_text().strip()
            if text and len(text) > 50:
                content_parts.append("--- Side Effects ---")
                content_parts.append(text)
                break
        
        # Approach 2: Look for headings containing "side effects"
        if not content_parts:
            headings = soup.find_all(['h1', 'h2', 'h3'], 
                                   string=lambda text: text and 'side effects' in text.lower())
            
            for heading in headings:
                content_parts.append(f"--- {heading.get_text()} ---")
                
                # Find following content
                current = heading.find_next_sibling()
                count = 0
                while current and count < 5:
                    if current.name in ['h1', 'h2', 'h3']:
                        break
                    
                    text = current.get_text().strip()
                    if text and len(text) > 20:
                        content_parts.append(text)
                    
                    current = current.find_next_sibling()
                    count += 1
                
                if content_parts:
                    break
        
        # Approach 3: Look for paragraphs with side effects keywords
        if not content_parts:
            keywords = ['side effect', 'adverse', 'reaction', 'emergency', 'call your doctor', 'serious', 'common', 'rare']
            paragraphs = soup.find_all('p')
            
            for p in paragraphs:
                text = p.get_text().strip()
                if text and len(text) > 30:
                    text_lower = text.lower()
                    if any(keyword in text_lower for keyword in keywords):
                        content_parts.append(text)
        
        # Approach 4: Look for lists that might contain side effects
        if not content_parts:
            lists = soup.find_all(['ul', 'ol'])
            for list_elem in lists:
                text = list_elem.get_text().strip()
                if text and len(text) > 50:
                    text_lower = text.lower()
                    if any(keyword in text_lower for keyword in ['side effect', 'adverse', 'reaction']):
                        content_parts.append("--- Side Effects List ---")
                        content_parts.append(text)
                        break
        
        if content_parts:
            content = '\n\n'.join(content_parts)
            content = self.clean_content(content)
            return content
        else:
            return f"No specific side effects content found for {medication}"
    
    def clean_content(self, content):
        """Clean and format the side effects content"""
        if not content:
            return ""
        
        # Remove excessive whitespace and unwanted elements
        lines = content.split('\n')
        cleaned_lines = []
        
        for line in lines:
            line = line.strip()
            if line and not any(skip in line.lower() for skip in [
                'advertisement', 'ads by', 'sponsored', 'cookie', 'privacy',
                'terms of use', 'about us', 'contact us', 'site map',
                'subscribe', 'newsletter', 'email', 'follow us'
            ]):
                cleaned_lines.append(line)
        
        # Join lines and remove excessive blank lines
        content = '\n'.join(cleaned_lines)
        
        # Remove multiple consecutive newlines
        while '\n\n\n' in content:
            content = content.replace('\n\n\n', '\n\n')
        
        return content.strip()
    
    def add_delay(self):
        """Add random delay between requests"""
        delay = random.uniform(2, 5)  # Shorter delays for HTTP requests
        print(f"  ⏰ Waiting {delay:.1f} seconds before next request...")
        time.sleep(delay)

def update_excel_with_side_effects(max_medications=None):
    """Update Excel file with side effects for all medications"""
    
    excel_path = '/Users/juanlu/Documents/Wye/scrapper/Analysis/main_diseases_analysis_final.xlsx'
    
    if not os.path.exists(excel_path):
        print(f"❌ Excel file not found: {excel_path}")
        return
    
    # Load workbook
    wb = load_workbook(excel_path)
    
    if "All Unique Medications" not in wb.sheetnames:
        print("❌ 'All Unique Medications' sheet not found")
        return
    
    ws = wb["All Unique Medications"]
    
    # Initialize scraper
    scraper = DrugsScraper()
    
    try:
        # Get all medications (starting from row 9 where actual data begins)
        medications = []
        for row in range(9, ws.max_row + 1):  # Start from row 9
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and isinstance(cell_value, str) and cell_value.strip():
                medications.append((row, cell_value.strip()))
        
        if max_medications:
            medications = medications[:max_medications]
        
        print(f"📊 Found {len(medications)} medications to process")
        
        # Process each medication
        for i, (row_num, medication) in enumerate(medications, 1):
            print(f"\n{'='*60}")
            print(f"Processing {i}/{len(medications)}: {medication}")
            print(f"{'='*60}")
            
            # Check if already processed
            try:
                current_value = ws.cell(row=row_num, column=2).value
                if current_value and isinstance(current_value, str) and len(current_value) > 100:
                    print(f"  ⏭️  Already processed: {medication}")
                    continue
            except Exception as e:
                print(f"  ⚠️  Error checking existing value: {e}")
            
            # Scrape side effects
            side_effects = scraper.scrape_side_effects(medication)
            
            # Update Excel
            try:
                ws.cell(row=row_num, column=2, value=side_effects)
                print(f"  ✅ Updated Excel for {medication}")
            except Exception as e:
                print(f"  ❌ Error updating Excel for {medication}: {e}")
            
            # Save progress every 5 medications
            if i % 5 == 0:
                try:
                    wb.save(excel_path)
                    print(f"  💾 Progress saved after {i} medications")
                except Exception as e:
                    print(f"  ❌ Error saving progress: {e}")
            
            # Add delay between requests
            if i < len(medications):
                scraper.add_delay()
        
        # Final save
        try:
            wb.save(excel_path)
            print(f"\n✅ Successfully processed {len(medications)} medications!")
            print(f"📄 Updated Excel file: {excel_path}")
        except Exception as e:
            print(f"❌ Error saving final file: {e}")
        
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        try:
            wb.save(excel_path)  # Save progress even on error
        except:
            pass

if __name__ == "__main__":
    print("🚀 Starting Drugs.com Side Effects Scraper - HTTP Version")
    print("="*60)
    print("🌐 Using HTTP requests with 2-5 second delays")
    print("="*60)
    
    # Process ALL medications
    update_excel_with_side_effects()  # No limit - process all
    
    print("\n" + "="*60)
    print("🎉 FULL SCRAPING COMPLETED!")
    print("="*60)
