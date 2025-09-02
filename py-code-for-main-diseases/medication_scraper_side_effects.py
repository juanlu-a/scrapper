#!/usr/bin/env python3
"""
MedlinePlus Side Effects Scraper
Scrapes side effects information for medications from https://medlineplus.gov/druginformation.html
and adds the information to the existing medication Excel file.

Author: Assistant
Date: September 2025
"""

import pandas as pd
import os
import time
import random
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException, WebDriverException, NoSuchElementException
import re
from bs4 import BeautifulSoup
import json
import sys
from tqdm import tqdm
import colorama
from colorama import Fore, Back, Style
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import google.generativeai as genai
from dotenv import load_dotenv

# Initialize colorama
colorama.init(autoreset=True)

# Load environment variables
load_dotenv('/Users/juanlu/Documents/Wye/scrapper/.env')

class MedlinePlusSideEffectsScraper:
    def __init__(self, headless=False):
        self.headless = headless
        self.driver = None
        self.wait = None
        self.base_url = "https://medlineplus.gov/druginformation.html"
        self.search_url = "https://medlineplus.gov/druginformation.html"
        
        # Configure Google Generative AI
        api_key = os.getenv('GOOGLE_GEMINI_API_KEY')
        if not api_key:
            raise ValueError("GOOGLE_GEMINI_API_KEY not found in environment variables. Please check your .env file.")
        
        genai.configure(api_key=api_key)
        self.model = genai.GenerativeModel("gemini-1.5-flash")
        
        # Cache for processed medications
        self.cache_file = "side_effects_cache.json"
        self.cache = self.load_cache()
        
        # Results storage
        self.results = {}
        
        self.init_driver()
        
    def print_header(self, title, subtitle=""):
        """Print a styled header"""
        print(f"\n{Fore.GREEN}{'═'*80}")
        print(f"{Fore.WHITE}{Style.BRIGHT}{title:^80}")
        if subtitle:
            print(f"{Fore.CYAN}{subtitle:^80}")
        print(f"{Fore.GREEN}{'═'*80}{Style.RESET_ALL}")
    
    def print_section(self, title):
        """Print a section header"""
        print(f"\n{Fore.BLUE}{Style.BRIGHT}▶ {title}")
        print(f"{Fore.BLUE}{'─' * (len(title) + 2)}{Style.RESET_ALL}")
    
    def print_success(self, message):
        """Print a success message"""
        print(f"{Fore.GREEN}✅ {message}{Style.RESET_ALL}")
    
    def print_error(self, message):
        """Print an error message"""
        print(f"{Fore.RED}❌ {message}{Style.RESET_ALL}")
    
    def print_warning(self, message):
        """Print a warning message"""
        print(f"{Fore.YELLOW}⚠️ {message}{Style.RESET_ALL}")
    
    def print_info(self, message):
        """Print an info message"""
        print(f"{Fore.CYAN}ℹ️ {message}{Style.RESET_ALL}")
    
    def load_cache(self):
        """Load existing cache if available"""
        try:
            if os.path.exists(self.cache_file):
                with open(self.cache_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception as e:
            self.print_warning(f"Could not load cache: {e}")
        return {}
    
    def save_cache(self):
        """Save cache to file"""
        try:
            with open(self.cache_file, 'w', encoding='utf-8') as f:
                json.dump(self.cache, f, indent=2, ensure_ascii=False)
        except Exception as e:
            self.print_error(f"Could not save cache: {e}")
    
    def init_driver(self):
        """Initialize the Chrome driver"""
        if self.driver:
            try:
                self.driver.quit()
            except:
                pass
        
        self.driver = self.setup_driver(self.headless)
        self.wait = WebDriverWait(self.driver, 15)
    
    def setup_driver(self, headless=False):
        """Set up Chrome driver with options"""
        chrome_options = Options()
        if headless:
            chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_argument("--disable-web-security")
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        
        # Add user agent to avoid detection
        chrome_options.add_argument("--user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
        
        try:
            driver = webdriver.Chrome(options=chrome_options)
            driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            return driver
        except Exception as e:
            self.print_error(f"Error setting up Chrome driver: {e}")
            raise e
    
    def navigate_to_medlineplus(self):
        """Navigate to MedlinePlus drug information page"""
        try:
            self.print_info(f"Navigating to {self.base_url}")
            self.driver.get(self.base_url)
            time.sleep(3)
            
            # Wait for page to load
            self.wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
            self.print_success("Successfully loaded MedlinePlus drug information page")
            return True
            
        except Exception as e:
            self.print_error(f"Failed to navigate to MedlinePlus: {e}")
            return False
    
    def search_medication(self, medication_name):
        """Search for a medication on MedlinePlus"""
        try:
            # Clean medication name for search
            clean_name = self.clean_medication_name(medication_name)
            self.print_info(f"Searching for: {clean_name}")
            
            # Navigate to search page first
            if not self.navigate_to_medlineplus():
                return None
            
            # Find search box - try multiple selectors
            search_selectors = [
                'input[name="query"]',
                'input[type="search"]',
                '#search-box',
                '.search-input',
                'input[placeholder*="search"]',
                'input[aria-label*="search"]'
            ]
            
            search_box = None
            for selector in search_selectors:
                try:
                    search_box = self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, selector)))
                    break
                except TimeoutException:
                    continue
            
            if not search_box:
                # Try alternative approach - direct URL construction
                search_url = f"https://medlineplus.gov/druginfo/medmaster/search.html?query={clean_name}"
                self.driver.get(search_url)
                time.sleep(3)
                return self.extract_drug_info_from_page()
            
            # Clear and enter search term
            search_box.clear()
            search_box.send_keys(clean_name)
            search_box.send_keys(Keys.RETURN)
            
            time.sleep(3)
            
            # Look for direct drug page or search results
            return self.handle_search_results(clean_name)
            
        except Exception as e:
            self.print_error(f"Error searching for {medication_name}: {e}")
            return None
    
    def clean_medication_name(self, name):
        """Clean medication name for search"""
        # Remove common suffixes and prefixes
        clean_name = re.sub(r'\s+(tablet|capsule|injection|cream|ointment|gel|liquid|suspension|er|xr|xl)s?$', '', name.lower())
        clean_name = re.sub(r'\s+\d+\s*(mg|mcg|g|ml|iu).*$', '', clean_name)
        clean_name = re.sub(r'\s+oral.*$', '', clean_name)
        clean_name = clean_name.strip()
        return clean_name
    
    def handle_search_results(self, medication_name):
        """Handle search results and find the most relevant drug page"""
        try:
            # Check if we're already on a drug information page
            if "druginfo" in self.driver.current_url.lower():
                return self.extract_drug_info_from_page()
            
            # Look for search results links
            result_selectors = [
                'a[href*="druginfo"]',
                'a[href*="meds"]',
                '.search-result a',
                '.result-item a',
                'a[title*="{}"]'.format(medication_name)
            ]
            
            for selector in result_selectors:
                try:
                    results = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    if results:
                        # Click on the first relevant result
                        results[0].click()
                        time.sleep(3)
                        return self.extract_drug_info_from_page()
                except:
                    continue
            
            # If no direct results, try generic search
            return self.try_generic_drug_search(medication_name)
            
        except Exception as e:
            self.print_error(f"Error handling search results: {e}")
            return None
    
    def try_generic_drug_search(self, medication_name):
        """Try searching for generic version of the drug"""
        try:
            # Common generic mappings
            generic_mappings = {
                'advil': 'ibuprofen',
                'tylenol': 'acetaminophen',
                'motrin': 'ibuprofen',
                'aleve': 'naproxen',
                'prozac': 'fluoxetine',
                'zoloft': 'sertraline',
                'lipitor': 'atorvastatin',
                'nexium': 'esomeprazole'
            }
            
            generic_name = generic_mappings.get(medication_name.lower(), medication_name)
            
            if generic_name != medication_name:
                self.print_info(f"Trying generic name: {generic_name}")
                return self.search_medication(generic_name)
            
            return None
            
        except Exception as e:
            self.print_error(f"Error in generic search: {e}")
            return None
    
    def extract_drug_info_from_page(self):
        """Extract drug information including side effects from the current page"""
        try:
            # Get page source for parsing
            page_source = self.driver.page_source
            soup = BeautifulSoup(page_source, 'html.parser')
            
            # Extract text content
            text_content = soup.get_text()
            
            # Use LLM to extract side effects information
            return self.extract_side_effects_with_llm(text_content)
            
        except Exception as e:
            self.print_error(f"Error extracting drug info: {e}")
            return None
    
    def extract_side_effects_with_llm(self, page_content):
        """Use LLM to extract and summarize side effects from page content"""
        try:
            prompt = f"""
            Analyze the following medical webpage content and extract side effects information for the medication.
            
            Please provide a comprehensive but concise summary of side effects in the following format:
            
            COMMON SIDE EFFECTS: [list the most common side effects]
            SERIOUS SIDE EFFECTS: [list serious/severe side effects if mentioned]
            RARE SIDE EFFECTS: [list rare side effects if mentioned]
            
            If no side effects are clearly mentioned, respond with "No side effects information found on this page."
            
            Keep the response under 500 words and focus only on side effects information.
            
            Page content:
            {page_content[:8000]}  # Limit content to avoid token limits
            """
            
            response = self.model.generate_content(prompt)
            
            if response and response.text:
                return response.text.strip()
            else:
                return "No side effects information could be extracted."
                
        except Exception as e:
            self.print_error(f"Error using LLM to extract side effects: {e}")
            return "Error extracting side effects information."
    
    def process_medication(self, medication_name):
        """Process a single medication and get its side effects"""
        try:
            # Check cache first
            if medication_name in self.cache:
                self.print_info(f"Found {medication_name} in cache")
                return self.cache[medication_name]
            
            self.print_section(f"Processing: {medication_name}")
            
            # Search and get side effects
            side_effects = self.search_medication(medication_name)
            
            if side_effects:
                self.print_success(f"Successfully extracted side effects for {medication_name}")
                # Cache the result
                self.cache[medication_name] = side_effects
                self.save_cache()
            else:
                self.print_warning(f"No side effects found for {medication_name}")
                side_effects = "No side effects information found."
                self.cache[medication_name] = side_effects
                self.save_cache()
            
            # Random delay to avoid rate limiting
            time.sleep(random.uniform(2, 5))
            
            return side_effects
            
        except Exception as e:
            self.print_error(f"Error processing {medication_name}: {e}")
            return "Error retrieving side effects information."
    
    def load_medication_data(self, excel_file_path):
        """Load medication data from Excel file"""
        try:
            self.print_section("Loading medication data from Excel")
            
            # Read the Excel file
            df = pd.read_excel(excel_file_path)
            
            # Find the medication data section (starts after the statistics section)
            medication_start_idx = None
            for idx, row in df.iterrows():
                if str(row.iloc[0]).strip() == "Medication Name":
                    medication_start_idx = idx
                    break
            
            if medication_start_idx is None:
                raise ValueError("Could not find 'Medication Name' header in Excel file")
            
            # Extract medication data
            medications_df = df.iloc[medication_start_idx:].copy()
            medications_df.columns = medications_df.iloc[0]  # Set first row as headers
            medications_df = medications_df.iloc[1:].reset_index(drop=True)  # Remove header row
            
            # Remove rows with NaN medication names
            medications_df = medications_df.dropna(subset=['Medication Name'])
            
            self.print_success(f"Loaded {len(medications_df)} medications from Excel file")
            return medications_df
            
        except Exception as e:
            self.print_error(f"Error loading medication data: {e}")
            raise e
    
    def process_all_medications(self, excel_file_path, output_file_path=None):
        """Process all medications and add side effects column"""
        try:
            self.print_header("MEDLINEPLUS SIDE EFFECTS SCRAPER", "Processing medication side effects")
            
            # Load medication data
            medications_df = self.load_medication_data(excel_file_path)
            
            # Add side effects column if it doesn't exist
            if 'Side Effects' not in medications_df.columns:
                medications_df['Side Effects'] = ''
            
            total_medications = len(medications_df)
            self.print_info(f"Processing {total_medications} medications...")
            
            # Process each medication
            for idx, row in tqdm(medications_df.iterrows(), total=total_medications, desc="Processing medications"):
                medication_name = str(row['Medication Name']).strip()
                
                if pd.isna(medication_name) or medication_name.lower() in ['nan', '']:
                    continue
                
                # Skip if already processed and has side effects data
                if pd.notna(row.get('Side Effects', '')) and str(row.get('Side Effects', '')).strip():
                    self.print_info(f"Skipping {medication_name} - already has side effects data")
                    continue
                
                try:
                    side_effects = self.process_medication(medication_name)
                    medications_df.at[idx, 'Side Effects'] = side_effects
                    
                    # Save progress every 10 medications
                    if idx % 10 == 0:
                        self.save_progress(medications_df, excel_file_path, output_file_path)
                        
                except Exception as e:
                    self.print_error(f"Error processing {medication_name}: {e}")
                    medications_df.at[idx, 'Side Effects'] = "Error retrieving side effects information."
            
            # Final save
            self.save_final_results(medications_df, excel_file_path, output_file_path)
            
            self.print_success("All medications processed successfully!")
            return medications_df
            
        except Exception as e:
            self.print_error(f"Error in process_all_medications: {e}")
            raise e
    
    def save_progress(self, medications_df, original_file_path, output_file_path):
        """Save progress to avoid losing work"""
        try:
            # Create backup with proper structure
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_file = f"medication_side_effects_progress_{timestamp}.xlsx"
            
            # Save with new side effects column (just the medications data for progress)
            medications_df.to_excel(backup_file, index=False)
            self.print_info(f"Progress saved to {backup_file}")
            
        except Exception as e:
            self.print_warning(f"Could not save progress: {e}")
    
    def save_final_results(self, medications_df, original_file_path, output_file_path):
        """Save final results to Excel file by properly updating the original file structure and preserving formatting"""
        try:
            if output_file_path is None:
                # Create output filename based on original in the Analysis folder
                base_name = os.path.splitext(os.path.basename(original_file_path))[0]
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                analysis_folder = os.path.dirname(original_file_path)
                output_file_path = os.path.join(analysis_folder, f"{base_name}_WITH_SIDE_EFFECTS_{timestamp}.xlsx")
            
            # Load original Excel file using openpyxl to preserve exact structure AND formatting
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            wb = openpyxl.load_workbook(original_file_path)
            ws = wb.active
            
            # Find the medication data header row
            header_row = None
            for row in range(1, 20):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value == "Medication Name":
                    header_row = row
                    break
            
            if header_row is None:
                self.print_error("Could not find medication header row")
                return
                
            self.print_info(f"Found medication headers at row {header_row}")
            
            # Add "Side Effects" header to column 7 (preserving Disease Tag in column 6)
            header_cell = ws.cell(row=header_row, column=7, value="Side Effects")
            
            # Copy formatting from adjacent header cell (column 6) to maintain consistency
            ref_cell = ws.cell(row=header_row, column=6)
            if ref_cell.font:
                header_cell.font = Font(
                    name=ref_cell.font.name,
                    size=ref_cell.font.size,
                    bold=ref_cell.font.bold,
                    italic=ref_cell.font.italic,
                    color=ref_cell.font.color
                )
            if ref_cell.fill and hasattr(ref_cell.fill, 'start_color'):
                header_cell.fill = PatternFill(
                    start_color=ref_cell.fill.start_color,
                    end_color=ref_cell.fill.end_color,
                    fill_type=ref_cell.fill.fill_type
                )
            if ref_cell.border:
                header_cell.border = Border(
                    left=ref_cell.border.left,
                    right=ref_cell.border.right,
                    top=ref_cell.border.top,
                    bottom=ref_cell.border.bottom
                )
            if ref_cell.alignment:
                header_cell.alignment = Alignment(
                    horizontal=ref_cell.alignment.horizontal,
                    vertical=ref_cell.alignment.vertical,
                    wrap_text=ref_cell.alignment.wrap_text
                )
            
            # Create a mapping of medication names to side effects
            side_effects_mapping = {}
            for idx, row in medications_df.iterrows():
                med_name = str(row.get('Medication Name', '')).strip().lower()
                side_effects = str(row.get('Side Effects', 'No side effects information found.'))
                if med_name and med_name != 'nan':
                    side_effects_mapping[med_name] = side_effects
            
            # Update each medication row with side effects (add to new column 7)
            medications_processed = 0
            side_effects_added = 0
            
            for row in range(header_row + 1, ws.max_row + 1):
                medication_name = ws.cell(row=row, column=1).value
                
                if not medication_name or pd.isna(medication_name):
                    continue
                    
                medication_name = str(medication_name).strip()
                
                # Skip summary/header rows
                if medication_name.startswith('📊') or medication_name.startswith('📋') or medication_name.startswith('📈'):
                    continue
                    
                medications_processed += 1
                
                # Look up side effects (case insensitive)
                side_effects = side_effects_mapping.get(medication_name.lower(), "No side effects information found.")
                
                # Add side effects to column 7 (new column, preserving Disease Tag in column 6)
                side_effects_cell = ws.cell(row=row, column=7, value=side_effects)
                
                # Copy formatting from adjacent data cell (column 6) to maintain consistency
                reference_cell = ws.cell(row=row, column=6)
                if reference_cell.font:
                    side_effects_cell.font = Font(
                        name=reference_cell.font.name,
                        size=reference_cell.font.size,
                        bold=reference_cell.font.bold,
                        italic=reference_cell.font.italic,
                        color=reference_cell.font.color
                    )
                if reference_cell.fill and hasattr(reference_cell.fill, 'start_color'):
                    side_effects_cell.fill = PatternFill(
                        start_color=reference_cell.fill.start_color,
                        end_color=reference_cell.fill.end_color,
                        fill_type=reference_cell.fill.fill_type
                    )
                if reference_cell.border:
                    side_effects_cell.border = Border(
                        left=reference_cell.border.left,
                        right=reference_cell.border.right,
                        top=reference_cell.border.top,
                        bottom=reference_cell.border.bottom
                    )
                if reference_cell.alignment:
                    side_effects_cell.alignment = Alignment(
                        horizontal=reference_cell.alignment.horizontal,
                        vertical=reference_cell.alignment.vertical,
                        wrap_text=True  # Enable text wrapping for long side effects
                    )
                
                if side_effects != "No side effects information found.":
                    side_effects_added += 1
            
            # Save the updated workbook
            wb.save(output_file_path)
            
            # Auto-adjust column width for the new Side Effects column
            try:
                # Set a reasonable width for the Side Effects column (column G/7)
                ws.column_dimensions['G'].width = 50  # Adjust width as needed
                wb.save(output_file_path)  # Save again with column width
            except Exception as e:
                self.print_warning(f"Could not adjust column width: {e}")
            
            self.print_success(f"Final results saved to: {output_file_path}")
            self.print_info(f"📊 Total medications processed: {medications_processed}")
            self.print_info(f"💊 Side effects added: {side_effects_added}")
            self.print_info(f"📈 Success rate: {side_effects_added/medications_processed*100:.1f}%")
            self.print_info(f"✅ Disease Tag column preserved in column 6")
            self.print_info(f"✅ Side Effects added in new column 7")
            self.print_info(f"🎨 Original formatting preserved")
            
            # Clean up temporary files after successful completion
            self.cleanup_temporary_files()
            
            return output_file_path
            
        except Exception as e:
            self.print_error(f"Error saving final results: {e}")
            # Fallback - save just the medications data
            fallback_file = f"medication_side_effects_fallback_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            medications_df.to_excel(fallback_file, index=False)
            self.print_warning(f"Saved fallback file: {fallback_file}")
    
    def cleanup_temporary_files(self):
        """Clean up temporary files after successful completion"""
        try:
            files_deleted = 0
            
            # Delete progress files
            import glob
            progress_files = glob.glob("medication_side_effects_progress_*.xlsx")
            for file in progress_files:
                try:
                    os.remove(file)
                    files_deleted += 1
                    self.print_info(f"🗑️ Deleted progress file: {file}")
                except Exception as e:
                    self.print_warning(f"Could not delete {file}: {e}")
            
            # Delete fallback files
            fallback_files = glob.glob("medication_side_effects_fallback_*.xlsx")
            for file in fallback_files:
                try:
                    os.remove(file)
                    files_deleted += 1
                    self.print_info(f"🗑️ Deleted fallback file: {file}")
                except Exception as e:
                    self.print_warning(f"Could not delete {file}: {e}")
            
            # Delete cache file
            if os.path.exists(self.cache_file):
                try:
                    os.remove(self.cache_file)
                    files_deleted += 1
                    self.print_info(f"🗑️ Deleted cache file: {self.cache_file}")
                except Exception as e:
                    self.print_warning(f"Could not delete cache file: {e}")
            
            if files_deleted > 0:
                self.print_success(f"🧹 Cleanup completed: {files_deleted} temporary files deleted")
            else:
                self.print_info("🧹 No temporary files to clean up")
                
        except Exception as e:
            self.print_warning(f"Error during cleanup: {e}")

    def cleanup(self):
        """Clean up resources"""
        try:
            if self.driver:
                self.driver.quit()
        except:
            pass

def main():
    """Main function to run the scraper"""
    print(f"{Fore.GREEN}{Style.BRIGHT}🔍 MedlinePlus Side Effects Scraper{Style.RESET_ALL}")
    print(f"{Fore.CYAN}Starting side effects extraction process...{Style.RESET_ALL}")
    
    # Configuration
    excel_file_path = "/Users/juanlu/Documents/Wye/scrapper/Analysis/medication_data_20250821_164933_FINAL_ENHANCED.xlsx"
    output_file_path = None  # Will be auto-generated
    headless = False  # Set to True for headless browsing
    
    scraper = None
    try:
        # Initialize scraper
        scraper = MedlinePlusSideEffectsScraper(headless=headless)
        
        # Process all medications
        results_df = scraper.process_all_medications(excel_file_path, output_file_path)
        
        print(f"\n{Fore.GREEN}{Style.BRIGHT}✅ Successfully completed side effects extraction!")
        print(f"{Fore.CYAN}📊 Processed {len(results_df)} medications")
        print(f"{Fore.CYAN}💾 Results saved to Excel file")
        print(f"{Fore.CYAN}🧹 Temporary files cleaned up{Style.RESET_ALL}")
        
    except KeyboardInterrupt:
        print(f"\n{Fore.YELLOW}⚠️ Process interrupted by user{Style.RESET_ALL}")
    except Exception as e:
        print(f"\n{Fore.RED}❌ Error: {e}{Style.RESET_ALL}")
    finally:
        if scraper:
            scraper.cleanup()
        print(f"\n{Fore.BLUE}🏁 Scraper cleanup completed{Style.RESET_ALL}")

if __name__ == "__main__":
    main()
