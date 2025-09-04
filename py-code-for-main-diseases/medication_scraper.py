import pandas as pd
import os
import glob
from datetime import datetime
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException, WebDriverException
import re
from bs4 import BeautifulSoup
import json
import sys
from tqdm import tqdm
import colorama
from colorama import Fore, Back, Style
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle

colorama.init(autoreset=True)

class MedicationScraper:
    def __init__(self):
        self.driver = None
        self.existing_data = {}
        self.cache_file = "scraping_cache.json"
        self.batch_size = 10
        self.enhanced_brand_database = self.load_enhanced_brand_database()
        
        self.brand_extraction_patterns = [
                    r'Brand name[s]?:\s*([^,\n\r]+)',
        r'Brand:\s*([^,\n\r]+)',
        r'<strong>Brand name[s]?:</strong>\s*([^<]+)',
        r'<td[^>]*>Brand name[s]?:</td>\s*<td[^>]*>([^<]+)</td>',
        r'<span[^>]*>Brand name[s]?:</span>\s*([^<]+)',
        r'<div[^>]*>Brand name[s]?:</div>\s*([^<]+)',
        r'Also known as:\s*([^,\n\r]+)',
        r'Alternative names?:\s*([^,\n\r]+)',
        r'Common brands?:\s*([^,\n\r]+)',
        r'Brand names?:\s*([^,\n\r]+)',
        r'Available as:\s*([^,\n\r]+)',
        r'Marketed as:\s*([^,\n\r]+)',
        r'Sold as:\s*([^,\n\r]+)',
        r'Known as:\s*([^,\n\r]+)',
        r'Proprietary name[s]?:\s*([^,\n\r]+)',
        r'Trade name[s]?:\s*([^,\n\r]+)',
        r'<h[1-6][^>]*>([^<]*?(?:brand|Brand)[^<]*)</h[1-6]>',
        r'<div[^>]*class="[^"]*brand[^"]*"[^>]*>([^<]+)</div>',
        r'<span[^>]*class="[^"]*brand[^"]*"[^>]*>([^<]+)</span>',
        r'<tr[^>]*>.*?Brand.*?</tr>',
        r'<td[^>]*>Brand</td>\s*<td[^>]*>([^<]+)</td>',
        r'<li[^>]*>([^<]*?(?:brand|Brand)[^<]*)</li>',
        r'<ul[^>]*>.*?Brand.*?</ul>',
        r'([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)\s+(?:tablet|capsule|pill|injection|cream|ointment)',
        r'([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)\s+(?:mg|mcg|g|ml|IU)',
        r'([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)\s+(?:oral|topical|inhalation)',
        ]
        
        # Comprehensive brand database
        self.comprehensive_brands = {
            'Pain & Fever': [
                'Bayer', 'Ecotrin', 'St. Joseph', 'Bufferin', 'Anacin', 'Excedrin',
                'Advil', 'Motrin', 'Aleve', 'Tylenol', 'Panadol', 'Calpol',
                'Nuprin', 'Brufen', 'Nurofen', 'Feldene', 'Voltaren', 'Celebrex',
                'Vioxx', 'Bextra', 'Diclofenac', 'Ibuprofen', 'Naproxen', 'Aspirin'
            ],
            'Heart & Blood Pressure': [
                'Lipitor', 'Zocor', 'Crestor', 'Plavix', 'Brilinta', 'Effient',
                'Eliquis', 'Xarelto', 'Pradaxa', 'Warfarin', 'Coumadin', 'Heparin',
                'Lovenox', 'Fragmin', 'Arixtra', 'Metformin', 'Glucophage', 'Januvia',
                'Invokana', 'Farxiga', 'Jardiance', 'Victoza', 'Trulicity', 'Ozempic',
                'Wegovy', 'Lantus', 'NovoLog', 'Humalog', 'Tresiba', 'Toujeo'
            ],
            'Mental Health': [
                'Zoloft', 'Prozac', 'Celexa', 'Lexapro', 'Paxil', 'Wellbutrin',
                'Xanax', 'Valium', 'Ativan', 'Klonopin', 'Ambien', 'Lunesta',
                'Sonata', 'Rozerem', 'Belsomra', 'Dayvigo', 'Quviviq', 'Abilify',
                'Seroquel', 'Risperdal', 'Zyprexa', 'Geodon', 'Latuda', 'Vraylar'
            ],
            'Respiratory': [
                'Albuterol', 'Proventil', 'Ventolin', 'ProAir', 'Xopenex', 'Proventil',
                'Fluticasone', 'Flonase', 'Nasonex', 'Rhinocort', 'Nasacort', 'Qnasl',
                'Montelukast', 'Singulair', 'Zafirlukast', 'Accolate', 'Zileuton',
                'Ipratropium', 'Atrovent', 'Tiotropium', 'Spiriva', 'Umeclidinium'
            ],
            'Gastrointestinal': [
                'Zantac', 'Prilosec', 'Nexium', 'Prevacid', 'Aciphex', 'Dexilant',
                'Omeprazole', 'Esomeprazole', 'Pantoprazole', 'Lansoprazole', 'Rabeprazole',
                'Pepcid', 'Tagamet', 'Axid', 'Carafate', 'Reglan', 'Zofran'
            ],
            'Diabetes': [
                'Glucophage', 'Metformin', 'Januvia', 'Invokana', 'Farxiga', 'Jardiance',
                'Victoza', 'Trulicity', 'Ozempic', 'Wegovy', 'Lantus', 'NovoLog',
                'Humalog', 'Tresiba', 'Toujeo', 'Levemir', 'Toujeo', 'Basaglar'
            ],
            'Cholesterol': [
                'Lipitor', 'Zocor', 'Crestor', 'Pravachol', 'Lescol', 'Mevacor',
                'Livalo', 'Zetia', 'Vytorin', 'Repatha', 'Praluent', 'Nexletol'
            ],
            'Antibiotics': [
                'Amoxicillin', 'Augmentin', 'Zithromax', 'Biaxin', 'Cipro', 'Levaquin',
                'Keflex', 'Doxycycline', 'Minocycline', 'Tetracycline', 'Bactrim', 'Septra'
            ],
            'Allergies': [
                'Claritin', 'Zyrtec', 'Allegra', 'Xyzal', 'Clarinex', 'Claritin-D',
                'Zyrtec-D', 'Allegra-D', 'Benadryl', 'Chlor-Trimeton', 'Tavist', 'Seldane'
            ]
        }
        
        # Flatten the comprehensive brands list for easier searching
        self.all_brands = []
        for category, brands in self.comprehensive_brands.items():
            self.all_brands.extend(brands)
    
    def print_header(self, title, subtitle=""):
        """Print a styled header with modern visual design"""
        print(f"\n{Fore.GREEN}{'═'*70}")
        print(f"{Fore.WHITE}{Style.BRIGHT}{title:^70}")
        if subtitle:
            print(f"{Fore.CYAN}{subtitle:^70}")
        print(f"{Fore.GREEN}{'═'*70}{Style.RESET_ALL}")
    
    def print_section(self, title):
        """Print a section header with modern styling"""
        print(f"\n{Fore.BLUE}{Style.BRIGHT}▶ {title}")
        print(f"{Fore.BLUE}{'─' * (len(title) + 2)}{Style.RESET_ALL}")
    
    def print_success(self, message):
        """Print a success message with green styling"""
        print(f"{Fore.GREEN}✅ {message}{Style.RESET_ALL}")
    
    def print_error(self, message):
        """Print an error message with red styling"""
        print(f"{Fore.RED}❌ {message}{Style.RESET_ALL}")
    
    def print_warning(self, message):
        """Print a warning message with yellow styling"""
        print(f"{Fore.YELLOW}⚠️ {message}{Style.RESET_ALL}")
    
    def print_info(self, message):
        """Print an info message with blue styling"""
        print(f"{Fore.CYAN}ℹ️ {message}{Style.RESET_ALL}")
    
    def print_progress(self, current, total, description=""):
        """Print a progress bar with modern styling"""
        percentage = (current / total) * 100
        bar_length = 40
        filled_length = int(bar_length * current // total)
        bar = '█' * filled_length + '░' * (bar_length - filled_length)
        
        if description:
            print(f"\r{Fore.MAGENTA}{description}: {bar} {percentage:5.1f}% ({current}/{total})", end='', flush=True)
        else:
            print(f"\r{Fore.MAGENTA}Progress: {bar} {percentage:5.1f}% ({current}/{total})", end='', flush=True)
        
        if current == total:
            print()  # New line when complete
    
    def print_brand_extraction_summary(self, data):
        """Print a beautiful summary of brand extraction results"""
        total_medications = len(data)
        brand_names_found = sum(1 for d in data.values() if d['brand_name'] != 'Not found')
        multiple_brands = sum(1 for d in data.values() if '|' in str(d['brand_name']))
        generic_found = sum(1 for d in data.values() if 'Generic' in str(d['brand_name']))
        
        print(f"\n{Fore.CYAN}{'═'*70}")
        print(f"{Fore.WHITE}{Style.BRIGHT}{'BRAND EXTRACTION SUMMARY':^70}")
        print(f"{Fore.CYAN}{'═'*70}")
        
        print(f"{Fore.GREEN}📊 Total Medications Analyzed: {total_medications}")
        print(f"{Fore.GREEN}✅ Brand Names Found: {brand_names_found} ({brand_names_found/total_medications*100:.1f}%)")
        print(f"{Fore.BLUE}🔗 Multiple Brands Identified: {multiple_brands}")
        print(f"{Fore.YELLOW}💊 Generic Medications: {generic_found}")
        
        # Show examples of multiple brands
        if multiple_brands > 0:
            print(f"\n{Fore.CYAN}🔗 Examples of Multiple Brand Names:")
            count = 0
            for medication, info in data.items():
                if '|' in str(info['brand_name']) and count < 5:
                    brands = info['brand_name'].split(' | ')
                    print(f"  {Fore.WHITE}{medication}: {Fore.GREEN}{', '.join(brands[:3])}")
                    count += 1
        
        print(f"{Fore.CYAN}{'═'*70}{Style.RESET_ALL}")
    
    def print_data_quality_metrics(self, data):
        """Print comprehensive data quality metrics"""
        total = len(data)
        
        # Calculate metrics
        dosage_found = sum(1 for d in data.values() if d['dosage'] != 'Not found')
        how_to_take_found = sum(1 for d in data.values() if d['how_to_take'] != 'Not found')
        when_to_take_found = sum(1 for d in data.values() if d['when_to_take'] != 'Not found')
        
        print(f"\n{Fore.MAGENTA}{'═'*70}")
        print(f"{Fore.WHITE}{Style.BRIGHT}{'DATA QUALITY METRICS':^70}")
        print(f"{Fore.MAGENTA}{'═'*70}")
        
        print(f"{Fore.GREEN}📊 Dosage Information: {dosage_found}/{total} ({dosage_found/total*100:.1f}%)")
        print(f"{Fore.BLUE}📋 How to Take: {how_to_take_found}/{total} ({how_to_take_found/total*100:.1f}%)")
        print(f"{Fore.YELLOW}⏰ When to Take: {when_to_take_found}/{total} ({when_to_take_found/total*100:.1f}%)")
        
        # Quality score
        quality_score = (dosage_found + how_to_take_found + when_to_take_found) / (total * 3) * 100
        if quality_score >= 80:
            quality_emoji = "🟢"
            quality_color = Fore.GREEN
        elif quality_score >= 60:
            quality_emoji = "🟡"
            quality_color = Fore.YELLOW
        else:
            quality_emoji = "🔴"
            quality_color = Fore.RED
        
        print(f"{quality_color}{quality_emoji} Overall Data Quality: {quality_score:.1f}%{Style.RESET_ALL}")
        print(f"{Fore.MAGENTA}{'═'*70}{Style.RESET_ALL}")
    
    def setup_driver(self):
        chrome_options = Options()
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        chrome_options.add_argument('--disable-popup-blocking')
        chrome_options.add_argument('--disable-notifications')
        chrome_options.add_argument('--disable-images')
        chrome_options.add_argument('--disable-extensions')
        chrome_options.add_argument('--disable-plugins')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--disable-web-security')
        chrome_options.add_argument('--disable-features=VizDisplayCompositor')
        chrome_options.add_argument('--timeout=15000')
        chrome_options.add_argument('--page-load-strategy=eager')
        chrome_options.add_argument('--window-size=1920,1080')
        chrome_options.add_argument('--remote-debugging-port=9222')
        chrome_options.add_argument('--disable-css')
        chrome_options.add_argument('--disable-javascript')
        
        self.driver = webdriver.Chrome(options=chrome_options)
        self.driver.set_page_load_timeout(15)
        self.driver.set_script_timeout(15)
        self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    
    def is_driver_working(self):
        try:
            self.driver.current_url
            return True
        except:
            return False
    
    def pause_for_manual_intervention(self):
        print("🤖 Automatic mode - continuing automatically...")
        time.sleep(2)
        return True
        
    def restart_driver(self):
        try:
            if self.driver:
                self.driver.quit()
        except:
            pass
        time.sleep(3)
        self.setup_driver()
    
    def load_cache(self):
        if os.path.exists(self.cache_file):
            try:
                with open(self.cache_file, 'r') as f:
                    return json.load(f)
            except:
                return {}
        return {}
    
    def load_enhanced_brand_database(self):
        """Load comprehensive brand database with all brand names for 189+ medications"""
        return {
            # Immunosuppressants & Biologics
            'abatacept': 'Orencia',
            'adalimumab': 'Humira | Amjevita | Yusimry | Cyltezo | Yuflyma | Hadlima | Hulio | Hyrimoz | Idacio | Abrilada | Simlandi',
            'alirocumab': 'Praluent',
            'baricitinib': 'Olumiant',
            'certolizumab': 'Cimzia',
            'donanemab': 'Kisunla',
            'etanercept': 'Enbrel | Erelzi | Eticovo',
            'evolocumab': 'Repatha',
            'golimumab': 'Simponi | Simponi Aria',
            'infliximab': 'Remicade | Inflectra | Renflexis | Ixifi | Avsola',
            'lecanemab': 'Leqembi',
            'rituximab': 'Rituxan | Truxima | Ruxience | Riabni',
            'tocilizumab': 'Actemra',
            'tofacitinib': 'Xeljanz | Xeljanz XR',
            
            # Pain & Anti-inflammatory
            'acetaminophen': 'Actamin | Anacin AF | Aurophen | Bromo Seltzer | Childrens Tylenol | Mapap | M-Pap | Pharbetol | Silapap Childrens | Tactinal | Tempra Quicklets | Tycolene | Tylenol | Vitapap',
            'aspirin': 'Bayer | Ecotrin | Bufferin | Ascriptin | Easprin | Halfprin | Heartline | Low Dose ASA',
            'celecoxib': 'Celebrex',
            'diclofenac': 'Voltaren | Cataflam | Zipsor | Cambia | Pennsaid | Solaraze | Flector',
            'ibuprofen': 'Advil | Motrin | Nuprin | Caldolor | NeoProfen | Ibu-Tab | Midol | Pamprin',
            'indomethacin': 'Indocin | Indocin SR | Tivorbex',
            'meloxicam': 'Mobic | Vivlodex | Qmiiz ODT',
            'naproxen': 'Aleve | Naprosyn | Anaprox | Anaprox DS | EC-Naprosyn | Naprelan | Naprosyn SR',
            'tramadol': 'Ultram | ConZip | Ryzolt | Ultram ER',
            
            # Respiratory & Asthma
            'albuterol': 'ProAir | Ventolin | Proventil | AccuNeb | ProAir HFA | ProAir RespiClick | Ventolin HFA',
            'beclomethasone': 'Qvar | Beclovent | Vanceril',
            'budesonide': 'Pulmicort | Rhinocort | Entocort EC | Uceris | Tarpeyo',
            'ciclesonide': 'Alvesco | Omnaris | Zetonna',
            'fluticasone': 'Flovent | Flonase | Veramyst | Arnuity Ellipta | Breo Ellipta | Advair | Trelegy Ellipta',
            'formoterol': 'Foradil | Perforomist | Symbicort | Dulera | Bevespi Aerosphere',
            'ipratropium': 'Atrovent | Atrovent HFA | Combivent | DuoNeb',
            'levalbuterol': 'Xopenex | Xopenex HFA',
            'mometasone': 'Asmanex | Elocon | Nasonex',
            'olodaterol': 'Striverdi Respimat',
            'salmeterol': 'Serevent | Advair | AirDuo RespiClick',
            'theophylline': 'Theo-24 | Elixophyllin | Uniphyl | Theo-Dur',
            'umeclidinium': 'Incruse Ellipta | Anoro Ellipta | Trelegy Ellipta',
            'vilanterol': 'Breo Ellipta | Anoro Ellipta | Trelegy Ellipta',
            
            # Cardiovascular
            'amlodipine': 'Norvasc | Katerzia',
            'apixaban': 'Eliquis',
            'atorvastatin': 'Lipitor',
            'bisoprolol': 'Zebeta | Bisoprolol Fumarate',
            'captopril': 'Capoten',
            'carvedilol': 'Coreg | Coreg CR',
            'cilostazol': 'Pletal',
            'clopidogrel': 'Plavix',
            'dabigatran': 'Pradaxa',
            'digoxin': 'Lanoxin | Digitek | Digox',
            'diltiazem': 'Cardizem | Cardizem CD | Cardizem LA | Cartia XT | Dilacor XR | Diltia XT | Taztia XT | Tiazac',
            'dipyridamole': 'Persantine | Aggrenox',
            'edoxaban': 'Savaysa',
            'enalapril': 'Vasotec | Epaned',
            'ezetimibe': 'Zetia | Vytorin | Liptruzet',
            'fenofibrate': 'Tricor | Antara | Fenoglide | Fibricor | Lipofen | Lofibra | Triglide | Trilipix',
            'fluvastatin': 'Lescol | Lescol XL',
            'furosemide': 'Lasix',
            'hydrochlorothiazide': 'Microzide | HydroDIURIL | Esidrix | Oretic | Aquazide H | Hydro Par | Hydrochlorothiazide',
            'irbesartan': 'Avapro | Avalide',
            'lisinopril': 'Prinivil | Zestril | Qbrelis',
            'losartan': 'Cozaar | Hyzaar',
            'lovastatin': 'Mevacor | Altoprev',
            'metoprolol': 'Lopressor | Toprol XL | Metoprolol Succinate | Metoprolol Tartrate',
            'nifedipine': 'Procardia | Procardia XL | Adalat CC | Afeditab CR | Nifediac CC | Nifedical XL',
            'nimodipine': 'Nimotop',
            'nitroglycerin': 'Nitrostat | Nitro-Dur | Nitrolingual | NitroMist | Rectiv | Transderm-Nitro',
            'pitavastatin': 'Livalo | Zypitamag',
            'pravastatin': 'Pravachol',
            'ramipril': 'Altace',
            'rivaroxaban': 'Xarelto',
            'rosuvastatin': 'Crestor | Ezallor Sprinkle',
            'simvastatin': 'Zocor | Flolipid | Simcor | Vytorin',
            'spironolactone': 'Aldactone | CaroSpir',
            'ticagrelor': 'Brilinta',
            'valsartan': 'Diovan | Diovan HCT | Exforge | Exforge HCT | Prexxartan',
            'verapamil': 'Calan | Calan SR | Covera-HS | Isoptin | Isoptin SR | Verelan | Verelan PM',
            'warfarin': 'Coumadin | Jantoven',
            
            # Mental Health & Neurological
            'alprazolam': 'Xanax | Xanax XR | Niravam',
            'amitriptyline': 'Elavil | Vanatrip',
            'amphetamine': 'Adderall | Adderall XR | Mydayis | Evekeo',
            'aripiprazole': 'Abilify | Abilify Maintena | Aristada | Aristada Initio',
            'atomoxetine': 'Strattera',
            'benzphetamine': 'Didrex',
            'bupropion': 'Wellbutrin | Wellbutrin SR | Wellbutrin XL | Aplenzin | Forfivo XL | Zyban',
            'citalopram': 'Celexa',
            'clonazepam': 'Klonopin',
            'clozapine': 'Clozaril | Versacloz | Fazaclo',
            'desipramine': 'Norpramin',
            'desvenlafaxine': 'Pristiq | Khedezla',
            'dextroamphetamine': 'Dexedrine | Dexedrine Spansule | Dextrostat | ProCentra | Zenzedi',
            'diethylpropion': 'Tenuate | Tenuate Dospan',
            'doxepin': 'Silenor | Sinequan | Zonalon',
            'duloxetine': 'Cymbalta | Drizalma Sprinkle | Irenka',
            'escitalopram': 'Lexapro',
            'fluoxetine': 'Prozac | Prozac Weekly | Sarafem | Selfemra',
            'haloperidol': 'Haldol | Haldol Decanoate',
            'imipramine': 'Tofranil | Tofranil-PM',
            'isocarboxazid': 'Marplan',
            'lamotrigine': 'Lamictal | Lamictal ODT | Lamictal XR | Subvenite',
            'levomilnacipran': 'Fetzima',
            'lithium': 'Lithobid | Eskalith | Lithonate | Lithotabs',
            'lorazepam': 'Ativan',
            'lorcaserin': 'Belviq | Belviq XR',
            'mirtazapine': 'Remeron | Remeron SolTab',
            'nefazodone': 'Serzone',
            'nortriptyline': 'Pamelor | Aventyl',
            'olanzapine': 'Zyprexa | Zyprexa Relprevv | Zyprexa Zydis',
            'paroxetine': 'Paxil | Paxil CR | Pexeva | Brisdelle',
            'phendimetrazine': 'Bontril | Bontril PDM | Bontril Slow Release',
            'phenelzine': 'Nardil',
            'phentermine': 'Adipex-P | Lomaira | Suprenza',
            'protriptyline': 'Vivactil',
            'quetiapine': 'Seroquel | Seroquel XR',
            'risperidone': 'Risperdal | Risperdal Consta | Risperdal M-Tab',
            'rivastigmine': 'Exelon | Exelon Patch',
            'selegiline': 'Eldepryl | Emsam | Zelapar',
            'sertraline': 'Zoloft',
            'tizanidine': 'Zanaflex | Tizanidine Hydrochloride',
            'tranylcypromine': 'Parnate',
            'trazodone': 'Desyrel | Oleptro',
            'trimipramine': 'Surmontil',
            'venlafaxine': 'Effexor | Effexor XR | Pristiq',
            'vilazodone': 'Viibryd',
            'vortioxetine': 'Trintellix | Brintellix',
            'ziprasidone': 'Geodon',
            
            # Antibiotics & Antimicrobials
            'amoxicillin': 'Amoxil | Moxatag | Trimox',
            'azithromycin': 'Zithromax | Z-Pak | Zmax',
            'cefpodoxime': 'Vantin',
            'ceftriaxone': 'Rocephin',
            'cefuroxime': 'Ceftin | Zinacef',
            'ciprofloxacin': 'Cipro | Cipro XR | Proquin XR',
            'clarithromycin': 'Biaxin | Biaxin XL',
            'doxycycline': 'Vibramycin | Doryx | Monodox | Oracea | Adoxa | Atridox | Acticlate | Doxteric | Doxy 100 | Doxy 200 | Morgidox | Oraxyl | Targadox | Vibra-Tabs',
            'erythromycin': 'Ery-Tab | Eryc | EryPed | Erythrocin | PCE | E.E.S. | Erycette | T-Stat',
            'imipenem': 'Primaxin',
            'levofloxacin': 'Levaquin | Levaquin Leva-Pak | Quixin | Iquix',
            'linezolid': 'Zyvox | Zyvoxam',
            'meropenem': 'Merrem',
            'piperacillin': 'Pipracil | Zosyn',
            'vancomycin': 'Vancocin | Vancocin HCl | Firvanq',
            
            # Diabetes & Metabolic
            'glimepiride': 'Amaryl',
            'glipizide': 'Glucotrol | Glucotrol XL',
            'insulin': 'Humalog | Novolog | Apidra | Lantus | Levemir | Toujeo | Tresiba | Basaglar | Semglee | Humulin | Novolin | Afrezza | Fiasp | Lyumjev',
            'liraglutide': 'Victoza | Saxenda',
            'metformin': 'Glucophage | Glucophage XR | Fortamet | Glumetza | Riomet',
            'pioglitazone': 'Actos | Actoplus Met | Duetact | Oseni',
            'rosiglitazone': 'Avandia | Avandamet | Avandaryl',
            'semaglutide': 'Ozempic | Rybelsus | Wegovy',
            'tirzepatide': 'Mounjaro | Zepbound',
            
            # Gastrointestinal
            'cholestyramine': 'Questran | Questran Light | Prevalite | Cholestyramine Light',
            'colesevelam': 'Welchol',
            'colestipol': 'Colestid',
            'omeprazole': 'Prilosec | Prilosec OTC | Zegerid | Zegerid OTC',
            'sucralfate': 'Carafate',
            'sulfasalazine': 'Azulfidine | Azulfidine EN-tabs | Sulfazine | Sulfazine EC',
            
            # Supplements & Natural
            'berberine': 'Berberine | GlucoVantage | Berberine-500',
            'chondroitin': 'Chondroitin Sulfate | Cosamin DS | Flex-a-min | Osteo Bi-Flex | Schiff Move Free',
            'garlic': 'Garlic | Kyolic | Kwai | Garlique | Garlicin',
            'glucosamine': 'Glucosamine Sulfate | Cosamin DS | Flex-a-min | Osteo Bi-Flex | Schiff Move Free',
            'niacin': 'Niacin | Niacor | Niaspan | Slo-Niacin | Vitamin B3',
            'omega': 'Omega-3 | Fish Oil | Lovaza | Vascepa | Epanova | Omtryg',
            'orlistat': 'Xenical | Alli',
            'policosanol': 'Policosanol | Octa-6 | Octa-10',
            'probucol': 'Lorelco',
            'psyllium': 'Metamucil | Konsyl | Fiberall | Hydrocil | Perdiem | Reguloid',
            
            # Other Medications
            'alteplase': 'Activase | Cathflo Activase',
            'alfacalcidol': '1-Alpha Leo | Acal | Adela | Afcal | Albonate | Albone-d | Alcadol | Alsiodol | Antebe | Ao Si Hui | A-Ostin-D3 | Axelanol | Baludol | Bestcal | Bon Care | Bon One | Bondiol | Calcijex | One-Alpha',
            'atropine': 'Atropine | Atropen | AtroPen',
            'calcitriol': 'Rocaltrol | Calcijex | Vectical',
            'cinacalcet': 'Sensipar | Mimpara',
            'colchicine': 'Colcrys | Mitigare | Gloperba',
            'diphenhydramine': 'Benadryl | Unisom | Sominex | Nytol | Banophen | Diphenhist | Siladryl | Sleepinal | Twilite',
            'donepezil': 'Aricept | Aricept ODT',
            'galantamine': 'Razadyne | Razadyne ER',
            'glycopyrrolate': 'Robinul | Robinul Forte | Cuvposa | Glyrx-PF | Seebri Neohaler | Lonhala Magnair',
            'heparin': 'Heparin | Hep-Lock | Hep-Lock U/P | Heparin Lock Flush',
            'hydroxychloroquine': 'Plaquenil | Quineprox',
            'levetiracetam': 'Keppra | Keppra XR | Roweepra | Spritam',
            'mannitol': 'Osmitrol | Aridol | Bronchitol',
            'memantine': 'Namenda | Namenda XR | Namzaric',
            'methylphenidate': 'Ritalin | Ritalin LA | Ritalin SR | Concerta | Metadate CD | Metadate ER | Methylin | Methylin ER | Quillivant XR | Quillichew ER | Aptensio XR | Cotempla XR | Jornay PM | Adhansia XR | Relexxii',
            'methylprednisolone': 'Medrol | Medrol Dosepak | Solu-Medrol | Depo-Medrol',
            'morphine': 'MS Contin | Oramorph SR | Kadian | Avinza | Embeda | Morphabond | Arymo ER | MorphaBond ER',
            'naltrexone': 'Vivitrol | Revia | Depade',
            'oseltamivir': 'Tamiflu',
            'prednisone': 'Deltasone | Rayos | Prednisone Intensol | Sterapred | Sterapred DS',
            'pregabalin': 'Lyrica | Lyrica CR',
            'sevelamer': 'Renvela | Renagel',
            'tenecteplase': 'TNKase',
            'topiramate': 'Topamax | Topamax Sprinkle | Qudexy XR | Trokendi XR',
            'zanamivir': 'Relenza'
        }
    
    def save_cache(self, cache_data):
        with open(self.cache_file, 'w') as f:
            json.dump(cache_data, f)
    
    def load_existing_data(self):
        self.print_section("LOADING EXISTING DATA")
        
        pattern = "medication_*.xlsx"
        files = glob.glob(pattern)
        
        if not files:
            self.print_warning("No existing medication files found")
            self.print_info("A new Excel file will be created")
            return {}, None
        
        files.sort(key=os.path.getmtime, reverse=True)
        latest_file = files[0]
        self.print_success(f"Most recent file found: {latest_file}")
        
        try:
            df = pd.read_excel(latest_file)
            existing_data = {}
            
            for _, row in df.iterrows():
                medication_name = str(row['Medication Name']).strip()
                if pd.notna(medication_name) and medication_name != 'nan':
                    existing_data[medication_name] = {
                        'brand_name': str(row['Brand Names']) if pd.notna(row['Brand Names']) else 'Not found',
                        'dosage': str(row['Dosage Forms']) if pd.notna(row['Dosage Forms']) else 'Not found',
                        'how_to_take': str(row['How to Take']) if pd.notna(row['How to Take']) else 'Not found',
                        'when_to_take': str(row['When to Take']) if pd.notna(row['When to Take']) else 'Not found'
                    }
            
            self.print_success(f"Loaded {len(existing_data)} existing medications")
            return existing_data, latest_file
            
        except Exception as e:
            self.print_error(f"Error loading existing data: {e}")
            return {}, None
    
    def get_disease_associations(self):
        """Get disease associations for medications from the main diseases Excel"""
        self.print_section("LOADING DISEASE ASSOCIATIONS")
        
        try:
            # Read the main diseases analysis file
            xl_file = pd.ExcelFile('../Analysis/main_diseases_analysis_final.xlsx')
            medication_to_diseases = {}
            
            # Look for the "All Unique Medications" sheet
            target_sheet = None
            for sheet_name in xl_file.sheet_names:
                if 'all unique medications' in sheet_name.lower():
                    target_sheet = sheet_name
                    break
            
            if target_sheet:
                self.print_info(f"Found medications sheet: {target_sheet}")
                sheet_df = pd.read_excel('../Analysis/main_diseases_analysis_final.xlsx', sheet_name=target_sheet)
                
                # The columns are unnamed, so we use indices:
                # Column 0: Medication Name
                # Column 5: Disease Tag (last column)
                
                self.print_success(f"Reading from columns: 0 (medication) and 5 (disease tag)")
                
                for _, row in sheet_df.iterrows():
                    medication = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
                    disease_tag = str(row.iloc[5]).strip() if pd.notna(row.iloc[5]) else ''
                    
                    # Skip header rows and invalid entries
                    if (medication and 
                        medication != 'nan' and 
                        medication.upper() not in ['MEDICATION NAME', 'INFORMATION', 'PURPOSE:', 'SOURCE:', 'ENHANCEMENT:'] and
                        len(medication) > 2 and
                        disease_tag and
                        disease_tag != 'nan'):
                        medication_to_diseases[medication] = disease_tag
                
                self.print_success(f"Loaded {len(medication_to_diseases)} disease associations from {target_sheet}")
                return medication_to_diseases
            else:
                self.print_warning("'All Unique Medications' sheet not found")
            
            # Fallback: try to extract from individual disease sheets
            self.print_info("Falling back to individual disease sheets...")
            
            target_diseases = [
                'Heart disease', 'Chronic kidney disease', 'COPD', 'Pneumonia', 'Stroke',
                'Dementia', 'Depression (major depressive disorder)', 'High cholesterol', 'Obesity', 'Arthritis'
            ]
            
            for disease in target_diseases:
                # Clean disease name for sheet matching
                clean_disease = disease.replace('(', '').replace(')', '').replace('/', '-')[:31]
                
                if clean_disease in xl_file.sheet_names:
                    self.print_info(f"Processing disease sheet: {clean_disease}")
                    disease_df = pd.read_excel('../Analysis/main_diseases_analysis_final.xlsx', sheet_name=clean_disease)
                    
                    # Look for medication names in the first column
                    for _, row in disease_df.iterrows():
                        medication = str(row.iloc[0]).strip() if len(row) > 0 else ''
                        
                        # Check if this looks like a medication name (not a header or empty)
                        if (medication and 
                            medication != 'nan' and 
                            medication.upper() not in ['MEDICATION NAME', 'DISEASE INFORMATION', 'DIAGNOSIS', 'TREATMENTS', 'DIAGNOSTIC TESTS', 'MEDICATIONS & DRUGS'] and
                            len(medication) > 2 and
                            not medication.startswith('Total medications')):
                            
                            if medication not in medication_to_diseases:
                                medication_to_diseases[medication] = disease
                            else:
                                # If medication already exists, append disease
                                existing = medication_to_diseases[medication]
                                if disease not in existing:
                                    medication_to_diseases[medication] = f"{existing}; {disease}"
            
            self.print_success(f"Loaded {len(medication_to_diseases)} disease associations from individual sheets")
            return medication_to_diseases
            
        except Exception as e:
            self.print_error(f"Error loading disease associations: {e}")
            return {}
    
    def read_original_medications(self):
        self.print_section("READING MEDICATIONS FROM ORIGINAL EXCEL")
        
        try:
            df = pd.read_excel('../Analysis/main_diseases_analysis_final.xlsx')
            
            self.print_info(f"Available columns: {list(df.columns)}")
            
            medication_column = None
            
            for col in df.columns:
                col_str = str(col).lower()
                if 'unique medications' in col_str or 'all unique' in col_str:
                    medication_column = col
                    break
            
            if medication_column is None:
                self.print_info("Searching in all Excel sheets...")
                xl_file = pd.ExcelFile('../Analysis/main_diseases_analysis_final.xlsx')
                
                for sheet_name in xl_file.sheet_names:
                    self.print_info(f"Checking sheet: {sheet_name}")
                    sheet_df = pd.read_excel('../Analysis/main_diseases_analysis_final.xlsx', sheet_name=sheet_name)
                    
                    for col in sheet_df.columns:
                        col_str = str(col).lower()
                        if 'unique medications' in col_str or 'all unique' in col_str or 'medication' in col_str:
                            medication_column = col
                            self.print_success(f"Column found in sheet '{sheet_name}': {col}")
                            df = sheet_df
                            break
                    if medication_column:
                        break
            
            if medication_column is None:
                self.print_warning("Medication column not found. Using the first column.")
                medication_column = df.columns[0]
            
            self.print_info(f"Selected column: {medication_column}")
            
            medications = []
            for _, row in df.iterrows():
                medication = str(row[medication_column]).strip()
                if pd.notna(medication) and medication != 'nan':
                    medications.append(medication)
            
            self.print_success(f"Total medications in original Excel: {len(medications)}")
            return medications
            
        except Exception as e:
            self.print_error(f"Error reading original Excel: {e}")
            return []
    
    def identify_missing_medications(self, original_medications):
        self.print_section("IDENTIFYING MISSING MEDICATIONS")
        
        excluded_terms = [
            'ENHANCED SUMMARY', 'Total Unique Medications: 196', 
            'Diseases Analyzed: 10', 'Next Steps: Populate columns B-E with medication data',
            'INFORMATION', 'Purpose:', 'Source:', 'Enhancement:', 'MEDICATION NAME',
            'Total Unique Medications:', 'Diseases Analyzed:', 'Next Steps:'
        ]
        
        valid_original_medications = []
        for medication in original_medications:
            is_valid = True
            for excluded in excluded_terms:
                if excluded.lower() in medication.lower():
                    is_valid = False
                    break
            if is_valid and len(medication.strip()) > 2:
                valid_original_medications.append(medication)
        
        self.print_success(f"Valid medications from original: {len(valid_original_medications)}")
        
        missing_medications = []
        for medication in valid_original_medications:
            if medication not in self.existing_data:
                missing_medications.append(medication)
        
        self.print_info(f"Missing medications: {len(missing_medications)}")
        self.print_success(f"Already existing medications: {len(valid_original_medications) - len(missing_medications)}")
        
        if missing_medications:
            self.print_section("MISSING MEDICATIONS LIST")
            for i, med in enumerate(missing_medications[:10], 1):
                self.print_info(f"{i:2d}. {med}")
            if len(missing_medications) > 10:
                self.print_info(f"... and {len(missing_medications) - 10} more")
        
        return missing_medications
    
    def find_medication_link(self, medication_name):
        try:
            time.sleep(2)
            
            clean_name = medication_name.lower().strip()
            clean_name_no_spaces = clean_name.replace(' ', '')
            
            selectors = [
                "a[href*='/drugs/']",
                "a[href*='/drug/']", 
                "a[href*='/medication/']",
                "a[href*='/drugs.com/']",
                ".search-results a",
                ".drug-results a",
                "a[href*='drugs.com']",
                ".drug-link a",
                ".result-item a",
                "a[href*='/drugs/']"
            ]
            
            for selector in selectors:
                try:
                    links = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    
                    for link in links:
                        try:
                            link_text = link.text.lower().strip()
                            href = link.get_attribute('href').lower()
                            
                            if (clean_name in link_text or 
                                clean_name in href or
                                clean_name_no_spaces in href or
                                any(word in link_text for word in clean_name.split() if len(word) > 2)):
                                
                                exclude_terms = ['side effects', 'español', 'spanish', 'interactions', 'pregnancy', 'breastfeeding', 'overdose']
                                if not any(exclude in link_text for exclude in exclude_terms):
                                    print(f"✅ Found link: {link_text} -> {href}")
                                    return link
                        except Exception as e:
                            continue
                except Exception as e:
                    continue
            
            page_text = self.driver.page_source.lower()
            if clean_name in page_text:
                all_links = self.driver.find_elements(By.TAG_NAME, "a")
                for link in all_links:
                    try:
                        href = link.get_attribute('href')
                        link_text = link.text.lower().strip()
                        
                        if href and (clean_name in href.lower() or clean_name in link_text):
                            exclude_terms = ['side effects', 'español', 'spanish', 'interactions', 'pregnancy', 'breastfeeding', 'overdose']
                            if not any(exclude in href.lower() for exclude in exclude_terms):
                                print(f"✅ Found link via text search: {link_text} -> {href}")
                                return link
                    except:
                        continue
            
            print(f"❌ No suitable link found for {medication_name}")
            return None
            
        except Exception as e:
            print(f"❌ Error searching for link: {e}")
            return None
    
    def process_medication(self, medication_name):
        max_retries = 3
        for attempt in range(max_retries):
            try:
                print(f"🔍 Processing: {medication_name} (attempt {attempt + 1}/{max_retries})")
                
                if not self.is_driver_working():
                    print("🔄 Driver not working, restarting...")
                    try:
                        self.restart_driver()
                    except Exception as e:
                        print(f"❌ Error restarting driver: {e}")
                        print("⏸️  Manual intervention needed...")
                        if not self.pause_for_manual_intervention():
                            return
                        try:
                            self.setup_driver()
                        except Exception as e2:
                            print(f"❌ Critical error: {e2}")
                            return
                
                try:
                    self.driver.get("https://www.drugs.com")
                    time.sleep(2)
                except Exception as e:
                    print(f"❌ Error loading drugs.com: {e}")
                    if attempt < max_retries - 1:
                        print(f"🔄 Retrying... (attempt {attempt + 2}/{max_retries})")
                        time.sleep(3)
                        continue
                    else:
                        print(f"⚠️ Failed to load drugs.com after {max_retries} attempts")
                    return None
                
                try:
                    search_box = WebDriverWait(self.driver, 5).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "input[name='searchterm']"))
                    )
                except:
                    search_selectors = [
                        "input[type='text']",
                        "input[placeholder*='search']",
                        "input[placeholder*='Search']",
                        "#search",
                        ".search-input",
                        "input[name='q']",
                        "input[type='search']",
                        ".search-box input",
                        "#searchbox"
                    ]
                    
                    search_box = None
                    for selector in search_selectors:
                        try:
                            search_box = self.driver.find_element(By.CSS_SELECTOR, selector)
                            print(f"✅ Found search box with selector: {selector}")
                            break
                        except:
                            continue
                    
                    if not search_box:
                        print(f"❌ Search box not found for {medication_name}")
                        if attempt < max_retries - 1:
                            print(f"🔄 Retrying... (attempt {attempt + 2}/{max_retries})")
                            time.sleep(1)
                            continue
                        else:
                            print(f"⚠️ Failed to find search box after {max_retries} attempts")
                        return None
                
                search_box.clear()
                search_box.send_keys(medication_name)
                time.sleep(0.5)
                
                try:
                    search_button = self.driver.find_element(By.CSS_SELECTOR, "button[type='submit']")
                    self.driver.execute_script("arguments[0].click();", search_button)
                    print("✅ Search button clicked")
                except:
                    try:
                        search_button_selectors = [
                            "button[type='submit']",
                            "input[type='submit']",
                            ".search-button",
                            ".btn-search",
                            "button:contains('Search')",
                            "input[value*='Search']"
                        ]
                        
                        for selector in search_button_selectors:
                            try:
                                search_button = self.driver.find_element(By.CSS_SELECTOR, selector)
                                self.driver.execute_script("arguments[0].click();", search_button)
                                print(f"✅ Search button clicked with selector: {selector}")
                                break
                            except:
                                continue
                        else:
                            search_box.send_keys(Keys.RETURN)
                            print("✅ Search executed with Enter key")
                    except Exception as e:
                        print(f"⚠️ Error with search button: {e}")
                        search_box.send_keys(Keys.RETURN)
                        print("✅ Search executed with Enter key (fallback)")
                
                time.sleep(2)
                
                medication_link = self.find_medication_link(medication_name)
                if medication_link:
                    try:
                        if not self.is_driver_working():
                            print("🔄 Driver not working before clicking, restarting...")
                            self.restart_driver()
                            continue
                        
                        self.driver.execute_script("arguments[0].click();", medication_link)
                        time.sleep(3)
                        
                        info = self.extract_medication_info(self.driver.page_source, medication_name)
                        
                        print(f"📊 Extracted data for {medication_name}:")
                        print(f"  Brand: {info['brand_name']}")
                        print(f"  Dosage: {info['dosage']}")
                        print(f"  How to Take: {info['how_to_take']}")
                        print(f"  When to Take: {info['when_to_take']}")
                        
                        return info
                    except Exception as e:
                        print(f"❌ Error clicking on link: {e}")
                        if attempt < max_retries - 1:
                            print(f"🔄 Retrying... (attempt {attempt + 2}/{max_retries})")
                            time.sleep(2)
                            continue
                        else:
                            print(f"⚠️ Failed to process {medication_name} after {max_retries} attempts")
                            return None
                else:
                    print(f"❌ No link found for {medication_name}")
                    if attempt < max_retries - 1:
                        print(f"🔄 Retrying... (attempt {attempt + 2}/{max_retries})")
                        time.sleep(2)
                        continue
                    else:
                        print(f"⚠️ Failed to find link after {max_retries} attempts")
                    return None
                    
            except Exception as e:
                print(f"❌ Error processing {medication_name}: {e}")
                if attempt < max_retries - 1:
                    print(f"🔄 Retrying... (attempt {attempt + 2}/{max_retries})")
                    time.sleep(2)
                    continue
                else:
                    print(f"⚠️ Failed to process {medication_name} after {max_retries} attempts")
                return None
        
        print(f"⚠️ Failed to process {medication_name} after {max_retries} attempts")
        return None
    
    def extract_medication_info(self, page_source, medication_name=None):
        # Safety check for page_source
        if not page_source or page_source is None:
            return {
                'brand_name': 'Not found',
                'dosage': 'Not found',
                'how_to_take': 'Not found',
                'when_to_take': 'Not found'
            }
        
        try:
            soup = BeautifulSoup(page_source, 'html.parser')
            
            brand_name = self.extract_brand_name(page_source, medication_name)
            dosage = self.extract_dosage(page_source)
            how_to_take = self.extract_how_to_take(page_source)
            when_to_take = self.extract_when_to_take(page_source)
            
            return {
                'brand_name': brand_name,
                'dosage': dosage,
                'how_to_take': how_to_take,
                'when_to_take': when_to_take
            }
        except Exception as e:
            print(f"❌ Error extracting medication info: {e}")
            return {
                'brand_name': 'Not found',
                'dosage': 'Not found',
                'how_to_take': 'Not found',
                'when_to_take': 'Not found'
            }
    
    def extract_brand_name(self, page_source, medication_name=None):
        """Enhanced brand name extraction focusing on real brand names only"""
        all_brands = []
        
        # Safety check for page_source
        if not page_source or page_source is None:
            return "Not found"
        
        page_lower = page_source.lower()
        
        # Strategy 0: Check enhanced brand database first (most reliable)
        if medication_name and medication_name.lower() in self.enhanced_brand_database:
            enhanced_brands = self.enhanced_brand_database[medication_name.lower()]
            return enhanced_brands
        
        # Strategy 1: Look for the main "Brand Names" section (most reliable)
        # This extracts from the bulleted list format shown in the image
        brand_names_section = self.extract_brand_names_from_section(page_source)
        if brand_names_section:
            all_brands.extend(brand_names_section)
        
        # Strategy 2: Look for the main "Brand names:" section (fallback)
        # This is the primary source of brand names on drugs.com
        brand_section_patterns = [
            r'Brand name[s]?:\s*([^<\n]+?)(?:\n|$|<)',
            r'Brand name[s]?:\s*([^<\n]+?)(?:\s*[,;]|\s*$|\s*\.|\s*<)',
        ]
        
        for pattern in brand_section_patterns:
            matches = re.findall(pattern, page_source, re.IGNORECASE | re.MULTILINE)
            for match in matches:
                if match:
                    # Clean and split the brand names
                    brand_text = match.strip()
                    if brand_text:
                        # Split by commas and clean each brand
                        brand_parts = [part.strip() for part in brand_text.split(',')]
                        for part in brand_parts:
                            if part and self.is_valid_brand_name(part):
                                all_brands.append(part)
    
    def extract_brand_names_from_section(self, page_source):
        """Extract brand names from the Brand Names section with bulleted list format"""
        brands = []
        
        # Look for the Brand Names section
        brand_section_match = re.search(r'Brand Names.*?(?=<h[1-6]|</div>|$)', page_source, re.IGNORECASE | re.DOTALL)
        if brand_section_match:
            section_content = brand_section_match.group(0)
            
            # Extract brand names from bulleted list format
            # Pattern: • Brand Name (Manufacturer, Country) or • Brand Name
            brand_patterns = [
                r'•\s*([A-Za-z0-9\s\-\+\[\]\/]+?)\s*\([^)]*\)',  # Brand Name (Manufacturer, Country)
                r'•\s*([A-Za-z0-9\s\-\+\[\]\/]+?)(?:\s*$|\s*<)',  # Brand Name at end of line
                r'•\s*([A-Za-z0-9\s\-\+\[\]\/]+?)(?:\s*\(|\s*<)',  # Brand Name before parenthesis
            ]
            
            for pattern in brand_patterns:
                matches = re.findall(pattern, section_content, re.MULTILINE)
                for match in matches:
                    brand_name = match.strip()
                    # Clean up the brand name
                    brand_name = re.sub(r'\s+', ' ', brand_name)  # Normalize whitespace
                    brand_name = brand_name.strip()
                    
                    if brand_name and self.is_valid_brand_name(brand_name):
                        brands.append(brand_name)
        
        return brands
        
        # Strategy 2: Look for parentheses with brand names
        # Pattern: medication (brand_name) or medication (brand1, brand2)
        paren_patterns = [
            r'\(([A-Z][a-zA-Z\s\-&,]+?)\)',  # General parentheses
            r'\(([A-Z][a-zA-Z\s\-&,]+?),\s*([A-Z][a-zA-Z\s\-&,]+?)\)',  # Multiple brands in parentheses
        ]
        
        for pattern in paren_patterns:
            matches = re.findall(pattern, page_source)
            for match in matches:
                if isinstance(match, tuple):
                    # Multiple brands in one match
                    for brand_part in match:
                        if brand_part and self.is_valid_brand_name(brand_part.strip()):
                            all_brands.append(brand_part.strip())
                else:
                    # Single brand
                    if match and self.is_valid_brand_name(match.strip()):
                        all_brands.append(match.strip())
        
        # Strategy 3: Search for known brands in comprehensive database
        for brand in self.all_brands:
            if brand.lower() in page_lower:
                # Check if brand appears in a meaningful context
                brand_context = re.search(rf'\b{re.escape(brand)}\b', page_source, re.IGNORECASE)
                if brand_context:
                    # Get surrounding context to verify it's actually a brand name
                    context_start = max(0, brand_context.start() - 100)
                    context_end = min(len(page_source), brand_context.end() + 100)
                    context = page_source[context_start:context_end].lower()
                    
                    # Check if context suggests this is a brand name
                    brand_indicators = [
                        'brand', 'trade', 'proprietary', 'marketed', 'sold as', 'known as',
                        'also known as', 'alternative name', 'common brand', 'available as',
                        'brand name', 'trade name', 'proprietary name'
                    ]
                    
                    # Check for negative context that suggests it's NOT a brand name
                    negative_indicators = [
                        'avoid', 'do not take', 'interaction', 'contraindication', 'warning',
                        'side effect', 'adverse', 'allergy', 'hypersensitivity', 'caution'
                    ]
                    
                    if any(indicator in context for indicator in brand_indicators):
                        # Double check it's not in negative context
                        if not any(neg in context for neg in negative_indicators):
                            all_brands.append(brand)
                    # Also check if it appears in parentheses after the generic name
                    elif re.search(rf'\([^)]*{re.escape(brand)}[^)]*\)', page_source, re.IGNORECASE):
                        # Make sure it's not in a warning or interaction context
                        if not any(neg in context for neg in negative_indicators):
                            all_brands.append(brand)
        
        # Strategy 4: Look for other brand name sections
        other_brand_patterns = [
            r'Also known as:\s*([A-Z][a-zA-Z\s\-&,]+?)(?:\s*[,;]|\s*$|\s*\.|\s*<)',
            r'Common brands?:\s*([A-Z][a-zA-Z\s\-&,]+?)(?:\s*[,;]|\s*$|\s*\.|\s*<)',
            r'Available as:\s*([A-Z][a-zA-Z\s\-&,]+?)(?:\s*[,;]|\s*$|\s*\.|\s*<)',
            r'Marketed as:\s*([A-Z][a-zA-Z\s\-&,]+?)(?:\s*[,;]|\s*$|\s*\.|\s*<)',
        ]
        
        for pattern in other_brand_patterns:
            matches = re.findall(pattern, page_source, re.IGNORECASE)
            for match in matches:
                if isinstance(match, tuple):
                    match = match[0]
                brand_text = self.clean_text(match)
                if brand_text:
                    # Split by common separators and validate each part
                    for separator in [',', ';', 'and', '&', '/', '|']:
                        if separator in brand_text:
                            brand_parts = brand_text.split(separator)
                            for part in brand_parts:
                                cleaned_part = part.strip()
                                if self.is_valid_brand_name(cleaned_part):
                                    all_brands.append(cleaned_part)
                            break
                    else:
                        # No separators found, validate the whole text
                        if self.is_valid_brand_name(brand_text):
                            all_brands.append(brand_text)
        
        # Strategy 5: Look for trademark symbols and registered marks
        trademark_patterns = [
            r'\b([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)\s*[®™©]\b',
            r'\b([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)\s+(?:trademark|brand|proprietary)\b'
        ]
        
        for pattern in trademark_patterns:
            match = re.search(pattern, page_source, re.IGNORECASE)
            if match:
                brand_name = self.clean_text(match.group(1))
                if self.is_valid_brand_name(brand_name):
                    all_brands.append(brand_name)
        
        # Strategy 6: Check for generic terms (only if no specific brands found)
        if not all_brands:
            generic_terms = ['aspirin', 'acetaminophen', 'ibuprofen', 'naproxen', 'metformin', 'lisinopril', 'amlodipine', 'omeprazole', 'simvastatin']
            for generic in generic_terms:
                if generic in page_lower:
                    all_brands.append("Generic")
                    break
        
        # Remove duplicates and clean up
        unique_brands = []
        seen = set()
        for brand in all_brands:
            cleaned_brand = brand.strip()
            if cleaned_brand and cleaned_brand not in seen:
                seen.add(cleaned_brand)
                unique_brands.append(cleaned_brand)
        
        # Filter out inappropriate brand names based on medication type
        filtered_brands = self.filter_inappropriate_brands(unique_brands, page_source)
        
        # Return all brands found, or "Not found" if none
        if filtered_brands:
            if len(filtered_brands) == 1:
                return filtered_brands[0]
            else:
                return " | ".join(filtered_brands[:5])  # Limit to 5 brands to avoid overwhelming
        
        return "Not found"
    
    def filter_inappropriate_brands(self, brands, page_source):
        """Filter out brand names that are inappropriate for the medication type"""
        if not brands:
            return brands
        
        # Get medication type from page content
        page_lower = page_source.lower()
        
        # Define medication categories and their appropriate brands
        medication_categories = {
            'thrombolytic': ['alteplase', 'tenecteplase', 'streptokinase'],
            'anticoagulant': ['warfarin', 'heparin', 'enoxaparin', 'dabigatran', 'rivaroxaban', 'apixaban'],
            'antibiotic': ['amoxicillin', 'ciprofloxacin', 'cefpodoxime', 'vancomycin'],
            'antidepressant': ['amitriptyline', 'sertraline', 'venlafaxine', 'trazodone'],
            'antipsychotic': ['risperidone', 'quetiapine', 'ziprasidone'],
            'statin': ['simvastatin', 'atorvastatin', 'rosuvastatin', 'pravastatin'],
            'ace_inhibitor': ['lisinopril', 'ramipril', 'captopril'],
            'calcium_channel_blocker': ['amlodipine', 'verapamil', 'diltiazem'],
            'diuretic': ['furosemide', 'hydrochlorothiazide', 'spironolactone'],
            'beta_blocker': ['metoprolol', 'atenolol', 'carvedilol'],
            'pain_reliever': ['acetaminophen', 'ibuprofen', 'aspirin', 'naproxen'],
            'anxiolytic': ['alprazolam', 'lorazepam', 'diazepam'],
            'antihistamine': ['diphenhydramine', 'loratadine', 'cetirizine']
        }
        
        # Pain reliever brands that should only appear with pain medications
        pain_brands = ['advil', 'motrin', 'aleve', 'tylenol', 'panadol', 'anacin', 'bayer', 'aspirin']
        
        # Determine if this is a pain medication
        is_pain_medication = any(med in page_lower for med in medication_categories['pain_reliever'])
        
        # Filter out pain reliever brands from non-pain medications
        filtered_brands = []
        for brand in brands:
            brand_lower = brand.lower()
            
            # If it's a pain reliever brand and this is not a pain medication, skip it
            if brand_lower in pain_brands and not is_pain_medication:
                continue
            
            # Additional specific filters
            if brand_lower == 'advil' and not is_pain_medication:
                continue
            if brand_lower == 'motrin' and not is_pain_medication:
                continue
            if brand_lower == 'aleve' and not is_pain_medication:
                continue
            if brand_lower == 'tylenol' and not is_pain_medication:
                continue
            if brand_lower == 'panadol' and not is_pain_medication:
                continue
            if brand_lower == 'anacin' and not is_pain_medication:
                continue
            
            filtered_brands.append(brand)
        
        return filtered_brands
    
    def is_valid_brand_name(self, brand_name):
        """Check if a string is a valid brand name with strict validation"""
        if not brand_name or len(brand_name) < 2 or len(brand_name) > 50:
            return False
        
        brand_lower = brand_name.lower().strip()
        
        # Check for generic terms that shouldn't be considered brand names
        generic_terms = [
            'generic', 'tablet', 'pill', 'capsule', 'liquid', 'injection', 'oral', 'topical', 
            'cream', 'ointment', 'gel', 'patch', 'spray', 'drops', 'solution', 'suspension', 
            'syrup', 'powder', 'granule', 'mg', 'mcg', 'g', 'ml', 'iu', 'units', 'not found', 
            'unknown', 'n/a', 'na', 'names', 'name', 'brand', 'brands', 'are', 'of', 'for', 
            'and', 'or', 'the', 'some', 'side', 'effects', 'may', 'occur', 'during', 'hour', 
            'after', 'receiving', 'is', 'an', 'intravenous', 'powder', 'biosimilars', 'water', 
            'others', 'methylene', 'blue', 'world', 'health', 'organization', 'microgram', 
            'identify', 'alpha', 'leo', 'prepare', 'do', 'not', 'heat', 'release', 'allow', 
            'orally', 'disintegrating', 'injectable', 'needle', 'approved', 'weekly', 'browse', 
            'hydrochloride', 'tablets', '10mg', 'chemical', 'abstracts', 'service', 'ldl-c',
            'throm-bo-lit-ik', 'ssc-ild', 'wi-nrs', 'si-nrs', 'dsm-iv-tr', 'b-al'
        ]
        
        if any(generic in brand_lower for generic in generic_terms):
            return False
        
        # Check for patterns that are clearly not brand names
        invalid_patterns = [
            r'^names?:\s*$',  # Just "names:" or "name:"
            r'^\.\.\.\s*\+\d+\s*more$',  # "... +23 more"
            r'^[a-z]+\s+[a-z]+$',  # All lowercase words
            r'^\d+',  # Starts with number
            r'^[^A-Z]',  # Doesn't start with capital letter
            r'[()]',  # Contains parentheses
            r'^\s*$',  # Just whitespace
            r'^[^a-zA-Z]*$',  # No letters at all
            r'^[a-z]+$',  # Single lowercase word
            r'^[A-Z][a-z]*\s+[a-z]+$',  # Capital word followed by lowercase (like "Some side")
            r'^[A-Z][a-z]*-[a-z]+-[a-z]+$',  # Hyphenated technical terms
            r'^[A-Z]{2,}-[A-Z]{2,}$',  # Acronyms with hyphens
        ]
        
        for pattern in invalid_patterns:
            if re.match(pattern, brand_name, re.IGNORECASE):
                return False
        
        # Check if it looks like a proper brand name (starts with capital letter, contains letters)
        if not re.match(r'^[A-Z][a-zA-Z\s\-&]+$', brand_name):
            return False
        
        # Must contain at least one letter
        if not re.search(r'[a-zA-Z]', brand_name):
            return False
        
        # Should not be too short or too long
        if len(brand_name.strip()) < 2 or len(brand_name.strip()) > 30:
            return False
        
        # Additional validation: must be a known brand or look like a real brand name
        # Check against our comprehensive brand database
        if brand_lower not in [b.lower() for b in self.all_brands]:
            # If not in our database, apply stricter rules
            # Must be a single word or two words maximum
            words = brand_name.split()
            if len(words) > 2:
                return False
            
            # Each word should be 2-15 characters
            for word in words:
                if len(word) < 2 or len(word) > 15:
                    return False
                # Should not contain numbers or special characters (except hyphens)
                if not re.match(r'^[A-Za-z\-]+$', word):
                    return False
        
        return True
    
    def extract_dosage(self, page_source):
        """Extract ALL possible dosage forms from the page"""
        all_dosage_forms = []
        
        # Safety check for page_source
        if not page_source or page_source is None:
            return "Not found"
        
        # Look for dosage information in specific sections first
        dosage_sections = [
            'dosage',
            'strength',
            'available as',
            'form',
            'administration',
            'how supplied',
            'product forms',
            'presentation'
        ]
        
        page_lower = page_source.lower()
        
        # Search in dosage-related sections
        for section in dosage_sections:
            if section in page_lower:
                # Find the section and extract dosage info
                section_start = page_lower.find(section)
                section_end = min(section_start + 3000, len(page_lower))
                section_text = page_source[section_start:section_end]
                
                # Look for dosage patterns in this section
                dosage_forms = self.find_all_dosage_forms_in_text(section_text)
                all_dosage_forms.extend(dosage_forms)
        
        # If not found in sections, search the entire page
        if not all_dosage_forms:
            all_dosage_forms = self.find_all_dosage_forms_in_text(page_source)
        
        # Remove duplicates and clean up
        unique_dosage_forms = []
        seen = set()
        for form in all_dosage_forms:
            cleaned_form = form.strip()
            if cleaned_form and cleaned_form not in seen and len(cleaned_form) > 2:
                seen.add(cleaned_form)
                unique_dosage_forms.append(cleaned_form)
        
        # Return all dosage forms found, or "Not found" if none
        if unique_dosage_forms:
            if len(unique_dosage_forms) == 1:
                return unique_dosage_forms[0]
            else:
                return " | ".join(unique_dosage_forms[:5])  # Limit to 5 forms
        
        return "Not found"
    
    def find_all_dosage_forms_in_text(self, text):
        """Find ALL possible dosage forms in text"""
        all_forms = []
        
        # Safety check for text
        if not text or text is None:
            return all_forms
        
        # Enhanced form patterns to catch multiple forms
        form_patterns = [
            # Oral forms
            r'(?:oral\s+)?(?:tablet|pill|capsule|liquid|suspension|syrup|solution|powder|granule)',
            r'(?:chewable|disintegrating|extended\s+release|effervescent|compounding)',
            r'(?:oral\s+)?(?:tablet|pill|capsule)(?:\s+extended\s+release)?',
            r'(?:oral\s+)?(?:liquid|suspension|syrup|solution)',
            r'(?:oral\s+)?(?:powder|granule|effervescent)',
            r'(?:oral\s+)?(?:drops|spray|lozenge|gum)',
            
            # Injection forms
            r'(?:injection|injectable|subcutaneous|intramuscular|intravenous)',
            r'(?:intravenous\s+solution|subcutaneous\s+injection|intramuscular\s+injection)',
            r'(?:prefilled\s+syringe|auto\s+injector|pen\s+injector)',
            
            # Inhalation forms
            r'(?:inhalation|inhaler|aerosol|nebulizer|powder\s+inhaler|metered\s+dose\s+inhaler)',
            r'(?:dry\s+powder\s+inhaler|soft\s+mist\s+inhaler)',
            
            # Topical forms
            r'(?:topical|cream|ointment|gel|patch|lotion|foam|spray|shampoo)',
            r'(?:transdermal|dermal|cutaneous)',
            
            # Other forms
            r'(?:rectal\s+suppository|ophthalmic|otic|intranasal|nasal\s+spray)',
            r'(?:vaginal|buccal|sublingual|intrauterine)',
            r'(?:ophthalmic\s+drops|ophthalmic\s+ointment|ophthalmic\s+gel)',
            r'(?:otic\s+drops|otic\s+suspension)',
            
            # Generic patterns
            r'(?:tablet|capsule|pill|liquid|suspension|syrup|solution|powder|granule)',
            r'(?:injection|inhalation|topical|rectal|ophthalmic|otic|nasal)',
            r'(?:drops|spray|lozenge|gum|suppository|implant|device)'
        ]
        
        # Look for specific dosage form mentions
        for pattern in form_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            for match in matches:
                if isinstance(match, tuple):
                    match = match[0]
                cleaned = self.clean_text(match)
                if cleaned and len(cleaned) > 2:
                    # Avoid false positives
                    if not any(exclude in cleaned.lower() for exclude in ['ear', 'eye', 'nose', 'mouth', 'skin', 'head', 'hand', 'foot']):
                        standardized = self.standardize_administration_form(cleaned)
                        if standardized not in all_forms:
                            all_forms.append(standardized)
        
        # Look for "Available as" or "Form" sections that list multiple forms
        availability_patterns = [
            r'Available as:\s*([^,\n\r]+)',
            r'Form[s]?:\s*([^,\n\r]+)',
            r'Presentation[s]?:\s*([^,\n\r]+)',
            r'How supplied:\s*([^,\n\r]+)',
            r'Product forms?:\s*([^,\n\r]+)'
        ]
        
        for pattern in availability_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                forms_text = self.clean_text(match.group(1))
                # Split by common separators and clean each form
                for separator in [',', ';', 'and', '&', '/', '|']:
                    if separator in forms_text:
                        form_parts = forms_text.split(separator)
                        for part in form_parts:
                            cleaned_part = part.strip()
                            if cleaned_part and len(cleaned_part) > 2:
                                standardized = self.standardize_administration_form(cleaned_part)
                                if standardized not in all_forms:
                                    all_forms.append(standardized)
                        break
                else:
                    # No separators found, add the whole text
                    if forms_text and len(forms_text) > 2:
                        standardized = self.standardize_administration_form(forms_text)
                        if standardized not in all_forms:
                            all_forms.append(standardized)
        
        return all_forms
    
    def find_dosage_in_text(self, text):
        """Find dosage information in text - focusing on administration form (legacy method)"""
        forms = self.find_all_dosage_forms_in_text(text)
        if forms:
            return self.standardize_administration_form(forms[0])
        return "Not found"
    
    def standardize_administration_form(self, form):
        form_lower = form.lower().strip()
        
        mappings = {
            'tablet': 'Oral tablet',
            'pill': 'Oral tablet',
            'capsule': 'Oral capsule',
            'liquid': 'Oral liquid',
            'suspension': 'Oral suspension',
            'syrup': 'Oral syrup',
            'solution': 'Oral solution',
            'powder': 'Oral powder',
            'granule': 'Oral granule',
            'chewable': 'Chewable tablet',
            'disintegrating': 'Oral tablet, disintegrating',
            'extended release': 'Oral tablet, extended release',
            'effervescent': 'Oral tablet, effervescent',
            'compounding': 'Compounding powder',
            'injection': 'Injection',
            'injectable': 'Injection',
            'subcutaneous': 'Subcutaneous injection',
            'intramuscular': 'Intramuscular injection',
            'intravenous': 'Intravenous solution',
            'inhalation': 'Inhalation',
            'inhaler': 'Inhalation',
            'aerosol': 'Inhalation aerosol',
            'nebulizer': 'Inhalation nebulizer',
            'topical': 'Topical',
            'cream': 'Topical cream',
            'ointment': 'Topical ointment',
            'gel': 'Topical gel',
            'patch': 'Topical patch',
            'rectal suppository': 'Rectal suppository',
            'ophthalmic': 'Ophthalmic',
            'otic': 'Otic',
            'intranasal': 'Intranasal',
            'nasal spray': 'Nasal spray'
        }
        
        for key, value in mappings.items():
            if key in form_lower:
                return value
        
        if any(word in form_lower for word in ['tablet', 'pill', 'capsule']):
            if 'chewable' in form_lower:
                return 'Chewable tablet'
            elif 'disintegrating' in form_lower:
                return 'Oral tablet, disintegrating'
            elif 'extended' in form_lower or 'release' in form_lower:
                return 'Oral tablet, extended release'
            elif 'effervescent' in form_lower:
                return 'Oral tablet, effervescent'
            elif 'oral' in form_lower:
                return 'Oral tablet'
            else:
                return 'Oral tablet'
        elif any(word in form_lower for word in ['liquid', 'suspension', 'syrup', 'solution']):
            return 'Oral liquid'
        elif any(word in form_lower for word in ['powder', 'granule']):
            return 'Oral powder'
        elif any(word in form_lower for word in ['injection', 'injectable']):
            return 'Injection'
        elif any(word in form_lower for word in ['inhalation', 'inhaler', 'aerosol']):
            return 'Inhalation'
        elif any(word in form_lower for word in ['topical', 'cream', 'ointment', 'gel', 'patch']):
            return 'Topical'
        else:
            return form.capitalize()
        
        return "Not found"
    

    
    def extract_how_to_take(self, page_source):
        all_instructions = []
        
        # Safety check for page_source
        if not page_source or page_source is None:
            return "Not found"
        
        how_to_sections = [
            'how to take',
            'how to use',
            'administration',
            'instructions',
            'directions',
            'usage',
            'method of use',
            'patient instructions',
            'dosing instructions',
            'proper use'
        ]
        
        page_lower = page_source.lower()
        
        for section in how_to_sections:
            if section in page_lower:
                section_start = page_lower.find(section)
                section_end = min(section_start + 3000, len(page_lower))
                section_text = page_source[section_start:section_end]
                
                instructions = self.find_all_how_to_take_in_text(section_text)
                all_instructions.extend(instructions)
        
        food_instructions = self.find_food_instructions(page_source)
        if food_instructions:
            all_instructions.extend(food_instructions)
        
        if not all_instructions:
            all_instructions = self.find_all_how_to_take_in_text(page_source)
        
        unique_instructions = []
        seen = set()
        for instruction in all_instructions:
            cleaned = instruction.strip()
            if cleaned and cleaned not in seen and len(cleaned) > 3:
                seen.add(cleaned)
                unique_instructions.append(cleaned)
        
        if unique_instructions:
            if len(unique_instructions) == 1:
                return unique_instructions[0]
            else:
                return " | ".join(unique_instructions[:3])
        
        return "Not found"
    
    def find_how_to_take_in_text(self, text):
        how_patterns = [
            r'(?:take|use)\s+(?:with|without)\s+(?:food|meals)',
            r'(?:take|use)\s+(?:on\s+)?(?:empty|full)\s+(?:stomach)',
            r'(?:with|without)\s+(?:food|meals)',
            r'(?:on\s+)?(?:empty|full)\s+(?:stomach)',
            r'(?:take|use)\s+(?:with|without)\s+(?:water)',
            r'(?:with|without)\s+(?:water)',
            r'(?:with\s+)?(?:a\s+)?(?:full\s+)?(?:glass\s+of\s+water)',
            r'swallow\s+(?:the\s+)?(?:tablet|capsule)\s+(?:whole|with\s+water)',
            r'swallow\s+(?:whole|with\s+water|with\s+food)',
            r'(?:take|use)\s+(?:orally|by\s+mouth)',
            r'(?:oral|injection|inhalation|topical)\s+(?:administration|use)',
            r'how\s+to\s+(?:take|use):\s*([^,\n]+)',
            r'instructions:\s*([^,\n]+)',
            r'directions:\s*([^,\n]+)'
        ]
        
        for pattern in how_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            if matches:
                for match in matches:
                    if isinstance(match, tuple):
                        match = match[0]
                    cleaned = self.clean_text(match)
                    if cleaned and len(cleaned) > 5:
                        return self.simplify_how_to_take(cleaned)
        
        if 'oral' in text.lower() or 'tablet' in text.lower() or 'capsule' in text.lower() or 'pill' in text.lower():
            return 'Oral'
        elif 'inhalation' in text.lower() or 'inhaler' in text.lower():
            return 'Inhalation'
        elif 'topical' in text.lower():
            return 'Topical'
        elif 'injection' in text.lower():
            injection_context = re.search(r'(?:take|use|administer|given)\s+(?:by\s+)?injection', text.lower())
            if injection_context:
                return 'Injection'
            else:
                return 'Oral'
        
        return "Not found"
    
    def simplify_how_to_take(self, text):
        text = text.lower().strip()
        
        mappings = {
            'with food': 'With food',
            'with meals': 'With food',
            'without food': 'Without food',
            'on empty stomach': 'On empty stomach',
            'with water': 'With water',
            'without water': 'Without water',
            'swallow whole': 'Swallow whole',
            'orally': 'Oral',
            'by mouth': 'Oral',
            'injection': 'Injection',
            'inhalation': 'Inhalation',
            'topical': 'Topical',
            'subcutaneous': 'Subcutaneous injection',
            'intramuscular': 'Intramuscular injection',
            'intravenous': 'Intravenous solution'
        }
        
        for key, value in mappings.items():
            if key in text:
                return value
        
        return text.capitalize()
    
    def extract_when_to_take(self, page_source):
        # Safety check for page_source
        if not page_source or page_source is None:
            return "Not found"
        
        # Strategy 1: Look for specific "when to take" sections first
        when_sections = [
            'when to take',
            'when to use',
            'timing',
            'schedule',
            'administration schedule',
            'dosing schedule',
            'frequency',
            'how often',
            'dosage and administration',
            'administration',
            'instructions',
            'directions for use',
            'patient instructions',
            'dosing instructions'
        ]
        
        page_lower = page_source.lower()
        
        for section in when_sections:
            if section in page_lower:
                section_start = page_lower.find(section)
                section_end = min(section_start + 2000, len(page_lower))
                section_text = page_source[section_start:section_end]
                
                when_to_take = self.find_when_to_take_in_text(section_text)
                if when_to_take != "Not found":
                    return when_to_take
        
        # Strategy 2: Look for dosage sections that might contain timing info
        dosage_sections = [
            'dosage',
            'dosing',
            'strength',
            'available as',
            'how supplied'
        ]
        
        for section in dosage_sections:
            if section in page_lower:
                section_start = page_lower.find(section)
                section_end = min(section_start + 1500, len(page_lower))
                section_text = page_source[section_start:section_end]
                
                when_to_take = self.find_when_to_take_in_text(section_text)
                if when_to_take != "Not found":
                    return when_to_take
        
        # Strategy 3: Search the entire page for timing patterns
        return self.find_when_to_take_in_text(page_source)
    
    def find_when_to_take_in_text(self, text):
        when_patterns = [
            # Frequency patterns
            r'(?:take|use|administer|give)\s+(?:every\s+)?(\d+\s+(?:hours?|days?|weeks?|months?))',
            r'(?:take|use|administer|give)\s+(?:once|twice|three\s+times|four\s+times|five\s+times)\s+(?:daily|per\s+day|a\s+day)',
            r'(?:take|use|administer|give)\s+(\d+)\s+times?\s+(?:daily|per\s+day|a\s+day)',
            r'(?:take|use|administer|give)\s+(?:once|twice|three\s+times|four\s+times|five\s+times)\s+(?:every\s+\d+\s+(?:hours?|days?))',
            
            # Time of day patterns
            r'(?:take|use|administer|give)\s+(?:in\s+)?(?:the\s+)?(?:morning|afternoon|evening|night|bedtime|at\s+bedtime)',
            r'(?:take|use|administer|give)\s+(?:with\s+)?(?:breakfast|lunch|dinner|meals|food)',
            r'(?:take|use|administer|give)\s+(?:before\s+)?(?:bed|sleep|going\s+to\s+bed)',
            r'(?:take|use|administer|give)\s+(?:on\s+an?\s+)?(?:empty\s+stomach|full\s+stomach)',
            r'(?:take|use|administer|give)\s+(?:at\s+)?(?:(\d{1,2}):(\d{2})\s*(?:AM|PM|am|pm)?)',
            r'(?:take|use|administer|give)\s+(?:at\s+)?(?:(\d{1,2})\s*(?:AM|PM|am|pm))',
            
            # As needed patterns
            r'(?:take|use|administer|give)\s+(?:as\s+)?(?:needed|required|necessary|prn)',
            r'(?:take|use|administer|give)\s+(?:when\s+)?(?:needed|required|necessary)',
            r'(?:take|use|administer|give)\s+(?:for\s+)?(?:pain|symptoms|discomfort)',
            
            # Specific timing patterns
            r'(?:take|use|administer|give)\s+(?:at\s+)?(?:the\s+)?(?:same\s+time\s+every\s+day)',
            r'(?:take|use|administer|give)\s+(?:at\s+)?(?:regular\s+intervals)',
            r'(?:take|use|administer|give)\s+(?:continuously|around\s+the\s+clock)',
            
            # Duration patterns
            r'(?:take|use|administer|give)\s+(?:for\s+)?(\d+\s+(?:days?|weeks?|months?|years?))',
            r'(?:take|use|administer|give)\s+(?:until\s+)?(?:symptoms\s+improve|pain\s+relief)',
            
            # Simple frequency patterns (standalone)
            r'(?:once|twice|three\s+times|four\s+times|five\s+times)\s+(?:daily|per\s+day|a\s+day)',
            r'(\d+)\s+times?\s+(?:daily|per\s+day|a\s+day)',
            r'(?:every\s+\d+\s+(?:hours?|days?|weeks?|months?))',
            r'(?:morning|afternoon|evening|night|bedtime)',
            r'(?:daily|regularly|continuously)',
            
            # With/without food patterns
            r'(?:with\s+food|without\s+food|on\s+an?\s+empty\s+stomach|on\s+a\s+full\s+stomach)',
            
            # Before/after patterns
            r'(?:before|after)\s+(?:meals|eating|food|breakfast|lunch|dinner)',
            r'(?:before|after)\s+(?:bedtime|sleep|going\s+to\s+bed)',
            
            # Specific conditions
            r'(?:when\s+)?(?:pain\s+occurs|symptoms\s+appear|needed\s+for\s+pain)',
            r'(?:as\s+directed\s+by\s+your\s+doctor|as\s+prescribed)',
            
            # Time-specific patterns
            r'(?:at\s+)?(?:(\d{1,2}):(\d{2})\s*(?:AM|PM|am|pm)?)',
            r'(?:at\s+)?(?:(\d{1,2})\s*(?:AM|PM|am|pm))',
            r'(?:every\s+(\d+)\s+to\s+(\d+)\s+hours?)',
            r'(?:every\s+(\d+)\s+to\s+(\d+)\s+days?)'
        ]
        
        for pattern in when_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            if matches:
                for match in matches:
                    if isinstance(match, tuple):
                        match = match[0]
                    cleaned = self.clean_text(match)
                    if cleaned and len(cleaned) > 3:
                        return self.standardize_when_to_take(cleaned)
        
        return "Not found"
    
    def standardize_when_to_take(self, when_to_take):
        if not when_to_take or when_to_take == "Not found":
            return "Not found"
        
        text = when_to_take.lower().strip()
        
        # Basic mappings
        mappings = {
            'morning': 'Morning',
            'afternoon': 'Afternoon', 
            'evening': 'Evening',
            'night': 'Night',
            'bedtime': 'Bedtime',
            'at bedtime': 'At bedtime',
            'bed': 'Bedtime',
            'breakfast': 'With breakfast',
            'lunch': 'With lunch',
            'dinner': 'With dinner',
            'food': 'With food',
            'meals': 'With meals',
            'empty stomach': 'On empty stomach',
            'full stomach': 'With food',
            'as needed': 'As needed',
            'when needed': 'As needed',
            'prn': 'As needed',
            'daily': 'Daily',
            'regularly': 'Regularly',
            'continuously': 'Continuously',
            'around the clock': 'Continuously',
            'same time every day': 'Same time daily',
            'regular intervals': 'Regular intervals',
            'as directed': 'As directed',
            'as prescribed': 'As prescribed'
        }
        
        for key, value in mappings.items():
            if key in text:
                return value
        
        # Enhanced frequency patterns
        frequency_patterns = [
            (r'every (\d+) hours?', r'Every \1 hours'),
            (r'every (\d+) days?', r'Every \1 days'),
            (r'every (\d+) weeks?', r'Every \1 weeks'),
            (r'every (\d+) months?', r'Every \1 months'),
            (r'(\d+) times? daily', r'\1 times daily'),
            (r'(\d+) times? per day', r'\1 times daily'),
            (r'(\d+) times? a day', r'\1 times daily'),
            (r'once daily', 'Once daily'),
            (r'twice daily', 'Twice daily'),
            (r'three times daily', 'Three times daily'),
            (r'four times daily', 'Four times daily'),
            (r'five times daily', 'Five times daily'),
            (r'once a day', 'Once daily'),
            (r'twice a day', 'Twice daily'),
            (r'three times a day', 'Three times daily'),
            (r'four times a day', 'Four times daily'),
            (r'five times a day', 'Five times daily'),
            (r'(\d+) times a day', r'\1 times daily'),
            (r'(\d+) times per day', r'\1 times daily'),
            (r'every (\d+) to (\d+) hours?', r'Every \1-\2 hours'),
            (r'every (\d+) to (\d+) days?', r'Every \1-\2 days')
        ]
        
        for pattern, replacement in frequency_patterns:
            if re.search(pattern, text):
                return re.sub(pattern, replacement, text, flags=re.IGNORECASE)
        
        # Time patterns
        time_patterns = [
            (r'(\d{1,2}):(\d{2})\s*(am|pm)', r'\1:\2 \3'),
            (r'(\d{1,2})\s*(am|pm)', r'\1:00 \2'),
            (r'at (\d{1,2}):(\d{2})', r'At \1:\2'),
            (r'at (\d{1,2})', r'At \1:00')
        ]
        
        for pattern, replacement in time_patterns:
            if re.search(pattern, text):
                return re.sub(pattern, replacement, text, flags=re.IGNORECASE)
        
        # Duration patterns
        duration_patterns = [
            (r'for (\d+) days?', r'For \1 days'),
            (r'for (\d+) weeks?', r'For \1 weeks'),
            (r'for (\d+) months?', r'For \1 months'),
            (r'for (\d+) years?', r'For \1 years'),
            (r'until symptoms improve', 'Until symptoms improve'),
            (r'until pain relief', 'Until pain relief')
        ]
        
        for pattern, replacement in duration_patterns:
            if re.search(pattern, text):
                return re.sub(pattern, replacement, text, flags=re.IGNORECASE)
        
        # Food-related patterns
        food_patterns = [
            (r'with food', 'With food'),
            (r'without food', 'Without food'),
            (r'on empty stomach', 'On empty stomach'),
            (r'on full stomach', 'With food'),
            (r'before meals', 'Before meals'),
            (r'after meals', 'After meals'),
            (r'before eating', 'Before meals'),
            (r'after eating', 'After meals'),
            (r'before breakfast', 'Before breakfast'),
            (r'after breakfast', 'After breakfast'),
            (r'before lunch', 'Before lunch'),
            (r'after lunch', 'After lunch'),
            (r'before dinner', 'Before dinner'),
            (r'after dinner', 'After dinner')
        ]
        
        for pattern, replacement in food_patterns:
            if re.search(pattern, text):
                return re.sub(pattern, replacement, text, flags=re.IGNORECASE)
        
        # Pain/symptom patterns
        symptom_patterns = [
            (r'when pain occurs', 'When pain occurs'),
            (r'when symptoms appear', 'When symptoms appear'),
            (r'for pain', 'For pain'),
            (r'for symptoms', 'For symptoms'),
            (r'for discomfort', 'For discomfort')
        ]
        
        for pattern, replacement in symptom_patterns:
            if re.search(pattern, text):
                return re.sub(pattern, replacement, text, flags=re.IGNORECASE)
        
        return when_to_take.strip().capitalize()
    
    def clean_text(self, text):
        if not text or text is None:
            return ""
        
        # Ensure text is a string
        text = str(text)
        
        text = re.sub(r'<[^>]+>', '', text)
        text = re.sub(r'&quot;', '', text)
        text = re.sub(r'&amp;', '&', text)
        text = re.sub(r'&lt;', '<', text)
        text = re.sub(r'&gt;', '>', text)
        text = re.sub(r'&nbsp;', ' ', text)
        text = re.sub(r'window\.\w+', '', text)
        text = re.sub(r'get\s*\(\s*\)', '', text)
        text = re.sub(r'function\s*\([^)]*\)', '', text)
        text = re.sub(r'var\s+\w+', '', text)
        text = re.sub(r'console\.\w+', '', text)
        text = re.sub(r'&quot;get&quot;', '', text)
        text = re.sub(r'get\s*\(\s*\)', '', text)
        text = re.sub(r'window\.sup_platform', '', text)
        text = re.sub(r'\s+', ' ', text)
        text = text.strip()
        
        if len(text) < 3 or text.lower() in ['get', 'window', 'function', 'var']:
            return ""
        
        return text
    
    def clean_and_format_data(self, data):
        cleaned_data = {}
        
        for medication, info in data.items():
            brand_name = info['brand_name']
            if brand_name and brand_name != "Not found":
                brand_name = re.sub(r'\([^)]*\)', '', brand_name)
                brand_name = brand_name.split(',')[0].strip()
                brand_name = brand_name.capitalize()
                if len(brand_name) > 50:
                    brand_name = brand_name[:50]
                if any(generic in brand_name.lower() for generic in ['generic', 'tablet', 'pill', 'capsule', 'liquid', 'injection', 'oral', 'pain']):
                    brand_name = "Generic"
            else:
                brand_name = "Not found"
            
            dosage = info['dosage']
            if dosage and dosage != "Not found":
                dosage = re.sub(r'(side effects|drugs|guide|form:|forms:|drug information)', '', dosage, flags=re.IGNORECASE)
                dosage = dosage.strip()
                if dosage and len(dosage) > 3:
                    dosage = re.sub(r'\s+', ' ', dosage)
                    if ';' in dosage:
                        parts = dosage.split(';')
                        unique_parts = list(dict.fromkeys([part.strip() for part in parts]))
                        dosage = '; '.join(unique_parts)
                    dosage = re.sub(r'(\b\w+\b)(?:\s*[;,]\s*\1)+', r'\1', dosage)
                    dosage = re.sub(r'(\b\w+\b)(?:\s+\1)+', r'\1', dosage)
                else:
                    dosage = "Not found"
            else:
                dosage = "Not found"
            
            how_to_take = info['how_to_take']
            if how_to_take and how_to_take != "Not found":
                how_to_take = self.simplify_instructions(how_to_take)
                if how_to_take and how_to_take != "Not found":
                    how_to_take = how_to_take.capitalize()
                else:
                    how_to_take = "Not found"
            else:
                how_to_take = "Not found"
            
            when_to_take = info['when_to_take']
            if when_to_take and when_to_take != "Not found":
                when_to_take = re.sub(r'\s+', ' ', when_to_take.strip())
                when_to_take = self.standardize_when_to_take(when_to_take)
            else:
                when_to_take = "Not found"
            
            cleaned_data[medication] = {
                'brand_name': brand_name,
                'dosage': dosage,
                'how_to_take': how_to_take,
                'when_to_take': when_to_take
            }
        
        return cleaned_data
    

    
    def simplify_instructions(self, text):
        if not text or text == "Not found":
            return "Not found"
        
        text = text.lower().strip()
        
        mappings = {
            'with food': 'With food',
            'with meals': 'With food',
            'without food': 'Without food',
            'on empty stomach': 'On empty stomach',
            'with water': 'With water',
            'with plenty of water': 'With water',
            'without water': 'Without water',
            'orally': 'Oral',
            'by mouth': 'Oral',
            'swallow whole': 'Swallow whole',
            'with a full glass of water': 'With water',
            'take with food': 'With food',
            'take without food': 'Without food',
            'take on empty stomach': 'On empty stomach',
            'take by mouth': 'Oral',
            'take orally': 'Oral',
            'before meals': 'Before meals',
            'after meals': 'After meals',
            'swallow the tablet': 'Swallow whole',
            'swallow the capsule': 'Swallow whole',
            'take with a full glass of water': 'With water',
            'take with plenty of water': 'With water',
            'with food or water': 'With food',
            'take with food or water': 'With food',
            'on full stomach': 'With food',
            'take on full stomach': 'With food',
            'swallow the tablet whole': 'Swallow whole',
            'swallow the capsule whole': 'Swallow whole',
            'take with meals': 'With food',
            'take without meals': 'Without food',
            'take with water': 'With water',
            'take without water': 'Without water',
            'injection': 'Injection',
            'injectable': 'Injection',
            'subcutaneous': 'Injection',
            'intramuscular': 'Injection',
            'intravenous': 'Injection',
            'inhalation': 'Inhalation',
            'inhale': 'Inhalation',
            'inhaler': 'Inhalation',
            'aerosol': 'Inhalation',
            'topical': 'Topical',
            'apply': 'Topical',
            'cream': 'Topical',
            'ointment': 'Topical'
        }
        
        for key, value in mappings.items():
            if key in text:
                return value
        
        if 'with food' in text or 'with meals' in text:
            return 'With food'
        elif 'without food' in text:
            return 'Without food'
        elif 'with water' in text:
            return 'With water'
        elif 'without water' in text:
            return 'Without water'
        elif 'empty stomach' in text:
            return 'On empty stomach'
        elif 'orally' in text or 'by mouth' in text:
            return 'Oral'
        elif 'swallow' in text:
            return 'Swallow whole'
        elif 'injection' in text or 'injectable' in text:
            return 'Injection'
        elif 'inhalation' in text or 'inhale' in text or 'inhaler' in text:
            return 'Inhalation'
        elif 'topical' in text or 'apply' in text:
            return 'Topical'
        
        valid_instruction_keywords = [
            'with', 'without', 'food', 'water', 'meal', 'stomach', 'empty', 'full',
            'swallow', 'take', 'before', 'after', 'during', 'whole'
        ]
        
        form_keywords = [
            'oral', 'injection', 'inhalation', 'topical', 'tablet', 'capsule', 'mouth'
        ]
        
        text_lower = text.lower()
        has_valid_keywords = any(keyword in text_lower for keyword in valid_instruction_keywords)
        has_form_keywords = any(keyword in text_lower for keyword in form_keywords)
        
        if has_valid_keywords and not has_form_keywords and len(text) > 5:
            return text.capitalize()
        else:
            return "Not found"
    
    def update_excel(self, new_data, existing_file):
        self.print_section("UPDATING EXCEL FILE")
        
        # Get disease associations
        disease_associations = self.get_disease_associations()
        
        try:
            if existing_file:
                df = pd.read_excel(existing_file)
                
                new_rows = []
                for medication, data in new_data.items():
                    # Get disease tag for this medication
                    disease_tag = disease_associations.get(medication, 'Unknown')
                    
                    new_rows.append({
                        'Medication Name': medication,
                        'Brand Names': data['brand_name'],
                        'Dosage Forms': data['dosage'],
                        'How to Take': data['how_to_take'],
                        'When to Take': data['when_to_take'],
                        'Disease Tag': disease_tag
                    })
                
                new_df = pd.DataFrame(new_rows)
                updated_df = pd.concat([df, new_df], ignore_index=True)
                updated_df.to_excel(existing_file, index=False)
                
                self.print_success(f"Excel updated: {existing_file}")
                self.print_success(f"New medications added: {len(new_data)}")
                self.print_success(f"Total medications in file: {len(updated_df)}")
                
                return existing_file
            else:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                new_filename = f"../Analysis/medication_data_{timestamp}.xlsx"
                
                new_rows = []
                for medication, data in new_data.items():
                    # Get disease tag for this medication
                    disease_tag = disease_associations.get(medication, 'Unknown')
                    
                    new_rows.append({
                        'Medication Name': medication,
                        'Brand Names': data['brand_name'],
                        'Dosage Forms': data['dosage'],
                        'How to Take': data['how_to_take'],
                        'When to Take': data['when_to_take'],
                        'Disease Tag': disease_tag
                    })
                
                new_df = pd.DataFrame(new_rows)
                new_df.to_excel(new_filename, index=False)
                
                self.print_success(f"New Excel created: {new_filename}")
                self.print_success(f"Medications added: {len(new_data)}")
                
                return new_filename
            
        except Exception as e:
            self.print_error(f"Error updating Excel: {e}")
            return None
    
    def update_how_to_take_only(self, cache, existing_file):
        self.print_section("UPDATING HOW TO TAKE COLUMN")
        
        try:
            df = pd.read_excel(existing_file)
            self.print_success(f"Total medications: {len(df)}")
            
            self.print_section("CURRENT HOW TO TAKE STATISTICS")
            how_to_take_counts = df['How to Take'].value_counts()
            for pattern, count in how_to_take_counts.head(10).items():
                percentage = (count / len(df)) * 100
                self.print_info(f"{pattern}: {count} ({percentage:.1f}%)")
            
            self.print_section("IMPROVING HOW TO TAKE COLUMN")
            improved_count = 0
            
            for index, row in df.iterrows():
                medication = row['Name']
                if medication in cache:
                    original = row['How to Take']
                    improved = self.simplify_instructions(original)
                    
                    if improved != original:
                        self.print_success(f"{medication}: '{original}' → '{improved}'")
                        df.at[index, 'How to Take'] = improved
                        improved_count += 1
            
            self.print_success(f"Improved {improved_count} entries")
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            new_filename = f"../Analysis/medication_data_{timestamp}.xlsx"
            df.to_excel(new_filename, index=False)
            
            self.print_success(f"Improved file saved: {new_filename}")
            
            self.print_section("NEW HOW TO TAKE STATISTICS")
            new_how_to_take_counts = df['How to Take'].value_counts()
            for pattern, count in new_how_to_take_counts.head(10).items():
                percentage = (count / len(df)) * 100
                self.print_info(f"{pattern}: {count} ({percentage:.1f}%)")
            
            self.print_section("MOST COMMON PATTERNS")
            patterns = {
                'With food': 0,
                'With water': 0,
                'On empty stomach': 0,
                'Oral': 0,
                'Injection': 0,
                'Inhalation': 0,
                'Topical': 0,
                'Swallow whole': 0,
                'Not found': 0
            }
            
            for value in df['How to Take']:
                if value in patterns:
                    patterns[value] += 1
            
            for pattern, count in patterns.items():
                if count > 0:
                    percentage = (count / len(df)) * 100
                    self.print_info(f"{pattern}: {count} ({percentage:.1f}%)")
                    
        except Exception as e:
            self.print_error(f"Error updating How to Take column: {e}")
    
    def create_enhanced_professional_excel(self):
        self.print_header("🎨 ENHANCED PROFESSIONAL EXCEL CREATOR", "Create Professional Excel with Enhanced Data")
        
        try:
            reprocessed_file = self.enhance_existing_data()
            if not reprocessed_file:
                return
            
            df = pd.read_excel(reprocessed_file)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_filename = f"../Analysis/medication_data_{timestamp}.xlsx"
            
            # Professional Excel creation disabled - only create clean Excel with Disease Tag
            # self.print_section("CREATING PROFESSIONAL EXCEL")
            # self.create_professional_excel(df, excel_filename)
            
            # self.print_success(f"Enhanced professional Excel created: {excel_filename}")
            # self.print_info("Features: Professional styling, alternating row colors, statistics dashboard")
            # self.print_info("Enhanced data quality analysis and visual presentation")
            
            return excel_filename
            
        except Exception as e:
            self.print_error(f"Error creating enhanced Excel: {e}")
            return None
    
    def run(self, limit=None):
        self.print_header("🚀 INTELLIGENT MEDICATION SCRAPING", "Enhanced Brand Name Extraction & Modern Visual Interface")
        
        try:
            self.existing_data, existing_file = self.load_existing_data()
            original_medications = self.read_original_medications()
            missing_medications = self.identify_missing_medications(original_medications)
            
            # Apply limit if specified
            if limit and limit > 0:
                missing_medications = missing_medications[:limit]
                self.print_info(f"🔬 TEST MODE: Limited to {limit} medications for testing")
            
            if missing_medications:
                self.print_section(f"PROCESSING {len(missing_medications)} MEDICATIONS")
                self.print_info(f"Total medications to process: {len(missing_medications)}")
            
            cache = self.load_cache()
            
            if not missing_medications and cache:
                self.print_section("UPDATING EXISTING DATA")
                self.print_success("All medications already processed. Updating 'How to Take' column with improved cleaning...")
                self.update_how_to_take_only(cache, existing_file)
                return
            elif not missing_medications:
                self.print_success("No missing medications. All medications are already in our results.")
                return
            
            cache = self.load_cache()
            self.setup_driver()
            
            scraped_data = {}
            total_missing = len(missing_medications)
            
            self.print_info(f"Processing {total_missing} missing medications in batches of {self.batch_size}...")
            
            for batch_start in range(0, total_missing, self.batch_size):
                batch_end = min(batch_start + self.batch_size, total_missing)
                batch_medications = missing_medications[batch_start:batch_end]
                
                self.print_section(f"BATCH {batch_start//self.batch_size + 1}: Processing medications {batch_start + 1}-{batch_end}")
                
                for i, medication in enumerate(batch_medications, 1):
                    global_index = batch_start + i
                    self.print_progress(global_index, total_missing, f"Processing {medication}")
                    
                    if medication in cache:
                        self.print_success(f"{medication}: Using cached data")
                        scraped_data[medication] = cache[medication]
                        continue
                    
                    try:
                        result = self.process_medication(medication)
                        if result:
                            scraped_data[medication] = result
                            cache[medication] = result
                            self.print_success(f"{medication}: {result['brand_name']} | {result['dosage']} | {result['how_to_take']} | {result['when_to_take']}")
                        else:
                            self.print_error(f"{medication}: Could not process")
                    except Exception as e:
                        self.print_error(f"Error processing {medication}: {e}")
                        time.sleep(1)
                        continue
                
                self.print_info(f"Restarting driver after batch {batch_start//self.batch_size + 1}...")
                
                try:
                    self.restart_driver()
                except Exception as e:
                    self.print_warning(f"Error restarting driver: {e}")
                    self.print_info("Attempting automatic restart...")
                    try:
                        if self.driver:
                            self.driver.quit()
                    except:
                        pass
                    time.sleep(3)
                    try:
                        self.setup_driver()
                    except Exception as e2:
                        self.print_error(f"Critical error setting up driver: {e2}")
                        self.print_info("Trying one more time...")
                        time.sleep(5)
                        try:
                            self.setup_driver()
                        except Exception as e3:
                            self.print_error(f"Final error setting up driver: {e3}")
                            self.print_error("Stopping script due to driver issues")
                            break
                
                if scraped_data:
                    cleaned_data = self.clean_and_format_data(scraped_data)
                    self.save_cache(cache)
                    self.print_success(f"Progress saved after batch {batch_start//self.batch_size + 1}")
                
                self.print_info("Batch completed. Continuing automatically...")
                time.sleep(1)
            
            if self.driver:
                self.driver.quit()
            
            if scraped_data:
                cleaned_data = self.clean_and_format_data(scraped_data)
                self.save_cache(cache)
                updated_file = self.update_excel(cleaned_data, existing_file)
                
                self.print_header("🎉 SCRAPING COMPLETED!", "Enhanced Multi-Brand Extraction Results")
                self.print_success(f"File updated: {updated_file}")
                self.print_success(f"Medications processed: {len(scraped_data)}")
                
                # Enhanced brand extraction summary
                self.print_brand_extraction_summary(cleaned_data)
                
                # Data quality metrics
                self.print_data_quality_metrics(cleaned_data)
                
                # Show brand name categories found
                self.print_section("BRAND NAME CATEGORIES EXTRACTED")
                brand_categories = {}
                for data in cleaned_data.values():
                    if data['brand_name'] != 'Not found':
                        # Handle multiple brands
                        if '|' in str(data['brand_name']):
                            brands = data['brand_name'].split(' | ')
                            for brand in brands:
                                for category, category_brands in self.comprehensive_brands.items():
                                    if brand in category_brands:
                                        brand_categories[category] = brand_categories.get(category, 0) + 1
                                        break
                        else:
                            for category, brands in self.comprehensive_brands.items():
                                if data['brand_name'] in brands:
                                    brand_categories[category] = brand_categories.get(category, 0) + 1
                                    break
                
                for category, count in sorted(brand_categories.items(), key=lambda x: x[1], reverse=True):
                    self.print_info(f"{category}: {count} brands")
                
                # Enhanced brand analysis
                self.analyze_brand_extraction_results(cleaned_data)
            else:
                self.print_warning("No new medications processed")
            
            # Always run enhance command after scraping (regardless of new medications)
            self.print_section("AUTOMATICALLY RUNNING ENHANCE COMMAND")
            self.print_info("Running enhanced Excel creation with Disease Tag column...")
            try:
                enhanced_file = self.create_enhanced_professional_excel()
                if enhanced_file:
                    self.print_success(f"Enhanced Excel created: {enhanced_file}")
                    self.print_info("Features: Disease Tag column added, no data modification")
                else:
                    self.print_warning("Enhanced Excel creation failed")
            except Exception as e:
                self.print_error(f"Error creating enhanced Excel: {e}")
            
        except Exception as e:
            self.print_error(f"Error in scraping: {e}")
            import traceback
            self.print_error(f"Full traceback: {traceback.format_exc()}")
            if self.driver:
                self.driver.quit()
    
    def analyze_brand_extraction_results(self, data):
        """Analyze and display detailed brand extraction results"""
        self.print_section("DETAILED BRAND EXTRACTION ANALYSIS")
        
        total_medications = len(data)
        brand_names_found = sum(1 for d in data.values() if d['brand_name'] != 'Not found')
        generic_found = sum(1 for d in data.values() if d['brand_name'] == 'Generic')
        not_found = sum(1 for d in data.values() if d['brand_name'] == 'Not found')
        
        # Success rates
        brand_success_rate = (brand_names_found / total_medications) * 100
        generic_success_rate = (generic_found / total_medications) * 100
        overall_success_rate = ((brand_names_found + generic_found) / total_medications) * 100
        
        self.print_success(f"Overall Brand Extraction Success: {overall_success_rate:.1f}%")
        self.print_info(f"Specific Brand Names: {brand_names_found} ({brand_success_rate:.1f}%)")
        self.print_info(f"Generic Identifications: {generic_found} ({generic_success_rate:.1f}%)")
        self.print_warning(f"Not Found: {not_found} ({not_found/total_medications*100:.1f}%)")
        
        # Top extracted brands
        brand_counts = {}
        for d in data.values():
            if d['brand_name'] not in ['Not found', 'Generic']:
                brand_counts[d['brand_name']] = brand_counts.get(d['brand_name'], 0) + 1
        
        if brand_counts:
            self.print_section("TOP EXTRACTED BRANDS")
            sorted_brands = sorted(brand_counts.items(), key=lambda x: x[1], reverse=True)
            for brand, count in sorted_brands[:10]:
                self.print_info(f"{brand}: {count} occurrences")
        
        # Category distribution
        self.print_section("BRAND CATEGORY DISTRIBUTION")
        category_counts = {}
        for d in data.values():
            if d['brand_name'] != 'Not found':
                for category, brands in self.comprehensive_brands.items():
                    if d['brand_name'] in brands:
                        category_counts[category] = category_counts.get(category, 0) + 1
                        break
        
        for category, count in sorted(category_counts.items(), key=lambda x: x[1], reverse=True):
            percentage = (count / brand_names_found) * 100 if brand_names_found > 0 else 0
            self.print_info(f"{category}: {count} brands ({percentage:.1f}%)")
        
        # Quality metrics
        self.print_section("DATA QUALITY METRICS")
        dosage_found = sum(1 for d in data.values() if d['dosage'] != 'Not found')
        how_to_take_found = sum(1 for d in data.values() if d['how_to_take'] != 'Not found')
        when_to_take_found = sum(1 for d in data.values() if d['when_to_take'] != 'Not found')
        
        self.print_info(f"Dosage Information: {dosage_found} ({dosage_found/total_medications*100:.1f}%)")
        self.print_info(f"How to Take: {how_to_take_found} ({how_to_take_found/total_medications*100:.1f}%)")
        self.print_info(f"When to Take: {when_to_take_found} ({when_to_take_found/total_medications*100:.1f}%)")
        
        # Recommendations
        self.print_section("IMPROVEMENT RECOMMENDATIONS")
        if brand_success_rate < 70:
            self.print_warning("Brand extraction success rate below 70%. Consider:")
            self.print_info("  - Adding more brand patterns to extraction")
            self.print_info("  - Expanding the comprehensive brand database")
            self.print_info("  - Reviewing failed extractions for patterns")
        
        if not_found > total_medications * 0.3:
            self.print_warning("High number of 'Not found' results. Consider:")
            self.print_info("  - Manual review of failed extractions")
            self.print_info("  - Adding medication-specific extraction rules")
            self.print_info("  - Checking source data quality")
        
        self.print_success("Brand extraction analysis complete!")
    
    def create_professional_excel(self, data, filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Medication Analysis"
        
        colors = {
            'header_bg': '1E3C72',
            'header_text': 'FFFFFF',
            'subheader_bg': '2A5298',
            'subheader_text': 'FFFFFF',
            'stats_bg': 'F8F9FA',
            'stats_border': 'E9ECEF',
            'row_alt1': 'F8F9FA',
            'row_alt2': 'FFFFFF',
            'success_green': '28A745',
            'info_blue': '2A5298',
            'warning_yellow': 'FFC107',
        }
        
        header_style = NamedStyle(name="header_style")
        header_style.font = Font(name='Arial', size=14, bold=True, color=colors['header_text'])
        header_style.fill = PatternFill(start_color=colors['header_bg'], end_color=colors['header_bg'], fill_type='solid')
        header_style.alignment = Alignment(horizontal='center', vertical='center')
        header_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        subheader_style = NamedStyle(name="subheader_style")
        subheader_style.font = Font(name='Arial', size=12, bold=True, color=colors['subheader_text'])
        subheader_style.fill = PatternFill(start_color=colors['subheader_bg'], end_color=colors['subheader_bg'], fill_type='solid')
        subheader_style.alignment = Alignment(horizontal='center', vertical='center')
        subheader_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        ws.merge_cells('A1:F1')
        ws['A1'] = 'MEDICATION COMPREHENSIVE ANALYSIS'
        ws['A1'].style = header_style
        ws.row_dimensions[1].height = 40
        
        ws.merge_cells('A2:F2')
        ws['A2'] = 'Enhanced Multi-Brand & Multi-Dosage Form Extraction with Disease Tags'
        ws['A2'].style = subheader_style
        ws.row_dimensions[2].height = 30
        
        stats_start_row = 4
        ws.merge_cells(f'A{stats_start_row}:F{stats_start_row}')
        ws[f'A{stats_start_row}'] = '📊 ANALYSIS STATISTICS'
        ws[f'A{stats_start_row}'].font = Font(name='Arial', size=12, bold=True)
        ws[f'A{stats_start_row}'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[stats_start_row].height = 25
        
        total_medications = len(data)
        brand_names_found = sum(1 for d in data['Brand Names'] if pd.notna(d) and str(d) != 'Not found')
        multiple_brands = sum(1 for d in data['Brand Names'] if pd.notna(d) and '|' in str(d))
        dosage_forms_found = sum(1 for d in data['Dosage Forms'] if pd.notna(d) and str(d) != 'Not found')
        multiple_dosage_forms = sum(1 for d in data['Dosage Forms'] if pd.notna(d) and '|' in str(d))
        how_to_take_found = sum(1 for d in data['How to Take'] if pd.notna(d) and str(d) != 'Not found')
        when_to_take_found = sum(1 for d in data['When to Take'] if pd.notna(d) and str(d) != 'Not found')
        disease_tags_found = sum(1 for d in data['Disease Tag'] if pd.notna(d) and str(d) != 'Unknown')
        
        stats_data = [
            ['Total Medications', total_medications, 'Brand Names Found', brand_names_found],
            ['Multiple Brands', multiple_brands, 'Dosage Forms Found', dosage_forms_found],
            ['Multiple Dosage Forms', multiple_dosage_forms, 'How to Take Found', how_to_take_found],
            ['When to Take Found', when_to_take_found, 'Disease Tags Found', disease_tags_found],
            ['Success Rate', f"{(brand_names_found/total_medications*100):.1f}%", '', '']
        ]
        
        for i, row_data in enumerate(stats_data):
            row_num = stats_start_row + 1 + i
            for j, value in enumerate(row_data):
                col = chr(ord('A') + j)
                cell = ws[f'{col}{row_num}']
                cell.value = value
                
                if j % 2 == 0:
                    cell.font = Font(name='Arial', size=10, bold=True)
                    cell.fill = PatternFill(start_color=colors['stats_bg'], end_color=colors['stats_bg'], fill_type='solid')
                else:
                    cell.font = Font(name='Arial', size=10)
                    cell.fill = PatternFill(start_color=colors['stats_bg'], end_color=colors['stats_bg'], fill_type='solid')
                
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin', color=colors['stats_border']),
                    right=Side(style='thin', color=colors['stats_border']),
                    top=Side(style='thin', color=colors['stats_border']),
                    bottom=Side(style='thin', color=colors['stats_border'])
                )
        
        table_start_row = stats_start_row + 6
        ws.merge_cells(f'A{table_start_row}:F{table_start_row}')
        ws[f'A{table_start_row}'] = '📋 MEDICATION DETAILS'
        ws[f'A{table_start_row}'].font = Font(name='Arial', size=12, bold=True)
        ws[f'A{table_start_row}'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[table_start_row].height = 25
        
        headers = ['Medication Name', 'Brand Names', 'Dosage Forms', 'How to Take', 'When to Take', 'Disease Tag']
        header_row = table_start_row + 1
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.style = subheader_style
            ws.column_dimensions[chr(ord('A') + col - 1)].width = 25
        
        for i, (_, row) in enumerate(data.iterrows()):
            row_num = header_row + 1 + i
            
            if i % 2 == 0:
                row_color = colors['row_alt1']
            else:
                row_color = colors['row_alt2']
            
            for col, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col)
                cell.value = value
                
                cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type='solid')
                cell.font = Font(name='Arial', size=10)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                cell.border = Border(
                    left=Side(style='thin', color=colors['stats_border']),
                    right=Side(style='thin', color=colors['stats_border']),
                    top=Side(style='thin', color=colors['stats_border']),
                    bottom=Side(style='thin', color=colors['stats_border'])
                )
                
                if col == 2 and '|' in str(value):
                    cell.font = Font(name='Arial', size=10, bold=True, color=colors['info_blue'])
                elif col == 3 and '|' in str(value):
                    cell.font = Font(name='Arial', size=10, bold=True, color=colors['success_green'])
                elif col == 6 and ';' in str(value):
                    cell.font = Font(name='Arial', size=10, bold=True, color=colors['info_blue'])
                elif 'Generic' in str(value):
                    cell.font = Font(name='Arial', size=10, italic=True, color='6C757D')
        

        
        wb.save(filename)
        return filename
    
    def enhance_existing_data(self):
        files = [f for f in os.listdir('../Analysis') if f.startswith('medication_data_') and f.endswith('.xlsx')]
        
        if not files:
            self.print_error("No medication data files found in Analysis directory!")
            return
        
        files.sort(key=lambda x: os.path.getmtime(os.path.join('../Analysis', x)), reverse=True)
        
        # Find the original file with all the data (not a processed one)
        original_file = None
        for file in files:
            if 'medication_data_20250903_190600.xlsx' in file:
                original_file = file
                break
        
        if not original_file:
            # Fallback to most recent file
            original_file = files[0]
            self.print_warning("Original file not found, using most recent file")
        
        self.print_success(f"Processing: {original_file}")
        
        df = pd.read_excel(f'../Analysis/{original_file}')
        self.print_info(f"Loaded {len(df)} medications")
        
        # Get disease associations
        disease_associations = self.get_disease_associations()
        
        # Simply add Disease Tag column to existing data without modifying anything else
        enhanced_data = []
        
        for index, row in df.iterrows():
            # Get all existing data EXACTLY as-is (no modifications)
            # The original file has columns: ['MEDICATION COMPREHENSIVE ANALYSIS', 'Unnamed: 1', 'Unnamed: 2', 'Unnamed: 3', 'Unnamed: 4']
            medication = row.iloc[0]  # First column is medication name
            
            # Skip non-medication rows (headers, statistics, etc.)
            if pd.isna(medication) or str(medication).startswith('📊') or str(medication).startswith('Enhanced') or str(medication) == 'Medication Name':
                continue
                
            brand_name = row.iloc[1] if len(row) > 1 else None
            current_dosage = row.iloc[2] if len(row) > 2 else None
            current_how_to_take = row.iloc[3] if len(row) > 3 else None
            when_to_take = row.iloc[4] if len(row) > 4 else None
            
            # Get disease tag for this medication (ONLY addition)
            disease_tag = disease_associations.get(medication, 'Unknown')
            
            # Keep all existing data exactly as-is, just add Disease Tag
            enhanced_data.append({
                'Medication Name': medication,
                'Brand Names': brand_name,
                'Dosage Forms': current_dosage,
                'How to Take': current_how_to_take,
                'When to Take': when_to_take,
                'Disease Tag': disease_tag
            })
        
        enhanced_df = pd.DataFrame(enhanced_data)
        
        # Keep all data as-is, no filtering needed since we're working with clean data
        clean_df = enhanced_df.copy()
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        table_filename = f"../Analysis/medication_data_{timestamp}.xlsx"
        
        # Apply professional formatting to clean data with full statistics
        self.create_professional_excel(clean_df, table_filename)
        
        self.print_success(f"Enhanced data saved: {table_filename}")
        self.print_info("Added Disease Tag column without modifying existing data")
        
        # Simple statistics without enhancement counts
        self.print_simple_statistics(clean_df, disease_associations)
        
        return table_filename
    
    def print_simple_statistics(self, df, disease_associations):
        """Print simple statistics for the enhanced data"""
        total_medications = len(df)
        disease_tags_found = sum(1 for d in df['Disease Tag'] if d != 'Unknown')
        
        print(f"\n{'═'*70}")
        print(f"{'📊 DISEASE TAG ADDITION RESULTS':^70}")
        print(f"{'═'*70}")
        print(f"📋 Total Medications:           {total_medications}")
        print(f"🏷️ Disease Tags Found:         {disease_tags_found} ({disease_tags_found/total_medications*100:.1f}%)")
        print(f"❓ Unknown Disease Tags:       {total_medications - disease_tags_found} ({(total_medications - disease_tags_found)/total_medications*100:.1f}%)")
        print(f"📊 Total Disease Associations: {len(disease_associations)}")
        print(f"{'═'*70}")
        
        if disease_tags_found > 0:
            print(f"\n✅ Successfully added Disease Tag column with {disease_tags_found} associations!")
        else:
            print(f"\n⚠️ No disease associations found. Check if main_diseases_analysis_final.xlsx exists.")
    
    def create_clean_professional_excel(self, df, filename):
        """Create a clean, professional Excel with just the medication data"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Medication Data"
        
        # Filter out non-medication rows (statistics, headers, etc.)
        medication_data = df.copy()
        
        # Remove rows that are not actual medications
        medication_data = medication_data[
            (medication_data['Medication Name'].notna()) & 
            (medication_data['Medication Name'].astype(str).str.len() > 2) &
            (~medication_data['Medication Name'].astype(str).str.contains('Total|Multiple|When to Take|MEDICATION DETAILS|ANALYSIS STATISTICS', case=False, na=False))
        ]
        
        # Header styling
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Set column headers
        headers = ['Medication Name', 'Brand Names', 'Dosage Forms', 'How to Take', 'When to Take', 'Disease Tag']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data rows (only medication data)
        for row_idx, (_, row) in enumerate(medication_data.iterrows(), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                # Alternate row colors
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        # Set column widths
        ws.column_dimensions['A'].width = 25  # Medication Name
        ws.column_dimensions['B'].width = 30  # Brand Names
        ws.column_dimensions['C'].width = 35  # Dosage Forms
        ws.column_dimensions['D'].width = 30  # How to Take
        ws.column_dimensions['E'].width = 25  # When to Take
        ws.column_dimensions['F'].width = 40  # Disease Tag
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
        
        wb.save(filename)
        self.print_success(f"Clean professional Excel created: {filename} with {len(medication_data)} medications")
    
    def find_additional_dosage_forms(self, medication_name, current_dosage):
        additional_forms = []
        
        medication_patterns = {
            'aspirin': ['Oral tablet', 'Chewable tablet', 'Oral liquid', 'Rectal suppository'],
            'acetaminophen': ['Oral tablet', 'Oral liquid', 'Oral suspension', 'Rectal suppository'],
            'ibuprofen': ['Oral tablet', 'Oral liquid', 'Oral suspension', 'Topical gel'],
            'metformin': ['Oral tablet', 'Oral solution', 'Oral suspension'],
            'lisinopril': ['Oral tablet', 'Oral solution'],
            'amlodipine': ['Oral tablet', 'Oral suspension'],
            'albuterol': ['Inhalation aerosol', 'Inhalation solution', 'Oral tablet', 'Oral syrup'],
            'fluticasone': ['Nasal spray', 'Inhalation aerosol', 'Topical cream', 'Topical ointment'],
            'omeprazole': ['Oral capsule', 'Oral tablet', 'Oral suspension'],
            'simvastatin': ['Oral tablet', 'Oral suspension'],
            'amoxicillin': ['Oral capsule', 'Oral tablet', 'Oral suspension', 'Oral liquid'],
            'prednisone': ['Oral tablet', 'Oral solution', 'Oral suspension'],
            'furosemide': ['Oral tablet', 'Oral solution', 'Injection'],
            'hydrochlorothiazide': ['Oral tablet', 'Oral capsule', 'Oral solution'],
            'atenolol': ['Oral tablet', 'Oral solution'],
            'carvedilol': ['Oral tablet', 'Oral capsule', 'Oral solution'],
            'metoprolol': ['Oral tablet', 'Oral solution', 'Oral suspension'],
            'losartan': ['Oral tablet', 'Oral solution'],
            'valsartan': ['Oral tablet', 'Oral suspension'],
            'amlodipine': ['Oral tablet', 'Oral suspension']
        }
        
        for pattern, forms in medication_patterns.items():
            if pattern.lower() in medication_name.lower():
                for form in forms:
                    if form.lower() != current_dosage.lower() and form not in additional_forms:
                        additional_forms.append(form)
        
        return additional_forms[:3]
    
    def extract_how_to_take_from_context(self, medication_name, dosage_forms):
        instructions = []
        
        if 'Oral' in dosage_forms:
            instructions.append('Oral')
        if 'Injection' in dosage_forms:
            instructions.append('Injection')
        if 'Inhalation' in dosage_forms:
            instructions.append('Inhalation')
        if 'Topical' in dosage_forms:
            instructions.append('Topical')
        if 'Rectal' in dosage_forms:
            instructions.append('Rectal')
        if 'Ophthalmic' in dosage_forms:
            instructions.append('Ophthalmic')
        if 'Otic' in dosage_forms:
            instructions.append('Otic')
        if 'Nasal' in dosage_forms:
            instructions.append('Nasal')
        
        if 'Oral' in dosage_forms:
            if 'tablet' in dosage_forms.lower() or 'capsule' in dosage_forms.lower():
                instructions.append('Swallow whole')
            if 'liquid' in dosage_forms.lower() or 'suspension' in dosage_forms.lower():
                instructions.append('Shake well')
        
        if 'Inhalation' in dosage_forms:
            instructions.append('Prime inhaler')
        
        if 'Topical' in dosage_forms:
            instructions.append('Apply to affected area')
        
        common_with_food = ['aspirin', 'ibuprofen', 'naproxen', 'diclofenac', 'metformin', 'metoprolol']
        if any(med in medication_name.lower() for med in common_with_food):
            instructions.append('Take with food')
        
        unique_instructions = []
        seen = set()
        for instruction in instructions:
            if instruction not in seen:
                seen.add(instruction)
                unique_instructions.append(instruction)
        
        if unique_instructions:
            if len(unique_instructions) == 1:
                return unique_instructions[0]
            else:
                return " | ".join(unique_instructions[:3])
        
        return None
    
    def generate_reprocessing_statistics(self, df, dosage_enhanced, how_to_take_enhanced):
        total_medications = len(df)
        
        brand_names_found = sum(1 for d in df['Brand Names'] if d != 'Not found')
        multiple_brands = sum(1 for d in df['Brand Names'] if '|' in str(d))
        dosage_forms_found = sum(1 for d in df['Dosage Forms'] if d != 'Not found')
        multiple_dosage_forms = sum(1 for d in df['Dosage Forms'] if '|' in str(d))
        how_to_take_found = sum(1 for d in df['How to Take'] if d != 'Not found')
        when_to_take_found = sum(1 for d in df['When to Take'] if d != 'Not found')
        disease_tags_found = sum(1 for d in df['Disease Tag'] if d != 'Unknown')
        
        print(f"\n{'═'*70}")
        print(f"{'📊 REPROCESSING RESULTS':^70}")
        print(f"{'═'*70}")
        
        print(f"{'📋 Total Medications:':<30} {total_medications}")
        print(f"{'✅ Brand Names Found:':<30} {brand_names_found} ({brand_names_found/total_medications*100:.1f}%)")
        print(f"{'🔗 Multiple Brands:':<30} {multiple_brands}")
        
        print(f"\n{'💊 DOSAGE FORMS:':<30}")
        print(f"{'  Forms Found:':<30} {dosage_forms_found} ({dosage_forms_found/total_medications*100:.1f}%)")
        print(f"{'  Multiple Forms:':<30} {multiple_dosage_forms}")
        print(f"{'  Enhanced:':<30} {dosage_enhanced}")
        
        print(f"\n{'📋 INSTRUCTIONS:':<30}")
        print(f"{'  How to Take:':<30} {how_to_take_found} ({how_to_take_found/total_medications*100:.1f}%)")
        print(f"{'  Enhanced:':<30} {how_to_take_enhanced}")
        print(f"{'  When to Take:':<30} {when_to_take_found} ({when_to_take_found/total_medications*100:.1f}%)")
        
        print(f"\n{'🏷️ DISEASE TAGS:':<30}")
        print(f"{'  Disease Tags Found:':<30} {disease_tags_found} ({disease_tags_found/total_medications*100:.1f}%)")
        
        quality_score = (brand_names_found + dosage_forms_found + how_to_take_found + when_to_take_found + disease_tags_found) / (total_medications * 5) * 100
        print(f"\n🟢 Overall Data Quality: {quality_score:.1f}%")
        print(f"{'═'*70}")
        
        if multiple_dosage_forms > 0:
            self.print_section("EXAMPLES OF MULTIPLE DOSAGE FORMS")
            count = 0
            for _, row in df.iterrows():
                if '|' in str(row['Dosage Forms']) and count < 10:
                    self.print_info(f"{row['Medication Name']}: {row['Dosage Forms']}")
                    count += 1
        
        if how_to_take_enhanced > 0:
            self.print_section("EXAMPLES OF ENHANCED HOW TO TAKE")
            count = 0
            for _, row in df.iterrows():
                if '|' in str(row['How to Take']) and count < 10:
                    self.print_info(f"{row['Medication Name']}: {row['How to Take']}")
                    count += 1
    
    def find_all_how_to_take_in_text(self, text):
        """Find ALL possible how to take instructions in text"""
        all_instructions = []
        
        # Safety check for text
        if not text or text is None:
            return all_instructions
        
        # Enhanced patterns for comprehensive instruction extraction
        enhanced_patterns = [
            # Food-related instructions
            r'(?:take|use|administer)\s+(?:with|without)\s+(?:food|meals)',
            r'(?:take|use|administer)\s+(?:on\s+)?(?:empty|full)\s+(?:stomach)',
            r'(?:with|without)\s+(?:food|meals)',
            r'(?:on\s+)?(?:empty|full)\s+(?:stomach)',
            
            # Water and liquid instructions
            r'(?:take|use)\s+(?:with|without)\s+(?:water|liquid)',
            r'(?:with|without)\s+(?:water|liquid)',
            r'(?:with\s+)?(?:a\s+)?(?:full\s+)?(?:glass\s+of\s+water)',
            
            # Swallowing instructions
            r'swallow\s+(?:the\s+)?(?:tablet|capsule|pill)\s+(?:whole|with\s+water|with\s+food)',
            r'swallow\s+(?:whole|with\s+water|with\s+food)',
            r'(?:chew|crush|break)\s+(?:the\s+)?(?:tablet)',
            
            # Administration route instructions
            r'(?:take|use)\s+(?:orally|by\s+mouth|sublingually|buccally)',
            r'(?:oral|injection|inhalation|topical|sublingual|buccal)\s+(?:administration|use)',
            
            # Specific instruction patterns
            r'how\s+to\s+(?:take|use):\s*([^,\n\r]+)',
            r'instructions:\s*([^,\n\r]+)',
            r'directions:\s*([^,\n\r]+)',
            r'administration:\s*([^,\n\r]+)',
            
            # Dosage form specific instructions
            r'(?:tablet|capsule|pill)\s+(?:should\s+be\s+)?(?:swallowed|chewed|crushed)',
            r'(?:liquid|suspension|syrup)\s+(?:should\s+be\s+)?(?:shaken|measured)',
            r'(?:inhaler|aerosol)\s+(?:should\s+be\s+)?(?:primed|shaken)',
            r'(?:cream|ointment|gel)\s+(?:should\s+be\s+)?(?:applied|rubbed)',
            
            # Timing instructions
            r'(?:take|use)\s+(?:at\s+)?(?:the\s+same\s+time|regular\s+intervals)',
            r'(?:take|use)\s+(?:before|after|during)\s+(?:meals|food)',
            r'(?:take|use)\s+(?:in\s+the\s+)?(?:morning|evening|bedtime)'
        ]
        
        for pattern in enhanced_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            if matches:
                for match in matches:
                    if isinstance(match, tuple):
                        match = match[0]
                    cleaned = self.clean_text(match)
                    if cleaned and len(cleaned) > 5:
                        standardized = self.simplify_how_to_take(cleaned)
                        if standardized not in all_instructions:
                            all_instructions.append(standardized)
        
        # Look for specific instruction sections
        instruction_sections = [
            r'Available as:\s*([^,\n\r]+)',
            r'Form[s]?:\s*([^,\n\r]+)',
            r'Administration:\s*([^,\n\r]+)',
            r'Instructions:\s*([^,\n\r]+)'
        ]
        
        for pattern in instruction_sections:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                instruction_text = self.clean_text(match.group(1))
                if instruction_text and len(instruction_text) > 5:
                    standardized = self.simplify_how_to_take(instruction_text)
                    if standardized not in all_instructions:
                        all_instructions.append(standardized)
        
        return all_instructions
    
    def reprocess_brand_names(self):
        """Reprocess existing data to extract multiple brand names"""
        self.print_header("🔄 BRAND NAME REPROCESSING", "Extract Multiple Brand Names from Existing Data")
        
        try:
            # Load the most recent medication data file
            files = [f for f in os.listdir('../Analysis') if f.startswith('medication_data_') and f.endswith('.xlsx')]
            if not files:
                self.print_error("No medication data files found!")
                return
            
            files.sort(key=lambda x: os.path.getmtime(os.path.join('../Analysis', x)), reverse=True)
            latest_file = files[0]
            
            self.print_success(f"Processing: {latest_file}")
            df = pd.read_excel(f'../Analysis/{latest_file}')
            
            # Filter to only medication rows (skip headers and statistics)
            medication_df = df[
                (df['Medication Name'].notna()) & 
                (~df['Medication Name'].str.contains('Total Medications', na=False)) &
                (~df['Medication Name'].str.contains('Multiple Brands', na=False)) &
                (~df['Medication Name'].str.contains('Multiple Dosage Forms', na=False)) &
                (~df['Medication Name'].str.contains('When to Take Found', na=False)) &
                (~df['Medication Name'].str.contains('Success Rate', na=False)) &
                (~df['Medication Name'].str.contains('📋 MEDICATION DETAILS', na=False)) &
                (~df['Medication Name'].str.contains('Medication Name', na=False))
            ].copy()
            
            self.print_info(f"Found {len(medication_df)} medications to reprocess")
            
            # Setup driver for web scraping
            self.setup_driver()
            
            improved_count = 0
            total_processed = 0
            
            for index, row in medication_df.iterrows():
                medication_name = row['Medication Name']
                current_brand = row['Brand Names']
                
                # Skip if already has multiple brands
                if '|' in str(current_brand):
                    continue
                
                self.print_info(f"Reprocessing: {medication_name}")
                
                try:
                    # Get the medication page
                    self.driver.get("https://www.drugs.com")
                    time.sleep(1)
                    
                    # Search for the medication
                    search_box = WebDriverWait(self.driver, 5).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "input[name='searchterm']"))
                    )
                    search_box.clear()
                    search_box.send_keys(medication_name)
                    search_box.send_keys(Keys.RETURN)
                    time.sleep(2)
                    
                    # Find and click the medication link
                    medication_link = self.find_medication_link(medication_name)
                    if medication_link:
                        self.driver.execute_script("arguments[0].click();", medication_link)
                        time.sleep(3)
                        
                        # Extract improved brand names
                        improved_brand = self.extract_brand_name(self.driver.page_source, medication_name)
                        
                        if improved_brand != current_brand and improved_brand != "Not found":
                            self.print_success(f"{medication_name}: '{current_brand}' → '{improved_brand}'")
                            medication_df.at[index, 'Brand Names'] = improved_brand
                            improved_count += 1
                        else:
                            self.print_info(f"{medication_name}: No improvement found")
                    
                    total_processed += 1
                    
                    # Restart driver every 10 medications
                    if total_processed % 10 == 0:
                        self.print_info("Restarting driver...")
                        self.restart_driver()
                    
                except Exception as e:
                    self.print_error(f"Error processing {medication_name}: {e}")
                    continue
            
            if self.driver:
                self.driver.quit()
            
            # Save the improved data
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            improved_filename = f"../Analysis/medication_data_reprocessed_{timestamp}.xlsx"
            
            # Apply professional formatting
            self.create_clean_professional_excel(medication_df, improved_filename)
            
            self.print_success(f"Reprocessed data saved: {improved_filename}")
            self.print_success(f"Improved {improved_count} out of {total_processed} medications")
            
            # Show statistics
            multiple_brands = len(medication_df[medication_df['Brand Names'].str.contains('|', na=False)])
            self.print_info(f"Total medications with multiple brand names: {multiple_brands}")
            
        except Exception as e:
            self.print_error(f"Error in reprocessing: {e}")
            if self.driver:
                self.driver.quit()
    
    def find_food_instructions(self, text):
        """Find food-related administration instructions"""
        food_patterns = [
            r'(?:take|use|administer)\s+(?:with|without)\s+(?:food|meals)',
            r'(?:take|use|administer)\s+(?:on\s+)?(?:empty|full)\s+(?:stomach)',
            r'(?:with|without)\s+(?:food|meals)',
            r'(?:on\s+)?(?:empty|full)\s+(?:stomach)',
            r'(?:take|use)\s+(?:before|after|during)\s+(?:meals|food)'
        ]
        
        food_instructions = []
        for pattern in food_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            for match in matches:
                if isinstance(match, tuple):
                    match = match[0]
                cleaned = self.clean_text(match)
                if cleaned and len(cleaned) > 5:
                    standardized = self.simplify_how_to_take(cleaned)
                    if standardized not in food_instructions:
                        food_instructions.append(standardized)
        
        return food_instructions

def main():
    if len(sys.argv) > 1:
        command = sys.argv[1].lower()
        
        if command == "scrape":
            scraper = MedicationScraper()
            scraper.run()
        elif command == "test":
            # Test with 10 medications
            scraper = MedicationScraper()
            scraper.run(limit=10)
        elif command == "reprocess":
            scraper = MedicationScraper()
            scraper.reprocess_brand_names()
        elif command == "help":
            print("Available commands:")
            print("  scrape    - Run full medication scraping process (includes automatic professional formatting)")
            print("  test      - Test with 10 medications to verify brand name extraction")
            print("  reprocess - Reprocess existing data to extract multiple brand names")
            print("  help      - Show this help message")
        else:
            print(f"Unknown command: {command}")
            print("Use 'help' to see available commands")
    else:
        scraper = MedicationScraper()
        scraper.run()

if __name__ == "__main__":
    main() 