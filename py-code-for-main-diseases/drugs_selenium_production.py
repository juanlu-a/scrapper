import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
from webdriver_manager.chrome import ChromeDriverManager
import time
import random
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

class DrugsSfxProductionScraper:
    def __init__(self):
        self.base_url = "https://www.drugs.com"
        self.sfx_url = "https://www.drugs.com/sfx/"
        self.driver = None
        self.success_count = 0
        self.error_count = 0
        
    def setup_driver(self):
        """Setup Chrome driver with robust options"""
        try:
            chrome_options = Options()
            
            # Add arguments for stability
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--disable-blink-features=AutomationControlled")
            chrome_options.add_argument("--disable-extensions")
            chrome_options.add_argument("--disable-plugins")
            chrome_options.add_argument("--disable-images")  # Faster loading
            chrome_options.add_argument("--disable-javascript")  # Faster loading, less detection
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_experimental_option('useAutomationExtension', False)
            chrome_options.add_argument("--user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
            chrome_options.add_argument("--window-size=1920,1080")
            
            # For production, run headless
            chrome_options.add_argument("--headless")
            
            # Setup driver
            service = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            
            # Set timeouts
            self.driver.set_page_load_timeout(30)
            self.driver.implicitly_wait(10)
            
            # Execute script to remove webdriver property
            self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            
            print("✓ Chrome driver initialized successfully")
            return True
            
        except Exception as e:
            print(f"✗ Error setting up Chrome driver: {str(e)}")
            return False
    
    def restart_driver(self):
        """Restart the driver if it crashes"""
        try:
            if self.driver:
                self.driver.quit()
                time.sleep(2)
            
            return self.setup_driver()
            
        except Exception as e:
            print(f"✗ Error restarting driver: {str(e)}")
            return False
    
    def search_side_effects(self, medication_name):
        """Search for side effects with robust error handling"""
        max_retries = 3
        
        for attempt in range(max_retries):
            try:
                # Check if driver is still working
                if not self.driver or not self.is_driver_alive():
                    print(f"  Driver not alive, restarting... (attempt {attempt + 1})")
                    if not self.restart_driver():
                        continue
                
                print(f"🔍 Searching side effects for: {medication_name} (attempt {attempt + 1})")
                
                # Try direct SFX URL
                clean_name = medication_name.lower().replace(' ', '-').replace('(', '').replace(')', '')
                direct_url = f"{self.sfx_url}{clean_name}.html"
                
                print(f"  Trying: {direct_url}")
                
                self.driver.get(direct_url)
                time.sleep(random.uniform(3, 6))
                
                # Check if page loaded successfully
                page_source = self.driver.page_source
                
                if ("404" not in page_source.lower() and 
                    "not found" not in page_source.lower() and
                    "403" not in page_source and
                    len(page_source) > 1000):
                    
                    print(f"  ✓ Page loaded successfully")
                    result = self.extract_side_effects_content(medication_name, direct_url)
                    
                    if result['status'] == 'Success':
                        self.success_count += 1
                        return result
                
                # If direct URL failed, try alternatives
                print(f"  Direct SFX failed, trying alternatives...")
                return self.try_alternatives(medication_name)
                
            except WebDriverException as e:
                print(f"  ✗ WebDriver error (attempt {attempt + 1}): {str(e)}")
                if attempt < max_retries - 1:
                    time.sleep(5)
                    continue
                else:
                    break
                    
            except Exception as e:
                print(f"  ✗ General error (attempt {attempt + 1}): {str(e)}")
                if attempt < max_retries - 1:
                    time.sleep(5)
                    continue
                else:
                    break
        
        # If all attempts failed
        self.error_count += 1
        return {
            'medication': medication_name,
            'status': 'Error',
            'full_information': f'Failed to access information for {medication_name} after {max_retries} attempts',
            'source_url': 'N/A'
        }
    
    def is_driver_alive(self):
        """Check if driver is still alive"""
        try:
            self.driver.current_url
            return True
        except:
            return False
    
    def try_alternatives(self, medication_name):
        """Try alternative URLs"""
        clean_name = medication_name.lower().replace(' ', '-').replace('(', '').replace(')', '')
        
        alternatives = [
            f"{self.base_url}/{clean_name}.html",
            f"{self.base_url}/mtm/{clean_name}.html",
            f"{self.base_url}/otc/{clean_name}.html"
        ]
        
        for url in alternatives:
            try:
                print(f"    Trying alternative: {url}")
                self.driver.get(url)
                time.sleep(random.uniform(2, 4))
                
                page_source = self.driver.page_source
                
                if ("404" not in page_source.lower() and 
                    "not found" not in page_source.lower() and
                    "403" not in page_source and
                    len(page_source) > 1000):
                    
                    print(f"    ✓ Alternative page loaded: {url}")
                    result = self.extract_side_effects_content(medication_name, url)
                    
                    if result['status'] == 'Success':
                        self.success_count += 1
                        return result
                        
            except Exception as e:
                print(f"    ✗ Alternative failed: {str(e)}")
                continue
        
        # If all alternatives failed
        self.error_count += 1
        return {
            'medication': medication_name,
            'status': 'Not Found',
            'full_information': f'No accessible page found for {medication_name}',
            'source_url': 'N/A'
        }
    
    def extract_side_effects_content(self, medication_name, source_url):
        """Extract side effects content from current page"""
        try:
            content_parts = []
            content_parts.append(f"=== {medication_name} Side Effects Information ===\n")
            content_parts.append(f"Source: {source_url}\n")
            
            # Get page source for parsing
            page_source = self.driver.page_source
            
            # Use BeautifulSoup for better parsing
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(page_source, 'html.parser')
            
            # Method 1: Look for side effects headings
            side_effects_found = False
            
            headings = soup.find_all(['h1', 'h2', 'h3', 'h4'], string=re.compile(r'side effects?', re.IGNORECASE))
            
            for heading in headings:
                side_effects_found = True
                content_parts.append(f"\n--- {heading.get_text(strip=True)} ---\n")
                
                # Get following content
                current = heading.next_sibling
                while current and current.name not in ['h1', 'h2', 'h3', 'h4']:
                    if hasattr(current, 'get_text'):
                        if current.name == 'p':
                            text = current.get_text(strip=True)
                            if text and len(text) > 15:
                                content_parts.append(text)
                        elif current.name in ['ul', 'ol']:
                            for li in current.find_all('li'):
                                li_text = li.get_text(strip=True)
                                if li_text:
                                    content_parts.append(f"• {li_text}")
                        elif current.name == 'div' and current.get_text(strip=True):
                            div_text = current.get_text(strip=True)
                            if len(div_text) > 15:
                                content_parts.append(div_text)
                    
                    current = current.next_sibling
                
                content_parts.append("")  # Add spacing
            
            # Method 2: Look for paragraphs with side effects keywords
            if not side_effects_found:
                paragraphs = soup.find_all('p')
                for para in paragraphs:
                    text = para.get_text(strip=True)
                    if (text and len(text) > 20 and 
                        ('side effect' in text.lower() or 
                         'adverse' in text.lower() or 
                         'reaction' in text.lower())):
                        content_parts.append(text)
                        content_parts.append("")
            
            # Method 3: Extract main content if still limited
            if len(content_parts) < 5:
                main_content = soup.find('main') or soup.find('div', class_='main-content') or soup.find('article')
                if main_content:
                    text_content = main_content.get_text()
                    if text_content and len(text_content) > 200:
                        # Split into manageable chunks
                        chunks = text_content.split('\n\n')
                        for chunk in chunks:
                            chunk = chunk.strip()
                            if chunk and len(chunk) > 30:
                                content_parts.append(chunk)
                                content_parts.append("")
            
            # Return result
            if len(content_parts) > 3:
                result_text = '\n'.join(content_parts)
                print(f"  ✓ Extracted {len(result_text)} characters")
                
                return {
                    'medication': medication_name,
                    'status': 'Success',
                    'full_information': result_text,
                    'source_url': source_url
                }
            else:
                return {
                    'medication': medication_name,
                    'status': 'Limited Info',
                    'full_information': f'Limited information found for {medication_name}.',
                    'source_url': source_url
                }
                
        except Exception as e:
            print(f"  ✗ Error extracting content: {str(e)}")
            return {
                'medication': medication_name,
                'status': 'Error',
                'full_information': f'Error extracting content for {medication_name}: {str(e)}',
                'source_url': source_url
            }
    
    def close(self):
        """Close the browser"""
        try:
            if self.driver:
                self.driver.quit()
                print("✓ Browser closed")
        except:
            pass

def update_excel_with_selenium_scraper(max_medications=10):
    """Update Excel file using Selenium scraper"""
    
    # Load the existing Excel file
    excel_path = '/Users/juanlu/Documents/Wye/scrapper/Analysis/main_diseases_analysis_final.xlsx'
    
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found at {excel_path}")
        return
    
    # Load the workbook
    wb = load_workbook(excel_path)
    
    # Get the unique medications sheet
    if "All Unique Medications" not in wb.sheetnames:
        print("Error: 'All Unique Medications' sheet not found")
        return
    
    medications_ws = wb["All Unique Medications"]
    
    # Read medications from the sheet
    medications = []
    for row in medications_ws.iter_rows(min_row=9, max_col=1, values_only=True):
        if row[0] and row[0].strip():
            medications.append(row[0].strip())
    
    # Limit for testing
    medications = medications[:max_medications]
    
    print(f"Processing {len(medications)} medications...")
    
    # Initialize scraper
    scraper = DrugsSfxProductionScraper()
    
    if not scraper.setup_driver():
        print("Failed to initialize scraper")
        return
    
    # Add new column header for "Full Information" if not exists
    if not medications_ws['I8'].value:
        medications_ws['I8'] = 'SELENIUM SCRAPED DATA'
        # Style the header
        medications_ws['I8'].font = Font(bold=True, color="FFFFFF")
        medications_ws['I8'].fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
        medications_ws['I8'].border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        medications_ws['I8'].alignment = Alignment(horizontal='center', vertical='center')
        medications_ws.column_dimensions['I'].width = 80
    
    try:
        # Process each medication
        for i, medication in enumerate(medications):
            print(f"\n[{i+1}/{len(medications)}] Processing: {medication}")
            
            # Scrape side effects
            result = scraper.search_side_effects(medication)
            
            # Add to Excel
            row_num = 9 + i
            medications_ws[f'I{row_num}'] = result['full_information']
            
            # Add border and formatting
            cell = medications_ws[f'I{row_num}']
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Color based on status
            if result['status'] == 'Success':
                cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            elif result['status'] == 'Error':
                cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            
            # Save progress every 5 medications
            if (i + 1) % 5 == 0:
                wb.save(excel_path)
                print(f"  ✓ Saved progress: {i+1} medications processed")
                print(f"  Success rate: {scraper.success_count}/{i+1} ({scraper.success_count/(i+1)*100:.1f}%)")
            
            # Add delay between requests
            time.sleep(random.uniform(4, 8))
    
    finally:
        scraper.close()
        
        # Final save
        wb.save(excel_path)
        print(f"\n✅ Completed! Processed {len(medications)} medications")
        print(f"Success rate: {scraper.success_count}/{len(medications)} ({scraper.success_count/len(medications)*100:.1f}%)")
        print(f"Updated Excel file: {excel_path}")

if __name__ == "__main__":
    print("Drugs.com Selenium Side Effects Scraper - Production Version")
    print("=" * 60)
    
    # Process first 10 medications as test
    update_excel_with_selenium_scraper(max_medications=10)
