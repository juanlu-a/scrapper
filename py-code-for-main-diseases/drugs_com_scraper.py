import pandas as pd
import requests
from bs4 import BeautifulSoup
import time
import random
from urllib.parse import urljoin, quote
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

class DrugsComScraper:
    def __init__(self):
        self.base_url = "https://www.drugs.com"
        self.session = requests.Session()
        # Add headers to mimic a real browser
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.5',
            'Accept-Encoding': 'gzip, deflate',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
        })
        
    def search_medication(self, medication_name):
        """Search for a medication on Drugs.com and return the URL of the medication page"""
        try:
            # Clean the medication name for search
            clean_name = medication_name.strip().lower()
            
            # Try direct URL first (most common pattern)
            direct_url = f"{self.base_url}/{clean_name}.html"
            
            # Test if direct URL works
            response = self.session.get(direct_url, timeout=10)
            if response.status_code == 200 and "side-effects" in response.text.lower():
                print(f"✓ Found direct URL for {medication_name}: {direct_url}")
                return direct_url
            
            # If direct URL doesn't work, use search
            search_url = f"{self.base_url}/search.php"
            params = {
                'searchterm': medication_name
            }
            
            response = self.session.get(search_url, params=params, timeout=10)
            if response.status_code != 200:
                print(f"✗ Search failed for {medication_name}: HTTP {response.status_code}")
                return None
            
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Look for the first drug result link
            # Drugs.com search results typically have links in specific patterns
            drug_links = soup.find_all('a', href=True)
            
            for link in drug_links:
                href = link.get('href')
                if href and (href.startswith('/') or href.startswith('http')):
                    # Convert relative URLs to absolute
                    if href.startswith('/'):
                        full_url = urljoin(self.base_url, href)
                    else:
                        full_url = href
                    
                    # Check if this looks like a drug page
                    if (medication_name.lower() in href.lower() or 
                        medication_name.lower() in link.get_text().lower()):
                        if '.html' in href and 'search' not in href:
                            print(f"✓ Found search result for {medication_name}: {full_url}")
                            return full_url
            
            print(f"✗ No drug page found for {medication_name}")
            return None
            
        except Exception as e:
            print(f"✗ Error searching for {medication_name}: {str(e)}")
            return None
    
    def get_side_effects_url(self, drug_url):
        """Get the side effects URL for a drug page"""
        try:
            if not drug_url:
                return None
                
            # Most drugs.com side effects pages follow the pattern: drugname.html#side-effects
            # or drugname-side-effects.html
            
            base_url = drug_url.replace('.html', '')
            
            # Try the anchor link first
            side_effects_url = f"{base_url}.html#side-effects"
            
            # Also try the separate page pattern
            side_effects_page = f"{base_url}-side-effects.html"
            
            # Test which one works
            response = self.session.get(side_effects_url, timeout=10)
            if response.status_code == 200:
                return side_effects_url
            
            response = self.session.get(side_effects_page, timeout=10)
            if response.status_code == 200:
                return side_effects_page
                
            # If neither works, return the main drug page
            return drug_url
            
        except Exception as e:
            print(f"✗ Error getting side effects URL: {str(e)}")
            return drug_url
    
    def scrape_side_effects(self, medication_name):
        """Scrape side effects information for a medication"""
        try:
            print(f"🔍 Searching for {medication_name}...")
            
            # Step 1: Find the medication page
            drug_url = self.search_medication(medication_name)
            if not drug_url:
                return {
                    'medication': medication_name,
                    'status': 'Not Found',
                    'full_information': f'Medication "{medication_name}" not found on Drugs.com'
                }
            
            # Step 2: Get the side effects page
            side_effects_url = self.get_side_effects_url(drug_url)
            
            # Step 3: Scrape the side effects content
            response = self.session.get(side_effects_url, timeout=10)
            if response.status_code != 200:
                return {
                    'medication': medication_name,
                    'status': 'Error',
                    'full_information': f'Failed to access side effects page for {medication_name}'
                }
            
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Extract side effects content
            side_effects_content = self.extract_side_effects_content(soup, medication_name)
            
            print(f"✓ Successfully scraped {medication_name}")
            
            return {
                'medication': medication_name,
                'status': 'Success',
                'full_information': side_effects_content,
                'source_url': side_effects_url
            }
            
        except Exception as e:
            print(f"✗ Error scraping {medication_name}: {str(e)}")
            return {
                'medication': medication_name,
                'status': 'Error',
                'full_information': f'Error scraping {medication_name}: {str(e)}'
            }
    
    def extract_side_effects_content(self, soup, medication_name):
        """Extract all side effects content from the page"""
        content_parts = []
        
        try:
            # Look for side effects section
            side_effects_section = None
            
            # Try different selectors to find side effects content
            selectors = [
                'div[id*="side-effects"]',
                'section[id*="side-effects"]',
                'div.side-effects',
                'section.side-effects',
                'div[class*="side-effects"]',
                'div[class*="sideeffects"]'
            ]
            
            for selector in selectors:
                side_effects_section = soup.select_one(selector)
                if side_effects_section:
                    break
            
            # If no specific section found, look for headings containing "side effects"
            if not side_effects_section:
                headings = soup.find_all(['h1', 'h2', 'h3', 'h4'], string=re.compile(r'side effects', re.IGNORECASE))
                if headings:
                    side_effects_section = headings[0].parent
            
            # Extract content
            if side_effects_section:
                content_parts.append(f"=== {medication_name} Side Effects Information ===\n")
                
                # Get all text content, preserving structure
                for element in side_effects_section.find_all(['p', 'ul', 'ol', 'li', 'div', 'h1', 'h2', 'h3', 'h4']):
                    text = element.get_text(strip=True)
                    if text and len(text) > 10:  # Skip very short text
                        # Add structure indicators
                        if element.name in ['h1', 'h2', 'h3', 'h4']:
                            content_parts.append(f"\n--- {text} ---\n")
                        elif element.name == 'li':
                            content_parts.append(f"• {text}")
                        else:
                            content_parts.append(text)
                        content_parts.append("")  # Add line break
            
            # If still no content, try to find any text about side effects
            if not content_parts:
                # Look for any text containing "side effects" 
                all_text = soup.get_text()
                if "side effects" in all_text.lower():
                    # Extract paragraphs containing side effects information
                    paragraphs = soup.find_all('p')
                    for p in paragraphs:
                        text = p.get_text(strip=True)
                        if 'side effect' in text.lower() and len(text) > 20:
                            content_parts.append(text)
                            content_parts.append("")
            
            # Join all content
            if content_parts:
                return '\n'.join(content_parts)
            else:
                return f"No side effects information found for {medication_name} on the page"
                
        except Exception as e:
            return f"Error extracting side effects content for {medication_name}: {str(e)}"
    
    def add_delay(self):
        """Add a random delay to avoid being blocked"""
        delay = random.uniform(1, 3)  # Random delay between 1-3 seconds
        time.sleep(delay)

def update_excel_with_side_effects():
    """Update the Excel file with side effects information"""
    
    # Load the existing Excel file
    excel_path = '/Users/juanlu/Documents/Wye/scrapper/Analysis/main_diseases_analysis_final.xlsx'
    
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found at {excel_path}")
        return
    
    # Load the workbook
    wb = load_workbook(excel_path)
    
    # Get the unique medications sheet
    if "All Unique Medications" not in wb.sheetnames:
        print("Error: 'All Unique Medications' sheet not found")
        return
    
    medications_ws = wb["All Unique Medications"]
    
    # Read medications from the sheet
    medications = []
    for row in medications_ws.iter_rows(min_row=9, max_col=1, values_only=True):  # Start from row 9 (after headers)
        if row[0] and row[0].strip():  # Skip empty rows
            medications.append(row[0].strip())
    
    print(f"Found {len(medications)} medications to process")
    
    # Initialize the scraper
    scraper = DrugsComScraper()
    
    # Add new column header for "Full Information"
    medications_ws['F8'] = 'FULL INFORMATION'
    # Style the header
    medications_ws['F8'].font = Font(bold=True, color="FFFFFF")
    medications_ws['F8'].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    medications_ws['F8'].border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    medications_ws['F8'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Set column width
    medications_ws.column_dimensions['F'].width = 60
    
    # Process each medication
    processed_count = 0
    
    for i, medication in enumerate(medications):
        if not medication:
            continue
            
        print(f"\n[{i+1}/{len(medications)}] Processing: {medication}")
        
        # Scrape side effects
        result = scraper.scrape_side_effects(medication)
        
        # Add to Excel
        row_num = 9 + i  # Start from row 9
        medications_ws[f'F{row_num}'] = result['full_information']
        
        # Add border and formatting
        cell = medications_ws[f'F{row_num}']
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Alternate row colors
        if i % 2 == 0:
            cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        processed_count += 1
        
        # Add delay between requests
        scraper.add_delay()
        
        # Save progress every 10 medications
        if processed_count % 10 == 0:
            wb.save(excel_path)
            print(f"✓ Saved progress: {processed_count} medications processed")
    
    # Final save
    wb.save(excel_path)
    print(f"\n✅ Completed! Processed {processed_count} medications")
    print(f"Updated Excel file: {excel_path}")

if __name__ == "__main__":
    print("Starting Drugs.com Side Effects Scraper...")
    print("This will update the Excel file with side effects information")
    print("=" * 60)
    
    update_excel_with_side_effects()
