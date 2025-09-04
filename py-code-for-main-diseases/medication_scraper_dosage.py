#!/usr/bin/env python3
"""
WebMD Dosage Information Scraper
Scrapes dosage information for medications from https://www.webmd.com/drugs/2/index
and adds the information to the existing medication Excel file.

Author: Assistant
Date: September 2025
"""

import pandas as pd
import os
import time
import random
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException, WebDriverException, NoSuchElementException
import re
from bs4 import BeautifulSoup
import json
import sys
from tqdm import tqdm
import colorama
from colorama import Fore, Back, Style
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import google.generativeai as genai
from dotenv import load_dotenv
import glob

# Initialize colorama
colorama.init(autoreset=True)

# Load environment variables
load_dotenv('/Users/juanlu/Documents/Wye/scrapper/.env')

class WebMDDosageScraper:
    def __init__(self, headless=False):
        self.headless = headless
        self.driver = None
        self.wait = None
        self.base_url = "https://www.webmd.com/drugs/2/index"
        
        # Configure Google Generative AI
        api_key = os.getenv('GOOGLE_GEMINI_API_KEY')
        if not api_key:
            raise ValueError("GOOGLE_GEMINI_API_KEY not found in environment variables. Please check your .env file.")
        
        genai.configure(api_key=api_key)
        self.model = genai.GenerativeModel("gemini-1.5-flash")
        
        # Cache for processed medications
        self.cache_file = "dosage_cache.json"
        self.cache = self.load_cache()
        
        # Results storage
        self.results = {}
        
        self.init_driver()
        
    def print_header(self, title, subtitle=""):
        """Print a styled header"""
        print(f"\n{Fore.GREEN}{'═'*80}")
        print(f"{Fore.WHITE}{Style.BRIGHT}{title:^80}")
        if subtitle:
            print(f"{Fore.CYAN}{subtitle:^80}")
        print(f"{Fore.GREEN}{'═'*80}{Style.RESET_ALL}")
    
    def print_section(self, title):
        """Print a section header"""
        print(f"\n{Fore.BLUE}{Style.BRIGHT}▶ {title}")
        print(f"{Fore.BLUE}{'─' * (len(title) + 2)}{Style.RESET_ALL}")
    
    def print_success(self, message):
        """Print a success message"""
        print(f"{Fore.GREEN}✅ {message}{Style.RESET_ALL}")
    
    def print_error(self, message):
        """Print an error message"""
        print(f"{Fore.RED}❌ {message}{Style.RESET_ALL}")
    
    def print_warning(self, message):
        """Print a warning message"""
        print(f"{Fore.YELLOW}⚠️ {message}{Style.RESET_ALL}")
    
    def print_info(self, message):
        """Print an info message"""
        print(f"{Fore.CYAN}ℹ️ {message}{Style.RESET_ALL}")
    
    def load_cache(self):
        """Load existing cache if available"""
        try:
            if os.path.exists(self.cache_file):
                with open(self.cache_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception as e:
            self.print_warning(f"Could not load cache: {e}")
        return {}
    
    def save_cache(self):
        """Save cache to file"""
        try:
            with open(self.cache_file, 'w', encoding='utf-8') as f:
                json.dump(self.cache, f, indent=2, ensure_ascii=False)
        except Exception as e:
            self.print_error(f"Could not save cache: {e}")
    
    def init_driver(self):
        """Initialize the Chrome driver"""
        if self.driver:
            try:
                self.driver.quit()
            except:
                pass
        
        self.driver = self.setup_driver(self.headless)
        self.wait = WebDriverWait(self.driver, 15)
    
    def setup_driver(self, headless=False):
        """Set up Chrome driver with options"""
        chrome_options = Options()
        if headless:
            chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_argument("--disable-web-security")
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        
        # Add user agent to avoid detection
        chrome_options.add_argument("--user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
        
        try:
            driver = webdriver.Chrome(options=chrome_options)
            driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            return driver
        except Exception as e:
            self.print_error(f"Error setting up Chrome driver: {e}")
            raise e
    
    def navigate_to_webmd(self):
        """Navigate to WebMD drugs page"""
        try:
            self.print_info(f"Navigating to {self.base_url}")
            self.driver.get(self.base_url)
            time.sleep(3)
            
            # Wait for page to load
            self.wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
            self.print_success("Successfully loaded WebMD drugs page")
            return True
            
        except Exception as e:
            self.print_error(f"Failed to navigate to WebMD: {e}")
            return False
    
    def search_medication(self, medication_name):
        """Search for a medication on WebMD"""
        try:
            # Clean medication name for search
            clean_name = self.clean_medication_name(medication_name)
            self.print_info(f"Searching for: {clean_name}")
            
            # Navigate to search page first
            if not self.navigate_to_webmd():
                return None
            
            # Find search box - try multiple selectors
            search_selectors = [
                'input[placeholder="Enter medication name to search"]',
                'input[name="query"]',
                'input[type="search"]',
                '#search-query',
                '.search-input',
                'input[placeholder*="search"]',
                'input[aria-label*="search"]',
                '#drug-search',
                '.webmd-input__inner'
            ]
            
            search_box = None
            for selector in search_selectors:
                try:
                    search_box = self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, selector)))
                    break
                except TimeoutException:
                    continue
            
            if not search_box:
                # Try alternative approach - direct search URL
                search_url = f"https://www.webmd.com/drugs/2/search?type=drugs&query={clean_name}"
                self.driver.get(search_url)
                time.sleep(3)
                return self.handle_search_results(clean_name)
            
            # Clear and enter search term
            search_box.clear()
            search_box.send_keys(clean_name)
            search_box.send_keys(Keys.RETURN)
            
            time.sleep(3)
            
            # Look for direct drug page or search results
            return self.handle_search_results(clean_name)
            
        except Exception as e:
            self.print_error(f"Error searching for {medication_name}: {e}")
            return None
    
    def clean_medication_name(self, name):
        """Clean medication name for search"""
        # Remove common suffixes and prefixes
        clean_name = re.sub(r'\s+(tablet|capsule|injection|cream|ointment|gel|liquid|suspension|er|xr|xl)s?$', '', name.lower())
        clean_name = re.sub(r'\s+\d+\s*(mg|mcg|g|ml|iu).*$', '', clean_name)
        clean_name = re.sub(r'\s+oral.*$', '', clean_name)
        clean_name = clean_name.strip()
        return clean_name
    
    def handle_search_results(self, medication_name):
        """Handle search results and find the most relevant drug page"""
        try:
            # Check if we're already on a drug information page
            if "drug" in self.driver.current_url.lower() and "drug-" in self.driver.current_url.lower():
                return self.extract_dosage_info_from_page()
            
            # Look for search results links
            result_selectors = [
                'a[href*="/drugs/2/drug-"]',
                'a[href*="/drugs/"]',
                '.search-result a',
                '.result-item a',
                'a[title*="{}"]'.format(medication_name),
                '.drug-search-result a'
            ]
            
            for selector in result_selectors:
                try:
                    results = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    if results:
                        # Click on the first relevant result
                        results[0].click()
                        time.sleep(3)
                        return self.extract_dosage_info_from_page()
                except:
                    continue
            
            # If no direct results, try generic search
            return self.try_generic_drug_search(medication_name)
            
        except Exception as e:
            self.print_error(f"Error handling search results: {e}")
            return None
    
    def try_generic_drug_search(self, medication_name):
        """Try searching for generic version of the drug"""
        try:
            # Common generic mappings
            generic_mappings = {
                'advil': 'ibuprofen',
                'tylenol': 'acetaminophen',
                'motrin': 'ibuprofen',
                'aleve': 'naproxen',
                'prozac': 'fluoxetine',
                'zoloft': 'sertraline',
                'lipitor': 'atorvastatin',
                'nexium': 'esomeprazole'
            }
            
            generic_name = generic_mappings.get(medication_name.lower(), medication_name)
            
            if generic_name != medication_name:
                self.print_info(f"Trying generic name: {generic_name}")
                return self.search_medication(generic_name)
            
            return None
            
        except Exception as e:
            self.print_error(f"Error in generic search: {e}")
            return None
    
    def extract_dosage_info_from_page(self):
        """Extract dosage information from the current page"""
        try:
            # Get page source for parsing
            page_source = self.driver.page_source
            soup = BeautifulSoup(page_source, 'html.parser')
            
            # Extract text content
            text_content = soup.get_text()
            
            # Use LLM to extract dosage information
            return self.extract_dosage_with_llm(text_content)
            
        except Exception as e:
            self.print_error(f"Error extracting dosage info: {e}")
            return None
    
    def extract_dosage_with_llm(self, page_content):
        """Use LLM to extract and summarize dosage information from page content"""
        try:
            prompt = f"""
            Analyze the following medical webpage content and extract ONLY the most essential dosage information for the medication.
            
            Provide a BRIEF summary (maximum 2-3 lines) with only the key dosage information:
            - Standard adult dose (mg, frequency)
            - Maximum daily dose if mentioned
            - Key administration notes (with/without food, timing)
            
            Format example: "Adults: 500-1000mg every 4-6 hours. Maximum: 4000mg/day. Take with food."
            
            If no clear dosage information is found, respond with "No dosage information found."
            
            Keep response under 150 characters and focus only on practical dosing information.
            
            Page content:
            {page_content[:8000]}  # Limit content to avoid token limits
            """
            
            response = self.model.generate_content(prompt)
            
            if response and response.text:
                # Clean up the response and ensure it's brief
                result = response.text.strip()
                # If response is too long, truncate it
                if len(result) > 200:
                    result = result[:197] + "..."
                return result
            else:
                return "No dosage information could be extracted."
                
        except Exception as e:
            self.print_error(f"Error using LLM to extract dosage: {e}")
            return "Error extracting dosage information."
    
    def process_medication(self, medication_name):
        """Process a single medication and get its dosage information"""
        try:
            # Check cache first
            if medication_name in self.cache:
                self.print_info(f"Found {medication_name} in cache")
                return self.cache[medication_name]
            
            self.print_section(f"Processing: {medication_name}")
            
            # Search and get dosage information
            dosage_info = self.search_medication(medication_name)
            
            if dosage_info:
                self.print_success(f"Successfully extracted dosage for {medication_name}")
                # Cache the result
                self.cache[medication_name] = dosage_info
                self.save_cache()
            else:
                self.print_warning(f"No dosage information found for {medication_name}")
                dosage_info = "No dosage information found."
                self.cache[medication_name] = dosage_info
                self.save_cache()
            
            # Random delay to avoid rate limiting
            time.sleep(random.uniform(2, 5))
            
            return dosage_info
            
        except Exception as e:
            self.print_error(f"Error processing {medication_name}: {e}")
            return "Error retrieving dosage information."
    
    def load_medication_data(self, excel_file_path):
        """Load medication data from Excel file"""
        try:
            self.print_section("Loading medication data from Excel")
            
            # Read the Excel file using openpyxl to preserve structure
            import openpyxl
            wb = openpyxl.load_workbook(excel_file_path)
            ws = wb.active
            
            # Find the medication data header row
            header_row = None
            for row in range(1, 20):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value == "Medication Name":
                    header_row = row
                    break
            
            if header_row is None:
                raise ValueError("Could not find 'Medication Name' header in Excel file")
            
            # Extract headers
            headers = []
            for col in range(1, ws.max_column + 1):
                header_value = ws.cell(row=header_row, column=col).value
                if header_value:
                    headers.append(header_value)
                else:
                    headers.append(f"Unnamed: {col-1}")
            
            # Extract medication data
            medications_data = []
            for row in range(header_row + 1, ws.max_row + 1):
                row_data = {}
                medication_name = ws.cell(row=row, column=1).value
                
                if not medication_name or pd.isna(medication_name):
                    continue
                    
                medication_name = str(medication_name).strip()
                
                # Skip summary/header rows
                if medication_name.startswith('📊') or medication_name.startswith('📋') or medication_name.startswith('📈'):
                    continue
                
                # Collect all column data for this row
                for col in range(1, len(headers) + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    row_data[headers[col-1]] = cell_value
                
                medications_data.append(row_data)
            
            medications_df = pd.DataFrame(medications_data)
            
            self.print_success(f"Loaded {len(medications_df)} medications from Excel file")
            return medications_df
            
        except Exception as e:
            self.print_error(f"Error loading medication data: {e}")
            raise e
    
    def process_all_medications(self, excel_file_path, output_file_path=None):
        """Process all medications and add dosage column"""
        try:
            self.print_header("WEBMD DOSAGE SCRAPER", "Processing medication dosage information")
            
            # Load medication data
            medications_df = self.load_medication_data(excel_file_path)
            
            # Add dosage column if it doesn't exist
            if 'Dosage' not in medications_df.columns:
                medications_df['Dosage'] = ''
            
            total_medications = len(medications_df)
            self.print_info(f"Processing {total_medications} medications...")
            
            # Process each medication
            for idx, row in tqdm(medications_df.iterrows(), total=total_medications, desc="Processing medications"):
                medication_name = str(row['Medication Name']).strip()
                
                if pd.isna(medication_name) or medication_name.lower() in ['nan', '']:
                    continue
                
                # Skip if already processed and has dosage data
                if pd.notna(row.get('Dosage', '')) and str(row.get('Dosage', '')).strip():
                    self.print_info(f"Skipping {medication_name} - already has dosage data")
                    continue
                
                try:
                    dosage_info = self.process_medication(medication_name)
                    medications_df.at[idx, 'Dosage'] = dosage_info
                    
                    # Save progress every 10 medications
                    if idx % 10 == 0:
                        self.save_progress(medications_df, excel_file_path, output_file_path)
                        
                except Exception as e:
                    self.print_error(f"Error processing {medication_name}: {e}")
                    medications_df.at[idx, 'Dosage'] = "Error retrieving dosage information."
            
            # Final save
            self.save_final_results(medications_df, excel_file_path, output_file_path)
            
            self.print_success("All medications processed successfully!")
            return medications_df
            
        except Exception as e:
            self.print_error(f"Error in process_all_medications: {e}")
            raise e
    
    def save_progress(self, medications_df, original_file_path, output_file_path):
        """Save progress to avoid losing work"""
        try:
            # Create backup with proper structure
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_file = f"medication_dosage_progress_{timestamp}.xlsx"
            
            # Save with new dosage column (just the medications data for progress)
            medications_df.to_excel(backup_file, index=False)
            self.print_info(f"Progress saved to {backup_file}")
            
        except Exception as e:
            self.print_warning(f"Could not save progress: {e}")
    
    def save_final_results(self, medications_df, original_file_path, output_file_path):
        """Save final results to Excel file by properly updating the original file structure and preserving formatting"""
        try:
            if output_file_path is None:
                # Create output filename based on original in the Analysis folder
                base_name = os.path.splitext(os.path.basename(original_file_path))[0]
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                analysis_folder = os.path.dirname(original_file_path)
                output_file_path = os.path.join(analysis_folder, f"{base_name}_WITH_DOSAGE_{timestamp}.xlsx")
            
            # Load original Excel file using openpyxl to preserve exact structure AND formatting
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            wb = openpyxl.load_workbook(original_file_path)
            ws = wb.active
            
            # Find the medication data header row
            header_row = None
            for row in range(1, 20):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value == "Medication Name":
                    header_row = row
                    break
            
            if header_row is None:
                self.print_error("Could not find medication header row")
                return
                
            self.print_info(f"Found medication headers at row {header_row}")
            
            # Add "Dosage" header to the next available column (column 8 if Side Effects is in column 7)
            next_column = ws.max_column + 1
            header_cell = ws.cell(row=header_row, column=next_column, value="Dosage")
            
            # Copy formatting from adjacent header cell to maintain consistency
            ref_cell = ws.cell(row=header_row, column=next_column - 1)  # Previous column
            if ref_cell.font:
                header_cell.font = Font(
                    name=ref_cell.font.name,
                    size=ref_cell.font.size,
                    bold=ref_cell.font.bold,
                    italic=ref_cell.font.italic,
                    color=ref_cell.font.color
                )
            if ref_cell.fill and hasattr(ref_cell.fill, 'start_color'):
                header_cell.fill = PatternFill(
                    start_color=ref_cell.fill.start_color,
                    end_color=ref_cell.fill.end_color,
                    fill_type=ref_cell.fill.fill_type
                )
            if ref_cell.border:
                header_cell.border = Border(
                    left=ref_cell.border.left,
                    right=ref_cell.border.right,
                    top=ref_cell.border.top,
                    bottom=ref_cell.border.bottom
                )
            if ref_cell.alignment:
                header_cell.alignment = Alignment(
                    horizontal=ref_cell.alignment.horizontal,
                    vertical=ref_cell.alignment.vertical,
                    wrap_text=ref_cell.alignment.wrap_text
                )
            
            # Create a mapping of medication names to dosage information
            dosage_mapping = {}
            for idx, row in medications_df.iterrows():
                med_name = str(row.get('Medication Name', '')).strip().lower()
                dosage_info = str(row.get('Dosage', 'No dosage information found.'))
                if med_name and med_name != 'nan':
                    dosage_mapping[med_name] = dosage_info
            
            # Update each medication row with dosage information (add to new column)
            medications_processed = 0
            dosage_added = 0
            
            for row in range(header_row + 1, ws.max_row + 1):
                medication_name = ws.cell(row=row, column=1).value
                
                if not medication_name or pd.isna(medication_name):
                    continue
                    
                medication_name = str(medication_name).strip()
                
                # Skip summary/header rows
                if medication_name.startswith('📊') or medication_name.startswith('📋') or medication_name.startswith('📈'):
                    continue
                    
                medications_processed += 1
                
                # Look up dosage information (case insensitive)
                dosage_info = dosage_mapping.get(medication_name.lower(), "No dosage information found.")
                
                # Add dosage to new column (preserving all existing columns)
                dosage_cell = ws.cell(row=row, column=next_column, value=dosage_info)
                
                # Copy formatting from adjacent data cell to maintain consistency
                reference_cell = ws.cell(row=row, column=next_column - 1)
                if reference_cell.font:
                    dosage_cell.font = Font(
                        name=reference_cell.font.name,
                        size=reference_cell.font.size,
                        bold=reference_cell.font.bold,
                        italic=reference_cell.font.italic,
                        color=reference_cell.font.color
                    )
                if reference_cell.fill and hasattr(reference_cell.fill, 'start_color'):
                    dosage_cell.fill = PatternFill(
                        start_color=reference_cell.fill.start_color,
                        end_color=reference_cell.fill.end_color,
                        fill_type=reference_cell.fill.fill_type
                    )
                if reference_cell.border:
                    dosage_cell.border = Border(
                        left=reference_cell.border.left,
                        right=reference_cell.border.right,
                        top=reference_cell.border.top,
                        bottom=reference_cell.border.bottom
                    )
                if reference_cell.alignment:
                    dosage_cell.alignment = Alignment(
                        horizontal=reference_cell.alignment.horizontal,
                        vertical=reference_cell.alignment.vertical,
                        wrap_text=True  # Enable text wrapping for long dosage information
                    )
                
                if dosage_info != "No dosage information found.":
                    dosage_added += 1
            
            # Auto-adjust column width for the new Dosage column
            try:
                # Set a reasonable width for the Dosage column (shorter since text is brief)
                column_letter = chr(64 + next_column)  # Convert column number to letter
                ws.column_dimensions[column_letter].width = 35  # Smaller width for brief dosage
            except Exception as e:
                self.print_warning(f"Could not adjust column width: {e}")
            
            # Save the updated workbook
            wb.save(output_file_path)
            
            self.print_success(f"Final results saved to: {output_file_path}")
            self.print_info(f"📊 Total medications processed: {medications_processed}")
            self.print_info(f"💊 Dosage information added: {dosage_added}")
            self.print_info(f"📈 Success rate: {dosage_added/medications_processed*100:.1f}%")
            self.print_info(f"✅ All existing columns preserved")
            self.print_info(f"✅ Dosage added in new column {next_column}")
            self.print_info(f"🎨 Original formatting preserved")
            
            # Clean up temporary files after successful completion
            self.cleanup_temporary_files()
            
            return output_file_path
            
        except Exception as e:
            self.print_error(f"Error saving final results: {e}")
            # Fallback - save just the medications data
            fallback_file = f"medication_dosage_fallback_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            medications_df.to_excel(fallback_file, index=False)
            self.print_warning(f"Saved fallback file: {fallback_file}")
    
    def cleanup_temporary_files(self):
        """Clean up temporary files after successful completion"""
        try:
            files_deleted = 0
            
            # Delete progress files
            progress_files = glob.glob("medication_dosage_progress_*.xlsx")
            for file in progress_files:
                try:
                    os.remove(file)
                    files_deleted += 1
                    self.print_info(f"🗑️ Deleted progress file: {file}")
                except Exception as e:
                    self.print_warning(f"Could not delete {file}: {e}")
            
            # Delete fallback files
            fallback_files = glob.glob("medication_dosage_fallback_*.xlsx")
            for file in fallback_files:
                try:
                    os.remove(file)
                    files_deleted += 1
                    self.print_info(f"🗑️ Deleted fallback file: {file}")
                except Exception as e:
                    self.print_warning(f"Could not delete {file}: {e}")
            
            # Delete cache file
            if os.path.exists(self.cache_file):
                try:
                    os.remove(self.cache_file)
                    files_deleted += 1
                    self.print_info(f"🗑️ Deleted cache file: {self.cache_file}")
                except Exception as e:
                    self.print_warning(f"Could not delete cache file: {e}")
            
            if files_deleted > 0:
                self.print_success(f"🧹 Cleanup completed: {files_deleted} temporary files deleted")
            else:
                self.print_info("🧹 No temporary files to clean up")
                
        except Exception as e:
            self.print_warning(f"Error during cleanup: {e}")

    def cleanup(self):
        """Clean up resources"""
        try:
            if self.driver:
                self.driver.quit()
        except:
            pass

def main():
    """Main function to run the scraper"""
    print(f"{Fore.GREEN}{Style.BRIGHT}💊 WebMD Dosage Information Scraper{Style.RESET_ALL}")
    print(f"{Fore.CYAN}Starting dosage information extraction process...{Style.RESET_ALL}")
    
    # Configuration - automatically find the latest side effects output file
    analysis_folder = "/Users/juanlu/Documents/Wye/scrapper/Analysis/"
    
    # Find the most recent side effects file
    side_effects_files = glob.glob(os.path.join(analysis_folder, "*WITH_SIDE_EFFECTS*.xlsx"))
    if not side_effects_files:
        print(f"{Fore.RED}❌ No side effects files found in {analysis_folder}")
        print(f"{Fore.YELLOW}Please run medication_scraper_side_effects.py first{Style.RESET_ALL}")
        return
    
    # Get the most recent file
    excel_file_path = max(side_effects_files, key=os.path.getctime)
    print(f"{Fore.CYAN}📁 Using input file: {os.path.basename(excel_file_path)}{Style.RESET_ALL}")
    
    output_file_path = None  # Will be auto-generated
    headless = False  # Set to True for headless browsing
    
    scraper = None
    try:
        # Initialize scraper
        scraper = WebMDDosageScraper(headless=headless)
        
        # Process all medications
        results_df = scraper.process_all_medications(excel_file_path, output_file_path)
        
        print(f"\n{Fore.GREEN}{Style.BRIGHT}✅ Successfully completed dosage extraction!")
        print(f"{Fore.CYAN}📊 Processed {len(results_df)} medications")
        print(f"{Fore.CYAN}💾 Results saved to Excel file")
        print(f"{Fore.CYAN}🧹 Temporary files cleaned up{Style.RESET_ALL}")
        
    except KeyboardInterrupt:
        print(f"\n{Fore.YELLOW}⚠️ Process interrupted by user{Style.RESET_ALL}")
    except Exception as e:
        print(f"\n{Fore.RED}❌ Error: {e}{Style.RESET_ALL}")
    finally:
        if scraper:
            scraper.cleanup()
        print(f"\n{Fore.BLUE}🏁 Scraper cleanup completed{Style.RESET_ALL}")

if __name__ == "__main__":
    main()
