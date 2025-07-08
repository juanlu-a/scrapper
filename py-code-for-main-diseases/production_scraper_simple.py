from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import time
import random
from openpyxl import load_workbook
import os

class DrugsScraper:
    def __init__(self, headless=True):
        self.driver = self.setup_driver(headless)
        self.wait = WebDriverWait(self.driver, 15)
        
    def setup_driver(self, headless=True):
        """Set up Chrome driver with minimal configuration"""
        chrome_options = Options()
        if headless:
            chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--window-size=1920,1080")
        chrome_options.add_argument("--user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
        
        # Use system ChromeDriver
        try:
            driver = webdriver.Chrome(service=Service("/opt/homebrew/bin/chromedriver"), options=chrome_options)
        except:
            # Fallback to default
            driver = webdriver.Chrome(options=chrome_options)
        
        return driver
    
    def scrape_side_effects(self, medication, max_retries=2):
        """Scrape side effects for a specific medication"""
        print(f"\n🔍 Processing: {medication}")
        
        for attempt in range(max_retries):
            try:
                if attempt > 0:
                    print(f"  🔄 Retry attempt {attempt + 1}/{max_retries}")
                
                # Step 1: Navigate to drugs.com
                self.driver.get("https://www.drugs.com")
                time.sleep(3)
                
                # Step 2: Find search box and search
                search_box = self.wait.until(
                    EC.presence_of_element_located((By.NAME, "searchterm"))
                )
                search_box.clear()
                search_box.send_keys(medication)
                search_box.send_keys(Keys.RETURN)
                
                print(f"  ✅ Searched for: {medication}")
                time.sleep(4)
                
                # Step 3: Find and click first result
                first_result = self.find_first_result(medication)
                if not first_result:
                    print(f"  ❌ Could not find result for {medication}")
                    continue
                
                # Click on first result
                try:
                    first_result.click()
                    print(f"  ✅ Clicked first result for {medication}")
                except Exception as e:
                    self.driver.execute_script("arguments[0].click();", first_result)
                    print(f"  ✅ Clicked first result (JS) for {medication}")
                
                time.sleep(4)
                
                # Step 4: Look for side effects content
                content = self.extract_side_effects_content(medication)
                
                if content and len(content) > 50:
                    print(f"  ✅ Successfully processed {medication} ({len(content)} characters)")
                    return content
                else:
                    print(f"  ❌ No substantial side effects content found for {medication}")
                    if attempt < max_retries - 1:
                        continue
                    else:
                        return f"No detailed side effects information found for {medication}"
                
            except Exception as e:
                error_msg = f"❌ Error processing {medication} (attempt {attempt + 1}): {str(e)}"
                print(error_msg)
                if attempt < max_retries - 1:
                    time.sleep(3)
                else:
                    return f"Error: {str(e)}"
        
        return f"Failed to process {medication} after {max_retries} attempts"
    
    def find_first_result(self, medication):
        """Find the first relevant result"""
        try:
            # Look for links containing the medication name
            results = self.driver.find_elements(By.CSS_SELECTOR, "a[href*='.html']")
            
            for result in results:
                href = result.get_attribute('href')
                text = result.text.strip().lower()
                
                # Simple check - if medication name appears in link text
                if medication.lower() in text and '.html' in href:
                    print(f"  ✅ Found result: {text[:50]}...")
                    return result
            
            # If no exact match, try first result
            if results:
                print(f"  ⚠️  Using first available result")
                return results[0]
            
        except Exception as e:
            print(f"  ❌ Error finding result: {e}")
        
        return None
    
    def extract_side_effects_content(self, medication):
        """Extract side effects content from the current page"""
        print(f"  📄 Extracting side effects content for {medication}")
        
        try:
            time.sleep(3)
            
            # Try to find side effects content
            content_parts = []
            
            # Approach 1: Look for side effects in page text
            page_text = self.driver.find_element(By.TAG_NAME, "body").text.lower()
            
            if 'side effects' in page_text or 'side effect' in page_text:
                # Try to find specific side effects sections
                try:
                    # Look for paragraphs or divs containing side effects
                    elements = self.driver.find_elements(By.XPATH, "//*[contains(text(), 'side effects') or contains(text(), 'Side Effects')]")
                    
                    for element in elements:
                        # Get parent element to capture more context
                        parent = element.find_element(By.XPATH, "./..")
                        text = parent.text.strip()
                        
                        if text and len(text) > 50:
                            content_parts.append(text)
                            break
                
                except Exception as e:
                    pass
                
                # If no specific sections found, look for general content
                if not content_parts:
                    try:
                        # Look for paragraphs with side effects keywords
                        paragraphs = self.driver.find_elements(By.TAG_NAME, "p")
                        keywords = ['side effect', 'adverse', 'reaction', 'emergency', 'call your doctor']
                        
                        for p in paragraphs:
                            text = p.text.strip()
                            if text and len(text) > 30:
                                text_lower = text.lower()
                                if any(keyword in text_lower for keyword in keywords):
                                    content_parts.append(text)
                                    if len(content_parts) >= 3:  # Limit to avoid too much content
                                        break
                    
                    except Exception as e:
                        pass
            
            if content_parts:
                content = '\n\n'.join(content_parts)
                content = self.clean_content(content)
                return content
            else:
                return f"No specific side effects content found for {medication}"
                
        except Exception as e:
            return f"Error extracting side effects content: {str(e)}"
    
    def clean_content(self, content):
        """Clean and format the side effects content"""
        if not content:
            return ""
        
        # Remove excessive whitespace
        lines = content.split('\n')
        cleaned_lines = []
        
        for line in lines:
            line = line.strip()
            if line and not any(skip in line.lower() for skip in [
                'advertisement', 'ads by', 'sponsored', 'cookie', 'privacy',
                'terms of use', 'about us', 'contact us', 'site map'
            ]):
                cleaned_lines.append(line)
        
        content = '\n'.join(cleaned_lines)
        
        # Remove multiple consecutive newlines
        while '\n\n\n' in content:
            content = content.replace('\n\n\n', '\n\n')
        
        return content.strip()
    
    def add_delay(self):
        """Add random delay between requests"""
        delay = random.uniform(3, 6)
        print(f"  ⏰ Waiting {delay:.1f} seconds before next request...")
        time.sleep(delay)
    
    def close(self):
        """Close the browser"""
        self.driver.quit()

def update_excel_with_side_effects(max_medications=None):
    """Update Excel file with side effects for all medications"""
    
    excel_path = '/Users/juanlu/Documents/Wye/scrapper/Analysis/main_diseases_analysis_final.xlsx'
    
    if not os.path.exists(excel_path):
        print(f"❌ Excel file not found: {excel_path}")
        return
    
    # Load workbook
    wb = load_workbook(excel_path)
    
    if "All Unique Medications" not in wb.sheetnames:
        print("❌ 'All Unique Medications' sheet not found")
        return
    
    ws = wb["All Unique Medications"]
    
    # Initialize scraper
    scraper = DrugsScraper(headless=False)  # Use visible browser for debugging
    
    try:
        # Get all medications (starting from row 9 where actual data begins)
        medications = []
        for row in range(9, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and isinstance(cell_value, str) and cell_value.strip():
                medications.append((row, cell_value.strip()))
        
        if max_medications:
            medications = medications[:max_medications]
        
        print(f"📊 Found {len(medications)} medications to process")
        
        # Process each medication
        for i, (row_num, medication) in enumerate(medications, 1):
            print(f"\n{'='*60}")
            print(f"Processing {i}/{len(medications)}: {medication}")
            print(f"{'='*60}")
            
            # Check if already processed (skip if more than 100 chars)
            try:
                current_value = ws.cell(row=row_num, column=2).value
                if current_value and isinstance(current_value, str) and len(current_value) > 100:
                    # But skip if it's an error message
                    if not current_value.startswith("❌") and not current_value.startswith("Error"):
                        print(f"  ⏭️  Already processed: {medication}")
                        continue
            except Exception as e:
                print(f"  ⚠️  Error checking existing value: {e}")
            
            # Scrape side effects
            side_effects = scraper.scrape_side_effects(medication)
            
            # Update Excel
            try:
                ws.cell(row=row_num, column=2, value=side_effects)
                print(f"  ✅ Updated Excel for {medication}")
            except Exception as e:
                print(f"  ❌ Error updating Excel for {medication}: {e}")
            
            # Save progress every 3 medications
            if i % 3 == 0:
                try:
                    wb.save(excel_path)
                    print(f"  💾 Progress saved after {i} medications")
                except Exception as e:
                    print(f"  ❌ Error saving progress: {e}")
            
            # Add delay between requests
            if i < len(medications):
                scraper.add_delay()
        
        # Final save
        try:
            wb.save(excel_path)
            print(f"\n✅ Successfully processed {len(medications)} medications!")
            print(f"📄 Updated Excel file: {excel_path}")
        except Exception as e:
            print(f"❌ Error saving final file: {e}")
        
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        try:
            wb.save(excel_path)
        except:
            pass
    
    finally:
        scraper.close()

if __name__ == "__main__":
    print("🚀 Starting Drugs.com Side Effects Scraper - Simple Chrome Version")
    print("="*60)
    print("🔍 Using visible Chrome browser with 3-6 second delays")
    print("="*60)
    
    # Process ALL medications
    update_excel_with_side_effects()
    
    print("\n" + "="*60)
    print("🎉 FULL SCRAPING COMPLETED!")
    print("="*60)
