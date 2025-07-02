import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def create_main_diseases_analysis_v3():
    """
    Create an Excel file with comprehensive analysis of main diseases from final_diseases_complete.csv
    Each disease gets its own sheet with structured data and enhanced medication information
    """
    
    # Target diseases to extract (use exact names from CSV)
    target_diseases = [
        'Heart disease',
        'Chronic kidney disease',
        'COPD',
        'Pneumonia',
        'Stroke',
        'Dementia',
        'Depression (major depressive disorder)',
        'High cholesterol',
        'Obesity', 
        'Arthritis'
    ]
    
    # Read the CSV file
    csv_path = '/Users/juanlu/Documents/Wye/scrapper/CSV/final_diseases_complete.csv'
    df = pd.read_csv(csv_path)
    
    # Read the drug data
    drug_data_path = '/Users/juanlu/Documents/Wye/scrapper/Analysis/drug_data_analysis.xlsx'
    drug_df = pd.read_excel(drug_data_path, sheet_name='All Drugs')
    
    # Create workbook
    wb = Workbook()
    
    # Create summary sheet first
    summary_ws = wb.active
    summary_ws.title = "Summary"
    create_summary_sheet(summary_ws, target_diseases)
    
    # Process each target disease
    processed_diseases = set()
    created_sheets = []
    
    for disease in target_diseases:
        # Find matching rows with more specific matching
        if disease == 'Heart disease':
            disease_data = df[df['Disease_Name_English'].str.contains('^Heart disease$', case=False, na=False, regex=True)]
        elif disease == 'Obesity':
            disease_data = df[df['Disease_Name_English'].str.contains('^Obesity$', case=False, na=False, regex=True)]
        elif disease == 'Stroke':
            disease_data = df[df['Disease_Name_English'].str.contains('^Stroke$', case=False, na=False, regex=True)]
        else:
            disease_data = df[df['Disease_Name_English'].str.contains(f'^{disease}$', case=False, na=False, regex=True)]
        
        if disease_data.empty:
            print(f"No exact match found for {disease}, trying partial match...")
            disease_data = df[df['Disease_Name_English'].str.contains(disease, case=False, na=False, regex=False)]
            
        if disease_data.empty:
            print(f"No data found for {disease}")
            continue
            
        # Get the first match
        disease_row = disease_data.iloc[0]
        disease_name = disease_row['Disease_Name_English']
        
        # Skip if we've already processed this disease
        if disease_name in processed_diseases:
            continue
        processed_diseases.add(disease_name)
        
        # Create sheet name (remove special characters)
        sheet_name = disease_name.replace('(', '').replace(')', '').replace('/', '-')[:31]
        ws = wb.create_sheet(title=sheet_name)
        
        # Set up the sheet structure with enhanced medication data
        setup_disease_sheet_v3(ws, disease_row, disease_name, drug_df)
        created_sheets.append((disease, disease_name))
        print(f"✓ Created sheet for: {disease_name}")
    
    # Update summary sheet with actual results
    update_summary_sheet(summary_ws, created_sheets)
    
    # Save the workbook
    output_path = '/Users/juanlu/Documents/Wye/scrapper/Analysis/main_diseases_analysis_final.xlsx'
    wb.save(output_path)
    print(f"\nAnalysis saved to: {output_path}")
    
    return output_path

def create_summary_sheet(ws, target_diseases):
    """Create a summary sheet with overview information"""
    
    # Header styling
    header_font = Font(bold=True, size=16, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    subheader_font = Font(bold=True, size=12, color="FFFFFF")
    subheader_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    
    # Title
    ws['A1'] = 'MAIN DISEASES COMPREHENSIVE ANALYSIS'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Date and info
    ws['A3'] = 'Analysis Date:'
    ws['B3'] = '2025-07-02'
    ws['A4'] = 'Source Data:'
    ws['B4'] = 'final_diseases_complete.csv + drug_data_analysis.xlsx'
    ws['A5'] = 'Total Target Diseases:'
    ws['B5'] = len(target_diseases)
    
    # Target diseases list
    ws['A7'] = 'TARGET DISEASES'
    ws['A7'].font = subheader_font
    ws['A7'].fill = subheader_fill
    ws.merge_cells('A7:D7')
    ws['A7'].alignment = Alignment(horizontal='center')
    
    ws['A8'] = 'Disease Name'
    ws['B8'] = 'Status'
    ws['C8'] = 'Matched Name'
    ws['D8'] = 'Spanish Name'
    
    # Style header row
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}8'].font = Font(bold=True)
        ws[f'{col}8'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Add target diseases (will be updated later)
    for i, disease in enumerate(target_diseases, 9):
        ws[f'A{i}'] = disease
        ws[f'B{i}'] = 'Processing...'
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 25

def update_summary_sheet(ws, created_sheets):
    """Update the summary sheet with actual results"""
    
    # Create a mapping of created sheets
    created_mapping = {original: matched for original, matched in created_sheets}
    
    # Read CSV to get Spanish names
    csv_path = '/Users/juanlu/Documents/Wye/scrapper/CSV/final_diseases_complete.csv'
    df = pd.read_csv(csv_path)
    
    # Update the status for each disease
    row = 9
    target_diseases = [
        'Heart disease', 'Chronic kidney disease', 'COPD', 'Pneumonia', 'Stroke',
        'Dementia', 'Depression (major depressive disorder)', 'High cholesterol',
        'Obesity', 'Arthritis'
    ]
    
    for disease in target_diseases:
        if disease in created_mapping:
            matched_name = created_mapping[disease]
            ws[f'B{row}'] = '✓ Found'
            ws[f'C{row}'] = matched_name
            
            # Get Spanish name
            spanish_name = df[df['Disease_Name_English'] == matched_name]['Disease_Name_Spanish'].iloc[0]
            ws[f'D{row}'] = spanish_name
            
            # Color the row green
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        else:
            ws[f'B{row}'] = '✗ Not Found'
            ws[f'C{row}'] = 'No match'
            ws[f'D{row}'] = '-'
            
            # Color the row red
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        
        row += 1
    
    # Add statistics
    ws[f'A{row+2}'] = 'STATISTICS'
    ws[f'A{row+2}'].font = Font(bold=True, size=12)
    ws[f'A{row+3}'] = f'Diseases Found: {len(created_sheets)}'
    ws[f'A{row+4}'] = f'Diseases Not Found: {len(target_diseases) - len(created_sheets)}'
    ws[f'A{row+5}'] = f'Success Rate: {len(created_sheets)/len(target_diseases)*100:.1f}%'

def setup_disease_sheet_v3(ws, disease_row, disease_name, drug_df):
    """Set up each disease sheet with structured data and enhanced medication information"""
    
    # Header styling
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_font = Font(bold=True, size=12, color="FFFFFF")
    subheader_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    
    # Border styling
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Row 1: Disease Title
    ws['A1'] = f'{disease_name.upper()} - COMPREHENSIVE ANALYSIS'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    # Row 3: Disease Names
    ws['A3'] = 'DISEASE INFORMATION'
    ws['A3'].font = subheader_font
    ws['A3'].fill = subheader_fill
    ws.merge_cells('A3:F3')
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Disease name details
    ws['A4'] = 'English Name:'
    ws['B4'] = disease_row['Disease_Name_English']
    ws['A5'] = 'Spanish Name:'
    ws['B5'] = disease_row['Disease_Name_Spanish']
    
    # Style the info cells
    for row in [4, 5]:
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Row 7: Diagnosis Section
    ws['A7'] = 'DIAGNOSIS'
    ws['A7'].font = subheader_font
    ws['A7'].fill = subheader_fill
    ws.merge_cells('A7:F7')
    ws['A7'].alignment = Alignment(horizontal='center')
    
    # Diagnosis information
    diagnosis_text = disease_row['Diagnosis'] if pd.notna(disease_row['Diagnosis']) else 'No diagnosis information available'
    ws['A8'] = 'Diagnosis Process:'
    ws['A8'].font = Font(bold=True)
    
    # Put all diagnosis text in one cell (B8) - user will manually merge cells for readability
    ws['B8'] = diagnosis_text
    ws['B8'].alignment = Alignment(wrap_text=True, vertical='top')
    # Set a taller row height for the diagnosis text
    ws.row_dimensions[8].height = max(60, min(200, len(diagnosis_text) // 10))
    
    # Calculate next row (just one row for diagnosis now)
    next_row = 10
    
    # Treatments Section
    ws[f'A{next_row}'] = 'TREATMENTS'
    ws[f'A{next_row}'].font = subheader_font
    ws[f'A{next_row}'].fill = subheader_fill
    ws.merge_cells(f'A{next_row}:F{next_row}')
    ws[f'A{next_row}'].alignment = Alignment(horizontal='center')
    
    treatments = disease_row['Treatments'] if pd.notna(disease_row['Treatments']) else 'No treatment information available'
    
    ws[f'A{next_row+1}'] = 'Available Treatments:'
    ws[f'A{next_row+1}'].font = Font(bold=True)
    # Put all treatment text in one cell - user will manually merge cells for readability
    ws[f'B{next_row+1}'] = treatments
    ws[f'B{next_row+1}'].alignment = Alignment(wrap_text=True, vertical='top')
    # Set row height based on content length
    ws.row_dimensions[next_row+1].height = max(60, min(200, len(treatments) // 10))
    
    # Calculate next row (just one row for treatments now)
    next_row = next_row + 3
    
    # Tests Section  
    ws[f'A{next_row}'] = 'DIAGNOSTIC TESTS'
    ws[f'A{next_row}'].font = subheader_font
    ws[f'A{next_row}'].fill = subheader_fill
    ws.merge_cells(f'A{next_row}:F{next_row}')
    ws[f'A{next_row}'].alignment = Alignment(horizontal='center')
    
    tests = disease_row['Tests'] if pd.notna(disease_row['Tests']) else 'No test information available'
    
    ws[f'A{next_row+1}'] = 'Diagnostic Tests:'
    ws[f'A{next_row+1}'].font = Font(bold=True)
    # Put all test text in one cell - user will manually merge cells for readability
    ws[f'B{next_row+1}'] = tests
    ws[f'B{next_row+1}'].alignment = Alignment(wrap_text=True, vertical='top')
    # Set row height based on content length
    ws.row_dimensions[next_row+1].height = max(60, min(200, len(tests) // 10))
    
    # Calculate next row (just one row for tests now)
    next_row = next_row + 3
    
    # Enhanced Medications Section
    ws[f'A{next_row}'] = 'MEDICATIONS & DRUGS - DETAILED INFORMATION'
    ws[f'A{next_row}'].font = subheader_font
    ws[f'A{next_row}'].fill = subheader_fill
    ws.merge_cells(f'A{next_row}:F{next_row}')
    ws[f'A{next_row}'].alignment = Alignment(horizontal='center')
    
    medications = disease_row['Medications_Drugs'] if pd.notna(disease_row['Medications_Drugs']) else 'No medication information available'
    medication_list = medications.split(';') if pd.notna(disease_row['Medications_Drugs']) else ['No medications listed']
    
    # Create enhanced medication table
    med_row = next_row + 2
    ws[f'A{med_row}'] = 'Medication Name'
    ws[f'B{med_row}'] = 'What Is'
    ws[f'C{med_row}'] = 'Side Effects'
    ws[f'D{med_row}'] = 'Disease Tag'
    
    # Style header row
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{med_row}'].font = Font(bold=True)
        ws[f'{col}{med_row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        ws[f'{col}{med_row}'].border = thin_border
    
    # Add ALL medications with detailed information (no limit)
    for i, medication in enumerate(medication_list):
        med_name = medication.strip()
        
        # Look up medication in drug database
        drug_info = lookup_drug_info(med_name, drug_df)
        
        ws[f'A{med_row+1+i}'] = med_name
        ws[f'B{med_row+1+i}'] = drug_info['what_is']
        ws[f'C{med_row+1+i}'] = drug_info['side_effects']
        ws[f'D{med_row+1+i}'] = disease_name
        
        # Add borders and text wrapping
        for col in ['A', 'B', 'C', 'D']:
            cell = ws[f'{col}{med_row+1+i}']
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Add summary note
    ws[f'A{med_row+1+len(medication_list)+1}'] = f"Total medications for {disease_name}: {len(medication_list)}"
    ws[f'A{med_row+1+len(medication_list)+1}'].font = Font(bold=True, italic=True)
    ws.merge_cells(f'A{med_row+1+len(medication_list)+1}:D{med_row+1+len(medication_list)+1}')
    
    # Set column widths for better display of ALL medications
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50  # Increased for longer descriptions
    ws.column_dimensions['C'].width = 45  # Increased for longer side effects
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

def lookup_drug_info(med_name, drug_df):
    """
    Look up detailed information for a medication from the drug database
    """
    # Clean the medication name for better matching
    clean_med_name = med_name.strip().lower()
    
    # Try exact match first
    exact_match = drug_df[drug_df['Drug Name'].str.lower() == clean_med_name]
    
    if not exact_match.empty:
        row = exact_match.iloc[0]
        return {
            'what_is': truncate_text(row['What Is'], 300),  # Increased length
            'side_effects': truncate_text(row['Side Effects'], 250)  # Increased length
        }
    
    # Try partial match
    partial_match = drug_df[drug_df['Drug Name'].str.lower().str.contains(clean_med_name, na=False)]
    
    if not partial_match.empty:
        row = partial_match.iloc[0]
        return {
            'what_is': truncate_text(row['What Is'], 300),  # Increased length
            'side_effects': truncate_text(row['Side Effects'], 250)  # Increased length
        }
    
    # If no match found, return default
    return {
        'what_is': 'Information not available in database',
        'side_effects': 'Side effects information not available'
    }

def truncate_text(text, max_length):
    """
    Truncate text to a maximum length and add ellipsis if needed
    Keep more useful information by being smarter about truncation
    """
    if pd.isna(text):
        return 'Information not available'
    
    text = str(text).strip()
    if len(text) <= max_length:
        return text
    
    # Try to truncate at sentence boundaries for better readability
    sentences = text.split('. ')
    result = ""
    
    for sentence in sentences:
        if len(result + sentence + '. ') <= max_length - 3:
            result += sentence + '. '
        else:
            break
    
    if result:
        return result.strip() + '...'
    else:
        # If even the first sentence is too long, just truncate
        return text[:max_length-3] + '...'

def split_text_into_chunks(text, chunk_size):
    """Split long text into smaller chunks for better Excel display"""
    if len(text) <= chunk_size:
        return [text]
    
    chunks = []
    words = text.split()
    current_chunk = ""
    
    for word in words:
        if len(current_chunk + " " + word) <= chunk_size:
            current_chunk += " " + word if current_chunk else word
        else:
            if current_chunk:
                chunks.append(current_chunk)
            current_chunk = word
    
    if current_chunk:
        chunks.append(current_chunk)
    
    return chunks

if __name__ == "__main__":
    print("Creating Complete Main Diseases Analysis with ALL Medications...")
    output_file = create_main_diseases_analysis_v3()
    print(f"Complete analysis with ALL medications finished! File saved at: {output_file}")
